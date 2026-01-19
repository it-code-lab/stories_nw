"""
contentplanner_worker.py

Purpose
-------
Utilities to pull records from ContentPlanner DB (via website APIs) and populate:
1) media_jobs.xlsx (Image generation jobs)
2) heygen_submit_videos.xlsx (HeyGen submission jobs) - clear from row 10 onward
3) master_shorts_uploader_data.xlsx (Uploader master sheet) - clear from row 80 onward

Designed to be:
- Callable from Flask endpoints (no input() required)
- Runnable as a CLI tool

Before using
------------
1) Set BASE_URL and TOKEN below.
2) Ensure these APIs exist on your contentplanner website:
   - /api/worker_get_topic_id.php?youtube_channel_name=...
   - /api/worker_fetch_image_jobs.php?topic_id=...
   - /api/worker_fetch_heygen_submit_jobs.php?topic_id=...
   - /api/worker_fetch_upload_jobs.php?topic_id=...
"""

import argparse
import sys
from pathlib import Path
from typing import Dict, Any, List, Optional

import requests
from openpyxl import Workbook, load_workbook
import re

from assemble_from_videos import _safe_out_name   

# =============================
# Config
# =============================
BASE_URL = "https://contentplanner.readernook.com"  # <-- change if needed
TOKEN = "CHANGE_ME_TO_A_LONG_RANDOM_SECRET"         # <-- set your worker token

HEADERS = {"X-Worker-Token": TOKEN}

# Image jobs excel
EXCEL_FILE = Path("media_jobs.xlsx")
SHEET_NAME = "Jobs"

# HeyGen submit excel
HEYGEN_FILE = "heygen_submit_videos.xlsx"
HEYGEN_SHEET = "Sheet1"
HEYGEN_START_ROW = 10  # clear & paste from row 10 onward
HEYGEN_REQUIRED_COLS = [
    "HeyGen_Template_url",
    "story_text",
    "video_name",
    "status",
    "message",
    "submitted_at",
]

# Uploader master excel
UPLOADER_MASTER_FILE = "master_shorts_uploader_data.xlsx"
SECTION_ORDER_FILE = r"edit_vid_input/order.xlsx"
UPLOADER_START_ROW = 80  # clear & paste from row 80 onward
UPLOADER_REQUIRED_COLS = [
    "media_file", "yt_title", "yt_description", "youtube_status", "youTubeChannel",
    "media_type", "future", "yt_playlist", "yt_schedule_date", "yt_tags",
    "section_id", "avatar_img",  # useful for DB linkage
]

SECTION_ORDER_REQUIRED_COLS = ["filename","title","section_title"]
SECTION_ORDER_START_ROW = 2  # clear & paste from row 2 onward

HEYGEN_BULK_BG_FILE = "heygen_bulk_bg.xlsx"
# =============================
# Low-level HTTP helpers
# =============================
def _get_json(path: str, params: dict, timeout: int = 60) -> dict:
    url = f"{BASE_URL}{path}"
    r = requests.get(url, params=params, headers=HEADERS, timeout=timeout)
    r.raise_for_status()
    return r.json()


def resolve_topic_id_by_channel(youtube_channel_name: str) -> int:
    """
    Resolves topic_id using the channel name from topic.youtube_channel_name.
    Requires /api/worker_get_topic_id.php on your website.
    """
    name = (youtube_channel_name or "").strip()
    if not name:
        raise ValueError("youtube_channel_name is required")

    data = _get_json(
        "/api/worker_get_topic_id.php",
        {"youtube_channel_name": name},
        timeout=30,
    )
    if not data.get("ok") or not data.get("topic_id"):
        raise RuntimeError(f"Channel not found in contentplanner DB: {name}")
    return int(data["topic_id"])


# =============================
# Fetchers (DB → JSON)
# =============================
def fetch_image_jobs(topic_id: int, limit: int = 5000) -> List[dict]:
    # expects /api/worker_fetch_image_jobs.php
    data = _get_json("/api/worker_fetch_image_jobs.php", {"topic_id": topic_id, "limit": limit})
    return data.get("items", [])


def fetch_heygen_submit_jobs(topic_id: int, limit: int = 5000) -> List[dict]:
    # expects /api/worker_fetch_heygen_submit_jobs.php
    data = _get_json("/api/worker_fetch_heygen_submit_jobs.php", {"topic_id": topic_id, "limit": limit})
    return data.get("items", [])


def fetch_scheduled_rows(topic_id: int, only_due_now: bool = False, limit: int = 5000) -> List[dict]:
    # expects /api/worker_fetch_upload_jobs.php
    data = _get_json(
        "/api/worker_fetch_upload_jobs.php",
        {"topic_id": topic_id, "only_due_now": 1 if only_due_now else 0, "limit": limit},
    )
    return data.get("items", [])

def fetch_ordered_section_title_rows(topic_id: int, only_due_now: bool = False, limit: int = 5000) -> List[dict]:
    # expects /api/worker_fetch_ordered_section_title_rows.php
    data = _get_json(
        "/api/worker_fetch_ordered_section_title_rows.php",
        {"topic_id": topic_id, "only_due_now": 1 if only_due_now else 0, "limit": limit},
    )
    return data.get("items", [])
# =============================
# Excel helpers
# =============================
def ensure_headers(ws, required_cols: List[str]) -> Dict[str, int]:
    """
    Ensures required columns exist in row 1. Returns mapping col_name -> col_index (1-based).
    """
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [h if h is not None else "" for h in headers]
    header_map = {h: i + 1 for i, h in enumerate(headers) if h}

    for col in required_cols:
        if col not in header_map:
            ws.cell(row=1, column=len(headers) + 1, value=col)
            headers.append(col)
            header_map[col] = len(headers)

    return header_map


def clear_rows_from(ws, start_row: int, max_col: int):
    """
    Clears cell values from start_row to end (keeps formatting).
    """
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def format_schedule_date(dt_str: str) -> str:
    """
    Uploader expects a date-like string. We output YYYY-MM-DD.
    dt_str from DB is typically "YYYY-MM-DD HH:MM:SS".
    """
    if not dt_str:
        return ""
    return dt_str[:10]


# =============================
# 1) Populate Image Jobs Excel
# =============================
def write_image_jobs_excel(items: List[dict], image_provider: str = "", image_orientation: str = "") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Matches your image creation excel expectations + adds section_id + image_name
    headers = [
        "prompt", "account_id", "image_path", "video_cmd", "video_path", "status",
        "account_id_1", "account_id_2", "section_id", "image_name","image_orientation","image_provider"
    ]
    ws.append(headers)

    for it in items:
        ws.append([
            it.get("image_prompt", ""),
            "",  # account_id left blank (your script distributes across profiles)
            "",  # image_path blank so script processes it
            "", "", "", "", "",
            it.get("section_id", ""),
            it.get("image_name", ""),
            image_orientation,
            image_provider,
        ])

    wb.save(EXCEL_FILE)


def populate_image_jobs_excel_for_channel(youtube_channel_name: str, image_provider: str = "", image_orientation: str = "") -> Dict[str, Any]:
    topic_id = resolve_topic_id_by_channel(youtube_channel_name)
    items = fetch_image_jobs(topic_id=topic_id, limit=5000)

    if not items:
        return {"ok": True, "count": 0, "message": "No image jobs found (image_status=generating)."}

    write_image_jobs_excel(items, image_provider=image_provider, image_orientation=image_orientation)
    return {"ok": True, "count": len(items), "message": f"Wrote {len(items)} image job(s) to {EXCEL_FILE.resolve()}."}


# =============================
# 2) Populate HeyGen Submit Excel (row 10+)
# =============================
def populate_heygen_submit_excel_for_channel(youtube_channel_name: str) -> Dict[str, Any]:
    topic_id = resolve_topic_id_by_channel(youtube_channel_name)
    items = fetch_heygen_submit_jobs(topic_id=topic_id, limit=5000)

    if not items:
        return {"ok": True, "count": 0, "message": "No eligible rows found for HeyGen submit."}

    wb = load_workbook(HEYGEN_FILE)
    ws = wb[HEYGEN_SHEET] if HEYGEN_SHEET in wb.sheetnames else wb.active

    colmap = ensure_headers(ws, HEYGEN_REQUIRED_COLS)
    max_col = max(colmap.values())

    # Clear rows from 10 onward
    clear_rows_from(ws, start_row=HEYGEN_START_ROW, max_col=max_col)

    row = HEYGEN_START_ROW
    missing_template = 0

    for it in items:
        template_url = (it.get("template_url") or "").strip()
        text = it.get("section_text") or ""
        name = it.get("image_name") or ""

        if not template_url:
            missing_template += 1

        ws.cell(row=row, column=colmap["HeyGen_Template_url"], value=template_url)
        ws.cell(row=row, column=colmap["story_text"], value=text)
        # Concatenate video_name with ".mp4" if not already present
        
        # if name and not name.lower().endswith(".mp4"):
        #     name += ".mp4"
        ws.cell(row=row, column=colmap["video_name"], value=name) 

        # Reset so heygen_submit_videos.py processes it
        ws.cell(row=row, column=colmap["status"], value="")
        ws.cell(row=row, column=colmap["message"], value="")
        ws.cell(row=row, column=colmap["submitted_at"], value="")

        row += 1

    wb.save(HEYGEN_FILE)

    msg = f"Wrote {len(items)} HeyGen row(s) into {HEYGEN_FILE} starting at row {HEYGEN_START_ROW} (cleared old rows from {HEYGEN_START_ROW}+)."
    if missing_template:
        msg += f" WARNING: {missing_template} row(s) missing template_url (set story default avatar)."

    return {"ok": True, "count": len(items), "message": msg}


# =============================
# 3) Populate Upload Excel (row 80+)
# =============================
def ensure_uploader_columns(ws) -> Dict[str, int]:
    headers = [c.value for c in ws[1]]
    headers = [h if h is not None else "" for h in headers]
    header_map = {h: i + 1 for i, h in enumerate(headers) if h}

    for col in UPLOADER_REQUIRED_COLS:
        if col not in header_map:
            ws.cell(row=1, column=len(headers) + 1, value=col)
            headers.append(col)
            header_map[col] = len(headers)

    return header_map

def ensure_section_order_file_columns(ws) -> Dict[str, int]:
    headers = [c.value for c in ws[1]]
    headers = [h if h is not None else "" for h in headers]
    header_map = {h: i + 1 for i, h in enumerate(headers) if h}

    for col in SECTION_ORDER_REQUIRED_COLS:
        if col not in header_map:
            ws.cell(row=1, column=len(headers) + 1, value=col)
            headers.append(col)
            header_map[col] = len(headers)

    return header_map

def ensure_heygen_bulk_bg_file_columns(ws) -> Dict[str, int]:
    headers = [c.value for c in ws[1]]
    headers = [h if h is not None else "" for h in headers]
    header_map = {h: i + 1 for i, h in enumerate(headers) if h}

    for col in ["heygen_video","bg","status"]:
        if col not in header_map:
            ws.cell(row=1, column=len(headers) + 1, value=col)
            headers.append(col)
            header_map[col] = len(headers)

    return header_map

def clear_uploader_from_row(ws, header_map: Dict[str, int], start_row: int):
    max_row = ws.max_row
    max_col = max(header_map.values()) if header_map else ws.max_column
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None

# Used for shorts 
def populate_upload_excel_for_channel(youtube_channel_name: str, only_due_now: bool = False) -> Dict[str, Any]:
    topic_id = resolve_topic_id_by_channel(youtube_channel_name)
    items = fetch_scheduled_rows(topic_id=topic_id, only_due_now=only_due_now, limit=5000)

    if not items:
        return {"ok": True, "count": 0, "message": "No scheduled rows found for this channel."}

    wb = load_workbook(UPLOADER_MASTER_FILE)
    ws = wb.active

    header_map = ensure_uploader_columns(ws)

    # Clear from row 80 onward
    clear_uploader_from_row(ws, header_map, UPLOADER_START_ROW)

    row_idx = UPLOADER_START_ROW

    for it in items:
        section_id = it.get("section_id", "")
        image_name = it.get("image_name") or ""
        # Optional: if you have final_video_path in API, it will be filled automatically
        # media_file = (it.get("final_video_path") or "")
        # if not media_file:
        #     # If your final videos are named by image_name and stored in a known folder,
        #     # you can build the path here. Otherwise leave blank and fill later.
        #     media_file = ""

        media_file = "out/" + image_name + ".mp4"

        story_title = it.get("section_title") or ""
        story_desc = it.get("section_text") or ""
        section_title = it.get("section_title") or ""
        section_desc = it.get("section_text") or ""
        playlist = it.get("youtube_playlist_name") or ""
        scheduled_at = it.get("scheduled_at") or ""

        yt_title = section_title.strip() if section_title.strip() else story_title.strip()
        yt_description = section_desc.strip() if section_desc.strip() else story_desc.strip()

        youTubeChannel = it.get("youtube_channel_name") or youtube_channel_name

        ws.cell(row=row_idx, column=header_map["section_id"], value=section_id)
        ws.cell(row=row_idx, column=header_map["media_type"], value="video")
        ws.cell(row=row_idx, column=header_map["future"], value="")
        ws.cell(row=row_idx, column=header_map["youtube_status"], value="")
        ws.cell(row=row_idx, column=header_map["youTubeChannel"], value=youTubeChannel)

        ws.cell(row=row_idx, column=header_map["media_file"], value=media_file)
        ws.cell(row=row_idx, column=header_map["yt_title"], value=yt_title)
        ws.cell(row=row_idx, column=header_map["yt_description"], value=yt_description)
        ws.cell(row=row_idx, column=header_map["yt_playlist"], value=playlist)
        ws.cell(row=row_idx, column=header_map["yt_schedule_date"], value=format_schedule_date(scheduled_at))
        ws.cell(row=row_idx, column=header_map["yt_tags"], value="")

        row_idx += 1

    wb.save(UPLOADER_MASTER_FILE)
    return {
        "ok": True,
        "count": len(items),
        "message": f"Wrote {len(items)} upload row(s) into {UPLOADER_MASTER_FILE} starting at row {UPLOADER_START_ROW} (cleared old rows from {UPLOADER_START_ROW}+)."
    }


def _safe_out_name_duplicate(title: str, max_len: int = 120) -> str:
    title = (title or "").strip()
    if not title:
        title = "story"
    # Windows-safe filename
    title = re.sub(r'[\\/:*?"<>|]+', "", title)
    title = re.sub(r"\s+", " ", title).strip()
    # Keep unicode letters/digits/_/-/space
    title = re.sub(r"[^\w\s\-]+", "", title, flags=re.UNICODE)
    title = re.sub(r"\s+", "-", title).strip("-_")
    return (title[:max_len].rstrip("-_")) or "story"


def populate_upload_excel_long_for_channel(youtube_channel_name: str, only_due_now: bool = False) -> Dict[str, Any]:
    topic_id = resolve_topic_id_by_channel(youtube_channel_name)
    items = fetch_scheduled_rows(topic_id=topic_id, only_due_now=only_due_now, limit=5000)

    if not items:
        return {"ok": True, "count": 0, "message": "No scheduled rows found for this channel."}

    wb = load_workbook(UPLOADER_MASTER_FILE)
    ws = wb.active

    header_map = ensure_uploader_columns(ws)

    # Clear from row 80 onward
    clear_uploader_from_row(ws, header_map, UPLOADER_START_ROW)

    row_idx = UPLOADER_START_ROW
    seen = set()
    written = 0
    skipped_dupes = 0

    for it in items:
        # section_id = it.get("section_id", "")

        # Optional: if you have final_video_path in API, it will be filled automatically
        # media_file = (it.get("final_video_path") or "")
        # if not media_file:
        #     # If your final videos are named by image_name and stored in a known folder,
        #     # you can build the path here. Otherwise leave blank and fill later.
        #     media_file = ""

        story_title = it.get("story_title") or ""
        safe_name = _safe_out_name(story_title)
        dedupe_key = it.get("story_id") or it.get("story_key") or safe_name

        if not dedupe_key:
            # Nothing to dedupe on; skip (or choose a different fallback)
            continue

        if dedupe_key in seen:
            skipped_dupes += 1
            continue
        seen.add(dedupe_key)

        media_file = f"edit_vid_output/{safe_name}.mp4"

        story_desc = (it.get("story_description") or "").strip()
        playlist = it.get("youtube_playlist_name") or ""
        scheduled_at = it.get("scheduled_at") or ""
        youTubeChannel = it.get("youtube_channel_name") or youtube_channel_name
        avatar_img = it.get("avatar_img") or ""

        ws.cell(row=row_idx, column=header_map["media_type"], value="video")
        ws.cell(row=row_idx, column=header_map["future"], value="")
        ws.cell(row=row_idx, column=header_map["youtube_status"], value="")
        ws.cell(row=row_idx, column=header_map["youTubeChannel"], value=youTubeChannel)

        ws.cell(row=row_idx, column=header_map["media_file"], value=media_file)
        ws.cell(row=row_idx, column=header_map["yt_title"], value=story_title)
        ws.cell(row=row_idx, column=header_map["yt_description"], value=story_desc)
        ws.cell(row=row_idx, column=header_map["yt_playlist"], value=playlist)
        ws.cell(row=row_idx, column=header_map["avatar_img"], value=avatar_img)
        ws.cell(row=row_idx, column=header_map["yt_schedule_date"], value=format_schedule_date(scheduled_at))
        ws.cell(row=row_idx, column=header_map["yt_tags"], value="")

        row_idx += 1
        written += 1

    wb.save(UPLOADER_MASTER_FILE)
    return {
        "ok": True,
        "count": written,
        "message": (
            f"Wrote {written} unique upload row(s) into {UPLOADER_MASTER_FILE} starting at row {UPLOADER_START_ROW} "
            f"(skipped {skipped_dupes} duplicate item(s); cleared old rows from {UPLOADER_START_ROW}+)."
        )
    }

def populate_section_order_excel_from_db(youtube_channel_name: str, only_due_now: bool = False) -> Dict[str, Any]:
    topic_id = resolve_topic_id_by_channel(youtube_channel_name)
    items = fetch_ordered_section_title_rows(topic_id=topic_id, only_due_now=only_due_now, limit=5000)

    if not items:
        return {"ok": True, "count": 0, "message": "No rows found for this channel."}

    wb = load_workbook(SECTION_ORDER_FILE)
    ws = wb.active

    header_map = ensure_section_order_file_columns(ws)

    # Clear from row 80 onward
    clear_uploader_from_row(ws, header_map, SECTION_ORDER_START_ROW)

    row_idx = SECTION_ORDER_START_ROW

    for it in items:
        story_title = it.get("story_title") or ""

        # Append .mp4 if not present        
        filename = it.get("file_name") or ""
        if filename and not filename.lower().endswith(".mp4"):
            filename += ".mp4"

        section_title = it.get("section_title") or ""
        
        ws.cell(row=row_idx, column=header_map["filename"], value=filename)
        ws.cell(row=row_idx, column=header_map["title"], value=story_title)
        ws.cell(row=row_idx, column=header_map["section_title"], value=section_title)

        row_idx += 1

    wb.save(SECTION_ORDER_FILE)
    return {
        "ok": True,
        "count": len(items),
        "message": f"Wrote {len(items)} upload row(s) into {SECTION_ORDER_FILE} starting at row {SECTION_ORDER_START_ROW} (cleared old rows from {SECTION_ORDER_START_ROW}+)."
    }


def populate_heygen_bulk_bg_excel_from_db(youtube_channel_name: str, only_due_now: bool = False) -> Dict[str, Any]:
    topic_id = resolve_topic_id_by_channel(youtube_channel_name)
    items = fetch_ordered_section_title_rows(topic_id=topic_id, only_due_now=only_due_now, limit=5000)

    if not items:
        return {"ok": True, "count": 0, "message": "No rows found for this channel."}

    wb = load_workbook(HEYGEN_BULK_BG_FILE)
    ws = wb.active

    header_map = ensure_heygen_bulk_bg_file_columns(ws)

    # Clear from row 80 onward
    clear_uploader_from_row(ws, header_map, 2)

    row_idx = 2

    for it in items:
        # story_title = it.get("story_title") or ""

        # Append .mp4 if not present        
        filename = it.get("file_name") or ""
        if filename and not filename.lower().endswith(".mp4"):
            filename += ".mp4"

        # append "heygen_downloads/" prefix
        ws.cell(row=row_idx, column=header_map["heygen_video"], value="heygen_downloads/" + filename)


        ws.cell(row=row_idx, column=header_map["bg"], value="downloads/" +filename)

        row_idx += 1

    wb.save(HEYGEN_BULK_BG_FILE)
    return {
        "ok": True,
        "count": len(items),
        "message": f"Wrote {len(items)} upload row(s) into {HEYGEN_BULK_BG_FILE} starting at row 2 (cleared old rows from 2+)."
    }


# =============================
# CLI (optional)
# =============================
def _print_result(res: Dict[str, Any]) -> None:
    print(res.get("message", ""))
    if "count" in res:
        print(f"count={res['count']}")


def main_cli():
    p = argparse.ArgumentParser(description="ContentPlanner DB → Excel exporter")
    p.add_argument("--channel", required=True, help="YouTube channel name (topic.youtube_channel_name)")
    p.add_argument("--mode", required=True, choices=["images", "heygen", "uploads", "all"], help="Which export to run")
    p.add_argument("--due-now", action="store_true", help="For uploads: export only rows due now (scheduled_at <= now)")
    args = p.parse_args()

    channel = args.channel.strip()

    if args.mode in ("images", "all"):
        _print_result(populate_image_jobs_excel_for_channel(channel))

    if args.mode in ("heygen", "all"):
        _print_result(populate_heygen_submit_excel_for_channel(channel))

    if args.mode in ("uploads", "all"):
        _print_result(populate_upload_excel_for_channel(channel, only_due_now=args.due_now))


if __name__ == "__main__":
    # If you run this file directly, it will work as CLI:
    # python contentplanner_worker.py --channel "Cash Flow Hacks" --mode all
    try:
        main_cli()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
