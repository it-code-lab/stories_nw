# multi_profile_media_agent.py  (Excel-enabled, with account_id_1/account_id_2)
# ------------------------------------------------------------
# Two-pass workflow:
#   1) RUN_MODE="images" -> read prompts from Excel, generate images, write image_path (+account_id_1)
#   2) RUN_MODE="videos" -> read image_path + video_cmd, render, write video_path (+account_id_2)
#
# Sheet: Jobs
# Columns (case-insensitive): prompt, account_id, image_path, video_cmd, video_path, status,
#                             account_id_1, account_id_2
# ------------------------------------------------------------

import asyncio
import json
import os
import random
import re
import shutil
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import uuid
from playwright.async_api import async_playwright, BrowserContext, Page, expect


async def _wait_for_enter(message: str = "Press Enter to continue...") -> None:
    """Non-blocking 'input()' for async code."""
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, lambda: input(message))


async def bootstrap_profile_logins(pw, account: Dict[str, str]) -> None:
    """Open a persistent profile and let the user login once. Press Enter to move on."""
    browser_type = getattr(pw, BROWSER_NAME)
    print(f"[bootstrap:{account['id']}] Launching profile: {account['profile']}")
    ctx_kwargs = dict(
        user_data_dir=account["profile"],
        headless=HEADLESS,
        executable_path=CHROME_EXECUTABLE,
        device_scale_factor=DEVICE_SCALE_FACTOR,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--start-maximized",
            "--no-default-browser-check",
            "--no-first-run",
        ],
        accept_downloads=True,
    )
    ctx: BrowserContext = await browser_type.launch_persistent_context(**ctx_kwargs)
    page: Page = await ctx.new_page()
    await page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    for site_key in BOOTSTRAP_SITES:
        url = _get_site_url(account, site_key)
        if not url:
            continue
        print(f"[bootstrap:{account['id']}] Open {site_key}: {url}")
        try:
            await page.goto(url)
        except Exception as e:
            print(f"[bootstrap:{account['id']}] Failed to open {site_key}: {e}")
            continue
        await _wait_for_enter(f"Login if needed for {account['id']} ({site_key}), then press Enter... ")

    await ctx.close()
# ===== Excel =====
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import hashlib

# from contentplanner_worker import db_report_image
# =========================
# GLOBAL CONFIG (edit here)
# =========================

# =========================
# RETRY CONFIG
# =========================

ENABLE_RETRY = True        # <--- main flag
MAX_RETRY_ATTEMPTS = 5     # max retries per run
RETRY_SLEEP_SECONDS = 10  # pause between retries

# =========================
RUN_MODE = "videos"   # "images" or "videos"

EXCEL_FILE  = r"media_jobs.xlsx"
SHEET_NAME  = "Jobs"

# default video rendering if video_cmd is blank
DEFAULT_VIDEO_CMD = (
    'ffmpeg -y -loop 1 -i "{image}" -t 8 -r 30 -pix_fmt yuv420p "{out}"'
)

# "Video account" label when videos are created locally (not by a browser account)
VIDEO_ACCOUNT_ID = "local_ffmpeg"

# Video rendering defaults for the local helper (not used by DEFAULT_VIDEO_CMD directly)
RESOLUTION = "1080x1920"
DURATION   = 8
FPS        = 30
KENBURNS   = True

# Playwright / Browser
BROWSER_NAME      = "chromium"
CHROME_EXECUTABLE = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
HEADLESS          = False
POLITE_MIN_WAIT   = 0.8
POLITE_MAX_WAIT   = 1.8

# Make screenshots crisper (also helps UI fallbacks)
DEVICE_SCALE_FACTOR = 2

# Downloads isolation (prevents Chrome saving into OS default Downloads)
ISOLATE_DOWNLOADS = True

# Work distribution
MAX_GGL_IMG_PROMPTS_PER_ACCOUNT = 40
MAX_META_VID_PROMPTS_PER_ACCOUNT = 60
SHUFFLE_PROMPTS         = False
DRY_RUN                 = False

# ---- Account audit columns (do not rename unless you also update ensure_columns) ----
ACCOUNT_ID1_COL = "account_id_1"   # who generated the image
ACCOUNT_ID2_COL = "account_id_2"   # who generated the video

# ============ ACCOUNTS / PROFILES ============
# This script supports MANY isolated login profiles.
#
# Recommended approach:
#   - Create one folder per identity under ./browser_profiles/<profile_dir>
#   - Run with BOOTSTRAP_LOGIN=True once to login to the sites you need
#   - After that, sessions persist in those folders.
#
# Profiles are loaded from ./profiles.json (created on first run).
# Each item supports:
#   {
#     "id": "gmail_01",
#     "profile_dir": "gmail_01",              // relative -> ./browser_profiles/gmail_01
#     "out": "downloads/gmail_01",
#     "sites": {
#        "aistudio": "https://aistudio.google.com/",
#        "meta": "https://www.meta.ai/media/?nr=1",
#        "claude": "https://claude.ai/",
#        "grok": "https://x.com/i/grok",
#        "elevenlabs": "https://elevenlabs.io/app",
#        "heygen": "https://app.heygen.com/"
#     }
#   }
#
# Backward-compatible:
#   - "profile_dir" can also be an ABSOLUTE path to an existing Chrome profile folder
#   - legacy keys "profile", "google_url", "meta_url" are also accepted

PROFILES_BASE_DIR = Path(__file__).with_name("browser_profiles")
PROFILES_JSON     = Path(__file__).with_name("profiles.json")

# If profiles.json does not exist yet, we'll write a starter file based on your previous ACCOUNTS list.
LEGACY_ACCOUNTS: List[Dict[str, str]] = []

LEGACY_ACCOUNTS_old: List[Dict[str, str]] = [
    # {
    #     "id": "numero_uno", #not worked
    #     "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 1",
    #     "out": r"downloads",
    #     "google_url": "https://aistudio.google.com/prompts/1SHiNmxmlkmYTqHH8wseV4evAegV0pRvH",
    #     "meta_url":   "https://www.meta.ai/media/?nr=1",
    # },


    {
        "id": "s173m", # worked sm173
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 3",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/1Tyv17HS3-w0sVP6oPNQ8ULMDBzAMvIMF",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    },
    {
        "id": "mail2km", # worked
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 4",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/1CzRw-7NIzeaIqFCpYc5-vV_vM2kRMDXh",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    },
    {
        "id": "mail2kishnm", # worked - numero uno
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 6",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/1uV0ZBl9JX_kBB-iBk_2alI0Bco72Ck0n",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    },    
    {
        "id": "mail2sm",
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 7",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/18LNm8fsaraxHYYWCB92vGp4BC1-put7S",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    }, 
    {
        "id": "mail2nishm1",
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 8",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/14ghA9r3kNxMo0GCgX5mxcfA8i-fk7ee-",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    }, 
    {
        "id": "mail2nakshm",
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 9",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/1GhH0rXYgQtwTY7h9FtKUfYCWtT0n8oph",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    }, 
    {
        "id": "mummy",
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 10",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/1NvEngwCN7o0ghPVb_x-MNOsi0ESIc_5M",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    }, 
    # {
    #     "id": "papa",
    #     "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 11",
    #     "out": r"downloads",
    #     "google_url": "https://aistudio.google.com/prompts/1Dyaf1Zo_EzdUIiZGlJzasLD6k0WvQcnU",
    #     "meta_url":   "https://www.meta.ai/media/?nr=1",
    # },
]

# Bootstrap helper:
# Set to True to open each profile and let you login once (press Enter in terminal to continue).
BOOTSTRAP_LOGIN = False
# Which sites to open during bootstrap (keys from "sites" dict)
BOOTSTRAP_SITES = ["aistudio", "meta"]


def _is_probably_absolute_path(p: str) -> bool:
    if not p:
        return False
    # Windows drive paths like C:\...
    if re.match(r"^[A-Za-z]:\\", p):
        return True
    return Path(p).is_absolute()


def _resolve_profile_path(profile_dir: str) -> Path:
    if _is_probably_absolute_path(profile_dir):
        return Path(profile_dir)
    return PROFILES_BASE_DIR / profile_dir


def _get_site_url(account: Dict[str, str], key: str, fallback: Optional[str] = None) -> str:
    sites = account.get("sites") or {}
    # Back-compat mapping
    if key == "aistudio":
        return sites.get("aistudio") or account.get("google_url") or fallback or "https://aistudio.google.com/"
    if key == "meta":
        return sites.get("meta") or account.get("meta_url") or fallback or "https://www.meta.ai/media/?nr=1"
    return sites.get(key) or fallback or ""


def load_accounts() -> List[Dict[str, str]]:
    PROFILES_BASE_DIR.mkdir(parents=True, exist_ok=True)

    if PROFILES_JSON.exists():
        try:
            data = json.loads(PROFILES_JSON.read_text(encoding="utf-8"))
        except Exception as e:
            raise RuntimeError(f"Failed to parse profiles.json: {e}")
        if not isinstance(data, list):
            raise RuntimeError("profiles.json must be a JSON array (list) of profile objects.")
    else:
        # Write a starter profiles.json:
        #  - we keep your existing ABSOLUTE Chrome profile paths so nothing breaks
        #  - and we also include a template "gmail_01" using a managed folder
        starter = []
        for acc in LEGACY_ACCOUNTS:
            starter.append({
                "id": acc["id"],
                "profile_dir": acc["profile"],  # keep existing Chrome profile folder
                "out": str(Path("downloads") / acc["id"]),
                "sites": {
                    "aistudio": acc.get("google_url", "https://aistudio.google.com/"),
                    "meta": acc.get("meta_url", "https://www.meta.ai/media/?nr=1"),
                },
            })

        starter.append({
            "id": "gmail_01",
            "profile_dir": "gmail_01",  # will live in ./browser_profiles/gmail_01
            "out": "downloads/gmail_01",
            "sites": {
                "aistudio": "https://aistudio.google.com/",
                "meta": "https://www.meta.ai/media/?nr=1",
                "claude": "https://claude.ai/",
                "grok": "https://x.com/i/grok",
                "elevenlabs": "https://elevenlabs.io/app",
                "heygen": "https://app.heygen.com/",
            },
        })

        PROFILES_JSON.write_text(json.dumps(starter, indent=2), encoding="utf-8")
        print(f"[profiles] Created starter profiles.json at: {PROFILES_JSON}")
        print("[profiles] You can now add 10+ profiles by duplicating the gmail_01 entry and changing id/profile_dir.")
        data = starter

    accounts: List[Dict[str, str]] = []
    seen_ids = set()
    for entry in data:
        if not isinstance(entry, dict):
            raise RuntimeError("Each item in profiles.json must be an object/dict.")
        acc = dict(entry)  # copy

        acc_id = acc.get("id")
        if not acc_id or not isinstance(acc_id, str):
            raise RuntimeError("Each profile must have a string 'id'.")
        if acc_id in seen_ids:
            raise RuntimeError(f"Duplicate profile id in profiles.json: {acc_id}")
        seen_ids.add(acc_id)

        profile_dir = acc.get("profile_dir") or acc.get("profile")
        if not profile_dir:
            raise RuntimeError(f"Profile '{acc_id}' must include 'profile_dir' (recommended) or legacy 'profile'.")

        profile_path = _resolve_profile_path(str(profile_dir))
        profile_path.mkdir(parents=True, exist_ok=True)
        acc["profile"] = str(profile_path)

        # Default output folder per account (recommended)
        acc["out"] = acc.get("out") or str(Path("downloads") / acc_id)

        # Back-compat: keep google_url/meta_url in sync
        acc["google_url"] = acc.get("google_url") or _get_site_url(acc, "aistudio", "https://aistudio.google.com/")
        acc["meta_url"] = acc.get("meta_url") or _get_site_url(acc, "meta", "https://www.meta.ai/media/?nr=1")

        accounts.append(acc)

    # refuse duplicate profile paths (can't open same profile in parallel)
    profiles = [a["profile"] for a in accounts]
    dupes = {p for p in profiles if profiles.count(p) > 1}
    if dupes:
        raise RuntimeError(f"Duplicate profile folders in profiles.json (one profile can't be reused concurrently): {dupes}")

    return accounts


ACCOUNTS: List[Dict[str, str]] = load_accounts()
META_ACCOUNTS: List[Dict[str, str]] = [
    {
        "id": "mail2sm",
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 1",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/1SHiNmxmlkmYTqHH8wseV4evAegV0pRvH",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    }, 
    {
        "id": "numero_uno",
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 3",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/18LNm8fsaraxHYYWCB92vGp4BC1-put7S",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    },
    {
        "id": "mail2km",
        "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 4",
        "out": r"downloads",
        "google_url": "https://aistudio.google.com/prompts/1CzRw-7NIzeaIqFCpYc5-vV_vM2kRMDXh",
        "meta_url":   "https://www.meta.ai/media/?nr=1",
    },
]


CHATGPT_ACCOUNT = {
    "id": "chatgpt_primary",
    "profile": r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 4",
    "out": r"downloads",
    "url": "https://chatgpt.com/images/",
}

# ======== Helpers ========

def ensure_dir(p: str | Path):
    Path(p).mkdir(parents=True, exist_ok=True)

def safe_filename_nohash(name: str, ext: str = "png", maxlen: int = 180) -> str:
    raw = (name or "").strip()
    base = re.sub(r"[^\w\-]+", "_", raw).strip("_").rstrip(" .")
    if not base:
        base = "asset"
    base = base[:maxlen].rstrip(" .")
    return f"{base}.{ext}"

def split_image_jobs(rows):
    chatgpt_jobs = []
    google_jobs = []

    for r in rows:
        provider = (r.get("image_provider") or "").lower()
        if provider == "chatgpt":
            chatgpt_jobs.append(r)
        else:
            google_jobs.append(r)

    return chatgpt_jobs, google_jobs


def safe_basename(name: str, ext: str = "png", idx: int | None = None, maxlen: int = 70) -> str:
    """
    Windows-safe filename:
    - strips unsafe chars
    - truncates to maxlen
    - appends short hash to preserve uniqueness
    """
    raw = (name or "").strip()
    base = re.sub(r"[^\w\-]+", "_", raw).strip("_")

    # avoid empty names
    if not base:
        base = "asset"

    # add uniqueness hash (based on full prompt, not truncated)
    h = hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest()[:8]

    # reserve room for suffixes
    suffix = f"_{idx}" if idx is not None else ""
    tail = f"{suffix}_{h}.{ext}"

    # truncate base so final filename length stays reasonable
    # (also avoids trailing dot/space issues)
    keep = max(1, maxlen - len(tail))
    base = base[:keep].rstrip(" .")

    return f"{base}{tail}"

def safe_basename_old2(name: str, ext: str = "png", idx: int | None = None) -> str:
    base = re.sub(r"[^\w\-]+", "_", name.strip()).strip("_")
    return f"{base}_{idx}.{ext}" if idx is not None else f"{base}.{ext}"

# def safe_basename(name: str, ext: str) -> str:
#     base = re.sub(r"[^\w\-]+", "_", name.strip()).strip("_")
#     return f"{base}.{ext}"

def safe_basename_old(s: str, maxlen: int = 50) -> str:
    s = re.sub(r"[^a-zA-Z0-9._-]+", "_", s).strip("_")
    base = s[:maxlen] or "asset"
    # DND - Working
    # ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    ts = datetime.now().strftime("%M%S")
    return f"{ts}_{base}"

def open_wb_with_retry(path: str, attempts: int = 5, delay: float = 0.5):
    for i in range(attempts):
        try:
            return load_workbook(path)
        except PermissionError:
            if i == attempts - 1:
                raise
            import time; time.sleep(delay)

def save_wb_with_retry(wb, path: str, attempts: int = 5, delay: float = 0.5):
    for i in range(attempts):
        try:
            wb.save(path)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            import time; time.sleep(delay)

def ensure_columns(ws, required_headers: List[str]) -> None:
    """Create any missing headers at the end of row 1."""
    existing = { (str(c.value).strip().lower() if c.value else ""): j
                 for j, c in enumerate(ws[1], start=1) }
    max_col = ws.max_column
    for name in required_headers:
        key = name.lower()
        if key not in existing or existing.get(key) is None:
            max_col += 1
            ws.cell(1, max_col).value = name
            existing[key] = max_col

def colmap_from_headers(ws) -> Dict[str, int]:
    """Map header name (lower) -> column index (1-based)."""
    header = {}
    for j, cell in enumerate(ws[1], start=1):
        if cell.value:
            header[str(cell.value).strip().lower()] = j
    return header

def safe_move_and_overwrite(src: Path, dest: Path):
    """Moves a file, overwriting the destination if it exists."""
    try:
        dest.parent.mkdir(parents=True, exist_ok=True)
        if dest.exists():
            dest.unlink() # Delete existing file
        shutil.move(str(src), str(dest))
    except Exception as e:
        print(f"Error moving {src} to {dest}: {e}")
        raise

def read_jobs_from_excel_for_images() -> List[Dict]:
    """Rows needing image generation: prompt set AND image_path empty."""
    wb = open_wb_with_retry(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # Ensure columns exist (adds audit columns if missing)
    ensure_columns(ws, [
        "prompt", "account_id", "image_path", "video_cmd", "video_path", "status",
        ACCOUNT_ID1_COL, ACCOUNT_ID2_COL
    ])
    h = colmap_from_headers(ws)

    rows = []
    for i in range(2, ws.max_row + 1):
        prompt = (ws.cell(i, h["prompt"]).value or "").strip()
        img    = (ws.cell(i, h["image_path"]).value or "").strip()
        acct   = (ws.cell(i, h["account_id"]).value or "").strip()
        section_id = int(ws.cell(i, h.get("section_id")).value or 0) if "section_id" in h else 0
        image_name = (ws.cell(i, h.get("image_name")).value or "").strip() if "image_name" in h else ""

        # image_provider = (ws.cell(i, h.get("image_provider")).value or "").strip().lower()

        col_idx = h.get("image_provider")
        image_provider = (ws.cell(i, col_idx).value or "").strip().lower() if col_idx else ""
        image_orientation = (ws.cell(i, h.get("image_orientation")).value or "").strip().lower()

        if prompt and not img:
            rows.append({
                "row": i, 
                "prompt": prompt, 
                "account_id": acct, 
                "section_id": section_id, 
                "image_name": image_name,
                "image_provider": image_provider,          # NEW
                "image_orientation": image_orientation,    # NEW                
                })
    save_wb_with_retry(wb, EXCEL_FILE)  # in case we added headers
    wb.close()
    return rows

def read_jobs_from_excel_for_videos() -> List[Dict]:
    """Rows needing video generation: image_path set AND video_path empty."""
    wb = open_wb_with_retry(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # Ensure columns exist
    ensure_columns(ws, [
        "prompt", "account_id", "image_path", "video_cmd", "video_path", "status",
        ACCOUNT_ID1_COL, ACCOUNT_ID2_COL
    ])
    h = colmap_from_headers(ws)

    rows = []
    for i in range(2, ws.max_row + 1):
        image_path = (ws.cell(i, h["image_path"]).value or "").strip()
        vout  = (ws.cell(i, h["video_path"]).value or "").strip()
        vcmd  = (ws.cell(i, h["video_cmd"]).value or "").strip()
        acct       = (ws.cell(i, h["account_id"]).value or "").strip()

        if image_path and not vout:
            rows.append({
                "row": i,
                "account_id": acct,
                "image_path": image_path,
                "video_cmd": vcmd,  # we'll treat this as the Meta animation prompt
            })
    save_wb_with_retry(wb, EXCEL_FILE)
    wb.close()
    return rows

def write_image_result(row_idx: int, image_path: str, account_id_used: str, status: str = "ok"):
    wb = open_wb_with_retry(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    ensure_columns(ws, [
        "prompt", "account_id", "image_path", "video_cmd", "video_path", "status",
        ACCOUNT_ID1_COL, ACCOUNT_ID2_COL
    ])
    h = colmap_from_headers(ws)

    # Write image path + status
    ws.cell(row_idx, h["image_path"]).value = image_path
    ws.cell(row_idx, h["status"]).value = status

    # If account_id_1 is empty, record who actually generated the image
    if ws.cell(row_idx, h[ACCOUNT_ID1_COL]).value in (None, ""):
        ws.cell(row_idx, h[ACCOUNT_ID1_COL]).value = account_id_used

    # Optional: if assignment account_id is blank, backfill with the one we used
    if ws.cell(row_idx, h["account_id"]).value in (None, ""):
        ws.cell(row_idx, h["account_id"]).value = account_id_used

    save_wb_with_retry(wb, EXCEL_FILE)
    wb.close()

def write_video_result(row_idx: int, video_path: str, account_id_used: str, status: str = "ok"):
    wb = open_wb_with_retry(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    ensure_columns(ws, [
        "prompt", "account_id", "image_path", "video_cmd", "video_path", "status",
        ACCOUNT_ID1_COL, ACCOUNT_ID2_COL
    ])
    h = colmap_from_headers(ws)

    # Write video path + status
    ws.cell(row_idx, h["video_path"]).value = video_path
    ws.cell(row_idx, h["status"]).value = status

    # If account_id_2 is empty, record who actually generated the video
    if ws.cell(row_idx, h[ACCOUNT_ID2_COL]).value in (None, ""):
        ws.cell(row_idx, h[ACCOUNT_ID2_COL]).value = account_id_used

    save_wb_with_retry(wb, EXCEL_FILE)
    wb.close()

# =========================
# Site-specific automation
# =========================

async def ensure_logged_in(page: Page, post_login_selector: str, site_name: str):
    print(f"[{site_name}] Waiting for post-login marker: {post_login_selector}")
    await page.wait_for_selector(post_login_selector, timeout=120_000)

async def generate_image_google_ai(page: Page, prompt: str, out_dir: Path, image_name: str = "") -> Path:
    """
    Uses selectors:
    - Prompt textbox role name: 'Enter a prompt to generate an'
    - Run button: role 'button' with exact name 'Run'
    - Newest gallery item: assumes append; grabs item at old_count index
    - Large view button label: 'Large view of this image'
    - Modal download button role name: 'Download'
    """
    # Post-login marker: the prompt textbox (adjust if the UI changes)
    textbox = page.get_by_role("textbox", name="Enter a prompt to generate an")
    await expect(textbox).to_be_visible(timeout=120_000)

    gallery_items = page.locator("ms-image-generation-gallery-image")
    before_cnt = await gallery_items.count()
    print(f"before_cnt = {before_cnt}")
    await textbox.fill(prompt)

    if DRY_RUN:
        # ph = out_dir / f"{safe_basename(prompt)}__DRYRUN.png"
        ph = out_dir / f"{safe_basename(prompt, 'png')}__DRYRUN.png"

        
        try:
            from PIL import Image; Image.new("RGB", (16, 16), (210, 210, 210)).save(ph)
        except Exception:
            ph.write_bytes(b"")
        return ph

    # run_btn = page.get_by_role("button", name="Run", exact=True)
    # await expect(run_btn).to_be_enabled()
    # await run_btn.click()

    # Removing exact=True allows for flexible whitespace handling
    # run_btn = page.get_by_role("button", name=re.compile(r"^\s*Run\s*$"))
    # await expect(run_btn).to_be_visible() # Good practice to check visibility too
    # await expect(run_btn).to_be_enabled()
    run_btn = page.locator("button.ms-button-primary", has_text="Run")
    await run_btn.click()

    await asyncio.sleep(15)
    print("waited 15 seconds")

    # Wait for a new gallery item
    async def gallery_increased():
        return await gallery_items.count() > before_cnt

    for _ in range(90):
        if await gallery_increased():
            break
        await asyncio.sleep(1.0)

    after_cnt = await gallery_items.count()
    if after_cnt <= before_cnt:
        raise RuntimeError("No new image appeared.")

    print(f"after_cnt = {after_cnt}")
    # Candidate: first new item (append behavior)
    candidate = gallery_items.nth(before_cnt)
    # out_path = out_dir / f"{safe_basename(prompt)}.png"
    
    # If Image name is blank, use safe_basename(prompt)
    if image_name.strip():
        # out_path = out_dir / f"{safe_basename(image_name, 'png')}"
        # out_path = out_dir / safe_filename_nohash(image_name, "png")
        out_path = out_dir / f"{image_name}.png"

    else:
        out_path = out_dir / safe_basename(prompt, "png")



    # Try large-view modal and proper download
    try:
        await candidate.get_by_label("Large view of this image").click()
        print("Large view button clicked")

        async with page.expect_download(timeout=30_000) as dl_info:
            await page.get_by_role("button", name="Download").click()
        dl = await dl_info.value

        # move to final path (avoid duplicate in default download dir)
        # suggested = dl.suggested_filename or out_path.name
        # dest = out_path.with_name(suggested)

        dest = out_path.with_name(out_path.name)
        src = await dl.path()
        # shutil.move(src, dest)
        safe_move_and_overwrite(Path(src), dest)


        # close modal if present
        try:
            await page.get_by_role("button", name="Close").click()
        except Exception:
            pass

        # Optional: quick dimension log (requires Pillow)
        try:
            from PIL import Image
            w, h = Image.open(dest).size
            print(f"[Google AI] Saved full-size {w}x{h}: {dest}")
        except Exception:
            print(f"[Google AI] Saved: {dest}")

        return dest

    except Exception:
        # Full-size fallback via modal image element (or thumb if needed)
        try:
            modal_img = page.get_by_role("img", name="Generated image").first
            await modal_img.wait_for(state="visible", timeout=30_000)

            # Try reading raw bytes from <img src>
            src_attr = await modal_img.get_attribute("src")
            if src_attr and src_attr.startswith("data:"):
                import base64
                m = re.match(r"data:(.*?);base64,(.*)", src_attr, re.DOTALL)
                if m:
                    raw = base64.b64decode(m.group(2))
                    out_path.write_bytes(raw)
                    await page.keyboard.press("Escape")
                    return out_path

            # Else screenshot as last resort
            await modal_img.screenshot(path=str(out_path))
            await page.keyboard.press("Escape")
            return out_path

        except Exception:
            # last resort: thumbnail screenshot
            img_inside = candidate.locator("img").first
            if await img_inside.count() > 0:
                await img_inside.screenshot(path=str(out_path))
                return out_path
            await candidate.screenshot(path=str(out_path))
            return out_path


def apply_image_orientation(prompt: str, orientation: str | None) -> str:
    """
    Injects size/aspect hints into the prompt.
    ChatGPT Images responds best to descriptive ratios, not exact pixels.
    """
    if not orientation:
        return prompt

    orientation = orientation.lower().strip()

    if orientation == "portrait":
        suffix = " -- vertical portrait orientation, 2:3 aspect ratio"
    elif orientation == "landscape":
        suffix = " -- horizontal landscape orientation, 16:9 aspect ratio"
    elif orientation == "square":
        suffix = " -- square composition, 1:1 aspect ratio"
    else:
        return prompt

    return f"{prompt.rstrip('.')}.{suffix}"

async def generate_image_router(
    *,
    page: Page,
    job: dict,
    out_dir: Path
) -> Path:
    """
    Routes image generation to ChatGPT Images or Google AI
    based on Excel flag.
    """

    provider = job.get("image_provider", "").lower()
    prompt = job["prompt"]
    image_name = job.get("image_name", "")
    orientation = job.get("image_orientation", "")

    # Inject orientation hint
    final_prompt = apply_image_orientation(prompt, orientation)

    if provider == "chatgpt":
        return await generate_consistent_image_chatgpt(
            page=page,
            prompt=final_prompt,
            out_dir=out_dir,
            image_name=image_name or safe_basename(prompt, "png").replace(".png", "")
        )

    # Default (existing behavior)
    return await generate_image_google_ai(
        page=page,
        prompt=final_prompt,
        out_dir=out_dir,
        image_name=image_name
    )

async def _click_if_visible(locator, *, timeout=1500) -> bool:
    try:
        if await locator.count() == 0:
            return False
        el = locator.first
        await el.wait_for(state="visible", timeout=timeout)
        try:
            await el.click(timeout=timeout)
        except Exception:
            await el.click(timeout=timeout, force=True)
        return True
    except Exception:
        return False

async def handle_new_chatgpt_ui(page: Page, ui_stamp: str) -> bool:
    """
    Handles ChatGPT UI overlays that block the download button.
    Bypasses visibility checks because these elements are often 
    physically off-screen or 'hidden' until focused.
    """

    # 1. Target the 'Image 1 is better' / Feedback buttons
    # We look for the specific text or the button structure
    feedback_btn = page.locator('button.btn-secondary').filter(
        has_text=re.compile(r"Image \d+( is better)?", re.I)
    ).first

    if await feedback_btn.count() > 0:
        try:
            # Use JS click to bypass 'visibility' and 'intercepted click' checks
            await feedback_btn.evaluate("el => el.click()")
            print("[ChatGPT Images] Clicked feedback button via JS")
            return True
        except Exception:
            pass

    # 2. Target the 'Skip to content' link
    # This element is usually at translate-y-[-100lvh] (off-screen)
    skip_link = page.locator('a[data-skip-to-content][href="#main"]').first
    
    if await skip_link.count() > 0:
        try:
            # Forcing a click on an off-screen element requires JS
            await skip_link.evaluate("el => el.click()")
            print("[ChatGPT Images] Clicked skip-to-content via JS")
            return True
        except Exception:
            pass

    return False

async def handle_new_chatgpt_ui_old2(page: Page, ui_stamp: str) -> bool:
    """
    Clicks NEW (unstamped) UI that sometimes appears after sending prompt:
    - 'Image 1 is better' / 'Image 2 is better' feedback buttons
    - 'Skip to content' link (hidden off-screen usually)
    Returns True if it clicked something.
    """

    # 1. Handle Feedback Buttons (Standard visibility logic works here)
    feedback_btn = page.locator(
        f'button.btn.btn-secondary:not([data-ui-stamp="{ui_stamp}"])'
    ).filter(has_text=re.compile(r"\bImage\s*\d+(\s+is\s+better)?\b", re.I))

    if await _click_if_visible(feedback_btn, timeout=2000):
        print("[ChatGPT Images] Clicked feedback UI")
        return True

    # 2. Handle Skip Link (Modified Logic)
    # The CSS class '-translate-y-[-100lvh]' makes this element effectively "invisible" 
    # to Playwright's strict checks unless focused.
    skip_link = page.locator(
        f'a[data-skip-to-content][href="#main"]:not([data-ui-stamp="{ui_stamp}"])'
    )
    
    # Check count only, do NOT wait for visibility
    if await skip_link.count() > 0:
        try:
            # Option A: The most robust way for off-screen links is executing JS directly
            await skip_link.first.evaluate("el => el.click()")
            
            # Option B: Alternatively, bring it into view via focus, then click
            # await skip_link.first.focus() 
            # await skip_link.first.click()

            print("[ChatGPT Images] Clicked skip-to-content")
            return True
        except Exception:
            # If interaction fails, just ignore it so we don't crash the run
            pass

    return False

async def handle_new_chatgpt_ui_old(page: Page, ui_stamp: str) -> bool:
    """
    Clicks NEW (unstamped) UI that sometimes appears after sending prompt:
    - 'Image 1 is better' / 'Image 2 is better' feedback buttons
    - 'Skip to content' link (rare / usually focus-only)
    Returns True if it clicked something.
    """

    # NEW feedback buttons (robust: match either "is better" OR just "Image 1"/"Image 2")
    feedback_btn = page.locator(
        f'button.btn.btn-secondary:not([data-ui-stamp="{ui_stamp}"])'
    ).filter(has_text=re.compile(r"\bImage\s*\d+(\s+is\s+better)?\b", re.I))

    if await _click_if_visible(feedback_btn, timeout=2000):
        print("[ChatGPT Images] Clicked feedback UI")
        return True

    # NEW skip-to-content link
    skip_link = page.locator(
        f'a[data-skip-to-content][href="#main"]:not([data-ui-stamp="{ui_stamp}"])'
    )
    if await _click_if_visible(skip_link, timeout=1000):
        print("[ChatGPT Images] Clicked skip-to-content")
        return True

    return False
    
async def generate_consistent_image_chatgpt(
    page: Page,
    prompt: str,
    out_dir: Path,
    image_name: str
) -> Path:
    """
    Generates an image in the SAME ChatGPT Images conversation
    and downloads ONLY the newly created image.
    """

    out_dir.mkdir(parents=True, exist_ok=True)

    # 1) Build a robust locator (CSS is often more stable than role+name here)
    dl_all = page.locator('button[aria-label="Download this image"]')
    dl_visible = page.locator('button[aria-label="Download this image"]:visible')

    # Debug counts (optional, but super helpful)
    before_all = await dl_all.count()
    before_vis = await dl_visible.count()
    print(f"[ChatGPT Images] Download buttons before prompt: all={before_all}, visible={before_vis}")

    # Stamp everything currently present
    stamp = f"seen-{uuid.uuid4().hex}"
    await page.evaluate(
        """(stamp) => {
            document
            .querySelectorAll('button[aria-label="Download this image"]')
            .forEach(b => b.setAttribute('data-stamp', stamp));
        }""",
        stamp
    )

    # Stamp any existing "popup UI" elements BEFORE sending prompt
    ui_stamp = f"ui-seen-{uuid.uuid4().hex}"
    await page.evaluate(
        """(stamp) => {
            const nodes = [
            ...document.querySelectorAll('a[data-skip-to-content][href="#main"]'),
            ...document.querySelectorAll('button.btn.btn-secondary')
            ];
            nodes.forEach(n => n.setAttribute('data-ui-stamp', stamp));
        }""",
        ui_stamp
    )

    # 2) Enter prompt

    # If paragraph is not found redirect to account["url"]

    try:
        prompt_box = page.get_by_role("paragraph")
        await expect(prompt_box).to_be_visible(timeout=10_000)
        await prompt_box.click()
    except Exception:
        account = CHATGPT_ACCOUNT
        print("[ChatGPT Images] Prompt box not found, reloading page...")
        await page.goto(account["url"], wait_until="networkidle")

        # 1) Build a robust locator (CSS is often more stable than role+name here)
        dl_all = page.locator('button[aria-label="Download this image"]')
        dl_visible = page.locator('button[aria-label="Download this image"]:visible')

        # Debug counts (optional, but super helpful)
        before_all = await dl_all.count()
        before_vis = await dl_visible.count()
        print(f"[ChatGPT Images] Download buttons before prompt: all={before_all}, visible={before_vis}")

        # Stamp everything currently present
        stamp = f"seen-{uuid.uuid4().hex}"
        await page.evaluate(
            """(stamp) => {
                document
                .querySelectorAll('button[aria-label="Download this image"]')
                .forEach(b => b.setAttribute('data-stamp', stamp));
            }""",
            stamp
        )

        # Stamp any existing "popup UI" elements BEFORE sending prompt
        ui_stamp = f"ui-seen-{uuid.uuid4().hex}"
        await page.evaluate(
            """(stamp) => {
                const nodes = [
                ...document.querySelectorAll('a[data-skip-to-content][href="#main"]'),
                ...document.querySelectorAll('button.btn.btn-secondary')
                ];
                nodes.forEach(n => n.setAttribute('data-ui-stamp', stamp));
            }""",
            ui_stamp
        )


        prompt_box = page.get_by_role("paragraph")
        await expect(prompt_box).to_be_visible(timeout=10_000)
        await prompt_box.click()
        
    



    # scope to the composer form that contains the send button
    # send_btn = page.get_by_test_id("send-button")
    # composer = send_btn.locator("xpath=ancestor::form[1]")
    # prompt_box = composer.locator("#prompt-textarea[contenteditable='true']").first
    # await expect(prompt_box).to_be_visible(timeout=30_000)
    # await prompt_box.click()

    await page.keyboard.type(prompt, delay=10)
    
    # prompt_box = page.locator('textarea#prompt-textarea, [contenteditable="true"][role="textbox"]').first
    # await prompt_box.wait_for(state="visible", timeout=30_000)
    # await prompt_box.click()
    # await prompt_box.fill(prompt)

    print(f"[ChatGPT Images] Prompt entered.")

    # 3) Send
    send_btn = page.get_by_test_id("send-button")
    # await expect(send_btn).to_be_enabled()
    await send_btn.click()
    print(f"[ChatGPT Images] Prompt sent, waiting for new image...")

    # 4) Wait for a NEW (unstamped) download button
    new_dl = page.locator(
        f'button[aria-label="Download this image"]:not([data-stamp="{stamp}"])'
    )

    await handle_new_chatgpt_ui(page, ui_stamp)

    for _ in range(240):
        await handle_new_chatgpt_ui(page, ui_stamp)
        if await new_dl.count() > 0:
            break
        await asyncio.sleep(1)

    if await new_dl.count() == 0:
        # raise RuntimeError("No new image appeared")
        print("[ChatGPT Images] No new image appeared, reloading page and retrying...")
        account = CHATGPT_ACCOUNT
        await page.goto(account["url"], wait_until="networkidle")        
        raise RuntimeError("No new image appeared reloaded page")

    # 5) Use the newest unstamped button
    newest_button = new_dl.last
    await expect(newest_button).to_be_visible(timeout=30_000)

    # 6) Download ONLY this image
    async with page.expect_download(timeout=60_000) as dl_info:
        await newest_button.click()

    download = await dl_info.value
    tmp_path = await download.path()

    # 7) Save with safe name
    final_path = out_dir / f"{image_name}.png"
    i = 1
    while final_path.exists():
        final_path = out_dir / f"{image_name}_{i}.png"
        i += 1

    # shutil.move(tmp_path, final_path)
    safe_move_and_overwrite(Path(tmp_path), final_path)
    print(f"[ChatGPT Images] Saved: {final_path}")

    return final_path


async def generate_consistent_image_chatgpt_working_exp_for1(
    page: Page,
    prompt: str,
    out_dir: Path,
    image_name: str
) -> Path:
    """
    Generates an image in the SAME ChatGPT Images conversation
    and downloads ONLY the newly created image.
    """

    out_dir.mkdir(parents=True, exist_ok=True)

    # 1) Count existing download buttons
    download_buttons = page.get_by_role("button", name="Download this image")
    before_count = await download_buttons.count()

    print(f"[ChatGPT Images] Existing download buttons before new image prompt: {before_count}")

    # 2) Enter prompt
    prompt_box = page.get_by_role("paragraph")
    await expect(prompt_box).to_be_visible(timeout=30_000)
    await prompt_box.click()
    await page.keyboard.type(prompt, delay=10)
    print(f"[ChatGPT Images] Prompt entered.")

    # 3) Send
    send_btn = page.get_by_test_id("send-button")
    # await expect(send_btn).to_be_enabled()
    await send_btn.click()
    print(f"[ChatGPT Images] Prompt sent, waiting for new image...")

    # 4) Wait for NEW image to appear
    async def new_image_ready():
        return await download_buttons.count() > before_count

    for _ in range(240):  # up to 4 minutes
        if await new_image_ready():
            break
        await asyncio.sleep(1)

    after_count = await download_buttons.count()
    if after_count <= before_count:
        raise RuntimeError("No new image appeared")

    # 5) Select newest download button
    newest_button = download_buttons.nth(after_count - 1)
    await expect(newest_button).to_be_visible()

    # 6) Download ONLY this image
    async with page.expect_download(timeout=60_000) as dl_info:
        await newest_button.click()

    download = await dl_info.value
    tmp_path = await download.path()

    # 7) Save with safe name
    final_path = out_dir / f"{image_name}.png"
    i = 1
    while final_path.exists():
        final_path = out_dir / f"{image_name}_{i}.png"
        i += 1

    # shutil.move(tmp_path, final_path)
    safe_move_and_overwrite(Path(tmp_path), final_path)
    print(f"[ChatGPT Images] Saved: {final_path}")

    return final_path


async def generate_video_meta_ai(page: Page, imagePath: str, prompt: str, out_dir: Path, url) -> Path:
    """
    Automates Meta AI 'Media' -> Video-from-Image flow.

    It will:
      - Switch to 'Video' tab (from Image)
      - Click 'Upload image' and attach the file
      - Fill 'Describe your animation...' with `prompt`
      - Click 'Animate'
      - Wait for rendering to finish
      - Click 'Download media' and MOVE the .mp4 into out_dir (no OS Downloads duplicate)

    Returns: Path to the saved .mp4 in out_dir
    """

    # await page.goto(url)
    await page.goto("https://www.meta.ai/")
    print(f"[Meta AI] Navigated to Meta AI page.")
    # Ensure absolute paths
    image_path = str(Path(imagePath).resolve())
    ensure_dir(out_dir)

    # 0) Be sure the meta media page is loaded (your caller already navigates here)
    # Try to land on the media composer area by clicking the "Media" (image/video) entry if present
    # try:
    #     # Some UIs have an icon cluster first; clicking first item opens the composer
    #     await page.locator(".x78zum5.xdt5ytf.x1qughib").first.click(timeout=3000)
    # except Exception:
    #     pass  # optional, UI-dependent

    # 1) Switch to "Video" mode (recorded flows showed two variants)
    # switched = False
    # try:
    #     # Variant A: a toggle button then literal "Video" text
    #     await page.get_by_role("button", name="Image").click(timeout=3000)
    #     await page.get_by_text("Video", exact=True).click(timeout=3000)
    #     switched = True
    # except Exception:
    #     try:
    #         # Variant B: direct "Video" tab
    #         await page.get_by_text("Video", exact=True).click(timeout=3000)
    #         switched = True
    #     except Exception:
    #         pass

    # if not switched:
    #     # Some shells show a small tab container; try again by clicking the tab region
    #     try:
    #         await page.locator(".xt9c220").click(timeout=3000)
    #         await page.get_by_text("Video", exact=True).click(timeout=3000)
    #     except Exception:
    #         # If we still can’t switch, proceed—the upload button usually scopes the correct panel
    #         pass

    # 2) Click "Upload image" and feed the file chooser (prefer file chooser to avoid targeting wrong input)
    # Define your strategies in the order you want to try them
    # Each item is a tuple: (Description, Locator)

    strategies = [
        ("Animate my photo link", page.get_by_text("Animate my photo", exact=False)),
        ("Animate my photo menuitem", page.get_by_role("menuitem", name="Animate my photo")),
        ("Create a video from my photo link", page.get_by_text("a video from my photo", exact=False)),
    ]

    success = False

    # 1. Try the File Chooser strategies
    for description, locator in strategies:
        try:
            print(f"Trying: {description}")
            async with page.expect_file_chooser(timeout=5000) as fc_info:
                await locator.click(timeout=3000)
            
            fc = await fc_info.value
            await fc.set_files(image_path)
            success = True
            break # Exit loop on success
        except Exception:
            continue # Try next strategy

    # 2. Final Fallback (Direct Input) if nothing else worked
    if not success:
        try:
            print("Trying direct file input fallback")
            await page.set_input_files("input[type='file']", image_path)
            success = True
        except Exception as e:
            raise RuntimeError(f"Could not upload image after all attempts: {e}")

    # 3) Fill the animation prompt (textbox labeled "Describe your animation...")

    default_text = prompt or "Animate this image smoothly with camera move."

    # Define input targets: (Description, Locator)
    input_strategies = [
        ("Composer Input (Test ID)", page.get_by_test_id("composer-input").and_(page.get_by_role("textbox"))), 
        ("Describe your animation textbox", page.get_by_role("textbox", name="Describe your animation...")),
        ("Ask anything textbox", page.get_by_role("textbox", name="Ask anything...")),
    ]
 
    # 2. Execution
    try:
        # Optional: Initial Focus click (swallow error if it fails)
        try:
            await page.get_by_role("paragraph").click(timeout=1500)
        except Exception:
            pass

        success = False

        for description, box in input_strategies:
            try:
                print(f"Trying: {description}")
                # Wait for visibility before attempting fill
                await expect(box).to_be_visible(timeout=5000) 
                await box.fill(default_text)
                
                print(f"Successfully filled prompt via: {description}")
                success = True
                break # Exit loop once successful
            except Exception:
                continue # Move to the next locator in the list

        if not success:
            raise RuntimeError("No recognized input boxes were found on the page.")

    except Exception as e:
        raise RuntimeError(f"Could not set the animation description: {e}")


    # 0) Create a unique stamp for THIS render
    stamp = f"seen-{uuid.uuid4().hex}"

    # 1) Tag ALL existing download buttons with data-stamp="<stamp>"
    #    We tag both common patterns:
    #    - ARIA labeled buttons: [role="button"][aria-label="Download media"]
    #    - Any role="button" containing 'download' in its text (fallback)
    # print(f"Stamping existing download buttons with: {stamp}")
    # await page.evaluate("""
    # (stamp) => {
    # const tag = (btn) => { try { btn.setAttribute('data-stamp', stamp); } catch(_){} };
    # const byAria = document.querySelectorAll('[role="button"][aria-label="Download media"]');
    # byAria.forEach(tag);

    # // Fallback: text-based (case-insensitive contains 'download')
    # const allButtons = document.querySelectorAll('[role="button"]');
    # allButtons.forEach(b => {
    #     const t = (b.textContent || '').toLowerCase();
    #     const aria = (b.getAttribute('aria-label') || '').toLowerCase();
    #     if (t.includes('download') || aria.includes('download')) tag(b);
    # });
    # }
    # """, stamp)

    # 0) Create a unique stamp
    stamp = f"seen-{uuid.uuid4().hex}"

    print(f"Stamping existing download buttons with: {stamp}")

    # 1) Execute the improved JavaScript logic
    await page.evaluate("""
    (stamp) => {
        const tag = (btn) => { 
            try { btn.setAttribute('data-stamp', stamp); } catch(_){} 
        };

        // Use a broader selector: find native <button> AND [role="button"]
        const allButtons = document.querySelectorAll('button, [role="button"]');

        allButtons.forEach(b => {
            const text = (b.textContent || '').toLowerCase();
            const aria = (b.getAttribute('aria-label') || '').toLowerCase();
            
            // Match if 'download' appears in text OR the aria-label (catches "Download")
            if (text.includes('download') || aria.includes('download')) {
                tag(b);
            }
        });
    }
    """, stamp) 


    new_dl = page.locator(
        f'button[aria-label*="download" i]:not([data-stamp="{stamp}"]), '
        f'[role="button"][aria-label*="download" i]:not([data-stamp="{stamp}"])'
    )

    # 2) Now click Animate (as you already do)
    #animate_btn = page.get_by_role("button", name="Animate")
    # animate_btn = page.get_by_role("button", name="Animate", exact=True)
    #animate_btn = page.get_by_role("button", name="Animate").locator(":enabled")
    

    print("Count of download buttons before creating new media:", await new_dl.count())

    # wait 5 seconds
    await asyncio.sleep(10.0)
    button_strategies = [
        # ("Animate button", page.get_by_role("button", name="Animate", exact=False)),
        ("Animate button (testid)", page.get_by_test_id("composer-animate-button")),
        ("Send button", page.get_by_role("button", name="Send")),
        # You can also use the animate_btn variable if defined elsewhere:
        # ("Animate variable", animate_btn) 
    ]

    success = False

    # 2. Iterate through buttons
    for description, btn_locator in button_strategies:
        try:
            print(f"Trying: {description}")
            # Wait for the button to be enabled before clicking
            await expect(btn_locator).to_be_enabled(timeout=5000)
            await btn_locator.click()
            
            print(f"Successfully clicked: {description}")
            success = True
            print("Waiting for rendering to start (50 seconds)...") 
            await asyncio.sleep(50.0) 
            break
        except Exception:
            continue

    # 3. Final error handling
    if not success:
        raise RuntimeError("Could not find an active Send or Animate button.")



    # 3) Wait for a *new* download button that does NOT have our stamp
    #    First prefer the precise ARIA label, then the text fallback.
    # new_dl = page.locator(f'[role="button"][aria-label="Download media"]:not([data-stamp="{stamp}"])')
    # fallback_dl = page.locator(f'[role="button"]:not([data-stamp="{stamp}"])').filter(has_text=re.compile(r'download', re.I))

    # # Poll up to ~4 minutes for either to appear
    # deadline_seconds = 240
    # found = False
    # for _ in range(deadline_seconds):
    #     if await new_dl.count() > 0:
    #         found = True
    #         target_btn = new_dl.first
    #         break
    #     if await fallback_dl.count() > 0:
    #         found = True
    #         target_btn = fallback_dl.first
    #         break
    #     await asyncio.sleep(1.0)

    # if not found:
    #     raise RuntimeError("Timeout: no new 'Download media' button appeared after Animate.")

    # 3) Wait for a *new* download button that does NOT have our stamp
    
    # Case-insensitive aria-label contains "download"
    # new_dl = page.locator(
    #     f'button[aria-label*="download" i]:not([data-stamp="{stamp}"]), '
    #     f'[role="button"][aria-label*="download" i]:not([data-stamp="{stamp}"])'
    # )

    print("Count of new download buttons with ARIA label:", await new_dl.count())

    # If you still want a "fallback", broaden it a bit (optional)
    fallback_dl = page.locator(
        f'[aria-label*="download" i]:not([data-stamp="{stamp}"])'
    )

    deadline_seconds = 240 
    found = False

    for _ in range(deadline_seconds):
        if await new_dl.count() > 0:
            print("Count of new download buttons with ARIA label:", await new_dl.count())
            target_btn = new_dl.first
            found = True
            print("Found new download button with ARIA label.")
            # Temporary
            await asyncio.sleep(5.0)  
            break

        if await fallback_dl.count() > 0:
            print("Count of new fallback download buttons with ARIA label:", await fallback_dl.count())
            target_btn = fallback_dl.first
            found = True
            print("Found new download button with fallback locator.")
            break

        await asyncio.sleep(1.0) 

    if not found:
        raise RuntimeError("Timeout: no new Download button appeared.")

    
    # 4) Click THAT specific new button, bind expect_download to it, and MOVE the file
    await target_btn.scroll_into_view_if_needed()

    # await expect(target_btn).to_be_visible(timeout=15_000)
    # await expect(target_btn).to_be_enabled(timeout=15_000)

    # # base = Path(imagePath).resolve().stem
    # # ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    # # out_path = out_dir / f"{ts}_{base}_meta.mp4"

    # base = Path(imagePath).stem
    # out_path = out_dir / f"{base}.mp4"

    # async with page.expect_download(timeout=90_000) as dl_info:
    #     await target_btn.click(force=True) 


    # Increase the timeout or use a more robust check if the UI is slow
    # print("Waiting for the download button to be visible and enabled...")
    # await expect(target_btn).to_be_visible(timeout=15_000)

    base = Path(imagePath).stem
    out_path = out_dir / f"{base}.mp4"

    try:
        print("Attempting to click the download button...") 
        async with page.expect_download(timeout=90_000) as dl_info:
            # force=True is the key here to ignore the canvas overlay
            await target_btn.click(force=True)
    except Exception as e:
        print(f"Standard click failed, trying dispatch_event: {e}")
        async with page.expect_download(timeout=90_000) as dl_info:
            await target_btn.dispatch_event('click')

    dl = await dl_info.value
    src_path = await dl.path()
    # shutil.move(src_path, out_path)
    safe_move_and_overwrite(Path(src_path), out_path)

    # # Try to close any overlay
    # for sel in [
    #     ('role=button[name="Close"]', lambda: page.get_by_role("button", name="Close")),
    # ]:
    #     try:
    #         await sel[1]().click(timeout=800)
    #     except Exception:
    #         pass

    return out_path

        
# =========================
# Per-account worker (images pass)
# =========================

async def run_account_images(pw, account: Dict[str, str], jobs: List[Dict]):
    out_dir = Path(account["out"])
    ensure_dir(out_dir)

    browser_type = getattr(pw, BROWSER_NAME)
    print(f"[{account['id']}] Launching with profile: {account['profile']}")
    ctx_kwargs = dict(
        user_data_dir=account["profile"],
        headless=HEADLESS,
        executable_path=CHROME_EXECUTABLE,
        device_scale_factor=DEVICE_SCALE_FACTOR,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--start-maximized",
            "--no-default-browser-check",
            "--no-first-run",
        ],
        accept_downloads=True,
    )
    if ISOLATE_DOWNLOADS:
        ensure_dir(out_dir / "__tmp_downloads")
        ctx_kwargs["downloads_path"] = str(out_dir / "__tmp_downloads")

    ctx: BrowserContext = await browser_type.launch_persistent_context(**ctx_kwargs)
    page: Page = await ctx.new_page()
    await page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    # open AI Studio once per account
    aistudio_url = _get_site_url(account, "aistudio", fallback="https://aistudio.google.com/prompts/new_image?model=imagen-4.0-generate-001")
    await page.goto(aistudio_url)

    # Wait for 10 seconds
    await asyncio.sleep(60)

    for job in jobs:
        row_idx = job["row"]
        prompt  = job["prompt"]
        image_name = job.get("image_name","")
        section_id = job.get("section_id", 0)
        try:
            # img_path = await generate_image_google_ai(page, prompt, out_dir, image_name)
            img_path = await generate_image_router(
                page=page,
                job=job,
                out_dir=out_dir
            )

            write_image_result(row_idx, str(Path(img_path).resolve()), account_id_used=account["id"], status="ok")
            # if section_id:
            #     db_report_image(section_id, ok=True, image_path=img_path)
            print(f"[{account['id']}] Row {row_idx} -> {img_path}")
            await asyncio.sleep(random.uniform(POLITE_MIN_WAIT, POLITE_MAX_WAIT))
            await asyncio.sleep(10)
            print("Waited 10 seconds before generating next image")
        except Exception as e:
            write_image_result(row_idx, "", account_id_used=account["id"], status=f"error: {e}")
            # if section_id:
            #     db_report_image(section_id, ok=False, error=str(e))
            print(f"[{account['id']}] Row {row_idx} ERROR: {e}")

    await ctx.close()
    print(f"[{account['id']}] Images pass done.")


# =========================
# Per-account worker (videos pass)
# =========================

async def run_account_videos(pw, account: Dict[str, str], jobs: List[Dict]):
    out_dir = Path(account["out"])
    ensure_dir(out_dir)

    browser_type = getattr(pw, BROWSER_NAME)
    print(f"[{account['id']}] Launching with profile: {account['profile']}")
    ctx_kwargs = dict(
        user_data_dir=account["profile"],
        headless=HEADLESS,
        executable_path=CHROME_EXECUTABLE,
        device_scale_factor=DEVICE_SCALE_FACTOR,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--start-maximized",
            "--no-default-browser-check",
            "--no-first-run",
        ],
        accept_downloads=True,
    )
    if ISOLATE_DOWNLOADS:
        ensure_dir(out_dir / "__tmp_downloads")
        ctx_kwargs["downloads_path"] = str(out_dir / "__tmp_downloads")

    ctx: BrowserContext = await browser_type.launch_persistent_context(**ctx_kwargs)
    page: Page = await ctx.new_page()
    await page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    # open Meta AI once per account
    # await page.goto(account["meta_url"])

    for job in jobs:
        row_idx = job["row"]
        prompt  = job["video_cmd"]
        imagePath = job["image_path"]
        try:
            vid_path = await generate_video_meta_ai(page, imagePath, prompt, out_dir, _get_site_url(account, "meta", fallback="https://www.meta.ai/media/?nr=1"))
            write_video_result(row_idx, str(Path(vid_path).resolve()), account_id_used=account["id"], status="ok")
            print(f"[{account['id']}] Row {row_idx} -> {vid_path}")
            await asyncio.sleep(random.uniform(POLITE_MIN_WAIT, POLITE_MAX_WAIT))
            #await asyncio.sleep(10)
            #print("Waited 10 seconds before generating next image")
        except Exception as e:
            write_video_result(row_idx, "", account_id_used=account["id"], status=f"error: {e}")
            print(f"[{account['id']}] Row {row_idx} ERROR: {e}")

    await ctx.close()
    print(f"[{account['id']}] Images pass done.")

# =========================
# Videos pass (shell command or default)
# =========================

def run_shell(cmd: str):
    print("RUN:", cmd)
    completed = subprocess.run(cmd, shell=True)
    if completed.returncode != 0:
        raise RuntimeError(f"Command failed with code {completed.returncode}")

def build_default_cmd(image: str, out: str) -> str:
    return DEFAULT_VIDEO_CMD.format(image=image, out=out)

def derive_video_out(image_path: str) -> str:
    p = Path(image_path)
    return str(p.with_suffix(".mp4"))

def process_videos_from_excel():
    rows = read_jobs_from_excel_for_videos()
    for job in rows:
        row_idx  = job["row"]
        image    = job["image"]
        video_out = derive_video_out(image)
        cmd_tpl  = job["video_cmd"] or ""
        cmd      = (cmd_tpl if cmd_tpl.strip() else build_default_cmd(image, video_out))
        # Replace common placeholders
        cmd = (cmd
               .replace("{image}", image)
               .replace("{out}", video_out))

        try:
            run_shell(cmd)
            write_video_result(row_idx, str(Path(video_out).resolve()), account_id_used=VIDEO_ACCOUNT_ID, status="ok")
            print(f"[videos] Row {row_idx} -> {video_out}")
        except Exception as e:
            write_video_result(row_idx, "", account_id_used=VIDEO_ACCOUNT_ID, status=f"error: {e}")
            print(f"[videos] Row {row_idx} ERROR: {e}")

# =========================
# Coordinator
# =========================

async def main_async_videos():
    # refuse duplicate profiles
    profiles = [acc["profile"] for acc in META_ACCOUNTS]
    dupes = {p for p in profiles if profiles.count(p) > 1}
    if dupes:
        raise RuntimeError(f"Duplicate profile folders detected (profiles.json): {dupes}")

    attempt = 1

    while True:
        print(f"\n🎬 VIDEO PASS — Attempt {attempt}")   
        # read Excel rows (prompt + empty image_path)
        all_rows = read_jobs_from_excel_for_videos()
        if not all_rows:
            print("No rows need videos.")
            return

        # optional shuffle
        if SHUFFLE_PROMPTS:
            random.shuffle(all_rows)

        # partition rows per account:
        # if account_id present in Excel, bind to that account; else round-robin
        rows_per_account: Dict[str, List[Dict]] = {a["id"]: [] for a in META_ACCOUNTS}
        free_rows = []
        for r in all_rows:
            acct = r.get("account_id", "")
            if acct and acct in rows_per_account:
                rows_per_account[acct].append(r)
            else:
                free_rows.append(r)
        # round-robin free rows
        acc_ids = list(rows_per_account.keys())
        k = 0
        for r in free_rows:
            rows_per_account[acc_ids[k % len(acc_ids)]].append(r)
            k += 1

        # prune empties and limit per account if desired
        jobs = []
        for acc in META_ACCOUNTS:
            bucket = rows_per_account[acc["id"]]
            if MAX_META_VID_PROMPTS_PER_ACCOUNT:
                bucket = bucket[:MAX_META_VID_PROMPTS_PER_ACCOUNT]
            if bucket:
                jobs.append((acc, bucket))

        async with async_playwright() as pw:

            # Optional: run once to bootstrap logins for each profile (see BOOTSTRAP_LOGIN/BOOTSTRAP_SITES)
            if BOOTSTRAP_LOGIN:
                for acc in ACCOUNTS:
                    await bootstrap_profile_logins(pw, acc)


            await asyncio.gather(*[run_account_videos(pw, acc, bucket) for acc, bucket in jobs])

        if not ENABLE_RETRY or attempt >= MAX_RETRY_ATTEMPTS:
            print("⛔ Video retries exhausted.")
            return
        
        attempt += 1
        print(f"🔁 Retrying remaining video jobs in {RETRY_SLEEP_SECONDS}s...")
        await asyncio.sleep(RETRY_SLEEP_SECONDS)

async def run_chatgpt_images(pw, jobs: list[dict]):
    """
    Runs ALL ChatGPT image jobs sequentially
    in ONE persistent browser context.
    """

    account = CHATGPT_ACCOUNT
    out_dir = Path(account["out"])
    ensure_dir(out_dir)

    browser_type = getattr(pw, BROWSER_NAME)

    ctx = await browser_type.launch_persistent_context(
        user_data_dir=account["profile"],
        headless=HEADLESS,
        executable_path=CHROME_EXECUTABLE,
        device_scale_factor=DEVICE_SCALE_FACTOR,
        accept_downloads=True,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--start-maximized",
        ],
    )

    page = await ctx.new_page()
    await page.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )

    # Open ChatGPT Images ONCE
    await page.goto(account["url"], wait_until="networkidle")
    # await asyncio.sleep(505)
    for idx, job in enumerate(jobs, start=1):
        row_idx = job["row"]

        try:
            img_path = await generate_image_router(
                page=page,
                job=job,
                out_dir=out_dir
            )
            write_image_result(
                row_idx,
                str(Path(img_path).resolve()),
                account_id_used=account["id"],
                status="ok"
            )

            print(f"[ChatGPT] Row {row_idx} -> {img_path}")
            # await asyncio.sleep(8)  # polite delay
            # 45–75s between images + jitter
            # await asyncio.sleep(random.uniform(45, 75))

        except Exception as e:
            write_image_result(
                row_idx,
                "",
                account_id_used=account["id"],
                status=f"error: {e}"
            )
            print(f"[ChatGPT] Row {row_idx} ERROR: {e}")
        finally:
            # long cooldown every 10 images
            if idx % 50 == 0:
                print("Taking 30 mins break after 50 img...")
                await asyncio.sleep(random.uniform(30*60, 31*60))   # 30–31 minutes   
                await page.goto(account["url"], wait_until="domcontentloaded", timeout=120_000)
                # wait for the prompt box to be ready again
                #await page.locator('textarea#prompt-textarea, [contenteditable="true"][role="textbox"]').first.wait_for(timeout=60_000)       
            if idx % 10 == 0:
                print("Taking 10 mins break after 10 img...")
                await asyncio.sleep(600)  # 10 minutes   
            else:
                print("Taking 45-75s break before next image...")
                await asyncio.sleep(random.uniform(45, 75))
    await ctx.close()


async def main_async_images():
    # refuse duplicate profiles
    profiles = [acc["profile"] for acc in ACCOUNTS]
    dupes = {p for p in profiles if profiles.count(p) > 1}
    if dupes:
        raise RuntimeError(f"Duplicate profile folders detected (profiles.json): {dupes}")

    attempt = 1
    ENABLE_RETRY = False # Set to False to disable retries for images
    while True:
        print(f"\n🖼️ IMAGE PASS — Attempt {attempt}")
        
        # read Excel rows (prompt + empty image_path)
        all_rows = read_jobs_from_excel_for_images()
        chatgpt_rows, google_rows = split_image_jobs(all_rows)

        # if not google_rows:
        #     print("No rows need images.")
        #     return

        # # optional shuffle
        # if SHUFFLE_PROMPTS:
        #     random.shuffle(google_rows)

        # # partition rows per account:
        # # if account_id present in Excel, bind to that account; else round-robin
        # rows_per_account: Dict[str, List[Dict]] = {a["id"]: [] for a in ACCOUNTS}
        # free_rows = []
        # for r in google_rows:
        #     acct = r.get("account_id", "")
        #     if acct and acct in rows_per_account:
        #         rows_per_account[acct].append(r)
        #     else:
        #         free_rows.append(r)
        # # round-robin free rows
        # acc_ids = list(rows_per_account.keys())
        # k = 0
        # for r in free_rows:
        #     rows_per_account[acc_ids[k % len(acc_ids)]].append(r)
        #     k += 1

        # # prune empties and limit per account if desired
        # jobs = []
        # for acc in ACCOUNTS:
        #     bucket = rows_per_account[acc["id"]]
        #     if MAX_GGL_IMG_PROMPTS_PER_ACCOUNT:
        #         bucket = bucket[:MAX_GGL_IMG_PROMPTS_PER_ACCOUNT]
        #     if bucket:
        #         jobs.append((acc, bucket))

        # async with async_playwright() as pw:

        #     await asyncio.gather(*[run_account_images(pw, acc, bucket) for acc, bucket in jobs])

        async with async_playwright() as pw:

            # Optional: run once to bootstrap logins for each profile (see BOOTSTRAP_LOGIN/BOOTSTRAP_SITES)
            if BOOTSTRAP_LOGIN:
                for acc in ACCOUNTS:
                    await bootstrap_profile_logins(pw, acc)


            tasks = []

            # Google AI → parallel, multi-profile
            if google_rows:
                google_jobs = partition_rows_by_account(google_rows, ACCOUNTS)
                tasks.extend(
                    run_account_images(pw, acc, bucket)
                    for acc, bucket in google_jobs
                )

            # ChatGPT Images → SINGLE profile, sequential
            if chatgpt_rows:
                tasks.append(run_chatgpt_images(pw, chatgpt_rows))

            await asyncio.gather(*tasks)


        if not ENABLE_RETRY or attempt >= MAX_RETRY_ATTEMPTS:
            print("⛔ Image retries exhausted.")
            return
        
        attempt += 1
        print(f"🔁 Retrying remaining image jobs in {RETRY_SLEEP_SECONDS}s...")
        await asyncio.sleep(RETRY_SLEEP_SECONDS)


from typing import Dict, List, Tuple


def partition_rows_by_account(
    rows: List[dict],
    accounts: List[dict],
    max_per_account: int | None = None
) -> List[Tuple[dict, List[dict]]]:
    """
    Partitions Excel job rows across accounts.

    Rules:
    1) If row.account_id matches an account, it is pinned there
    2) Remaining rows are distributed round-robin
    3) Optional max_per_account cap is enforced
    4) Empty buckets are omitted

    Returns:
        List of (account, [rows]) tuples
    """

    # Initialize buckets
    rows_per_account: Dict[str, List[dict]] = {
        acc["id"]: [] for acc in accounts
    }

    free_rows = []

    # Pass 1: pinned rows
    for row in rows:
        acct = (row.get("account_id") or "").strip()
        if acct and acct in rows_per_account:
            rows_per_account[acct].append(row)
        else:
            free_rows.append(row)

    # Pass 2: round-robin free rows
    acc_ids = [acc["id"] for acc in accounts]
    idx = 0

    for row in free_rows:
        target_id = acc_ids[idx % len(acc_ids)]
        rows_per_account[target_id].append(row)
        idx += 1

    # Pass 3: apply max cap and emit result
    result: List[Tuple[dict, List[dict]]] = []

    for acc in accounts:
        bucket = rows_per_account[acc["id"]]

        if max_per_account is not None:
            bucket = bucket[:max_per_account]

        if bucket:
            result.append((acc, bucket))

    return result

def createImages():
    asyncio.run(main_async_images())

def createVideos():
    asyncio.run(main_async_videos())

def main():
    if RUN_MODE.lower() == "images":
        asyncio.run(main_async_images())
    elif RUN_MODE.lower() == "videos":
        asyncio.run(main_async_videos())
        
        # process_videos_from_excel()
    else:
        raise SystemExit("RUN_MODE must be 'images' or 'videos'.")

if __name__ == "__main__":
    main()


# -----------------------------------------------------------------------------
# QUICK REMINDER: Multi-profile setup + usage
#
# 1) Profiles are configured in `profiles.json` (auto-created on first run).
#    - Each profile = one persistent browser identity (cookies/logins saved)
#    - "profile_dir": stored under ./browser_profiles/<profile_dir>  (managed profiles)
#    - You can also use an absolute Chrome profile path if you want "legacy" profiles.
#
# 2) To add a NEW profile (scalable to 10+):
#    - Open `profiles.json`
#    - Duplicate an existing profile block
#    - Change:
#        id          -> unique name (e.g., "gmail_11")
#        profile_dir -> folder name (e.g., "gmail_11")  => creates ./browser_profiles/gmail_11
#        out         -> optional per-profile downloads folder (recommended)
#        sites       -> ONLY include services this profile should use (omit others to skip)
#          Example (AI Studio only):
#            "sites": { "aistudio": "https://aistudio.google.com/" }
#
# 3) Bootstrap logins (login once, saved forever in that profile folder):
#    - Set BOOTSTRAP_LOGIN = True
#    - Set BOOTSTRAP_SITES = ["aistudio", "claude", "grok", "elevenlabs", "heygen", "meta"]
#    - Run the script and login manually when the page opens.
#    - Press Enter in the terminal to move to the next site/profile.
#    - The script skips any site not listed in that profile’s "sites".
#    - After onboarding, set BOOTSTRAP_LOGIN = False for normal runs.
#
# 4) Normal runs:
#    - The script launches the requested profile, reuses saved cookies, and proceeds.
#    - IMPORTANT: Don’t run the same profile in two sessions at once (profile lock conflicts).
#
# 5) Folder notes:
#    - ./browser_profiles/  => persistent identities (safe to backup)
#    - ./downloads/<id>/    => per-profile downloads to avoid file collisions
# -----------------------------------------------------------------------------
