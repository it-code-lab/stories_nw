import json
import csv
import os
import re
import time
from pathlib import Path
from datetime import datetime
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import traceback

from pin_overlay_batch import create_youtube_thumbnail  # to print detailed error info

# DND - To Run
# python youtube_uploader.py

UPLOAD_URL = "https://studio.youtube.com"

PROFILE_DIR = "C:\\Users\\mail2\\AppData\\Local\\Google\\Chrome\\User Data\\Profile 1"
CHROME_EXECUTABLE = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"

# INPUT_JSON = "upload_config.json"
# UPLOAD_LOG = "upload_log.csv"

EXCEL_FILE = "video_records.xlsx"

# def load_videos():
#     with open(INPUT_JSON, 'r', encoding='utf-8') as f:
#         return json.load(f)

def load_videos_from_excel():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    videos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        # Skip completely empty rows
        if record.get("video_path") and record.get("youtube_title") and record.get("youtube_channel_name"):
            videos.append(record)
    return wb, ws, videos

def is_excel_file_locked(file_path):
    try:
        with open(file_path, 'a'):
            pass
        return False
    except PermissionError:
        return True


def save_video_status(ws, row_idx, url, status):
    ws[f'N{row_idx}'] = url
    ws[f'L{row_idx}'] = status
    ws[f'M{row_idx}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

# def init_csv():
#     if not os.path.exists(UPLOAD_LOG):
#         with open(UPLOAD_LOG, 'w', newline='', encoding='utf-8') as f:
#             writer = csv.writer(f)
#             writer.writerow(['title', 'channel_name', 'video_url', 'status'])


# def log_upload(title, channel_name, video_url, status):
#     with open(UPLOAD_LOG, 'a', newline='', encoding='utf-8') as f:
#         writer = csv.writer(f)
#         writer.writerow([title, channel_name, video_url, status])


def switch_channel(page, target_channel):
    try:
        # page.locator('img#img[alt*="Avatar image"]').wait_for(timeout=10000)
        # page.locator('img#img[alt*="Avatar image"]').click()

        print("Waiting for avatar button...")
        page.locator('#avatar-btn').wait_for(timeout=10000)
        page.locator('#avatar-btn').click()
        print("Avatar button clicked.")
        time.sleep(1)

        # Locate the element containing the text "Switch account" and click it
        print("Waiting for 'Switch account' text...")
        page.get_by_text("Switch account").click()
        print("'Switch account' clicked.")
        time.sleep(1)

        # Construct a CSS selector to find the channel title
        channel_selector = f"#contents ytd-account-item-renderer tp-yt-paper-icon-item yt-formatted-string#channel-title:has-text('{target_channel}')"

        print(f"Waiting for channel title element with selector: '{channel_selector}'")
        page.wait_for_selector(channel_selector, timeout=10000)
        channel_title_element = page.locator(channel_selector)
        print("Channel title element found.")

        # Get the parent tp-yt-paper-icon-item (the clickable area)
        channel_item = channel_title_element.locator("xpath=ancestor::tp-yt-paper-icon-item")

        # Scroll the parent element into view if necessary and then click
        print("Scrolling channel item into view...")
        channel_item.scroll_into_view_if_needed()
        print("Clicking channel item...")
        channel_item.click()
        print(f"Clicked on channel: {target_channel}")


        # found = False
        # for _ in range(5):
        #     accounts_popup = page.locator('yt-multi-page-menu-section-renderer')
        #     if accounts_popup.locator(f'tp-yt-paper-item:has-text("{target_channel}")').count() > 0:
        #         accounts_popup.locator(f'tp-yt-paper-item:has-text("{target_channel}")').click()
        #         found = True
        #         break
        #     page.mouse.wheel(0, 300)
        #     time.sleep(1)

        # if not found:
        #     raise Exception(f"Channel '{target_channel}' not found for switching.")
        
        time.sleep(3)  # Let the channel switch fully
    except Exception as e:
        print(f"Error switching channel: {e}")
        raise


def upload_video(page, video_info):
    print("Received upload_video Arguments:", locals())
    page.goto(UPLOAD_URL)
    page.wait_for_load_state("networkidle")
    time.sleep(2)
    print("Calling switch_channel")
    switch_channel(page, video_info["youtube_channel_name"])

    try:
        page.get_by_text("Dashboard").click()
    except:
        pass
    
    #DND - Was working
    #page.locator('ytcp-button#create-icon').click()
    try:
        page.locator('#upload-icon').click()
    except:
        page.locator('ytcp-button#create-icon').click()
        page.get_by_text("Upload video").click()

    # try:        
    #     page.get_by_text("Upload video").click()
    # except:
    #     try:
    #         page.get_by_text("Upload videos").click()
    #     except:
    #         page.locator("#upload-button").click()



    # page.locator('#endpoint').wait_for(timeout=10000)
    # page.locator("#endpoint").click()


    # page.locator('tp-yt-paper-item[role="option"]:has-text("Upload videos")').wait_for(state="visible", timeout=10000)
    # page.locator('tp-yt-paper-item[role="option"]:has-text("Upload videos")').click()

    # Upload video file
    # Upload video file
    file_input = page.locator('input[type="file"]')
    raw_path = video_info["video_path"]

    # 1) If it's already an absolute path and exists, use it
    if os.path.isabs(raw_path) and os.path.exists(raw_path):
        absolute_video_path = raw_path
    else:
        # 2) Try treating it as a relative path (like "pinterest_pins\xyz.mp4")
        rel_candidate = os.path.abspath(raw_path)
        if os.path.exists(rel_candidate):
            absolute_video_path = rel_candidate
        else:
            # 3) Backward-compat: look under processed_videos folder
            video_filename = os.path.basename(raw_path)
            if not video_filename.lower().endswith(".mp4"):
                video_filename += ".mp4"

            processed_candidate = os.path.abspath(
                os.path.join("processed_videos", video_filename)
            )
            if not os.path.exists(processed_candidate):
                raise FileNotFoundError(
                    "Video file not found. Tried:\n"
                    f" - {rel_candidate}\n"
                    f" - {processed_candidate}"
                )

            absolute_video_path = processed_candidate

    print(f"Using video file: {absolute_video_path}")
    file_input.set_input_files(absolute_video_path)


    #file_input.set_input_files(os.path.abspath(video_info["video_path"]))

    # Fill Title
    # page.locator('textarea#title-textarea').wait_for(timeout=15000)
    # page.locator('textarea#title-textarea').fill(video_info["title"])


    safe_title = video_info["youtube_title"][:100]

    page.get_by_label("Add a title that describes your video (type @ to mention a channel)").fill(safe_title)
    print("Title Entered")

    description = video_info["youtube_description"]
    # Normalize description: remove leading/trailing blank lines, collapse multiple newlines
    description = re.sub(r'\n\s*\n+', '\n', description.strip())

    page.get_by_label("Tell viewers about your video (type @ to mention a channel)").fill(description)
    print("Description Entered")

    time.sleep(2)
    page.keyboard.press("Escape")
    time.sleep(1)

    if video_info.get("size", "").lower() == "landscape":
        thumb_path = video_info.get("thumbnail_path")
        if thumb_path:
            absolute_thumbnail_path = os.path.abspath(thumb_path)
            try:
                page.locator('input#file-loader').set_input_files(absolute_thumbnail_path)
                print("✅ Thumbnail uploaded successfully for landscape video.")
            except Exception as e:
                print(f"⚠️ Thumbnail upload skipped (element not found or disabled): {e}")
    else:
        print("ℹ️ Skipping thumbnail upload for portrait/Shorts video.")
    
    time.sleep(1)

    # Upload thumbnail image - DND working but leaving the file selector open
    # page.locator('button:has-text("Upload File")').click()
    # thumbnail_path = video_info["thumbnail_path"]
    # absolute_thumbnail_path = os.path.abspath(thumbnail_path)
    # file_input = page.locator('input[type="file"]')
    # file_input.set_input_files(absolute_thumbnail_path)
    # time.sleep(2)

    # Add to Playlist
    if "youtube_playlist_name" in video_info and video_info["youtube_playlist_name"]:
        #page.locator('ytcp-button:has-text("Select playlist")').click()
        
        #page.locator('button:has-text("Select")').click()
        try:
            page.locator(".right-container.style-scope.ytcp-dropdown-trigger").click()
            time.sleep(2)
            page.get_by_text(video_info["youtube_playlist_name"]).locator("xpath=ancestor::label").click()
            #page.locator(f'div[role="checkbox"]:has-text("{video_info["playlist_name"]}")').click()
            page.locator('ytcp-button:has-text("Done")').click()
            print("Playlist selected")
        except Exception as e:
            print(f"Error selecting playlist: {e}")
            time.sleep(1)
            #press escape to close playlist dialog
            page.keyboard.press("Escape")


    # Fill Description
    #page.locator('textarea#description-textarea').fill(video_info["description"])

    # Made for kids
    kids_flag = video_info.get("made_for_kids", False)
    if kids_flag:
        page.locator('tp-yt-paper-radio-button[name="VIDEO_MADE_FOR_KIDS_MFK"]').click()
        #page.locator('tp-yt-paper-radio-button[name="made_for_kids"]').click()
        #page.get_by_role("radio", name="VIDEO_MADE_FOR_KIDS_MFK").click()
    else:
        page.locator('tp-yt-paper-radio-button[name="VIDEO_MADE_FOR_KIDS_NOT_MFK"]').click()
        #page.locator('tp-yt-paper-radio-button[name="not_for_kids"]').click()
        #page.get_by_role("radio", name="VIDEO_MADE_FOR_KIDS_NOT_MFK").click()

    print("Kids option entered")



    # Show More
    #page.locator('ytcp-button:has-text("Show more")').click()
    page.locator('button:has-text("Show more")').click()

    time.sleep(1)

    #page.get_by_role("radio", name="VIDEO_HAS_ALTERED_CONTENT_NO").click()
    page.locator('tp-yt-paper-radio-button[name="VIDEO_HAS_ALTERED_CONTENT_NO"]').click()
    print("Altered content option entered")

    raw_tags = video_info.get("youtube_tags")

    if raw_tags is None:
        tags = ""
    elif isinstance(raw_tags, list):
        # join list into comma-separated string (or space-separated if you prefer)
        tags = ", ".join([str(t).strip() for t in raw_tags if str(t).strip()])
    else:
        tags = str(raw_tags).strip()

    tags = tags[:450]


    if tags:
        tags_input = page.locator('input[placeholder="Add tag"]')
        tag_list = [tag.strip() for tag in tags.split(",") if tag.strip()]
        # tags_input.fill(",".join(tags))
        truncated_tags = []
        current_length = 0

        if tag_list:
            current_length = 0
            for tag in tag_list:
                if current_length + len(tag) + 1 <= 500:
                    truncated_tags.append(tag)
                    current_length += len(tag) + 1
                else:
                    break
            tags_input.fill(",".join(truncated_tags))

        if truncated_tags:
            tags_input.fill(",".join(truncated_tags))
            print(f"Entered {len(truncated_tags)} tags (limited to 500 characters).")
        else:
            print("No tags could be added within the 500 character limit.")

    # Add Thumbnail if available
    if "thumbnail_path" in video_info and video_info["thumbnail_path"]:
        thumbnail_input = page.locator('input#file-loader')
        thumbnail_input.set_input_files(os.path.abspath(video_info["thumbnail_path"]))
        time.sleep(1)


    # Next Steps
    for _ in range(3):
        #page.locator('button:has-text("Next")').click()
        next_button = page.locator('button:has-text("Next")')
        next_button.wait_for(state="visible", timeout=60000)  # wait for visible
        wait_until_enabled(next_button, timeout=60)           # wait for enabled
        if"schedule_date" in video_info and video_info["schedule_date"]:
            time.sleep(2)  # Extra wait if scheduling
        next_button.click()
        time.sleep(1)

    # Schedule or Publish
    if "schedule_date" in video_info and video_info["schedule_date"]:

        schedule_date_value = video_info["schedule_date"]

        page.locator('#visibility-title:has-text("Schedule")').click()
        time.sleep(1)
        page.locator('ytcp-dropdown-trigger').click()
        page.get_by_label('Enter date').get_by_label('').press('Control+a')


        if isinstance(schedule_date_value, datetime):
            schedule_date_value = schedule_date_value.strftime('%B %d, %Y')  # Correct for YouTube
        else:
            # Try parsing string if not already datetime
            try:
                parsed_date = datetime.strptime(schedule_date_value, '%Y-%m-%d')
                schedule_date_value = parsed_date.strftime('%B %d, %Y')
                
            except:
                # Already in proper format or wrong format (let's try using it as-is)
                pass


        page.get_by_label('Enter date').get_by_label('').fill(schedule_date_value)

        print(f"Entered schedule date: {schedule_date_value}")

        page.locator('tp-yt-iron-overlay-backdrop').nth(2).click();
        # time.sleep(20) 
        page.locator(
            'span.progress-label:has-text("Checks complete")'
        ).wait_for(state="visible", timeout=500_000)

        page.locator('ytcp-button:has-text("Schedule")').click()
        print("Video scheduled successfully.")
    else:
        page.locator('tp-yt-paper-radio-button[name="PUBLIC"]').click()
        page.locator('ytcp-button:has-text("Publish")').click()
        print("Video published successfully.")

    # Get video URL
    time.sleep(5)
    try:
        video_link = page.locator('a.ytcp-video-info').get_attribute('href')
        if video_link:
            #full_url = f"https://www.youtube.com{video_link}"
            return video_link
    except:
        pass

    return "Unknown"

def wait_until_enabled(locator, timeout=60):
    """Wait until the locator becomes enabled, with timeout in seconds."""
    elapsed = 0
    while elapsed < timeout:
        if locator.is_enabled():
            return
        time.sleep(1)
        elapsed += 1
    raise Exception(f"Timeout: Element {locator} was not enabled after {timeout} seconds.")


def load_youtube_rows_from_master(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    # headers = [cell.value for cell in ws[1]]
    # header_map = {name: idx + 1 for idx, name in enumerate(headers)}

    headers = [cell.value for cell in ws[1]]
    headers = [h if h is not None else "" for h in headers]

    headers, header_map, _ = ensure_youtube_status_columns(ws, headers)

    required = ["media_file", "yt_title", "yt_description", "youtube_status","youTubeChannel","avatar_img","size"]
    for r in required:
        if r not in header_map:
            raise Exception(f"Missing column: {r}")

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        media_file = ws.cell(row=row_idx, column=header_map["media_file"]).value
        status_val = ws.cell(row=row_idx, column=header_map["youtube_status"]).value
        future_val = ws.cell(row=row_idx, column=header_map["future"]).value
        media_type = ws.cell(row=row_idx, column=header_map["media_type"]).value
        youTubeChannel = ws.cell(row=row_idx, column=header_map["youTubeChannel"]).value
        youTubePlaylist = ws.cell(row=row_idx, column=header_map["yt_playlist"]).value
        youTubeScheduleDate = ws.cell(row=row_idx, column=header_map.get("yt_schedule_date", 0)).value
        avatar_img = ws.cell(row=row_idx, column=header_map.get("avatar_img")).value
        size = ws.cell(row=row_idx, column=header_map.get("size")).value
        
        if not youTubeChannel:
            continue

        if not media_file:
            continue

        if media_type and media_type != "video":
            continue

        if status_val == "success":
            continue

        if future_val == "future":
            continue

        title = ws.cell(row=row_idx, column=header_map["yt_title"]).value
        desc = ws.cell(row=row_idx, column=header_map["yt_description"]).value


        if not title or not desc:
            continue

        rows.append({
            "row_idx": row_idx,
            "media_file": media_file,
            "youtube_title": title,
            "youtube_description": desc,
            "youtube_tags": ws.cell(row=row_idx, column=header_map.get("yt_tags", 0)).value,
            "youtube_playlist": youTubePlaylist,
            "made_for_kids": False,
            "youtube_channel_name": youTubeChannel,
            "youtube_schedule_date": youTubeScheduleDate,
            "avatar_img": avatar_img,
            "size": size,

        })

    return wb, ws, header_map, rows

def ensure_youtube_status_columns(ws, headers):
    """
    Ensure these columns exist:
      - youtube_url
      - youtube_status
      - youtube_error
    """
    extra_cols = [
        "youtube_url",
        "youtube_status",
        "youtube_error"
    ]
    changed = False
    for col_name in extra_cols:
        if col_name not in headers:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers), value=col_name)
            changed = True

    header_map = {name: idx + 1 for idx, name in enumerate(headers)}
    return headers, header_map, changed

def save_youtube_status(ws, header_map, row_idx, url, error=None):
    ws.cell(row=row_idx, column=header_map["youtube_url"], value=url or "")
    # ws.cell(row=row_idx, column=header_map["youtube_last_attempt"], value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    if error:
        ws.cell(row=row_idx, column=header_map["youtube_status"], value="error")
        ws.cell(row=row_idx, column=header_map["youtube_error"], value=str(error)[:500])
    else:
        ws.cell(row=row_idx, column=header_map["youtube_status"], value="success")
        ws.cell(row=row_idx, column=header_map["youtube_error"], value="")


def upload_shorts_from_master_file():
    wb, ws, header_map, youtube_rows = load_youtube_rows_from_master("master_shorts_uploader_data.xlsx")

    if not youtube_rows:
        print("No valid rows found in Excel for YouTube upload.")
        return
    
    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            PROFILE_DIR,
            headless=False,
            executable_path=CHROME_EXECUTABLE
        )
        page = browser.new_page()

        for row in youtube_rows:
            # media =  os.path.join("pinterest_uploads", row["media_file"])
            media = row["media_file"]
            
            avatar_img = row.get("avatar_img")
            size = (row.get("size") or "").lower()
            print(f"Processing row {row['row_idx']}: media={media}, avatar_img={avatar_img}, size={size}")
            if avatar_img :
                # If row["youtube_title"] contains :, split into title and subhead
                title_parts = row["youtube_title"].split(":", 1)
                if len(title_parts) == 2:
                    title_text = title_parts[0].strip()
                    subhead_text = title_parts[1].strip()
                else:
                    # If row["youtube_title"] contains title_text (subtext), split them using ()
                    title_parts = row["youtube_title"].split("(", 1)
                    if len(title_parts) == 2:
                        title_text = title_parts[0].strip()
                        subhead_text = title_parts[1].strip().rstrip(")")
                    else:
                        title_text = row["youtube_title"]
                        subhead_text = ""


                # # If row["youtube_title"] contains title_text (subtext), split them using ()
                # title_parts = row["youtube_title"].split("(", 1)
                # if len(title_parts) == 2:
                #     title_text = title_parts[0].strip()
                #     subhead_text = title_parts[1].strip().rstrip(")")
                # else:
                #     title_text = row["youtube_title"]
                #     subhead_text = ""

                print(f"Using avatar image for thumbnail: {avatar_img}")
                create_youtube_thumbnail(
                    base_image_path=f"avatar_thumbnails/{avatar_img}.png",
                    json_template_path=f"avatar_thumbnails/{avatar_img}_thumbnail.json",
                    title_text=title_text,
                    subhead_text=subhead_text,
                    output_name="avatar_thumbnails/final_thumbnail.jpg"
                )
                thumbnail_path = "avatar_thumbnails/final_thumbnail.jpg"
            else:
                print("No valid avatar image found, proceeding without it.")
                thumbnail_path = "";

            print(f"\n=== Uploading to YouTube: {media} ===")

            if row["youtube_channel_name"] is None or row["youtube_channel_name"].strip() == "":
                print(f"Skipping row {row['row_idx']} due to missing youTubeChannel.")
                save_youtube_status(ws, header_map, row["row_idx"], "", "Missing youTubeChannel")
                continue

            try:
                video_info = {
                    "youtube_title": row["youtube_title"],
                    "youtube_description": row["youtube_description"],
                    "youtube_tags": row["youtube_tags"],
                    "youtube_playlist_name": row["youtube_playlist"],
                    "video_path": media,
                    "made_for_kids": False,
                    "youtube_channel_name": row["youtube_channel_name"],
                    "size": size,
                    "schedule_date": row.get("youtube_schedule_date"),
                    "thumbnail_path": thumbnail_path,
                }

                url = upload_video(page, video_info)
                save_youtube_status(ws, header_map, row["row_idx"], url)

            except Exception as e:
                traceback.print_exc()
                save_youtube_status(ws, header_map, row["row_idx"], "", str(e))

    wb.save("master_shorts_uploader_data.xlsx")
    print("All YouTube uploads done.")


def upload_videos():
    # videos = load_videos()
    # init_csv()

    if is_excel_file_locked(EXCEL_FILE):
        print(f"Error: Please close '{EXCEL_FILE}' before running the uploader.")
        return  # Exit cleanly without trying uploads

    wb, ws, videos = load_videos_from_excel()

    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            PROFILE_DIR,
            headless=False,
            executable_path=CHROME_EXECUTABLE,
            args=["--disable-blink-features=AutomationControlled", "--start-maximized"],
        )
        page = browser.new_page()
        #page.set_viewport_size({"width": 1920, "height": 1080})

        # Hide automation fingerprint
        page.add_init_script("""Object.defineProperty(navigator, 'webdriver', {get: () => undefined})""")

        # for video in videos:
        #     try:
        #         print(f"\n=== Uploading: {video['title']} to {video['channel_name']} ===")
        #         video_url = upload_video(page, video)
        #         print(f"Uploaded successfully: {video_url}")
        #         log_upload(video["title"], video["channel_name"], video_url, "Success")
        #     except Exception as e:
        #         print(f"Error uploading {video['title']}: {e}")
        #         log_upload(video["title"], video["channel_name"], "", "Failed")


        # DND -  Working for top to bottom
        # for idx, video in enumerate(videos, start=2):
        #     if str(video.get("youtube_upload_status", "")).strip().lower() == "success":
        #         print(f"Skipping already uploaded: {video['youtube_title']}")
        #         continue

        #     try:
        #         print(f"\n=== Uploading: {video['youtube_title']} to {video['youtube_channel_name']} ===")
        #         video_url = upload_video(page, video)
        #         save_video_status(ws, idx, video_url, "Success")
        #     except Exception as e:
        #         print(f"Error uploading {video['youtube_title']}: {e}")
        #         save_video_status(ws, idx, "", str(e)[:500])  # Save first 200 characters of error

        # Start from the last row and move up
        for rev_idx, video in enumerate(reversed(videos), start=1):
            # Calculate actual Excel row number (considering header in row 1)
            idx = len(videos) - rev_idx + 2  # +2 because Excel rows start from 2

            if str(video.get("youtube_upload_status", "")).strip().lower() == "success":
                print(f"Skipping already uploaded: {video['youtube_title']}")
                continue

            try:
                print(f"\n=== Uploading: {video['youtube_title']} to {video['youtube_channel_name']} ===")
                video_url = upload_video(page, video)
                save_video_status(ws, idx, video_url, "Success")
            except Exception as e:
                print(f"Error uploading {video['youtube_title']}: {e}")
                save_video_status(ws, idx, "", str(e)[:500])

        wb.save(EXCEL_FILE)
        browser.close()


if __name__ == "__main__":
    # upload_videos()
    upload_shorts_from_master_file()
