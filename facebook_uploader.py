# facebook_uploader.py

import os
import time
import re
from pathlib import Path
from datetime import datetime

from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PWTimeoutError
from openpyxl import load_workbook
import subprocess
import json

# ================== CONFIG ==================

# Start on the Facebook home feed (we’ll open the composer from here)
FACEBOOK_UPLOAD_URL = "https://www.facebook.com/reels/create"

# Re-use the same Chrome profile you use for Pinterest/TikTok
PROFILE_DIR = r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 21"
CHROME_EXECUTABLE = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

# Re-use the same Excel for now
EXCEL_FILE = "master_shorts_uploader_data.xlsx"

BASE_DIR = Path(__file__).resolve().parent
# MEDIA_BASE = BASE_DIR / "pinterest_uploads"  # where media_file paths are relative to
MEDIA_BASE = BASE_DIR

VIDEO_EXTS = {".mp4", ".mov", ".mkv", ".webm"}




def focus_fb_composer(page, timeout=60000):
    # 1) Prefer textbox by role with partial/regex match
    candidates = [
        lambda: page.get_by_role("textbox", name=re.compile(r"What's on your mind", re.I)),
        lambda: page.get_by_role("textbox", name=re.compile(r"Write something", re.I)),
        # 2) Strong structural selector: contenteditable textbox
        lambda: page.locator('div[contenteditable="true"][role="textbox"]').first,
        # 3) aria-placeholder fallback
        lambda: page.locator('div[contenteditable="true"][aria-placeholder]').first,
    ]

    last_err = None
    for get in candidates:
        try:
            loc = get()
            loc.wait_for(state="visible", timeout=timeout)
            loc.click()
            return loc
        except Exception as e:
            last_err = e

    raise last_err

# ================== EXCEL HELPERS ==================



def is_reel_eligible(video_path):
    cmd = [
        "ffprobe", "-v", "error",
        "-select_streams", "v:0",
        "-show_entries", "stream=width,height,duration",
        "-of", "json",
        video_path
    ]
    out = subprocess.check_output(cmd).decode()
    data = json.loads(out)["streams"][0]

    width = data["width"]
    height = data["height"]
    duration = float(data.get("duration", 0))
    size_mb = os.path.getsize(video_path) / (1024 * 1024)

    aspect = width / height

    return (
        duration <= 90
        and height > width
        and 0.54 <= aspect <= 0.60
        and size_mb < 1000
        and video_path.lower().endswith(".mp4")
    )

def ensure_facebook_status_columns(ws, headers):
    """
    Ensure these columns exist:
      - facebook_video_url
      - facebook_upload_status
      - facebook_uploaded_at
      - facebook_caption   (optional override per row)
    """
    extra_cols = [
        "facebook_video_url",
        "facebook_upload_status",
        "facebook_uploaded_at",
        "facebook_caption",
    ]
    changed = False
    for col_name in extra_cols:
        if col_name not in headers:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers), value=col_name)
            changed = True

    header_map = {name: idx + 1 for idx, name in enumerate(headers)}
    return headers, header_map, changed


def load_videos_from_excel():
    """
    Load rows from EXCEL_FILE that have a video in 'media_file'.

    Returns: wb, ws, rows, header_map
    """
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    headers = [h if h is not None else "" for h in headers]

    headers, header_map, _ = ensure_facebook_status_columns(ws, headers)

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)}
        record["_row_idx"] = row_idx

        media_file = (record.get("media_file") or "").strip()
        if not media_file:
            continue

        ext = Path(media_file).suffix.lower()
        media_type = (record.get("media_type") or "").strip().lower()

        faceBookProfile = (record.get("faceBookProfile") or "").strip();
        if faceBookProfile == "":
            continue
        
        # Only take video rows
        # if media_type and media_type != "video":
        #     continue
        # if ext and ext not in VIDEO_EXTS:
        #     continue

        rows.append(record)

    return wb, ws, rows, header_map


def save_facebook_status(ws, header_map, row_idx, url, status, caption=""):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    url_col = header_map.get("facebook_video_url")
    status_col = header_map.get("facebook_upload_status")
    ts_col = header_map.get("facebook_uploaded_at")
    cap_col = header_map.get("facebook_caption")

    if url_col:
        ws.cell(row=row_idx, column=url_col, value=url)
    if status_col:
        ws.cell(row=row_idx, column=status_col, value=status)
    if ts_col:
        ws.cell(row=row_idx, column=ts_col, value=now_str)
    if cap_col and caption:
        ws.cell(row=row_idx, column=cap_col, value=caption)


def is_excel_file_locked(file_path):
    try:
        with open(file_path, "a"):
            pass
        return False
    except PermissionError:
        return True


# ================== FACEBOOK HELPERS ==================


def resolve_media_path(media_file: str) -> str:
    """
    Resolve media_file path relative to MEDIA_BASE.
    """
    rel = media_file.replace("\\", "/")
    full_path = (MEDIA_BASE / rel).resolve()
    return str(full_path)


def open_facebook_home(page):
    page.goto(FACEBOOK_UPLOAD_URL)
    # page.wait_for_load_state("networkidle")
    time.sleep(3)


def open_video_composer(page):
    """
    Open the 'Create post' / 'Reel' composer.

    IMPORTANT:
    These selectors are generic and might need small tweaks with Playwright inspector
    depending on your Facebook layout and language.
    """
    # 1) Try to click a "Reel" button (for Reels upload)
    try:
        reel_btn = page.get_by_role("button", name=re.compile(r"Reel", re.I)).first
        reel_btn.click()
        time.sleep(3)
        return
    except Exception:
        pass

    # 2) Fallback: open a normal post with "Photo/video"
    try:
        # Click the "What's on your mind" composer area
        composer = page.get_by_role("button", name=re.compile("What's on your mind", re.I)).first
        composer.click()
        time.sleep(2)
    except Exception:
        # Sometimes it's a div with that text instead of button
        try:
            page.get_by_text("What's on your mind", exact=False).first.click()
            time.sleep(2)
        except Exception as e:
            raise RuntimeError(
                f"Could not open Facebook composer. Update selectors in open_video_composer(). Details: {e}"
            )

    # Click "Photo/video" inside the composer
    try:
        pv_btn = page.get_by_text(re.compile(r"Photo/video", re.I)).first
        pv_btn.click()
        time.sleep(2)
    except Exception as e:
        raise RuntimeError(f"Could not find 'Photo/video' button. Update selectors. Details: {e}")


def upload_video(page, media_path: str):
    """
    Upload video file in the composer.

    We target input[type=file] which Facebook uses under the hood.
    """
    print(f"  -> Uploading video: {media_path}")

    file_input = page.locator('input[type="file"][accept*="video"]')
    if file_input.count() == 0:
        file_input = page.locator('input[type="file"]')

    count = file_input.count()
    if count == 0:
        raise RuntimeError("Could not find video upload input on Facebook page.")

    target = file_input.first
    if count > 1:
        for i in range(count):
            el = file_input.nth(i)
            if el.is_visible():
                target = el
                break

    target.set_input_files(media_path)

    # Give Facebook some time to process / generate thumbnail
    time.sleep(20)


def build_caption(row: dict) -> str:
    """
    Build a Facebook caption from row data if facebook_caption is not provided.
    Facebook supports long texts, but we’ll keep it under ~2,000 chars.
    """
    manual = (row.get("facebook_caption") or "").strip()
    if manual:
        caption = manual
    else:
        title = (row.get("pin_title") or "").strip()
        desc = (row.get("pin_description") or "").strip()
        tags = (row.get("pin_tags") or "").strip()

        # Use one main URL (prefer your site, then Amazon)
        url = (
            (row.get("pin_url_to_link") or "").strip()
            or (row.get("book_url") or "").strip()
        )

        parts = []
        if title:
            parts.append(title)
        if desc:
            parts.append(desc)
        if url:
            parts.append(url)
        if tags:
            parts.append(tags)

        caption = "\n\n".join(parts)

    if len(caption) > 2000:
        caption = caption[:1990].rstrip() + "…"

    return caption


def fill_caption(page, caption_text: str):
    """
    Fill the caption/description in the composer.

    We look for a contenteditable textbox – adjust the selectors once
    with Playwright inspector if needed.
    """
    caption_text = (caption_text or "").strip()
    if not caption_text:
        print("  [INFO] No caption provided; skipping caption fill.")
        return

    print("  -> Filling caption...")

    candidates = [
        # Reel composer often uses a generic contenteditable textbox
        'div[role="textbox"][contenteditable="true"]',
        # Fallbacks – ANY contenteditable textbox
        'div[contenteditable="true"][role="textbox"]',
        'div[contenteditable="true"]',
    ]

    editor = None
    for sel in candidates:
        loc = page.locator(sel)
        try:
            loc.first.wait_for(state="visible", timeout=5000)
            if loc.count() > 0:
                editor = loc.first
                break
        except Exception:
            continue

    if not editor:
        print("  [WARN] Could not find caption editor on Facebook page.")
        return

    editor.click()
    time.sleep(1)

    # Clear existing text if any
    try:
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
    except Exception:
        pass

    time.sleep(0.2)
    page.keyboard.type(caption_text, delay=5)


def click_post(page):
    """
    Click the 'Post' / 'Share now' / 'Share reel' button to publish.
    """
    print("  -> Clicking Post/Share...")

    labels = [
        r"Share now",
        r"Post",
        r"Share reel",
        r"Next",     # sometimes there's a Next then a final Share
    ]

    for pattern in labels:
        try:
            btn = page.get_by_role("button", name=re.compile(pattern, re.I)).first
            btn.wait_for(state="visible", timeout=10000)
            btn.click()
            time.sleep(5)
            # After clicking "Next", composer may show a final "Share" button
            if pattern.lower() == "next":
                continue
            return
        except Exception:
            continue

    raise RuntimeError("Could not find Post/Share button. Update selectors in click_post().")


def extract_post_url(page) -> str:
    """
    After posting, try to grab the post URL.

    Often FB stays on the feed, but sometimes you get redirected
    to the post. We try current URL and fall back to any link
    containing '/posts/' or '/reel/'.
    """
    url = page.url
    if "/posts/" in url or "/reel/" in url:
        return url

    try:
        link = page.locator('a[href*="/posts/"], a[href*="/reel/"]').first
        href = link.get_attribute("href") if link else None
        if href:
            if href.startswith("http"):
                return href
            else:
                return "https://www.facebook.com" + href
    except Exception:
        pass

    return "Unknown"


# ================== MAIN UPLOADER ==================


def upload_facebook_videos():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file not found: {EXCEL_FILE}")
        return

    if is_excel_file_locked(EXCEL_FILE):
        print(f"Error: Please close '{EXCEL_FILE}' before running the uploader.")
        return

    wb, ws, rows, header_map = load_videos_from_excel()

    if not rows:
        print("No valid video rows found in Excel for Facebook upload.")
        return

    print(f"Loaded {len(rows)} Facebook rows from {EXCEL_FILE}")

    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            PROFILE_DIR,
            headless=False,
            executable_path=CHROME_EXECUTABLE,
            args=["--disable-blink-features=AutomationControlled", "--start-maximized"],
        )
        page = browser.new_page()

        # Hide automation fingerprint
        page.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )

        # You should already be logged in to Facebook in this Chrome profile.
        profileLoaded  = False
        # total = len(rows)
        for row in reversed(rows):
            # row_idx = total - rev_idx + 2  # Excel row index (header is row 1)
            row_idx = row["_row_idx"]  

            status_val = str(row.get("facebook_upload_status") or "").strip().lower()
            future_val = str(row.get("future") or "").strip().lower()
            if status_val == "success":
                print(f"Skipping already uploaded Facebook video: {row.get('pin_title')}")
                continue

            if future_val == "future":
                print(f"Skipping future Facebook video row: {row.get('pin_title')}")
                continue

            if row["faceBookProfile"] == "":
                print(f"Skipping Facebook upload as no profile specified in row {row_idx}.")
                continue
            try:
                print(f"\n=== Facebook upload for media: {row.get('media_file')} (row {row_idx}) ===")

                page.goto("https://www.facebook.com/")
                # page.wait_for_load_state("networkidle")
                time.sleep(3)
                # open_facebook_home(page)

                if not profileLoaded:
                    faceBookProfile = row["faceBookProfile"]
                    page.get_by_label("Your profile").first.click(timeout=15000)
                    time.sleep(5)

                    try:
                        page.locator(f'span:has-text("{faceBookProfile}")').first.click()
                        time.sleep(5)
                    except:
                        try:
                            page.get_by_role("button", name="See all profiles").click()
                            time.sleep(5)
                            page.locator(f'span:has-text("{faceBookProfile}")').first.click()
                            time.sleep(5)
                            profileLoaded = True
                        except:
                            page.locator('span:has-text("See more profiles")').first.click()
                            time.sleep(5)
                            page.locator(f'span:has-text("{faceBookProfile}")').first.click()
                            time.sleep(5)

                    profileLoaded = True



                media_path = resolve_media_path(str(row["media_file"]))
                if not os.path.exists(media_path):
                    raise FileNotFoundError(f"Media file not found: {media_path}")

                if is_reel_eligible(media_path):
                    page.goto("https://www.facebook.com/reels/create")
                    time.sleep(5)

                    print("Eligible for Reel. Reels creation page loaded.")

                    # --- STEP 2: FILE UPLOAD ---
                    print(f"Uploading file: {media_path}")

                    # Wait for the main file input element to be available on the page
                    # file_input_locator = page.locator('input[type="file"]')
                    # file_input_locator.wait_for(state="attached", timeout=30000)                
                    # file_input_locator.set_input_files(media_path)

                    with page.expect_file_chooser() as fc_info:
                        page.get_by_role("button", name="Add video or drag and drop").click()
                    
                    file_chooser = fc_info.value
                    file_chooser.set_files(media_path)

                    print("File selected for upload; Clicking first 'Next' button....")
                    # Wait for the first 'Next' button to appear after upload
                    page.get_by_role("button", name="Next").wait_for(state="visible", timeout=5000)
                    # --- STEP 2: WIZARD STEPS (Two Next Clicks) ---
                    
                    # 1st Click: Usually for cropping/trimming
                    print("Clicking second 'Next' button...")
                    page.get_by_role("button", name="Next").click()
                    time.sleep(5)

                    # 2nd Click: Usually for editing/enhancements
                    print("Clicking third 'Next' button...")
                    page.get_by_role("button", name="Next").click()
                    time.sleep(5)

                    print("Getting paragraph to fill caption...")
                    para = page.get_by_role("paragraph")
                    para.wait_for(state="visible", timeout=60000)
                    para.click()
                    print("Para clicked for caption.")
                    time.sleep(5)

                    # caption = build_caption(row)
                    caption = row["fb_caption"]


                    # Some Facebook editors don't support .fill() directly; type instead:
                    # Clear existing text (just in case)
                    try:
                        # Try select-all + delete (Ctrl+A, Backspace)
                        page.keyboard.press("Control+A")
                        page.keyboard.press("Backspace")
                        time.sleep(0.5)
                    except Exception:
                        pass

                    # para.type(caption, delay=20)
                    page.keyboard.type(caption, delay=5)
                    print("Entered caption. Going to sleep 5 seconds. Not going to press Escape to defocus...")

                    
                    # time.sleep(5) # Small pause for "human" typing

                    # page.keyboard.press("Escape")

                    # print("Defocused caption area. Waiting 5 seconds before publishing...")
                    time.sleep(5)

                    # --- STEP 4: SHARE THE REEL ---
                    
                    print("  -> Clicking Publish/Post button...")

                    try:
                        publish_btn = page.get_by_role("button", name="Post").first
                        publish_btn.wait_for(state="visible", timeout=30000)
                        publish_btn.click()
                    except PWTimeoutError:
                        print("Post button not found; trying 'Publish' button instead...")
                        publish_btn = page.get_by_role("button", name="Publish").first
                        publish_btn.wait_for(state="visible", timeout=30000)
                        publish_btn.click()
                    time.sleep(8)  # give time for the post to be submitted
                    
                    print("✅ Upload Successful! Your Reel is being shared.")


                else:
                    page.goto("https://www.facebook.com")
                    time.sleep(5)

                    with page.expect_file_chooser() as fc_info:
                        page.get_by_role("button", name="Photo/video").click()
                    
                    file_chooser = fc_info.value
                    file_chooser.set_files(media_path)
                    print("File selected for upload. Sleeping 5 seconds for upload processing...")
                    time.sleep(5)  # wait for upload processing
                    print("Finding textbox to fill caption...")
                    # composer = focus_fb_composer(page, timeout=60000)
                    composer = page.locator('div[contenteditable="true"][role="textbox"]').first
                    composer.click()
                    print("Composer clicked.")

                    time.sleep(5)

                    # Clear existing text (if any)
                    page.keyboard.press("Control+A")
                    page.keyboard.press("Backspace")
                    time.sleep(5)

                    # caption = build_caption(row)
                    caption = row["fb_caption"]

                    page.keyboard.type(caption, delay=5)
                    print("Entered caption.")

                    print("Clicking first 'Next' button...")
                    page.get_by_role("button", name="Next").click()
                    time.sleep(5)
                    print("  -> Clicking Post button...")
                    time.sleep(5)
                    publish_btn = page.get_by_role("button", name="Post").first
                    publish_btn.wait_for(state="visible", timeout=60000)
                    publish_btn.click()
                    time.sleep(8)  # give time for the post to be submitted

                post_url = ""
                save_facebook_status(ws, header_map, row_idx, post_url, "Success", caption)
                print(f"✅ Facebook video posted & recorded for row {row_idx}: {post_url}")
            except Exception as e:
                err_msg = str(e)[:500]
                print(f"❌ Error uploading Facebook video for row {row_idx}: {err_msg}")
                save_facebook_status(ws, header_map, row_idx, "", err_msg)

            # Gentle pause between posts
            time.sleep(5)

        wb.save(EXCEL_FILE)
        browser.close()

    print("All Facebook uploads done.")


if __name__ == "__main__":
    upload_facebook_videos()
