#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import difflib
import shutil
from bs4 import BeautifulSoup, Tag
from moviepy.editor import VideoFileClip, TextClip, CompositeVideoClip
import requests
import whisper
import traceback
import json
#from scraper import safe_copy
from settings import sizes, background_music_options, font_settings
import csv
from datetime import datetime
from openpyxl import Workbook, load_workbook
#from transformers import pipeline

MAX_WORD_DURATION = 2.0  # seconds

# Extract Audio from Video
def extract_audio(video_path, audio_path):
    video = VideoFileClip(video_path)
    video.audio.write_audiofile(audio_path)

# Normalize Text with Enhanced Rules
def normalize_text(text, language="english"):
    text = text.lower()

    if language == "hindi":
        # Remove punctuation except Hindi characters and digits
        text = re.sub(r"[^\w\s\u0900-\u097F]", "", text)  # Keep Devanagari + digits + whitespace
    else:
        # English (default)
        text = re.sub(r"\s*-\s*", " ", text)
        text = re.sub(r"[^\w\s']", "", text)  # Keep letters, digits, apostrophes

    # Normalize spaces
    text = re.sub(r"\s+", " ", text).strip()
    return text

# Improved Subtitle Generation Logic
def generate_aligned_subtitles(audio_path, website_text, max_words):

    print("Received generate_aligned_subtitles Arguments:", locals())

    model = whisper.load_model("base")
    result = model.transcribe(audio_path, word_timestamps=True)
    grouped_subtitles = []
    current_caption = []
    current_start = None

    with open('temp/result.json', 'w') as f:
        json.dump(result, f,indent=4)

    for segment in result['segments']:
        for word_info in segment['words']:
            word = word_info['word']
            start = word_info['start']
            end = word_info['end']
            if current_start is None:
                current_start = start
            current_caption.append(word)
            if len(current_caption) >= max_words or word_info == segment['words'][-1]:
                caption_text = " ".join(current_caption)
                grouped_subtitles.append((current_start, end, caption_text))
                current_caption = []
                current_start = None

    #SM-DND - May be used in future
    # with open('temp/grouped_subtitles.json', 'w') as f:
    #     json.dump(grouped_subtitles, f,indent=4)

    return grouped_subtitles

# Enhanced Stylish Subtitle Creator
def create_stylish_subtitles(video, subtitles, style, fontsize, y_pos, font_settings):
    subtitle_clips = []
    style_options = font_settings[style]

    for start, end, text in subtitles:
        txt_clip = (
            TextClip(
                text,
                font=style_options['font'],
                fontsize=fontsize,
                color=style_options['color'],
                stroke_color="black",
                stroke_width=2,
                method="label",
                bg_color="rgba(0, 0, 0, 0.5)"
            )
            .set_position(("center", y_pos))
            .set_start(start)
            .set_duration(end - start)
        )
        subtitle_clips.append(txt_clip)

    return CompositeVideoClip([video] + subtitle_clips)

# Correct and Add Captions
import traceback

def add_captions(max_words, fontsize, y_pos, style, website_text, font_settings, input_video_path="final_video.mp4"):
    
    print("Received add_captions Arguments:", locals())

    audio_path = "audio.wav"
    output_video_path = "output_video.mp4"

    try:
        extract_audio(input_video_path, audio_path)
        print("Audio extracted successfully!")
    except Exception as e:
        print(f"Error extracting audio: {e}")
        traceback.print_exc()
        return
    
    try:
        audio_transcription = normalize_text(website_text)
        print("Text normalized successfully!")
    except Exception as e:
        print(f"Error normalizing text: {e}")
        traceback.print_exc()
        return
    
    try:
        # Improved Alignment with Website Text
        subtitles = generate_aligned_subtitles(audio_path, audio_transcription, max_words)
        print("Subtitles generated successfully!")
    except Exception as e:
        print(f"Error generating subtitles: {e}")
        traceback.print_exc()
        return
    
    try:
        # Create the Video with Captions
        video = VideoFileClip(input_video_path)
        styled_video = create_stylish_subtitles(video, subtitles, style, fontsize, y_pos, font_settings)
    except Exception as e:
        print(f"Error loading video: {e}")
        traceback.print_exc()
        return
    
    try:
        # Write output and ensure file closure
        styled_video.write_videofile(output_video_path, codec="libx264", audio_codec="aac")
    except Exception as e:
        print(f"Error writing video file: {e}")
        traceback.print_exc()
        return
    finally:
        # Ensure proper resource release
        styled_video.close()
        video.close()
    
    print("Captioning process completed successfully!")

def prepare_file_for_adding_captions_n_headings_thru_html(url, input_video_path="composed_video.mp4", base_file_name="output", language="english", story_text=""):
    
    #print("Received add_captions Arguments:", locals())

    audio_path = "audio.wav"


    try:
        extract_audio(input_video_path, audio_path)
        print("Audio extracted successfully!")
    except Exception as e:
        print(f"Error extracting audio: {e}")
        traceback.print_exc()
        return
    
    #model = whisper.load_model("base")
    #model = whisper.load_model("medium")

    if language == "hindi":
        # model = whisper.load_model("medium")
        # captions_data = model.transcribe(audio_path, word_timestamps=True, language="hi")
        model = whisper.load_model("large")
        captions_data = model.transcribe(
                            audio_path,
                            word_timestamps=True,
                            language="hi",     # or "en" for English
                            verbose=True,
                            fp16=False,         # Important on CPU
                            initial_prompt="यह एक कहानी है",
                            condition_on_previous_text=False,
                            temperature=(0.0, 0.2, 0.4)
                        )

    else:
        model = whisper.load_model("base")
        captions_data = model.transcribe(audio_path, word_timestamps=True)

    print("Detected language:", captions_data.get("language"))

    word_timestamps = []
    position_index = 0  # Track word positions

    for segment in captions_data["segments"]:
        for word_data in segment.get("words", []):
            word_timestamps.append({
                "word": word_data["word"].lower(),  # Normalize case
                "start": word_data["start"],
                "end": word_data["end"],
                "position": position_index,  # Assign position index
                "matched": False  # Initialize as not matched
            })
            position_index += 1

    for word in word_timestamps:
        start = word["start"]
        end = word["end"]
        if end - start > MAX_WORD_DURATION:
            word["end"] = start + MAX_WORD_DURATION

    with open('temp/word_timestamps.json', 'w', encoding='utf-8') as f:
        json.dump(word_timestamps, f,indent=4, ensure_ascii=False)

    shutil.copyfile('temp/word_timestamps.json', f"backup/word_timestamps_{base_file_name}.json")

    # Extract text with headings & list items
    if story_text != "":
        full_text = []
        position_index = 0
        element_type = "regular"

        words = re.findall(r'[\u0900-\u097F]+', story_text)


        for index, word in enumerate(words):
            full_text.append({
                "position": position_index,
                "word": word,
                "type": element_type,  # Regular text, heading, or list item
                "is_last_word": index == len(words) - 1  # True if it's the last word of the element
            })
            position_index += 1  # Move forward in word count

        with open('temp/full_text.txt', 'w', encoding='utf-8') as f:
            json.dump(full_text, f, indent=4, ensure_ascii=False)
    else:
        full_text = extract_full_text_with_positions(url)

    """Compare each word in sequence and set matched flag."""
    matched_results = []
    last_matched_index = 0  # Track the last matched word's position in word_timestamps

    i = 0  # Index for full_text

    SEARCH_WINDOW = 5  # Number of words after to check for a match
    """Compare each word in sequence and set matched flag, ensuring each word is mapped only once.
    After a match is found, the next word is looked up within the next SEARCH_WINDOW indices. If not found, mark as unmatched.
    """

    i = 0  # Manually control the index

    while i < len(full_text):  # Using full_text instead of splitting normalized_text
        word_data = full_text[i]
        word = word_data["word"]  # Extract word from full_text
        matched = False
        search_start = last_matched_index  # Start searching from the last matched index
        search_end = min(last_matched_index + SEARCH_WINDOW, len(word_timestamps))  # Search within next 10 indices

        for caption_index in range(search_start, search_end):
            cleaned_word_timestamps = normalize_text(word_timestamps[caption_index]["word"], language)

            if not word_timestamps[caption_index]["matched"] and cleaned_word_timestamps == word.strip():
                matched = True
                word_timestamps[caption_index]["matched"] = True
                last_matched_index = caption_index + 1  # Move the search window forward
                matched_results.append({
                    "position": word_timestamps[caption_index]["position"],
                    "word": word_timestamps[caption_index]["word"],
                    "start": word_timestamps[caption_index]["start"],
                    "end": word_timestamps[caption_index]["end"],
                    "matched": True,
                    "website_text_index": word_data["position"],  # Use full_text position
                    "type": word_data["type"]  # Store type for styling
                })
                break  # Move to the next website word

            # If current word didn't match, try words[i+1]
            elif i + 1 < len(full_text):
                cleaned_next_word = normalize_text(full_text[i+1]["word"], language)

                if not word_timestamps[caption_index]["matched"] and cleaned_word_timestamps == cleaned_next_word:
                    matched = True
                    word_timestamps[caption_index]["matched"] = True
                    last_matched_index = caption_index 

                    matched_results.append({
                        "position": None,  # No matching position
                        "word": full_text[i]["word"],
                        "start": None,
                        "end": None,
                        "matched": False,
                        "website_text_index": full_text[i]["position"],
                        "type": full_text[i]["type"]  # Store type for styling
                    })
                    matched_results.append({
                        "position": word_timestamps[caption_index]["position"],
                        "word": word_timestamps[caption_index]["word"],
                        "start": word_timestamps[caption_index]["start"],
                        "end": word_timestamps[caption_index]["end"],
                        "matched": True,
                        "website_text_index": full_text[i + 1]["position"],
                        "type": full_text[i+1]["type"]  # Store type for styling
                    })
                    i = i + 1  # Move to the next website word
                    break  # Stop looking once a match is found
        i += 1  # Move to the next word

        # If no match is found within the search window, mark as unmatched
        if not matched:
            matched_results.append({
                "position": None,  # No matching position
                "word": word,
                "start": None,
                "end": None,
                "matched": False,
                "website_text_index": word_data["position"],  # Use full_text position
                "type": word_data["type"]  # Store type for styling
            })
    with open('temp/matched_results.txt', 'w', encoding='utf-8') as f:
        json.dump(matched_results, f,indent=4, ensure_ascii=False)


    output_file_path = "temp/structured_output.json"


    # Find start & end word timings
    structured_output = find_timing_for_headings_list_items(full_text, matched_results)

    # Save structured output
    with open(output_file_path, "w", encoding="utf-8") as file:
        json.dump(structured_output, file, indent=4)

    shutil.copyfile('temp/structured_output.json', f"backup/structured_output_{base_file_name}.json")
    shutil.copyfile('composed_video.mp4', f"backup/composed_video_{base_file_name}.mp4")

    print(f"Headings & List Item Timings saved to: {output_file_path}")
    
    save_details_in_excel(captions_data, url, base_file_name)

    #DND - Working but not in use
    #save_details_in_csv(captions_data, url, base_file_name)

def save_details_in_excel(captions_data, url, base_file_name):
    """Safely write data to an Excel file, creating it if it doesn't exist."""
    excel_file = "video_records.xlsx"
    sheet_name = "Videos"

    # Only capture the simple text, not full captions_data
    captions_text = captions_data.get("text", "").strip()

    # Define columns
    columns = [
        'url',
        'video_path',
        'captions_text',
        'youtube_playlist_name',
        'youtube_title',
        'youtube_description',
        'youtube_tags',
        'youtube_channel_name',
        'video_is_about',
        'made_for_kids',
        'schedule_date',
        'youtube_upload_status',
        'last_update_date',
        'created_video_url'
    ]

    # Check if Excel file exists
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(columns)  # write header
        wb.save(excel_file)

    # Load workbook and sheet
    wb = load_workbook(excel_file)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
    else:
        ws = wb[sheet_name]

    # Prepare row data
    row = [
        url,
        base_file_name,
        captions_text,
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        'Pending',
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ''
    ]

    ws.append(row)
    wb.save(excel_file)

    print(f"Record for {base_file_name} added to {excel_file}")

#DND - Working but not in use
def save_details_in_csv(captions_data, url, base_file_name):
    """Safely write data to a CSV file, creating it if it doesn't exist."""
    csv_file = "video_records.csv"
    # Ensure captions_data is saved as JSON string (small, compact)
    captions_data_json = captions_data.get("text", "").strip()
    # Define standard fieldnames
    fieldnames = [
        'url',
        'video_path',
        'captions_text',
        'youtube_playlist_name',
        'youtube_title',
        'youtube_description',
        'youtube_tags',
        'youtube_channel_name',
        'video_is_about',
        'made_for_kids',
        'schedule_date',
        'youtube_upload_status',
        'last_update_date',
        'created_video_url'
    ]

    # Check if CSV exists
    file_exists = os.path.isfile(csv_file)

    with open(csv_file, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)

        # Write header if file is new
        if not file_exists:
            writer.writeheader()

        writer.writerow({
            'url': url,
            'video_path': base_file_name,
            'captions_text': captions_data_json,
            'youtube_playlist_name': '',
            'youtube_title': '',
            'youtube_description': '',
            'youtube_tags': '',
            'youtube_channel_name': '',
            'video_is_about': '',
            'made_for_kids': '',
            'schedule_date': '',
            'youtube_upload_status': 'Pending',
            'last_update_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'created_video_url': ''
        })

    print(f"Record for {base_file_name} added to {csv_file}")


def is_skippable(element):
    """
    Check if the element should be skipped for text extraction.
    """
    # Skip <span> only if it has onclick="toggleParentDetails(this)"
    if isinstance(element, Tag):  # Ensure it's a Tag before checking attributes
        if element.name == "span" and element.get("onclick") == "toggleParentDetails(this)":
            return True
        
        # Skip <button> elements
        if element.name == "button":
            return True
        
        # Skip <div> with specific classes
        skip_classes = {
            "audio-details", "image-props", "video-props", "video-listitem-props",
            "video-hdr-props", "audio-desc", "video1-desc", "image1-desc"
        }
        if element.name == "div" and any(cls in element.get("class", []) for cls in skip_classes):
            return True

    return False
            
def extract_full_text_with_positions(url):
    """Extracts full website text and marks positions of headings and list items."""

    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    full_text = []  # Store all words with their type (regular, heading, list_item)
    position_index = 0  # Track word positions

    for element in soup.select_one(".songLyrics").descendants:  # Capture text elements including headings and list items

        if isinstance(element, Tag):  # Only process HTML tags, ignore NavigableStrings directly
            if element.name in ("script", "style", "p", "button"):
                continue

            # Skip skippable elements
            if is_skippable(element):
                continue

        if isinstance(element, str) and element.strip():  # Process text nodes
            # Skip if any parent of this text node is skippable
            if not any(parent for parent in element.parents if isinstance(parent, Tag) and is_skippable(parent)):
                
                element_text = element.strip()
                words = re.findall(r'\b\w+\b', element_text.lower())  # Extract words

                # Find element type (only for Tags)
                element_type = "regular"
                if isinstance(element.parent, Tag):  # Check parent element instead of text node
                    if "video-hdr-inline-cls" in element.parent.get("class", []):
                        if "staying" in element.parent.get("class", []):
                            element_type = "staying-heading"
                        else:
                            element_type = "heading"
                    elif "video-listitem-inline-cls" in element.parent.get("class", []):
                        if "staying" in element.parent.get("class", []):
                            element_type = "staying-list-item"
                        else:
                            element_type = "list_item"

                for index, word in enumerate(words):
                    full_text.append({
                        "position": position_index,
                        "word": word,
                        "type": element_type,  # Regular text, heading, or list item
                        "is_last_word": index == len(words) - 1  # True if it's the last word of the element
                    })
                    position_index += 1  # Move forward in word count

    with open('temp/full_text.txt', 'w') as f:
        json.dump(full_text, f, indent=4)

    return full_text

def find_timing_for_headings_list_items(full_text, matched_results):
    """Finds start and end word positions for headings/list items using website_text_index."""
    structured_output = []
    i = 0

    while i < len(full_text):
        word_data = full_text[i]

        if word_data["type"] in ["heading", "list_item", "staying-heading", "staying-list-item"]:  # Process headings & list items
            start_position = word_data["position"]
            start_word = word_data["word"]

            # Group words until is_last_word is True
            grouped_words = [start_word]
            end_position = start_position

            j = i + 1
            while j < len(full_text) and not full_text[j - 1]["is_last_word"]:  # Continue until the last word of the element
                grouped_words.append(full_text[j]["word"])
                end_position = full_text[j]["position"]
                j += 1

            # 🔹 Find start timing using website_text_index
            start_timing = next((res["start"] for res in matched_results if res["website_text_index"] == start_position), None)

            # 🔹 Find end timing using website_text_index
            end_timing = next((res["end"] for res in matched_results if res["website_text_index"] == end_position), None)

            structured_output.append({
                "type": word_data["type"],
                "text": " ".join(grouped_words),  # Combine words into a full phrase
                "start_word_position": start_position,
                "start_word_start_timing": start_timing,
                "end_word_position": end_position,
                "end_word_end_timing": end_timing
            })

            i = j  # Skip to the next element
        else:
            i += 1  # Move forward normally

    return structured_output


#DND - Was working except when the two headings or list items are next to each other
def find_timing_for_headings_list_items_Old(full_text, matched_results):
    """Finds start and end word positions for headings/list items using website_text_index."""
    structured_output = []
    i = 0

    while i < len(full_text):
        word_data = full_text[i]
        current_type = word_data["type"]

        if current_type in ["heading", "list_item","staying-heading", "staying-list-item"]:  # Process headings & list items
            start_position = word_data["position"]
            start_word = word_data["word"]

            # Group words until type changes
            grouped_words = [start_word]
            end_position = start_position

            j = i + 1
            while j < len(full_text) and full_text[j]["type"] == current_type:  # Continue only if type matches
                grouped_words.append(full_text[j]["word"])
                end_position = full_text[j]["position"]
                j += 1

            # 🔹 Find start timing using website_text_index
            start_timing = next((res["start"] for res in matched_results if res["website_text_index"] == start_position), None)

            # 🔹 Find end timing using website_text_index
            end_timing = next((res["end"] for res in matched_results if res["website_text_index"] == end_position), None)

            structured_output.append({
                "type": current_type,
                "text": " ".join(grouped_words),  # Combine words into a full phrase
                "start_word_position": start_position,
                "start_word_start_timing": start_timing,
                "end_word_position": end_position,
                "end_word_end_timing": end_timing
            })

            i = j  # Skip to the next different type
        else:
            i += 1  # Move forward normally

    return structured_output


if __name__ == "__main__":

    # DND - Working    
    # input_video_path = "composed_video.mp4"
    # audio_path = "audio.wav"
    # output_video_path = "test_output_video.mp4"
    # file_path = "test_website_text.txt"
    # website_text = ""
    # max_words = 5
    # style = "Style 27"
    # fontsize = 90
    # y_pos = "bottom"
    # url = "https://readernook.com/topics/scary-stories/taxfiling"
    # try:
    #     with open(file_path, 'r', encoding='utf-8') as file:
    #         website_text =  file.read()
    # except Exception as e:
    #     print(f"Error reading file {file_path}: {e}")
    #     traceback.print_exc()

    #DND-Working
    #add_captions(max_words, fontsize, y_pos, style, website_text, font_settings, "composed_video.mp4")

    
    #prepare_file_for_adding_captions_n_headings_thru_html(url, "composed_video.mp4")
    prepare_file_for_adding_captions_n_headings_thru_html("", "composed_video.mp4", "output", "hindi", story_text="प्राचीन काल की बात है।भारतवर्ष के काशी नगर में एक ब्राह्मण रहता था।उसका जीवन अत्यंत कठिनाइयों से भरा हुआ था।घर में न चूल्हा जलता, न दीपक। न वस्त्र, न बर्तन।वह वृद्ध हो चला था, शरीर दुर्बल और आँखों में निरंतर चिंता का बादल।")
