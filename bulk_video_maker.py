from __future__ import annotations

import argparse
import os
import re
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from get_audio import get_audio_file

INPUT_EXCEL = "bulk_video_maker_input.xlsx"
INPUT_VIDEO_DIR = "edit_vid_input"
OUTPUT_VIDEO_DIR = "edit_vid_output"
GENERATED_AUDIO_DIR = "edit_vid_audio"
DEFAULT_BG_MUSIC_FILE = ""
DEFAULT_BG_MUSIC_VOLUME = 0.12
DEFAULT_LANGUAGE = "english"
DEFAULT_GENDER = "Male"
DEFAULT_TTS_ENGINE = "google"
DEFAULT_TTS_TYPE = "neural"
BG_MUSIC_FADE_OUT_SECONDS = 2.5

#Below status colums are added to the excel file if not already present, and used to track the processing status and results for each row. They can be customized as needed, but should match the constants defined here for the code to work correctly.
STATUS_COL = "video_build_status"
OUTPUT_VIDEO_COL = "video_output_file"
OUTPUT_AUDIO_COL = "generated_audio_file"
ERROR_COL = "video_build_error"
UPDATED_AT_COL = "video_build_updated_at"
NARRATION_DUR_COL = "narration_duration_sec"
SOURCE_VIDEO_DUR_COL = "source_video_duration_sec"
TTS_USED_COL = "tts_text_used"


def prepare_text_for_tts(text: str, language: str = "english") -> str:
    if not text:
        return ""

    emoji_pattern = re.compile(
        "["
        u"\U0001F600-\U0001F64F"
        u"\U0001F300-\U0001F5FF"
        u"\U0001F680-\U0001F6FF"
        u"\U0001F1E0-\U0001F1FF"
        u"\u2600-\u26FF"
        u"\u2700-\u27BF"
        u"\U0001F900-\U0001F9FF"
        u"\U0001FA70-\U0001FAFF"
        u"\u200d"
        u"\ufe0f"
        u"\u23cf"
        u"\u23e9"
        u"\u231a"
        u"\u3030"
        "]+",
        flags=re.UNICODE,
    )
    cleaned = emoji_pattern.sub("", str(text))

    if language.lower() == "hindi":
        cleaned = re.sub(r"[^\u0900-\u097Fa-zA-Z0-9.,!?\'\"\s-]", "", cleaned)
    else:
        cleaned = re.sub(r"[^a-zA-Z0-9.,!?\'\"\s-]", "", cleaned)

    cleaned = re.sub(r"\s+", " ", cleaned).strip()

    if cleaned and cleaned[-1] not in {".", "!", "?"}:
        cleaned += "."

    return cleaned


def normalize_path(value: str) -> Path:
    text = str(value).strip()
    text = text.replace("\\", os.sep).replace("/", os.sep)
    return Path(text)


def is_blank(value: object) -> bool:
    if value is None:
        return True
    return str(value).strip() == ""


def safe_stem(value: str, fallback: str) -> str:
    text = str(value).strip() if value else ""
    if not text:
        text = fallback
    text = re.sub(r'[<>:"/\\|?*]', "", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^\w.-]", "_", text, flags=re.UNICODE)
    return text[:120] or fallback


def ffprobe_duration(path: Path) -> float:
    cmd = [
        "ffprobe",
        "-v",
        "error",
        "-show_entries",
        "format=duration",
        "-of",
        "default=noprint_wrappers=1:nokey=1",
        str(path),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, check=True)
    return float(result.stdout.strip())


def get_cell_value(row_values: dict[str, object], candidates: list[str], default: Optional[str] = None) -> Optional[str]:
    lowered = {k.lower(): v for k, v in row_values.items()}
    for name in candidates:
        key = name.lower()
        if key in lowered and not is_blank(lowered[key]):
            return str(lowered[key]).strip()
    return default


def build_tts_text(row_values: dict[str, object]) -> str:
    explicit = get_cell_value(row_values, ["tts_text", "captions_text", "text"])
    if explicit:
        return explicit

    parts: list[str] = []
    for i in range(1, 16):
        field = f"text{i}"
        value = get_cell_value(row_values, [field])
        if is_blank(value):
            continue
        for part in str(value).split("^"):
            part = part.strip()
            if part:
                parts.append(part)
    return " ".join(parts)


def resolve_video_source(row_values: dict[str, object], base_dir: Path) -> Path:
    raw_value = get_cell_value(row_values, ["video_file_name", "video_path", "background_video_src"])
    if not raw_value:
        raise ValueError("Missing video source column. Use video_file_name, video_path, or background_video_src.")

    path = normalize_path(raw_value)
    if not path.is_absolute():
        path = base_dir / path
    return path


def resolve_bg_music(row_values: dict[str, object], base_dir: Path) -> Optional[Path]:
    raw_value = get_cell_value(row_values, ["background_music", "bg_music", "music_file"], DEFAULT_BG_MUSIC_FILE or None)
    if not raw_value:
        return None
    path = normalize_path(raw_value)
    if not path.is_absolute():
        path = base_dir / path
    return path


def generate_audio(text: str, output_audio: Path, language: str, gender: str) -> str:
    cleaned_text = prepare_text_for_tts(text, language)
    if not cleaned_text:
        raise ValueError("TTS text became empty after cleanup.")

    output_audio.parent.mkdir(parents=True, exist_ok=True)
    get_audio_file(
        text=cleaned_text,
        audio_file_name=str(output_audio),
        tts_engine=DEFAULT_TTS_ENGINE,
        language=language,
        gender=gender,
        type=DEFAULT_TTS_TYPE,
    )
    if not output_audio.exists():
        raise FileNotFoundError(f"Audio file was not created: {output_audio}")
    return cleaned_text


def build_ffmpeg_command(
    video_path: Path,
    narration_audio: Path,
    output_video: Path,
    narration_duration: float,
    video_duration: float,
    bg_music_path: Optional[Path],
    bg_music_volume: float,
) -> list[str]:
    pad_seconds = max(0.0, narration_duration - video_duration)
    filters = [f"[0:v]tpad=stop_mode=clone:stop_duration={pad_seconds:.3f},format=yuv420p[vout]"]

    cmd = ["ffmpeg", "-y", "-i", str(video_path), "-i", str(narration_audio)]
    audio_map = "1:a:0"

    if bg_music_path and bg_music_path.exists():
        fade_start = max(0.0, narration_duration - BG_MUSIC_FADE_OUT_SECONDS)
        cmd.extend(["-stream_loop", "-1", "-i", str(bg_music_path)])
        filters.append(
            f"[2:a]volume={bg_music_volume},atrim=0:{narration_duration:.3f},"
            f"afade=t=out:st={fade_start:.3f}:d={BG_MUSIC_FADE_OUT_SECONDS:.3f}[bgm]"
        )
        filters.append("[1:a][bgm]amix=inputs=2:duration=first:dropout_transition=2[aout]")
        audio_map = "[aout]"

    cmd.extend(
        [
            "-filter_complex",
            ";".join(filters),
            "-map",
            "[vout]",
            "-map",
            audio_map,
            "-t",
            f"{narration_duration:.3f}",
            "-c:v",
            "libx264",
            "-preset",
            "medium",
            "-crf",
            "18",
            "-c:a",
            "aac",
            "-b:a",
            "192k",
            "-movflags",
            "+faststart",
            str(output_video),
        ]
    )
    return cmd


def create_video(
    video_path: Path,
    narration_audio: Path,
    output_video: Path,
    bg_music_path: Optional[Path],
    bg_music_volume: float,
) -> tuple[float, float]:
    if not video_path.exists():
        raise FileNotFoundError(f"Video file not found: {video_path}")
    if bg_music_path and not bg_music_path.exists():
        raise FileNotFoundError(f"Background music file not found: {bg_music_path}")

    narration_duration = ffprobe_duration(narration_audio)
    video_duration = ffprobe_duration(video_path)

    output_video.parent.mkdir(parents=True, exist_ok=True)
    cmd = build_ffmpeg_command(
        video_path=video_path,
        narration_audio=narration_audio,
        output_video=output_video,
        narration_duration=narration_duration,
        video_duration=video_duration,
        bg_music_path=bg_music_path,
        bg_music_volume=bg_music_volume,
    )
    subprocess.run(cmd, check=True)
    return narration_duration, video_duration


def ensure_tracking_columns(ws) -> dict[str, int]:
    headers: list[str] = []
    for cell in ws[1]:
        value = "" if cell.value is None else str(cell.value).strip()
        headers.append(value)

    existing = {name.lower(): idx + 1 for idx, name in enumerate(headers) if name}

    required_columns = [
        STATUS_COL,
        OUTPUT_VIDEO_COL,
        OUTPUT_AUDIO_COL,
        ERROR_COL,
        UPDATED_AT_COL,
        NARRATION_DUR_COL,
        SOURCE_VIDEO_DUR_COL,
        TTS_USED_COL,
    ]

    for col_name in required_columns:
        if col_name.lower() not in existing:
            new_index = ws.max_column + 1
            ws.cell(row=1, column=new_index, value=col_name)
            existing[col_name.lower()] = new_index

    return existing


def row_to_dict(ws, row_idx: int) -> dict[str, object]:
    row_values: dict[str, object] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header is None:
            continue
        row_values[str(header).strip()] = ws.cell(row=row_idx, column=col_idx).value
    return row_values


def update_row_status(
    ws,
    col_map: dict[str, int],
    row_idx: int,
    *,
    status: str,
    output_video: str = "",
    output_audio: str = "",
    error_message: str = "",
    narration_duration: str = "",
    source_video_duration: str = "",
    tts_used: str = "",
) -> None:
    ws.cell(row=row_idx, column=col_map[STATUS_COL.lower()], value=status)
    ws.cell(row=row_idx, column=col_map[OUTPUT_VIDEO_COL.lower()], value=output_video)
    ws.cell(row=row_idx, column=col_map[OUTPUT_AUDIO_COL.lower()], value=output_audio)
    ws.cell(row=row_idx, column=col_map[ERROR_COL.lower()], value=error_message)
    ws.cell(row=row_idx, column=col_map[UPDATED_AT_COL.lower()], value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=row_idx, column=col_map[NARRATION_DUR_COL.lower()], value=narration_duration)
    ws.cell(row=row_idx, column=col_map[SOURCE_VIDEO_DUR_COL.lower()], value=source_video_duration)
    ws.cell(row=row_idx, column=col_map[TTS_USED_COL.lower()], value=tts_used)


def process_excel(excel_path: Path) -> None:
    bg_music_volume: float = 0.1
    force_retry_failed: bool = True
    wb = load_workbook(excel_path)
    ws = wb.active

    if ws.max_row < 2:
        raise ValueError("The Excel file is empty.")

    col_map = ensure_tracking_columns(ws)
    wb.save(excel_path)
    base_dir =  Path(__file__).resolve().parent
    output_video_dir = base_dir / OUTPUT_VIDEO_DIR
    audio_dir = base_dir / GENERATED_AUDIO_DIR
    vid_input_dir = base_dir / INPUT_VIDEO_DIR

    total_rows = ws.max_row - 1
    for row_idx in range(2, ws.max_row + 1):
        row_values = row_to_dict(ws, row_idx)
        row_number = row_idx
        current_status = (get_cell_value(row_values, [STATUS_COL]) or "").strip().lower()

        if current_status == "success":
            print(f"Skipping row {row_number}/{ws.max_row}: already marked success")
            continue
        if current_status == "processing" and not force_retry_failed:
            print(f"Skipping row {row_number}/{ws.max_row}: still marked processing")
            continue

        title = get_cell_value(row_values, ["output_file_name", "title", "youtube_title"], f"row_{row_number}") or f"row_{row_number}"
        file_stem = safe_stem(title, f"row_{row_number}")
        output_audio = audio_dir / f"{file_stem}_tts.mp3"
        output_video = output_video_dir / f"{file_stem}.mp4"

        print(f"\nProcessing row {row_number - 1}/{total_rows}: {title}")
        update_row_status(
            ws,
            col_map,
            row_idx,
            status="processing",
            output_video=str(output_video),
            output_audio=str(output_audio),
            error_message="",
        )
        wb.save(excel_path)

        try:
            tts_text = build_tts_text(row_values)
            if not tts_text:
                raise ValueError("No TTS text found. Add tts_text or text1..text15 columns.")

            source_video = resolve_video_source(row_values, vid_input_dir)
            bg_music = resolve_bg_music(row_values, base_dir)
            language = get_cell_value(row_values, ["language"], DEFAULT_LANGUAGE) or DEFAULT_LANGUAGE
            gender = get_cell_value(row_values, ["gender"], DEFAULT_GENDER) or DEFAULT_GENDER

            cleaned_text = generate_audio(tts_text, output_audio, language, gender)
            narration_duration, source_video_duration = create_video(
                video_path=source_video,
                narration_audio=output_audio,
                output_video=output_video,
                bg_music_path=bg_music,
                bg_music_volume=bg_music_volume,
            )

            update_row_status(
                ws,
                col_map,
                row_idx,
                status="success",
                output_video=str(output_video),
                output_audio=str(output_audio),
                error_message="",
                narration_duration=round(narration_duration, 3),
                source_video_duration=round(source_video_duration, 3),
                tts_used=cleaned_text,
            )
            wb.save(excel_path)
            print(f"Created: {output_video}")
        except Exception as exc:
            update_row_status(
                ws,
                col_map,
                row_idx,
                status="failed",
                output_video=str(output_video),
                output_audio=str(output_audio),
                error_message=str(exc),
            )
            wb.save(excel_path)
            print(f"Failed on row {row_number}: {exc}")

    wb.close()
    print(f"\nDone. Updated workbook: {excel_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Read Excel rows, create TTS audio, and merge with video.")
    parser.add_argument("--excel", default=INPUT_EXCEL, help="Input Excel file path")
    parser.add_argument("--base-dir", default=".", help="Base folder for relative video/music/output paths")
    parser.add_argument(
        "--bg-music-volume",
        type=float,
        default=DEFAULT_BG_MUSIC_VOLUME,
        help="Background music volume from 0.0 to 1.0",
    )
    parser.add_argument(
        "--retry-processing-rows",
        action="store_true",
        help="Also retry rows left in 'processing' state from an earlier interrupted run.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    process_excel(
        excel_path="bulk_video_maker_input.xlsx"
    )
