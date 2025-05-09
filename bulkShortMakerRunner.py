# bulkShortMakerRunner.py

import re
import subprocess
import json
import os
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
from mutagen.mp3 import MP3

from caption_generator import save_details_in_excel


# Constants
EXCEL_FILE = "video_records.xlsx"
SHEET_NAME = "Videos"
SHORTS_JSON_FILE = "bulkShorts_input.json"
OUTPUT_FOLDER = "processed_videos"
# Define columns
columns = [
    'url',
    'video_path',
    'captions_text',
    'youtube_playlist_name',
    'youtube_title',
    'youtube_description',
    'youtube_tags',
    'youtube_channel_name',
    'video_is_about',
    'made_for_kids',
    'schedule_date',
    'youtube_upload_status',
    'last_update_date',
    'created_video_url'
]
# Ensure output folder exists
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Prepare Excel if it doesn't exist
def setup_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(columns)
        wb.save(EXCEL_FILE)

# Save details in Excel
def save_details_in_excel_delit(video_path, title, description, tags, playlist):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    row = [
        "YTShorts",
        video_path,
        description,
        title,  #playlist
        '',# title,
        '', # description,
        '', # tags,
        '', # channel name,
        title, # video is about,
        '', #  made for kids,
        '', # schedule date,
        'Pending',
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ''
    ]

    ws.append(row)
    wb.save(EXCEL_FILE)
    print(f"✅ Saved record for {video_path}")

def get_total_audio_duration(short_index):
    total_duration = 0.0
    audio_folder = "generated_audio"  # or wherever your audio files are
    
    for i in range(100):  # up to 100 parts
        audio_file = os.path.join(audio_folder, f"short_{short_index}_{i}.mp3")
        if os.path.exists(audio_file):
            audio = MP3(audio_file)
            total_duration += audio.info.length
            total_duration += 0.4  # As 0.3 seconds delay is added between each part on the UI
        else:
            break  # Stop if no more parts
    
    return total_duration

# Main execution
def main():
    if is_excel_file_locked(EXCEL_FILE):
        print(f"Error: Please close '{EXCEL_FILE}' before running the application.")
        return  # Exit cleanly without trying uploads  
        
    # Load shorts data
    with open(SHORTS_JSON_FILE, 'r', encoding='utf-8') as f:
        shorts_data = json.load(f)

    setup_excel()

    for idx, short in enumerate(shorts_data):
        # if idx > 1:
        #     break

        # Save details in Excel
        title = short.get('title', '')
        saveAsName = re.sub(r'[<>:"/\\\\|?*]', '', title)  # Remove file-unsafe chars
        saveAsName = re.sub(r'[^a-zA-Z0-9_-]', '_', saveAsName)  # Replace special chars with underscore

        now = datetime.now()
        numeric_format = int(now.strftime("%Y%m%d%H%M%S"))

        print(f"\n▶️ Recording short {idx + 1}/{len(shorts_data)}")
        video_path = f"{idx}_{saveAsName}_{numeric_format}.mp4"
        output_video_name = f"{OUTPUT_FOLDER}/{video_path}"
        
        # output_video_name = f"short_{idx}.mp4"

        audio_duration = get_total_audio_duration(idx)
        buffer_seconds = 8  # or 5, depending on your subtext time

        recording_duration = audio_duration + buffer_seconds

        # Build Puppeteer command
        cmd = [
            "node", "puppeteer-bulkshorts-recorder.js",
            str(idx), str(output_video_name), str(recording_duration)
        ]

        print("Running Puppeteer with:", cmd)
        subprocess.run(cmd)


        description = short.get('description', '')
        tags = short.get('tags', '')  
        playlist = short.get('playlist', '')  
        channel = short.get('channel', '')
        schedule_date = short.get('schedule_date', '')

        save_details_in_excel(description, "YTShorts", video_path, description, tags, playlist, channel,title,schedule_date)
        
        #update status in excel

        

        # Small delay between recordings
        time.sleep(1)

    print("\n🎬 All shorts recorded successfully!")

def is_excel_file_locked(file_path):
    try:
        with open(file_path, 'a'):
            pass
        return False
    except PermissionError:
        return True
    
if __name__ == "__main__":
    main()
