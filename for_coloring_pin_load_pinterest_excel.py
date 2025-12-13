# generate_excel.py
import argparse
import os
import sys
import json
from pathlib import Path
import json
from bg_music_video import merge_video_with_bg_music_overwrite
from coloring_animation import _create_coloring_animation_by_color
from gemini_pool import GeminiPool  # same helper you use in get_seo_meta_data.py
import random
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageOps, ImageDraw, ImageFont, ImageFilter
import subprocess
from moviepy.editor import (
    ImageClip,
    TextClip,
    ColorClip,
    CompositeVideoClip,
    concatenate_videoclips,
)
from moviepy.video.fx import all as vfx


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp"}
VIDEO_EXTS = {".mp4", ".mov", ".mkv", ".avi"}

CONFIG_FILENAME = "pinterest_config.json"


# Gemini settings
DEFAULT_MODEL = "gemini-2.0-flash"
DEFAULT_TEMPERATURE = 0.3
# Increased because we now generate multi-platform metadata
DEFAULT_MAX_TOKENS = 512

# Global gemini_pool – initialised in main guard
gemini_pool = None

GENERIC_PROMO_SYSTEM_INSTRUCTION = """
You are a social media marketing assistant.

Goal:
Given a campaign and destination URL, generate:
- Pinterest title + description + tags
- Instagram caption + hashtags
- TikTok caption + hashtags
- YouTube title + description + tags
- X text + hashtags
- Facebook caption + hashtags

Rules:
- Platform-safe, family-friendly.
- No ALL CAPS clickbait.
- Make it relevant to destination_type:
  - product: highlight benefits, use cases, who it's for, soft CTA to check link
  - video: hook + what viewer will learn + CTA to watch
  - article: value summary + CTA to read
  - service: outcome + credibility + CTA to learn more
- Use "cta_lines" verbatim at the end of long descriptions where appropriate (YouTube/Facebook).
- Keep Pinterest title <= ~80 chars.
- Pinterest description 1–3 short sentences.
- Hashtags:
  - Pinterest: 5–12 as "#tag" style
  - Instagram/TikTok: 5–12
  - X: 3–8
Return STRICT JSON only with this schema:

{
  "pin_title": "string",
  "pin_description": "string",
  "tags": ["#tag1", "#tag2"],
  "instagram_caption": "string",
  "instagram_hashtags": ["#tag1"],
  "tiktok_caption": "string",
  "tiktok_hashtags": ["#tag1"],
  "youtube_title": "string",
  "youtube_description": "string",
  "youtube_tags": ["tag1","tag2"],
  "x_text": "string",
  "x_hashtags": ["#tag1"],
  "facebook_caption": "string",
  "facebook_hashtags": ["#tag1"]
}
"""

PINTEREST_SYSTEM_INSTRUCTION = """
You are a social media marketing assistant for digital coloring books sold on goodsandgift.com.

Goal:
Given information about a specific coloring book page (sample, cover, or activity page), generate:
- Pinterest Pin metadata
- Instagram caption + hashtags
- TikTok caption + hashtags
- YouTube title + description + tags
- X/Twitter text + hashtags
- Facebook caption + hashtags

Audience:
- Parents looking for printable coloring pages for kids
- Teachers and homeschoolers
- Adults who enjoy relaxing coloring (for some books)

General rules:
- Use simple, friendly language.
- Emphasize benefits: printable, instant download, number of pages, fun themes, age range if provided.
- Highlight that it’s a DIGITAL DOWNLOAD (no physical item shipped).
- Include keywords like "coloring page", "coloring book", "printable", "digital download", "kids activities", where appropriate.
- Avoid clickbait or ALL CAPS.
- Keep everything platform-safe (family friendly).

Pinterest title rules:
- Max ~80 characters.
- Include the main theme (e.g., farm animals, dinosaurs, mandalas, etc.).
- Mention "coloring page" or "coloring book".

Pinterest description rules:
- 1–3 short sentences.
- 140–300 characters is ideal.
- Mention what’s included (e.g., number of pages in the book, theme, target age if known).
- Encourage saving the Pin or downloading/printing later.

Hashtag rules (all platforms):
- 5–12 items per platform.
- Each tag should be a hashtag string, e.g. "#coloringpages".
- Mix broad and specific tags: e.g., "#coloringpages", "#printablecoloring", "#kidsactivities", "#mandalacoloring", "#farmanimals".
- No duplicates within one platform’s list.

Instagram:
- One caption string (few short lines ok).
- Hashtags can be appended at the end in the "instagram_hashtags" field.
- Caption should be friendly and slightly more expressive.

TikTok:
- Caption can be similar to Instagram but a bit punchier / hooky.
- Keep it relatively short.

YouTube:
- "youtube_title": strong, clear, SEO-friendly title for a Short about this coloring page/book.
- "youtube_description": 2–4 short lines + a soft CTA to download/print. You may mention that this is a vertical Short preview of the coloring page.
- "youtube_tags": mix of 5–15 SEO tags (NOT prefixed with "#").

X/Twitter:
- "x_text": short, tweet-length statement promoting the coloring page/book (~150 characters or less).
- "x_hashtags": 3–8 relevant hashtags.

Facebook:
- Caption: similar to Instagram but can be slightly longer if needed.
- Hashtags: 3–8 relevant hashtags.


Return STRICT JSON with this schema only:

{
  "pin_title": "string",
  "pin_description": "string",
  "tags": ["#tag1", "#tag2", "#tag3"],
  "instagram_caption": "string",
  "instagram_hashtags": ["#tag1", "#tag2"],
  "tiktok_caption": "string",
  "tiktok_hashtags": ["#tag1", "#tag2"],
  "youtube_title": "string",
  "youtube_description": "string",
  "youtube_tags": ["tag1", "tag2"],
  "x_text": "string",
  "x_hashtags": ["#tag1", "#tag2"],
  "facebook_caption": "string",
  "facebook_hashtags": ["#tag1", "#tag2"]
}

Do not include any commentary or text outside this JSON.
"""

def collect_media_files(root: Path, subfolder: str | None) -> list[Path]:
    """
    Collect image / video files from images_root / subfolder (recursive).
    """
    base = root / subfolder if subfolder else root
    if not base.exists():
        raise FileNotFoundError(f"Source folder does not exist: {base}")

    files = []
    for p in base.rglob("*"):
        if not p.is_file():
            continue
        if p.suffix.lower() in IMAGE_EXTS.union(VIDEO_EXTS):
            files.append(p)

    files.sort()
    return files

def add_ken_burns(clip, zoom: float = 1.06):
    """
    Apply a very gentle Ken Burns-style zoom to a clip over its duration.
    zoom = final scale at end (e.g. 1.06 = 6% larger).
    """
    dur = clip.duration or 1.0

    def scaler(t):
        # t goes from 0..dur, interpolate scale from 1.0 to zoom
        return 1.0 + (zoom - 1.0) * (t / dur)

    return clip.fx(vfx.resize, scaler)

def make_title_card(
    book_title: str,
    duration: float = 1.5,
    size: tuple[int, int] = (1080, 1920),
):
    """
    Simple intro/title card: solid background + centered text.
    """
    W, H = size
    bg = ColorClip(size=size, color=(20, 30, 60))  # deep bluish
    bg = bg.set_duration(duration)

    title_text = book_title.strip() or "Coloring Book Preview"

    try:
        title = TextClip(
            title_text,
            fontsize=60,
            color="white",
            method="caption",
            size=(int(W * 0.9), None),
        )
    except Exception:
        # If TextClip has issues (e.g. missing ImageMagick), just return bg
        return bg

    title = title.set_duration(duration)
    title = title.set_position(("center", "center"))

    return CompositeVideoClip([bg, title])


def make_outro_card(
    book_title: str,
    book_url: str,
    duration: float = 1.8,
    size: tuple[int, int] = (1080, 1920),
):
    """
    Simple outro/CTA card: brand color + CTA text.
    """
    W, H = size
    bg = ColorClip(size=size, color=(15, 80, 60))  # teal-ish
    bg = bg.set_duration(duration)

    line1 = book_title.strip() or "Cute Coloring Pages"
    line2 = "Download instantly at:"
    line3 = (book_url.strip() or "goodsandgift.com").replace("https://", "").replace("http://", "")

    texts = []
    try:
        t1 = TextClip(line1, fontsize=52, color="white", method="caption", size=(int(W * 0.9), None))
        t2 = TextClip(line2, fontsize=42, color="white", method="caption", size=(int(W * 0.9), None))
        t3 = TextClip(line3, fontsize=46, color="#facc15", method="caption", size=(int(W * 0.9), None))
        texts = [t1, t2, t3]
    except Exception:
        return bg

    # stack vertically
    total_h = sum(t.h for t in texts) + 40
    y_start = (H - total_h) // 2
    comps = [bg]
    y = y_start
    for t in texts:
        t = t.set_duration(duration)
        t = t.set_position(("center", y))
        comps.append(t)
        y += t.h + 10

    return CompositeVideoClip(comps)

def load_campaign_config(images_root: Path, source_subfolder: str | None) -> dict:
    """
    Load pinterest_config.json from:
      - images_root/source_subfolder/pinterest_config.json
      - images_root/source_subfolder/../pinterest_config.json

    Backward compatible with older keys:
      book_title -> campaign_name
      book_url   -> destination_url
    """

    base = (images_root / source_subfolder) if source_subfolder else images_root

    candidates = [
        base / CONFIG_FILENAME,
        base.parent / CONFIG_FILENAME
    ]

    print(f"[INFO] Looking for Pinterest config in:")
    for c in candidates:
        print(f"  {c}")

    for cfg in candidates:
        if not cfg.exists():
            continue

        try:
            with cfg.open("r", encoding="utf-8") as f:
                data = json.load(f)

            print(f"[INFO] Loaded Pinterest config from: {cfg}")

            # Backward compatibility
            book_title = str(data.get("book_title", "")).strip()
            book_url = str(data.get("book_url", "")).strip()

            campaign_name = str(data.get("campaign_name", "")).strip() or book_title
            destination_url = str(data.get("destination_url", "")).strip() or book_url

            return {
                # New generalized keys
                "campaign_name": campaign_name,
                "destination_url": destination_url,
                "destination_type": str(data.get("destination_type", "")).strip() or "product",
                "topic": str(data.get("topic", "")).strip(),

                "cta_text": str(data.get("cta_text", "")).strip(),
                "cta_lines": data.get("cta_lines", []) if isinstance(data.get("cta_lines", []), list) else [],
                "link_overrides": data.get("link_overrides", {}) if isinstance(data.get("link_overrides", {}), dict) else {},

                # Existing keys you already use
                "board_name": str(data.get("board_name", "")).strip(),
                "banner_text": str(data.get("banner_text", "")).strip(),
                "watermark_text": str(data.get("watermark_text", "")).strip(),
                "base_tags": str(data.get("base_tags", "")).strip(),

                # Keep old keys too (optional)
                "book_title": book_title,
                "book_url": book_url,
            }
        except Exception as e:
            print(f"[WARN] Failed to read config {cfg}: {e}")
            return {}

    print("[INFO] No pinterest_config.json found. Using only CLI/UI values.")
    return {}

def load_book_config(images_root: Path, source_subfolder: str | None) -> dict:
    """
    Look for pinterest_config.json inside source_subfolder and its parent.
    Returns dict with keys:
      book_title, book_url, board_name, banner_text, watermark_text
    Any missing file or key just means "no default" for that field.
    """

    if source_subfolder:
        base = images_root / source_subfolder
    else:
        base = images_root

    candidates = [
        base / CONFIG_FILENAME,
        base.parent / CONFIG_FILENAME
    ]

    print(f"[INFO] Looking for Pinterest config in:")
    for c in candidates:
        print(f"  {c}")
    
    for cfg in candidates:
        if cfg.exists():
            try:
                with cfg.open("r", encoding="utf-8") as f:
                    data = json.load(f)
                print(f"[INFO] Loaded Pinterest config from: {cfg}")
                return {
                    "book_title": str(data.get("book_title", "")).strip(),
                    "book_url": str(data.get("book_url", "")).strip(),
                    "board_name": str(data.get("board_name", "")).strip(),
                    "banner_text": str(data.get("banner_text", "")).strip(),
                    "watermark_text": str(data.get("watermark_text", "")).strip(),
                }
            except Exception as e:
                print(f"[WARN] Failed to read config {cfg}: {e}")
                return {}
    print("[INFO] No pinterest_config.json found. Using only CLI/UI values.")
    return {}

# threshold=245 works well for black lines with anti-aliasing on white.
# If it ever crops too aggressively, drop threshold to 235–240.
# If it leaves too much white, raise to 248–250.
def crop_to_content(img: Image.Image,
                    threshold: int = 245,
                    pad_ratio: float = 0.05) -> Image.Image:
    """
    Auto-crop a line-art coloring page so that we remove most of the white margin.

    - threshold: 0–255; higher = treat very light grays as background.
    - pad_ratio: extra padding around the detected drawing, as a fraction of bbox size.
    """
    # Work on grayscale
    gray = img.convert("L")

    # Create a mask: 255 where pixel is "ink", 0 where it's near-white
    # Anything darker than `threshold` is considered part of the drawing.
    bw = gray.point(lambda v: 255 if v < threshold else 0, mode="1")

    bbox = bw.getbbox()
    if not bbox:
        # No non-white content detected – return original
        return img

    left, upper, right, lower = bbox
    w, h = img.size

    # Add small padding around the drawing bbox
    box_w = right - left
    box_h = lower - upper
    pad_x = int(box_w * pad_ratio)
    pad_y = int(box_h * pad_ratio)

    left = max(0, left - pad_x)
    upper = max(0, upper - pad_y)
    right = min(w, right + pad_x)
    lower = min(h, lower + pad_y)

    return img.crop((left, upper, right, lower))

def make_pinterest_video_from_group(
    images: list[Path],
    out_dir: Path,
    video_style: str = "flipbook",
    duration: float = 10.0,  # default 10s now
    fps: int = 30,
    size: tuple[int, int] = (1080, 1920),
    book_title: str | None = None,
    book_url: str | None = None,
) -> Path:
    """
    Create a vertical MP4 from a group of Pinterest-ready images.

    - images: list of image paths (already branded + auto-cropped).
    - video_style: "flipbook" or "slideshow".
    - duration: total video duration in seconds (including intro/outro).
    """

    out_dir.mkdir(parents=True, exist_ok=True)
    if not images:
        raise ValueError("No images provided for video.")

    # Basic safety
    duration = max(duration, 4.0)

    # Reserve some time for intro + outro
    if video_style == "slideshow":
        intro_dur = 1.6
        outro_dur = 1.8
    else:  # flipbook
        intro_dur = 1.0
        outro_dur = 1.2

    main_duration = max(1.0, duration - intro_dur - outro_dur)
    n = len(images)

    # Per-page duration (cap for flipbook)
    per = main_duration / n
    if video_style == "flipbook":
        per = min(per, 0.45)
        per = max(per, 0.25)  # clamp between 0.25–0.45s
    else:
        per = max(per, 0.8)   # slideshow: at least 0.8s per page

    W, H = size
    clips = []

    for p in images:
        base = ImageClip(str(p))

        iw, ih = base.size
        scale = min(W / iw, H / ih)
        new_w, new_h = int(iw * scale), int(ih * scale)
        base = base.resize((new_w, new_h))
        base = base.on_color(size=size, color=(255, 255, 255), pos=("center", "center"))
        base = base.set_duration(per)

        if video_style == "slideshow":
            # Gentle Ken Burns zoom
            base = add_ken_burns(base, zoom=1.06)

        clips.append(base)

    # For now we skip intro/outro cards for speed
    video = concatenate_videoclips(clips, method="compose")

    # Fade-in & fade-out for the whole video
    video = video.fx(vfx.fadein, 0.25).fx(vfx.fadeout, 0.25)

    base_name = images[0].stem
    # out_name = f"{base_name}_{video_style}_pin.mp4"
    out_name = f"{base_name}.mp4"
    out_path = out_dir / out_name

    video.write_videofile(
        str(out_path),
        fps=fps,
        codec="libx264",
        audio=False,
        verbose=False,
        logger=None,
    )
    return out_path


def make_pinterest_video(
    src: Path,
    out_dir: Path,
    duration: float = 10.0,  # default 10s now
    fps: int = 30,
) -> Path:
    """
    Single-page video:
    - Use the Pinterest image as-is (no extra resize).
    - Apply gentle Ken Burns zoom.
    - Add soft fade-in / fade-out.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{src.stem}_pin.mp4"

    # Pinterest image is already the right aspect & size (e.g. 1000x1500)
    clip = ImageClip(str(src)).set_duration(duration)

    # Motion
    clip = add_ken_burns(clip, zoom=1.06)
    clip = clip.fx(vfx.fadein, 0.25).fx(vfx.fadeout, 0.25)

    clip.write_videofile(
        str(out_path),
        fps=fps,
        codec="libx264",
        audio=False,
        verbose=False,
        logger=None,
    )

    return out_path


def make_pinterest_image(
    src: Path,
    out_dir: Path,
    banner_text: str | None,
    watermark_text: str | None,
    target_size=(1000, 1500),
    fit_mode: str = "contain",          # "contain" (no crop) or "cover" (crop)
    bg_style: str = "white",            # "white" or "blur"
    text_shadow: bool = True,
    auto_crop_subject: bool = True,   # <--- NEW FLAG
) -> Path:
    """
    Create a Pinterest-friendly image:

    - target canvas size: 1000 x 1500
    - fit_mode:
        * "contain": whole image visible, letterboxed if needed (NO CROPPING)
        * "cover"  : image cropped to fill the canvas
    - bg_style:
        * "white": plain white background
        * "blur" : blurred version of the original image behind the main image
    - text_shadow: draws a subtle shadow behind banner & watermark text.
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    base_img = Image.open(src).convert("RGB")
    # NEW: auto-crop to drawing to reduce huge margins
    if auto_crop_subject:
        base_img = crop_to_content(base_img, threshold=245, pad_ratio=0.05)

    orig_w, orig_h = base_img.size
    tgt_w, tgt_h = target_size

    # -------- Background --------
    if bg_style == "blur":
        # Use a blurred "cover" version of the image as background
        scale_bg = max(tgt_w / orig_w, tgt_h / orig_h)
        bg_w = int(orig_w * scale_bg)
        bg_h = int(orig_h * scale_bg)
        bg = base_img.resize((bg_w, bg_h), Image.LANCZOS)

        left = (bg_w - tgt_w) // 2
        top = (bg_h - tgt_h) // 2
        bg = bg.crop((left, top, left + tgt_w, top + tgt_h))
        bg = bg.filter(ImageFilter.GaussianBlur(radius=18))
        # Slight darken so foreground pops
        dark_overlay = Image.new("RGBA", (tgt_w, tgt_h), (0, 0, 0, 60))
        bg_rgba = bg.convert("RGBA")
        bg_rgba.alpha_composite(dark_overlay, (0, 0))
        canvas = bg_rgba.convert("RGB")
    else:
        # Plain white background
        canvas = Image.new("RGB", (tgt_w, tgt_h), "white")

    # -------- Foreground (main page) --------
    if fit_mode == "cover":
        # Crop to fill
        main = base_img.copy()
        main = ImageOps.fit(main, (tgt_w, tgt_h), method=Image.LANCZOS)
        # Slight inset so banner/watermark don't cover important edges
        inset = 40
        main = main.resize((tgt_w - inset * 2, tgt_h - inset * 2), Image.LANCZOS)
        mx = inset
        my = inset
    else:
        # contain – NO CROPPING
        scale = min(tgt_w / orig_w, tgt_h / orig_h)
        new_w = int(orig_w * scale)
        new_h = int(orig_h * scale)
        main = base_img.resize((new_w, new_h), Image.LANCZOS)
        mx = (tgt_w - new_w) // 2
        my = (tgt_h - new_h) // 2

    canvas.paste(main, (mx, my))

    draw = ImageDraw.Draw(canvas)

    # Fonts
    try:
        font = ImageFont.truetype("arial.ttf", 36)
        font_small = ImageFont.truetype("arial.ttf", 28)
    except Exception:
        font = ImageFont.load_default()
        font_small = ImageFont.load_default()

    W, H = canvas.size

    # Helpers ----------------------------------------------------
    def measure(text: str, font_obj):
        if not text:
            return (0, 0)
        bbox = draw.textbbox((0, 0), text, font=font_obj)
        return (bbox[2] - bbox[0], bbox[3] - bbox[1])

    def draw_text_with_shadow(x, y, text, font_obj, fill="white"):
        if not text:
            return
        if text_shadow:
            # Simple dark shadow offset
            shadow_offset = 2
            draw.text((x + shadow_offset, y + shadow_offset),
                      text, font=font_obj, fill="black")
        draw.text((x, y), text, font=font_obj, fill=fill)

    # -------- Banner at top --------
    if banner_text:
        banner_h = int(H * 0.12)
        banner_overlay = Image.new("RGBA", (W, banner_h), (0, 0, 0, 150))
        canvas_rgba = canvas.convert("RGBA")
        canvas_rgba.alpha_composite(banner_overlay, (0, 0))
        canvas = canvas_rgba.convert("RGB")
        draw = ImageDraw.Draw(canvas)

        tw, th = measure(banner_text, font)
        tx = (W - tw) / 2
        ty = (banner_h - th) / 2
        draw_text_with_shadow(tx, ty, banner_text, font)

    # -------- Watermark at bottom --------
    if watermark_text:
        wm_h = int(H * 0.08)
        wm_overlay = Image.new("RGBA", (W, wm_h), (0, 0, 0, 130))
        canvas_rgba = canvas.convert("RGBA")
        canvas_rgba.alpha_composite(wm_overlay, (0, H - wm_h))
        canvas = canvas_rgba.convert("RGB")
        draw = ImageDraw.Draw(canvas)

        tw, th = measure(watermark_text, font_small)
        tx = (W - tw) / 2
        ty = H - wm_h + (wm_h - th) / 2
        draw_text_with_shadow(tx, ty, watermark_text, font_small)

    # -------- Save --------
    out_path = out_dir / (src.stem + "_pin.webp")
    canvas.save(out_path, "WEBP", quality=90)
    return out_path

def make_video_image(
    src: Path,
    out_dir: Path,
    fit_mode: str = "contain",
    bg_style: str = "white",
    auto_crop_subject: bool = True,
    size: tuple[int, int] = (1080, 1920),
) -> Path:
    """
    Create a clean vertical image for VIDEO USE ONLY.
    - No banner
    - No footer
    - No watermark
    - Just auto-cropped subject, centered and padded
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    im = Image.open(src).convert("RGB")

    # Optional subject auto-cropping
    if auto_crop_subject:
        im = crop_to_content(im, threshold=245, pad_ratio=0.05)

    # Resize while preserving aspect
    iw, ih = im.size
    W, H = size

    scale = min(W / iw, H / ih)
    new_w, new_h = int(iw * scale), int(ih * scale)
    im = im.resize((new_w, new_h), Image.LANCZOS)

    # Background
    if bg_style == "white":
        bg = Image.new("RGB", (W, H), "white")
    else:
        bg = Image.new("RGB", (W, H), bg_style)

    # Center object
    x = (W - new_w) // 2
    y = (H - new_h) // 2
    bg.paste(im, (x, y))

    # Output path
    # out_path = out_dir / f"{src.stem}_video.jpg"
    out_path = out_dir / f"{src.stem}.jpg"
    bg.save(out_path, "JPEG", quality=95)

    return out_path


def get_or_create_workbook(output_excel: Path, base_headers: list[str]):
    """
    If output Excel exists, open and reuse it (append new rows).
    If not, create a new workbook with the given base_headers.

    Returns: wb, ws, headers, header_map
      - headers: ordered list of header names in row 1
      - header_map: {header_name: 1-based column index}
    """
    if output_excel.exists():
        wb = load_workbook(output_excel)
        ws = wb.active

        # Existing headers in first row
        existing_headers = [cell.value or "" for cell in ws[1]]
        header_map = {name: idx + 1 for idx, name in enumerate(existing_headers)}

        # Ensure all base_headers exist; if missing, append new columns
        changed = False
        for h in base_headers:
            if h not in header_map:
                col_idx = len(existing_headers) + 1
                ws.cell(row=1, column=col_idx, value=h)
                existing_headers.append(h)
                header_map[h] = col_idx
                changed = True

        if changed:
            print("[INFO] Updated header row with new columns:", base_headers)

        return wb, ws, existing_headers, header_map

    # No existing file → create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pins"
    ws.append(base_headers)
    header_map = {name: idx + 1 for idx, name in enumerate(base_headers)}
    return wb, ws, base_headers, header_map

def resolve_destination_url(
    destination_url: str,
    pin_url_key: str,
    cfg: dict
) -> str:
    """
    Resolve final destination link:
    1) destination_url (from CLI/UI) if provided
    2) cfg.destination_url if provided
    3) cfg.link_overrides[pin_url_key] if present
    4) empty
    """
    destination_url = (destination_url or "").strip()
    if destination_url:
        return destination_url

    cfg_dest = str(cfg.get("destination_url", "")).strip()
    if cfg_dest:
        return cfg_dest

    overrides = cfg.get("link_overrides", {}) if isinstance(cfg.get("link_overrides", {}), dict) else {}
    override = str(overrides.get(pin_url_key, "")).strip()
    if override:
        return override

    return ""

def build_excel(
    images_root: Path,
    source_subfolder: str | None,
    output_excel: Path,
    media_type: str,
    max_pins: int | None,
    book_title: str,
    book_url: str,
    board_name: str,
    banner_text: str | None,
    watermark_text: str | None,
    fit_mode: str = "contain",
    bg_style: str = "white",
    text_shadow: bool = True,
    use_gemini: bool = False,
    base_tags: str | None = None,
    auto_crop_subject: bool = True,
    video_style: str = "single",
    pages_per_video: int = 10,
    video_duration: float = 10.0,   # default 10s for all platforms
    video_fps: int = 30,
    add_bg_music: bool = False,
    # pin_url: str = "",
) -> None:
    # 1) Load config (if present)
    # cfg = load_book_config(images_root, source_subfolder)
    cfg = load_campaign_config(images_root, source_subfolder)

    # 2) Resolve generalized campaign values (UI/CLI overrides config)
    campaign_name = (book_title or cfg.get("campaign_name") or cfg.get("book_title") or "Campaign").strip()


    # 2) Resolve final values (UI/CLI overrides config)
    book_title = (book_title or cfg.get("book_title") or "Coloring Book").strip()
    book_url = (book_url or cfg.get("book_url") or "").strip()
    # if pin_url == "amazon":
    #     book_url = "https://www.amazon.com/dp/B0G1TK51V4"
    # elif pin_url == "gumroad":
    #     pin_url = "https://kishna01.gumroad.com/"
    # elif pin_url == "readernook":
    #     book_url = "https://www.coloring.readernook.com/"
    pin_url = ""
    destination_url  = resolve_destination_url(
        destination_url=book_url,
        pin_url_key=pin_url,
        cfg=cfg
    )

    destination_type = (cfg.get("destination_type") or "product").strip()
    topic = (cfg.get("topic") or "").strip()

    board_name = (board_name or cfg.get("board_name") or "").strip()
    banner_text = (banner_text or cfg.get("banner_text") or "").strip() or None
    watermark_text = (watermark_text or cfg.get("watermark_text") or "").strip() or None
    base_tags = (base_tags or cfg.get("base_tags") or "").strip()

    cta_text = (cfg.get("cta_text") or "").strip()
    cta_lines = cfg.get("cta_lines", []) if isinstance(cfg.get("cta_lines", []), list) else []

    print(f"[INFO] Using values:")
    print(f"  campaign_name   = {campaign_name!r}")
    print(f"  destination_url = {destination_url!r}")
    print(f"  destination_type= {destination_type!r}")
    print(f"  topic           = {topic!r}")
    print(f"  board_name      = {board_name!r}")
    print(f"  banner_text     = {banner_text!r}")
    print(f"  watermark_text  = {watermark_text!r}")
    print(f"  use_gemini      = {use_gemini!r}")

    media_files = collect_media_files(images_root, source_subfolder)

    # Randomize the file order
    random.shuffle(media_files)

    groups: list[list[Path]]

    if media_type == "video" and video_style in ("flipbook", "slideshow"):
        # multi-page video: group images
        if pages_per_video < 1:
            pages_per_video = 8
        # limit images if max_pins requested
        if max_pins and max_pins > 0:
            media_files = media_files[: max_pins * pages_per_video]

        groups = [
            media_files[i : i + pages_per_video]
            for i in range(0, len(media_files), pages_per_video)
        ]
        if max_pins and max_pins > 0:
            groups = groups[:max_pins]

        print(f"[INFO] Creating {len(groups)} video pins ({video_style}), "
              f"{pages_per_video} pages per video (where available).")
    else:
        # image pins, or single-image video pins
        if max_pins and max_pins > 0:
            media_files = media_files[:max_pins]
        groups = [[f] for f in media_files]
        print(f"[INFO] Found {len(groups)} media items to process.")

    # Base headers we want to ensure exist (append-only).
    # These now include multi-platform fields.
    base_headers = [
        "pin_id",
        "media_file",       # final media file (image or video)
        "media_type",       # image / video
        "duration_sec",
        "aspect_ratio",
        "board_name",
        "book_title",
        "book_url",

        "campaign_name",
        "destination_url",
        "destination_type",
        "topic",
        # Platform toggles
        "platform_pinterest",
        "platform_instagram",
        "platform_tiktok",
        "platform_youtube",
        "platform_x",
        "platform_facebook",

        # Pinterest
        "pin_title",
        "pin_description",
        "pin_tags",
        "pin_url_to_link",

        # Instagram
        "ig_caption",
        "ig_hashtags",

        # TikTok
        "tiktok_caption",
        "tiktok_hashtags",

        # YouTube
        "yt_title",
        "yt_description",
        "yt_tags",
        "yt_playlist",

        # X / Twitter
        "x_text",
        "x_hashtags",

        # Facebook
        "fb_caption",
        "fb_hashtags",

        # Common visuals
        "banner_text",
        "watermark_text",
        "notes",
    ]

    # Open existing workbook if present, otherwise create new
    wb, ws, headers, header_map = get_or_create_workbook(output_excel, base_headers)

    # Determine next pin_id based on how many data rows already exist
    existing_rows = ws.max_row - 1  # excluding header row
    next_pin_id = existing_rows + 1
    print(f"[INFO] Existing rows: {existing_rows}, next_pin_id starts at {next_pin_id}")

    # Where to store generated Pinterest-ready images and videos
    pins_output_root = images_root.parent / "pinterest_pins"
    pins_output_root.mkdir(parents=True, exist_ok=True)

    for idx, group in enumerate(groups, start=1):
        try:
            primary = group[0]
            ext = primary.suffix.lower()
            is_image = ext in IMAGE_EXTS
            is_video = ext in VIDEO_EXTS

            if not (is_image or is_video):
                print(f"[SKIP] Unsupported file type: {primary}")
                continue

            print(f"[PROCESS] {idx}: {[g.name for g in group]}")

            if media_type == "image":
                # ---- IMAGE PINS (existing behaviour) ----
                if is_image:
                    try:
                        out_img = make_pinterest_image(
                            src=primary,
                            out_dir=pins_output_root,
                            banner_text=banner_text,
                            watermark_text=watermark_text,
                            fit_mode=fit_mode,
                            bg_style=bg_style,
                            text_shadow=text_shadow,
                            auto_crop_subject=auto_crop_subject,
                        )
                        media_path_for_excel = str(out_img.relative_to(images_root.parent))
                    except Exception as e:
                        print(f"[WARN] Failed to create Pinterest image for {primary}: {e}")
                        media_path_for_excel = str(primary.relative_to(images_root.parent))
                else:
                    media_path_for_excel = str(primary.relative_to(images_root.parent))

            elif media_type == "video":
                # ---- VIDEO PINS ----
                if video_style == "single":
                    # one image → one video
                    if is_image:
                        try:
                            out_img = make_pinterest_image(
                                src=primary,
                                out_dir=pins_output_root,
                                banner_text=banner_text,
                                watermark_text=watermark_text,
                                fit_mode=fit_mode,
                                bg_style=bg_style,
                                text_shadow=text_shadow,
                                auto_crop_subject=auto_crop_subject,
                            )
                            out_vid = make_pinterest_video(
                                src=out_img,
                                out_dir=pins_output_root,
                                duration=video_duration,
                                fps=video_fps,
                            )
                            media_path_for_excel = str(out_vid.relative_to(images_root.parent))
                        except Exception as e:
                            print(f"[WARN] Failed to create Pinterest VIDEO for {primary}: {e}")
                            media_path_for_excel = str(primary.relative_to(images_root.parent))
                    else:
                        media_path_for_excel = str(primary.relative_to(images_root.parent))

                else:
                    # ---- FLIPBOOK / SLIDESHOW (multi-page) ----
                    # 1) Create clean video images for all pages in the group
                    processed_imgs: list[Path] = []
                    for p in group:
                        if p.suffix.lower() not in IMAGE_EXTS:
                            continue
                        try:
                            proc = make_video_image(
                                src=p,
                                out_dir=pins_output_root,
                                fit_mode=fit_mode,
                                bg_style=bg_style,
                                auto_crop_subject=auto_crop_subject,
                            )
                            processed_imgs.append(proc)
                        except Exception as e:
                            print(f"[WARN] Failed to create image for {p}: {e}")

                    if not processed_imgs:
                        print(f"[WARN] No valid images for group {idx}")
                        media_path_for_excel = str(primary.relative_to(images_root.parent))
                    else:
                        try:
                            out_vid = make_pinterest_video_from_group(
                                images=processed_imgs,
                                out_dir=pins_output_root,
                                video_style=video_style,
                                duration=video_duration,
                                fps=video_fps,
                                book_title=book_title,
                                book_url=book_url,
                            )

                            media_path_for_excel = str(out_vid.relative_to(images_root.parent))

                            if add_bg_music:
                                merge_video_with_bg_music_overwrite("pinterest_uploads/" + media_path_for_excel, "background_music/dreamland.mp3", bg_volume=0.3, video_volume=1.0)

                        except Exception as e:
                            print(f"[WARN] Failed to create {video_style} video for group {idx}: {e}")
                            # Fallback: use first processed image
                            media_path_for_excel = str(processed_imgs[0].relative_to(images_root.parent))
            elif media_type == "coloring":
                # ---- COLORING ANIMATION VIDEOS ----
                if is_image:
                    try:
                        # Where to place the final coloring video
                        out_vid = pins_output_root / f"{primary.stem}_c.mp4"

                        input_path = str(primary)
                        output_path = str(out_vid)
                        target_size = (1080, 1920)  # vertical video like other pins

                        print(f"[INFO] Creating coloring animation for {primary.name}")
                        _create_coloring_animation_by_color(
                            input_path=input_path,
                            output_path=output_path,
                            fps=video_fps,
                            num_colors=7,
                            brush_steps_per_color=50,
                            hold_line_sec=1.2,
                            hold_end_sec=1.2,
                            target_size=target_size,
                            bg_color=(255, 255, 255),
                        )

                        media_path_for_excel = str(out_vid.relative_to(images_root.parent))

                        # Add background music exactly like other videos
                        if add_bg_music:
                            merge_video_with_bg_music_overwrite(
                                "pinterest_uploads/" + media_path_for_excel,
                                "background_music/dreamland.mp3",
                                bg_volume=0.3,
                                video_volume=1.0,
                            )

                    except Exception as e:
                        print(f"[WARN] Failed to create COLORING video for {primary}: {e}")
                        # Fallback: just log the original image path
                        media_path_for_excel = str(primary.relative_to(images_root.parent))
                else:
                    # Not an image – just log original path
                    media_path_for_excel = str(primary.relative_to(images_root.parent))

            else:
                # Unknown media_type
                media_path_for_excel = str(primary.relative_to(images_root.parent))

            # Simple label from filename
            page_label = primary.stem.replace("_", " ").replace("-", " ").title()

            error_msg = ""
            try:
                if use_gemini:
                    meta = generate_pin_meta_with_gemini(
                        campaign_name=campaign_name,
                        destination_url=destination_url,
                        destination_type=destination_type,
                        topic=topic,
                        page_label=page_label,
                        base_tags=base_tags,
                        pin_role="sample_page",
                        cta_lines=cta_lines,
                    )
                else:
                    meta = fallback_pin_meta(
                        campaign_name=campaign_name,
                        destination_url=destination_url,
                        destination_type=destination_type,
                        page_label=page_label,
                        base_tags=base_tags,
                        cta_lines=cta_lines,
                    )

            except Exception as e:
                print(f"[WARN] Gemini meta failed for {primary.name}: {e}")
                error_msg = f"Gemini meta failed: {e}"
                meta = fallback_pin_meta(book_title, book_url, page_label, base_tags)

            # Pinterest
            pin_title = meta["pin_title"]
            pin_description = meta["pin_description"]
            tags_str = meta["tags_str"]

            # Multi-platform fields from meta dict (with safe defaults)
            ig_caption = meta.get("ig_caption", pin_description)
            ig_hashtags = meta.get("ig_hashtags", tags_str)

            tiktok_caption = meta.get("tiktok_caption", ig_caption)
            tiktok_hashtags = meta.get("tiktok_hashtags", ig_hashtags)

            yt_title = meta.get("yt_title", pin_title)
            yt_description = meta.get("yt_description", pin_description)
            yt_tags = meta.get("yt_tags", "")

            x_text = meta.get("x_text", pin_title)
            x_hashtags = meta.get("x_hashtags", "")

            fb_caption = meta.get("fb_caption", ig_caption)
            fb_hashtags = meta.get("fb_hashtags", ig_hashtags)

            pin_url_to_link = destination_url or ""

            # Compute pin_id so it keeps incrementing across runs
            pin_id = next_pin_id
            next_pin_id += 1

            # Common fields
            duration_sec = video_duration if media_type in ("video", "coloring") else ""

            excel_media_type = "video" if media_type in ("video", "coloring") else media_type

            aspect_ratio = "9:16"  # matches 1080x1920 vertical

            # Platform toggles: default everything to "Y" (you can change in Excel)
            platform_flag = "Y"

            # Build a dict keyed by header name (only for the columns we care about)
            row_dict = {
                "pin_id": pin_id,
                "media_file": media_path_for_excel,
                "media_type": excel_media_type,
                "duration_sec": duration_sec,
                "aspect_ratio": aspect_ratio,
                "board_name": board_name,
                "book_title": book_title,
                "book_url": book_url,
                "campaign_name": campaign_name,
                "destination_url": destination_url,
                "destination_type": destination_type,
                "topic": topic,
                "platform_pinterest": platform_flag,
                "platform_instagram": platform_flag,
                "platform_tiktok": platform_flag,
                "platform_youtube": platform_flag,
                "platform_x": platform_flag,
                "platform_facebook": platform_flag,

                # Pinterest
                "pin_title": pin_title,
                "pin_description": pin_description,
                "pin_tags": tags_str,
                "pin_url_to_link": pin_url_to_link,

                # Instagram
                "ig_caption": ig_caption,
                "ig_hashtags": ig_hashtags,

                # TikTok
                "tiktok_caption": tiktok_caption,
                "tiktok_hashtags": tiktok_hashtags,

                # YouTube
                "yt_title": yt_title,
                "yt_description": yt_description,
                "yt_tags": yt_tags,
                "yt_playlist": "",  # fill manually if needed

                # X / Twitter
                "x_text": x_text,
                "x_hashtags": x_hashtags,

                # Facebook
                "fb_caption": fb_caption,
                "fb_hashtags": fb_hashtags,

                "banner_text": banner_text or "",
                "watermark_text": watermark_text or "",
                "notes": error_msg,
            }

            # Create a full row matching the current header order
            row_values = [None] * len(headers)
            for key, val in row_dict.items():
                if key in header_map:
                    col_idx = header_map[key] - 1  # 0-based index for Python list
                    row_values[col_idx] = val

            ws.append(row_values)
            wb.save(output_excel)
        except Exception as e:
            print(f"[ERROR] Failed group {idx}: {e}")
            wb.save(output_excel)  # Save progress even after failures
            continue            
    # output_excel.parent.mkdir(parents=True, exist_ok=True)
    # wb.save(output_excel)
    print(f"[OK] Excel written to: {output_excel}")


def generate_pin_meta_with_gemini(
    campaign_name: str,
    destination_url: str,
    destination_type: str,
    topic: str | None,
    page_label: str | None,
    base_tags: str | None,
    pin_role: str = "asset",
    cta_lines: list[str] | None = None,
) -> dict:
    """
    Calls Gemini via GeminiPool to generate pin_title, pin_description, tags[],
    plus multi-platform metadata. Returns dict; may raise or return fallback if something fails.
    """

    print(f"[INFO] Generating social meta with Gemini for page: {page_label!r}")

    global gemini_pool
    if gemini_pool is None:
        raise RuntimeError("gemini_pool is not initialised")

    base_tags = (base_tags or "").strip()

    cta_lines = cta_lines or []

    context = {
        "campaign_name": campaign_name,
        "destination_url": destination_url,
        "destination_type": destination_type,  # product | video | article | service | other
        "topic": topic or "",
        "creative_label": page_label or "",
        "pin_role": pin_role,
        "base_tags": (base_tags or "").strip(),
        "cta_lines": cta_lines,
    }

    prompt = f"""{GENERIC_PROMO_SYSTEM_INSTRUCTION}

    Context (JSON):
    {json.dumps(context, ensure_ascii=False, indent=2)}
    """

    raw = gemini_pool.generate_text(
        prompt=prompt,
        model=DEFAULT_MODEL,
        temperature=DEFAULT_TEMPERATURE,
        max_output_tokens=DEFAULT_MAX_TOKENS,
    )

    if not raw:
        raise RuntimeError("Empty response from Gemini for social meta")

    raw = raw.strip()

    if not raw.startswith("{"):
        first = raw.find("{")
        last = raw.rfind("}")
        if first != -1 and last != -1 and last > first:
            raw = raw[first:last + 1]

    data = json.loads(raw)

    # Helper to normalize hashtags or tag lists
    def join_hash_list(val, sep: str = " "):
        if isinstance(val, list):
            items = [str(t).strip() for t in val if str(t).strip()]
            return sep.join(items)
        return str(val).strip()

    def join_tag_list(val):
        if isinstance(val, list):
            items = [str(t).strip() for t in val if str(t).strip()]
            return ", ".join(items)
        return str(val).strip()

    pin_title = str(data.get("pin_title", "")).strip()
    pin_description = str(data.get("pin_description", "")).strip()
    tags_str = join_tag_list(data.get("tags", []))

    if not pin_title:
        raise RuntimeError("Gemini response missing pin_title")

    # Instagram
    ig_caption = str(data.get("instagram_caption", "")).strip()
    ig_hashtags = join_hash_list(data.get("instagram_hashtags", []))

    # TikTok
    tiktok_caption = str(data.get("tiktok_caption", "")).strip()
    tiktok_hashtags = join_hash_list(data.get("tiktok_hashtags", []))

    # YouTube
    yt_title = str(data.get("youtube_title", "")).strip()
    yt_description = str(data.get("youtube_description", "")).strip()
    yt_tags = join_tag_list(data.get("youtube_tags", []))

    # X / Twitter
    x_text = str(data.get("x_text", "")).strip()
    x_hashtags = join_hash_list(data.get("x_hashtags", []))

    # Facebook
    fb_caption = str(data.get("facebook_caption", "")).strip()
    fb_hashtags = join_hash_list(data.get("facebook_hashtags", []))

    # Fill reasonable fallbacks if any are missing
    if not pin_description:
        pin_description = (
            f"Download and print this coloring page from the '{book_title}' digital coloring book. "
            f"Perfect for creative, screen-free time at home or in the classroom. Instant digital download."
        )

    if not tags_str:
        tags_str = "#coloringpages, #printablecoloring, #kidsactivities, #digitaldownload"

    if not ig_caption:
        ig_caption = pin_description
    if not ig_hashtags:
        ig_hashtags = tags_str

    if not tiktok_caption:
        tiktok_caption = ig_caption
    if not tiktok_hashtags:
        tiktok_hashtags = ig_hashtags

    if not yt_title:
        yt_title = pin_title
    if not yt_description:
        yt_description = pin_description
    if not yt_tags:
        yt_tags = tags_str

    if not x_text:
        x_text = pin_title
    if not x_hashtags:
        x_hashtags = tags_str

    if not fb_caption:
        fb_caption = ig_caption
    if not fb_hashtags:
        fb_hashtags = ig_hashtags

    return {
        "pin_title": pin_title,
        "pin_description": pin_description,
        "tags_str": tags_str,
        "ig_caption": ig_caption,
        "ig_hashtags": ig_hashtags,
        "tiktok_caption": tiktok_caption,
        "tiktok_hashtags": tiktok_hashtags,
        "yt_title": yt_title,
        "yt_description": yt_description,
        "yt_tags": yt_tags,
        "x_text": x_text,
        "x_hashtags": x_hashtags,
        "fb_caption": fb_caption,
        "fb_hashtags": fb_hashtags,
    }


def fallback_pin_meta(
    campaign_name: str,
    destination_url: str,
    destination_type: str,
    page_label: str | None,
    base_tags: str | None,
    cta_lines: list[str] | None = None,
) -> dict:

    print("[INFO] Using fallback social metadata generation.")
    base_tags = (base_tags or "").strip()
    if base_tags:
        tags_str = base_tags
    else:
        tags_str = "#coloringpages, #printablecoloring, #kidsactivities, #digitaldownload"


    # Reuse across platforms
    # Reuse across platforms

    # y_tb_desc_add = """
    # 📥 Download the Coloring Book (Digital Soft Copy)

    # 👉 Get the PDF / Printable Pages Here:
    # https://goodsandgift.com/product-category/creative-crafting/

    # https://kishna01.gumroad.com/

    # 🖍️ Grab My Physical Coloring Books on Amazon

    # 📘 Creative Cubs Coloring Book Series:
    # https://www.amazon.com/dp/B0G1TK51V4

    # 🌐 Free Coloring Pages
    # Color online or download free pages at:
    # https://coloring.readernook.com/
    # """
    title = campaign_name
    cta_lines = cta_lines or []
    cta_block = "\n".join([line.strip() for line in cta_lines if str(line).strip()])

    if not cta_block and destination_url:
        cta_block = f"👉 Link: {destination_url}"

    desc = (
        f"Check out: '{campaign_name}'. "
        f"{'Preview: ' + page_label + '. ' if page_label else ''}"
        f"Learn more at the link."
    )

    yt_description = desc + "\n\n" + cta_block if cta_block else desc
    fb_caption = desc + "\n\n" + cta_block if cta_block else desc
    tiktok_caption = desc + ("\n" + destination_url if destination_url else "")

    ig_caption = desc
    ig_hashtags = tags_str
    tiktok_hashtags = tags_str
    yt_title = title
    yt_tags = "coloring pages, printable coloring, kids activities, digital download"
    x_text = title
    x_hashtags = "#coloringpages #printablecoloring #kidsactivities"
    fb_hashtags = tags_str

    return {
        "pin_title": title,
        "pin_description": desc,
        "tags_str": tags_str,
        "ig_caption": ig_caption,
        "ig_hashtags": ig_hashtags,
        "tiktok_caption": tiktok_caption,
        "tiktok_hashtags": tiktok_hashtags,
        "yt_title": yt_title,
        "yt_description": yt_description,
        "yt_tags": yt_tags,
        "x_text": x_text,
        "x_hashtags": x_hashtags,
        "fb_caption": fb_caption,
        "fb_hashtags": fb_hashtags,
    }

def batch_make_video_images_from_dir(
    input_dir: str | Path,
    target_size: str = "portrait",
    fit_mode: str = "contain",
    bg_style: str = "white",
    auto_crop_subject: bool = True,

) -> int:
    """
    Convert all images under `input_dir` (including subdirectories) into
    vertical "video images" using make_video_image, preserving the directory
    structure under a sibling folder named `edit_vid_output`.

    Example:
      input_dir = /path/to/downloads/1.Cute Farm Animals/pages
      output_root = /path/to/downloads/edit_vid_output

      /pages/a/b/cute_cow.png
      -> /edit_vid_output/a/b/cute_cow.jpg

    Returns:
        Number of images successfully processed.
    """

   
    if target_size == "portrait":
        size = (1080, 1920)
    elif target_size == "landscape":
        size = (1920, 1080)
    elif target_size == "square":
        size = (1080, 1080)
    elif target_size == "pinterest":
        size = (1000, 1500)


    input_dir = Path(input_dir).resolve()

    if not input_dir.is_dir():
        raise ValueError(f"Input path is not a directory: {input_dir}")

    # edit_vid_output will live next to the input_dir
    output_root = input_dir.parent / "edit_vid_output"

    print(f"[INFO] Batch video-image conversion")
    print(f"  input_dir   = {input_dir}")
    print(f"  output_root = {output_root}")

    processed_count = 0

    for img_path in input_dir.rglob("*"):
        if not img_path.is_file():
            continue

        if img_path.suffix.lower() not in IMAGE_EXTS:
            continue

        # Preserve relative structure under edit_vid_output
        rel_path = img_path.relative_to(input_dir)
        out_dir = output_root / rel_path.parent

        try:
            make_video_image(
                src=img_path,
                out_dir=out_dir,
                fit_mode=fit_mode,
                bg_style=bg_style,
                auto_crop_subject=auto_crop_subject,
                size=size,
            )
            processed_count += 1
            print(f"[OK] Converted: {img_path} -> {out_dir}")
        except Exception as e:
            print(f"[WARN] Failed to convert {img_path}: {e}")

    print(f"[DONE] Processed {processed_count} images from {input_dir}")
    return processed_count

def main(argv=None):
    parser = argparse.ArgumentParser(description="Generate multi-platform Excel from images/videos (Pinterest + others).")
    parser.add_argument("--images-root", required=True, help="Root images folder (e.g. application/downloads)")
    parser.add_argument("--source-subfolder", default="", help="Subfolder under images-root, e.g. '1.Cute Farm Animals/pages'")
    parser.add_argument("--output-excel", default="master_shorts_uploader_data.xlsx", help="Output Excel filename")
    parser.add_argument("--media-type", choices=["image", "video", "coloring"], default="image")
    # parser.add_argument("--pin-url", choices=["gng", "amazon", "gumroad","readernook"], default="gng")
    parser.add_argument("--max-pins", type=int, default=0, help="Max pins to generate (0 = all)")

    parser.add_argument(
        "--fit-mode",
        choices=["contain", "cover"],
        default="contain",
        help="How to fit the image into the Pinterest canvas (contain=no crop, cover=crop)."
    )
    parser.add_argument(
        "--bg-style",
        choices=["white", "blur"],
        default="white",
        help="Background style behind the page (white or blurred image)."
    )
    parser.add_argument(
        "--text-shadow",
        choices=["yes", "no"],
        default="yes",
        help="Enable or disable text shadow for banner and watermark."
    )

    parser.add_argument(
        "--use-gemini",
        choices=["yes", "no"],
        default="no",
        help="Use Gemini API to generate multi-platform titles/descriptions/tags.",
    )

    parser.add_argument(
        "--add-bg-music",
        choices=["yes", "no"],
        default="no",
        help="Add background music to video pins.",
    )

    parser.add_argument(
        "--auto-crop-subject",
        choices=["yes", "no"],
        default="yes",
        help="If 'yes', automatically crop around the drawing before creating the Pinterest image."
    )

    parser.add_argument(
        "--video-duration",
        type=float,
        default=10.0,
        help="Video duration in seconds for each vertical video.",
    )
    parser.add_argument(
        "--video-fps",
        type=int,
        default=30,
        help="Frames per second for video pins.",
    )

    parser.add_argument(
        "--video-style",
        choices=["single", "flipbook", "slideshow"],
        default="single",
        help="Style for video pins: single image, flipbook (multi-page), or slideshow.",
    )
    parser.add_argument(
        "--pages-per-video",
        type=int,
        default=10,
        help="Number of pages per video for flipbook/slideshow styles.",
    )

    # These are now OPTIONAL (can be provided by config file)
    parser.add_argument("--book-title", default="", help="Book title to use in metadata (overrides config)")
    parser.add_argument("--book-url", default="", help="URL to link from the post (overrides config)")
    parser.add_argument("--board-name", default="", help="Pinterest board name (overrides config)")
    parser.add_argument("--banner-text", default="", help="Text banner at top of image (overrides config)")
    parser.add_argument("--watermark-text", default="", help="Watermark text at bottom of image (overrides config)")

    args = parser.parse_args(argv)

    images_root = Path(args.images_root).resolve()
    source_subfolder = args.source_subfolder or None
    output_excel = Path(args.output_excel).resolve()

    max_pins = args.max_pins if args.max_pins > 0 else None

    print("[INFO] Starting multi-platform Excel generation...")
    print(f"  images_root      = {images_root}")
    print(f"  source_subfolder = {source_subfolder}")
    print(f"  output_excel     = {output_excel}")
    print(f"  media_type       = {args.media_type}")
    print(f"  max_pins         = {max_pins}")

    text_shadow = (args.text_shadow.lower() == "yes")
    auto_crop_subject = (args.auto_crop_subject.lower() == "yes")
    use_gemini = (args.use_gemini.lower() == "yes")
    add_bg_music = (args.add_bg_music.lower() == "yes")

    # pin_url = args.pin_url
    book_url = args.book_url



    global gemini_pool
    if use_gemini:
        GEM_STATE = Path(__file__).resolve().parent / ".gemini_pool_state_pinterest.json"
        gemini_pool = GeminiPool(
            api_keys=None,          # load from env/config, same as your SEO script
            per_key_rpm=25,
            state_path=str(GEM_STATE),
            autosave_every=3,
        )

    build_excel(
        images_root=images_root,
        source_subfolder=source_subfolder,
        output_excel=output_excel,
        media_type=args.media_type,
        max_pins=max_pins,
        book_title=args.book_title,
        book_url=book_url,
        board_name=args.board_name,
        banner_text=args.banner_text,
        watermark_text=args.watermark_text,
        fit_mode=args.fit_mode,
        bg_style=args.bg_style,
        text_shadow=text_shadow,
        use_gemini=use_gemini,
        base_tags=None,
        auto_crop_subject=auto_crop_subject,
        video_style=args.video_style,
        pages_per_video=args.pages_per_video,
        video_duration=args.video_duration,
        video_fps=args.video_fps,
        add_bg_music=add_bg_music,
        # pin_url=args.pin_url,
    )

    print("[DONE] Multi-platform Excel generation finished.")


if __name__ == "__main__":
    main()
