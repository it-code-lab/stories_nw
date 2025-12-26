# app.py
from flask import Flask, render_template, request, jsonify, send_from_directory
import os, json, uuid
from flask import request
import csv, io
import time
from pathlib import Path
from . import quiz_bp  # use the blueprint you defined

import  tempfile, os
from flask import Blueprint, jsonify, send_file
from werkzeug.utils import secure_filename

try:
    import openpyxl  # pip install openpyxl
except ImportError:
    openpyxl = None



# app = Flask(__name__)
APP_ROOT = os.path.dirname(os.path.abspath(__file__))
QUIZ_DIR = os.path.join(APP_ROOT, "quizzes")
UPLOAD_DIR = os.path.join(APP_ROOT, "uploads")

os.makedirs(QUIZ_DIR, exist_ok=True)
os.makedirs(os.path.join(UPLOAD_DIR, "images"), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_DIR, "audio"), exist_ok=True)

ALLOWED_IMPORTS = {'.xlsx', '.xls', '.csv'}

def _ext(name): 
    return os.path.splitext(name or '')[1].lower()

@quiz_bp.route('/api/template.csv', methods=['GET'])
def quiz_template():
    """
    Provide a simple CSV template with headers users can fill.
    """
    headers = [
        'type', 'question', 'opt1', 'opt2', 'opt3', 'opt4',
        'correct', 'difficulty', 'explanation', 'timer(sec)',
        'qImages', 'opt1_img', 'opt2_img', 'opt3_img', 'opt4_img'
    ]
    sample = [
        ['mcq','Which planet is known as the Red Planet?','Mercury','Venus','Earth','Mars','4','easy','','12','https://…','','','',''],
        ['single','Largest mammal on Earth?','','','','','','easy','','10','https://…','','','','','Blue whale'],
    ]
    mem = io.StringIO()
    w = csv.writer(mem)
    w.writerow(headers)
    for r in sample:
        # ensure length matches headers; pad if needed
        if len(r) < len(headers): r += ['']*(len(headers)-len(r))
        w.writerow(r)
    mem.seek(0)
    return send_file(
        io.BytesIO(mem.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='quiz_template.csv'
    )

@quiz_bp.route('/api/import_excel', methods=['POST'])
def quiz_import_excel():
    """
    Accept .xlsx/.xls/.csv and return parsed rows as `questions: [...]`.
    Column names handled in frontend mapper; here we produce per-row dicts keyed by header.
    """
    f = request.files.get('file')
    if not f or _ext(f.filename) not in ALLOWED_IMPORTS:
        return jsonify({'error':'Please upload .xlsx or .csv'}), 400

    ext = _ext(f.filename)
    rows = []

    if ext == '.csv':
        raw = f.stream.read()

        # Try UTF-8 first (CSV UTF-8), then fall back to Windows-1252 (Excel ANSI CSV)
        try:
            text = raw.decode('utf-8-sig')  # strict by default
        except UnicodeDecodeError:
            text = raw.decode('cp1252')     # common for Excel CSV on Windows

        reader = csv.DictReader(io.StringIO(text))

        def normalize_quotes(s: str) -> str:
            # Optional: normalize “smart quotes” to plain ones
            return (s.replace("’", "'")
                    .replace("‘", "'")
                    .replace("“", '"')
                    .replace("”", '"'))

        for row in reader:
            rows.append({
                (k or '').strip(): normalize_quotes((v or '').strip())
                for k, v in row.items()
            })

        return jsonify({'questions': rows})


    # Excel path
    if openpyxl is None:
        return jsonify({'error':'openpyxl not installed on server'}), 500

    with tempfile.NamedTemporaryFile(suffix=ext, delete=True) as tmp:
        f.save(tmp.name)
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        ws = wb.active

        # header row
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append((cell.value or '').strip() if isinstance(cell.value, str) else str(cell.value or '').strip())

        # subsequent rows
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = {}
            for i, h in enumerate(headers):
                val = r[i] if i < len(r) else ''
                if isinstance(val, (int, float)):
                    row[h] = str(val).strip()
                elif isinstance(val, str):
                    row[h] = val.strip()
                else:
                    row[h] = ''
            # ignore totally empty lines
            if any(v for v in row.values()):
                rows.append(row)

    return jsonify({'questions': rows})

@quiz_bp.route("/api/upload_background", methods=["POST"])
def upload_background():
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "Missing file"}), 400

    if not (f.mimetype or "").startswith("image/"):
        return jsonify({"ok": False, "error": "Only images allowed"}), 400

    bg_dir = Path(APP_ROOT) / "static" / "backgrounds"
    bg_dir.mkdir(parents=True, exist_ok=True)

    # always overwrite same file
    out_path = bg_dir / "uploaded_bg.jpg"
    f.save(str(out_path))

    v = int(time.time())
    return jsonify({
        "ok": True,
        "url": f"/quiz/static/backgrounds/uploaded_bg.jpg?v={v}"
    })

@quiz_bp.route("/api/media")
def list_media():
    bg_dir = os.path.join(APP_ROOT, "static/backgrounds")
    music_dir = os.path.join(APP_ROOT, "static/music")

    def list_files(path, exts):
        try:
            out = []
            for fn in os.listdir(path):
                low = fn.lower()
                if any(low.endswith(e) for e in exts):
                    out.append(f"/quiz/static/{os.path.basename(path)}/{fn}")
            return out
        except FileNotFoundError:
            return []

    return jsonify({
        "videos": list_files(bg_dir, [".mp4", ".webm", ".mov"]),
        "images": list_files(bg_dir, [".jpg", ".jpeg", ".png", ".webp"]),
        "music":  list_files(music_dir, [".mp3", ".wav"])
    })


@quiz_bp.post("/import/csv")
def import_csv():
    f = request.files.get('file')
    if not f:
        return {"error":"no file"}, 400
    text = f.stream.read().decode('utf-8', errors='ignore')
    rows = list(csv.DictReader(io.StringIO(text)))
    out = []
    for r in rows:
        t = (r.get('type') or 'mcq').strip().lower()
        base = {
        'id': f"q_{len(out)+1}",
        'type': t,
        'difficulty': (r.get('difficulty') or 'easy').lower(),
        'text': r.get('question') or '',
        'images': [p.strip() for p in (r.get('images') or '').split('|') if p.strip()],
        'timerSec': int(r.get('timerSec') or 0) or None,
        'explanation': r.get('explanation') or ''
    }
    if t=='mcq':
        opts = [r.get('optionA') or '', r.get('optionB') or '', r.get('optionC') or '', r.get('optionD') or '']
        cor = (r.get('correct') or 'A').strip().upper()
        idx = {'A':0,'B':1,'C':2,'D':3}.get(cor,0)
        base.update({'options':opts,'correctIndex':idx})
    else:
        base.update({'answer': r.get('answer') or ''})
        out.append(base)
        return {"questions": out}

@quiz_bp.get("/")
def dashboard():
    return render_template("dashboard.html")

@quiz_bp.get("/builder")
def builder():
    return render_template("builder.html")

@quiz_bp.post("/api/quizzes")
def save_quiz():
    data = request.get_json(force=True)
    qid = data.get("id") or f"quiz_{uuid.uuid4().hex}"
    data["id"] = qid
    with open(os.path.join(QUIZ_DIR, f"{qid}.json"), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True, "id": qid})

@quiz_bp.get("/api/quizzes/<qid>")
def get_quiz(qid):
    path = os.path.join(QUIZ_DIR, f"{qid}.json")
    if not os.path.exists(path):
        return jsonify({"error": "not found"}), 404
    with open(path, "r", encoding="utf-8") as f:
        return jsonify(json.load(f))

@quiz_bp.post("/api/upload")
def upload():
    file = request.files.get("file")
    kind = request.form.get("type", "images")
    dest_dir = os.path.join(UPLOAD_DIR, kind)
    os.makedirs(dest_dir, exist_ok=True)
    fname = f"{uuid.uuid4().hex}_{file.filename}"
    file.save(os.path.join(dest_dir, fname))
    return jsonify({"url": f"/quiz/uploads/{kind}/{fname}"})

@quiz_bp.get("/uploads/<kind>/<fname>")
def serve_upload(kind, fname):
    return send_from_directory(os.path.join(UPLOAD_DIR, kind), fname)

@quiz_bp.get("/play")
def play():
    # Accept ?id=quiz_... or let user choose a local JSON in the UI
    return render_template("player.html")

# if __name__ == "__main__":
#     app.run(debug=True)
