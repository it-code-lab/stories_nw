# pin_overlay_batch.py
import json
import os
import re
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List

from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_MEDIA_ROOT = Path("pinterest_uploads")


def _split_csv_tags(tag_string: str) -> List[str]:
    """
    Split comma-separated tags safely, trimming whitespace and ignoring empties.
    """
    if not tag_string:
        return []
    return [t.strip() for t in tag_string.split(",") if t.strip()]

def _normalize_tag(tag: str, keep_case: bool = False) -> str:
    """
    Normalize a tag:
    - trim
    - remove leading '#'
    - remove special chars (keep letters/numbers/underscore/space)
    - collapse spaces
    - convert spaces -> underscore for hashtags
    """
    tag = tag.strip()
    tag = tag.lstrip("#").strip()

    # Remove punctuation/special chars but keep word chars and spaces
    tag = re.sub(r"[^\w\s]", "", tag)

    # Collapse whitespace
    tag = re.sub(r"\s+", " ", tag).strip()

    if not keep_case:
        tag = tag.lower()

    return tag

def _to_hashtag(tag: str) -> str:
    """
    Convert normalized tag to a hashtag form: spaces -> underscore.
    """
    tag = tag.replace(" ", "_")
    return f"#{tag}" if tag else ""

def format_tags_for_platforms(
    comma_tags: str,
    *,
    ig_limit: int = 12,
    tt_limit: int = 5,
    fb_limit: int = 10,
    yt_limit: int = 15,
    pin_limit: int = 10,
) -> Dict[str, str]:
    """
    Returns correctly formatted tags/keywords for:
    Pinterest, Instagram, Facebook, TikTok, YouTube
    """

    raw = _split_csv_tags(comma_tags)

    # Deduplicate while preserving order
    seen = set()
    cleaned: List[str] = []
    for t in raw:
        n = _normalize_tag(t)
        if not n:
            continue
        if n in seen:
            continue
        seen.add(n)
        cleaned.append(n)

    # Pinterest / YouTube: keywords (no #), comma-separated
    pinterest_keywords = cleaned[:pin_limit]
    youtube_keywords = cleaned[:yt_limit]

    # Hashtag platforms: hashtags, space-separated
    ig_tags = [_to_hashtag(t) for t in cleaned[:ig_limit]]
    tt_tags = [_to_hashtag(t) for t in cleaned[:tt_limit]]
    fb_tags = [_to_hashtag(t) for t in cleaned[:fb_limit]]

    # Filter empty hashtags (just in case)
    ig_tags = [t for t in ig_tags if t]
    tt_tags = [t for t in tt_tags if t]
    fb_tags = [t for t in fb_tags if t]

    return {
        "pinterest": ", ".join(pinterest_keywords),
        "instagram": " ".join(ig_tags),
        "facebook": " ".join(fb_tags),
        "tiktok": " ".join(tt_tags),
        "youtube": ", ".join(youtube_keywords),
    }


# -----------------------------
# Helpers
# -----------------------------

def _derive_out_base(filename: str, title: str, fallback: str) -> str:
    """
    Priority:
      1) filename (strip folders, strip extension)
      2) title
      3) fallback (e.g., pin_12)
    """
    fn = (filename or "").strip()
    if fn:
        fn = fn.replace("\\", "/")
        base = Path(fn).name                 # keep only file name
        base = Path(base).stem               # remove extension
        base = base.strip()
        if base:
            return sanitize_filename(base)

    t = (title or "").strip()
    if t:
        return sanitize_filename(t)

    return sanitize_filename(fallback)

def sanitize_filename(name: str, max_len: int = 140) -> str:
    name = (name or "").strip()
    name = re.sub(r"[^\w\s\-.()]+", "", name, flags=re.UNICODE)
    name = re.sub(r"\s+", "_", name).strip("_")
    if not name:
        name = "pin"
    return name[:max_len]


def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)


def _as_float(x, default=0.0):
    try:
        return float(x)
    except Exception:
        return default


def _as_int(x, default=0):
    try:
        return int(float(x))
    except Exception:
        return default


def _truthy(x) -> bool:
    s = str(x).strip().lower()
    return s in ("1", "true", "yes", "y", "on")


def _hex_to_rgba(hex_color: str, alpha: float = 1.0) -> Tuple[int, int, int, int]:
    h = (hex_color or "#000000").strip().lstrip("#")
    if len(h) == 3:
        h = "".join([c + c for c in h])
    r = int(h[0:2], 16)
    g = int(h[2:4], 16)
    b = int(h[4:6], 16)
    a = int(max(0, min(1, alpha)) * 255)
    return (r, g, b, a)


def _clamp(n, a, b):
    return max(a, min(b, n))


def _find_font(font_family: str, font_weight: int, fonts_dir: Path) -> Optional[Path]:
    """
    Tries to find a .ttf in ./fonts. Put your fonts here if you want exact matching.
    Otherwise we fall back to PIL default.
    Suggested: add Oswald / Montserrat / Anton / BebasNeue to ./fonts.
    """
    fam = (font_family or "").strip().lower()
    candidates = []

    # Common naming patterns:
    # Oswald-Bold.ttf, Oswald-ExtraBold.ttf, Montserrat-Black.ttf, etc.
    if fam:
        candidates += list(fonts_dir.glob(f"*{fam}*.ttf"))
        candidates += list(fonts_dir.glob(f"*{fam}*.otf"))

    # Weight hint
    w = int(font_weight or 400)
    weight_keywords = []
    if w >= 900:
        weight_keywords = ["black", "extrabold", "ultrabold", "heavy"]
    elif w >= 700:
        weight_keywords = ["bold"]
    elif w >= 600:
        weight_keywords = ["semibold", "demibold"]
    else:
        weight_keywords = ["regular", "medium", "light"]

    scored = []
    for p in candidates:
        name = p.name.lower()
        score = 0
        for kw in weight_keywords:
            if kw in name:
                score += 2
        if "italic" in name:
            score -= 1
        scored.append((score, p))

    if scored:
        scored.sort(key=lambda t: t[0], reverse=True)
        return scored[0][1]

    return None


def _load_font(font_family: str, font_weight: int, size: int, fonts_dir: Path) -> ImageFont.FreeTypeFont:
    fp = _find_font(font_family, font_weight, fonts_dir)
    if fp and fp.exists():
        try:
            return ImageFont.truetype(str(fp), size=size)
        except Exception:
            pass
    # fallback
    try:
        return ImageFont.truetype("arial.ttf", size=size)
    except Exception:
        return ImageFont.load_default()


def _wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont, max_w: int) -> List[str]:
    """
    Simple word wrap by pixel width.
    """
    words = (text or "").split()
    if not words:
        return [""]

    lines = []
    cur = words[0]
    for w in words[1:]:
        test = f"{cur} {w}"
        if draw.textlength(test, font=font) <= max_w:
            cur = test
        else:
            lines.append(cur)
            cur = w
    lines.append(cur)
    return lines


def _fit_font_size(draw: ImageDraw.ImageDraw, text: str, font_family: str, font_weight: int,
                   box_w: int, box_h: int, start_size: int, min_size: int,
                   line_height: float, fonts_dir: Path) -> int:
    """
    Auto-fit font size so wrapped text fits inside box.
    """
    size = start_size
    while size >= min_size:
        font = _load_font(font_family, font_weight, size, fonts_dir)
        lines = _wrap_text(draw, text, font, max_w=box_w)
        # estimate height
        ascent, descent = font.getmetrics()
        lh_px = int((ascent + descent) * max(0.9, min(2.0, line_height)))
        total_h = lh_px * len(lines)
        if total_h <= box_h:
            return size
        size -= 2
    return max(min_size, 8)


# -----------------------------
# Core rendering
# -----------------------------
@dataclass
class RenderJob:
    bg_path: Path
    project_json_path: Path
    headline: str
    subhead: str
    logo_path: Optional[Path]
    title: str
    description: str
    pin_url: str
    board_name: str


def load_project_json(p: Path) -> Dict[str, Any]:
    return json.loads(p.read_text(encoding="utf-8"))

def _find_media_file_for_filename(folder: Path, filename: str, exts: set[str]) -> Optional[Path]:
    """
    filename may be:
      - "myvideo.mp4" (exact name)
      - "subdir/myvideo.mp4" (relative path inside folder)
      - full absolute path (we'll accept it if exists)
    Must include extension.
    """
    fn = (filename or "").strip()
    if not fn:
        return None

    fn_norm = fn.replace("\\", "/").strip()

    # Absolute path case
    p = Path(fn)
    if p.is_absolute() and p.exists() and p.is_file() and p.suffix.lower() in exts:
        return p

    # Relative path inside folder
    rel_candidate = (folder / Path(fn_norm)).resolve()
    try:
        # ensure it's under folder (avoid escaping)
        rel_candidate.relative_to(folder.resolve())
        if rel_candidate.exists() and rel_candidate.is_file() and rel_candidate.suffix.lower() in exts:
            return rel_candidate
    except Exception:
        pass

    # Exact filename match anywhere under folder
    target_name = Path(fn_norm).name  # keep only the last segment
    for p in folder.rglob("*"):
        if not p.is_file():
            continue
        if p.suffix.lower() not in exts:
            continue
        if p.name == target_name:
            return p

    return None

def apply_text_overrides(project: Dict[str, Any], headline: str, subhead: str, logo_path: Optional[Path]) -> Dict[str, Any]:
    """
    Override layers by name (case-insensitive):
      - text layer named "headline" gets headline
      - text layer named "subhead" gets subhead
      - image layer named "logo" or "logopath" gets logo_path
    Falls back to old behavior ONLY if named layers aren't found.
    """
    layers = project.get("layers", [])

    def norm(s: Any) -> str:
        return str(s or "").strip().lower()

    # Build quick lookup
    by_name = {}
    for ly in layers:
        n = norm(ly.get("name"))
        if n:
            by_name.setdefault(n, []).append(ly)

    # Apply by name
    found_any = False

    # headline
    if "headline" in by_name:
        for ly in by_name["headline"]:
            if ly.get("type") == "text":
                ly["text"] = headline or ly.get("text", "")
                found_any = True

    # subhead
    if "subhead" in by_name:
        for ly in by_name["subhead"]:
            if ly.get("type") == "text":
                ly["text"] = subhead or ly.get("text", "")
                found_any = True

    # logo / logopath
    logo_layer_names = ("logo", "logopath")
    if logo_path:
        logo_src = str(logo_path).replace("\\", "/")
        for nm in logo_layer_names:
            if nm in by_name:
                for ly in by_name[nm]:
                    if ly.get("type") == "image":
                        ly["imgSrc"] = logo_src
                        found_any = True

    # Optional fallback to legacy “by order” behavior (only if nothing matched by name)
    if not found_any:
        text_layers = [ly for ly in layers if ly.get("type") == "text"]
        img_layers = [ly for ly in layers if ly.get("type") == "image"]

        if text_layers:
            text_layers[0]["text"] = headline or text_layers[0].get("text", "")
        if len(text_layers) > 1:
            text_layers[1]["text"] = subhead or text_layers[1].get("text", "")

        if logo_path and img_layers:
            img_layers[0]["imgSrc"] = str(logo_path).replace("\\", "/")

    return project

def apply_text_overrides_old(project: Dict[str, Any], headline: str, subhead: str, logo_path: Optional[Path]) -> Dict[str, Any]:
    """
    Override the first 2 text layers (if present) with headline/subhead,
    and first image layer with logo_path if provided.
    Adjust this logic if you want to target layers by name/id.
    """
    layers = project.get("layers", [])
    text_layers = [ly for ly in layers if ly.get("type") == "text"]
    img_layers = [ly for ly in layers if ly.get("type") == "image"]

    if text_layers:
        text_layers[0]["text"] = headline or text_layers[0].get("text", "")
    if len(text_layers) > 1:
        text_layers[1]["text"] = subhead or text_layers[1].get("text", "")

    if logo_path and img_layers:
        # replace blob/null with actual path
        img_layers[0]["imgSrc"] = str(logo_path).replace("\\", "/")

    return project


def render_overlay_png(project: Dict[str, Any], out_png: Path, fonts_dir: Path) -> Tuple[int, int]:
    """
    Creates a transparent PNG containing ONLY overlays (text/logo boxes),
    sized exactly to project format (w,h).
    """
    fmt = project.get("format") or {}
    W = _as_int(fmt.get("w"), 1280)
    H = _as_int(fmt.get("h"), 720)

    overlay = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    for ly in project.get("layers", []):
        if not ly.get("visible", True):
            continue

        box = ly.get("box") or {}
        x = int(_as_float(box.get("x"), 0) * W)
        y = int(_as_float(box.get("y"), 0) * H)
        w = int(_as_float(box.get("w"), 0.5) * W)
        h = int(_as_float(box.get("h"), 0.2) * H)

        if w <= 1 or h <= 1:
            continue

        if ly.get("type") == "text":
            style = ly.get("style") or {}
            text = ly.get("text") or ""
            if _truthy(style.get("uppercase", False)):
                text = text.upper()

            font_family = style.get("fontFamily", "Arial")
            font_weight = _as_int(style.get("fontWeight"), 700)
            font_size = _as_int(style.get("fontSize"), 56)
            line_height = _as_float(style.get("lineHeight"), 1.05)
            auto_fit = _truthy(style.get("autoFit", True))

            fill = style.get("fill", "#ffffff")
            align = (style.get("align") or "center").lower()

            stroke_on = _truthy(style.get("strokeOn", False))
            stroke_color = style.get("strokeColor", "#000000")
            stroke_width = _as_int(style.get("strokeWidth"), 0)

            shadow_on = _truthy(style.get("shadowOn", False))
            shadow_color = style.get("shadowColor", "#000000")
            shadow_blur = _as_int(style.get("shadowBlur"), 0)  # blur not implemented (simple offset)
            shadow_dx = _as_int(style.get("shadowDx"), 4)
            shadow_dy = _as_int(style.get("shadowDy"), 4)

            box_on = _truthy(style.get("boxOn", False))
            box_color = style.get("boxColor", "#000000")
            box_opacity = _as_float(style.get("boxOpacity"), 0.35)
            box_pad = _as_int(style.get("boxPad"), 12)
            box_radius = _as_int(style.get("boxRadius"), 0)

            # background box behind text
            if box_on:
                bx0 = x
                by0 = y
                bx1 = x + w
                by1 = y + h
                rgba = _hex_to_rgba(box_color, box_opacity)
                if box_radius > 0:
                    draw.rounded_rectangle([bx0, by0, bx1, by1], radius=box_radius, fill=rgba)
                else:
                    draw.rectangle([bx0, by0, bx1, by1], fill=rgba)

                # shrink text area by padding
                tx0 = x + box_pad
                ty0 = y + box_pad
                tw = max(1, w - 2 * box_pad)
                th = max(1, h - 2 * box_pad)
            else:
                tx0, ty0, tw, th = x, y, w, h

            # auto-fit
            if auto_fit:
                fitted = _fit_font_size(
                    draw, text, font_family, font_weight,
                    box_w=tw, box_h=th,
                    start_size=font_size, min_size=18,
                    line_height=line_height,
                    fonts_dir=fonts_dir
                )
                font_size = fitted

            font = _load_font(font_family, font_weight, font_size, fonts_dir)

            lines = _wrap_text(draw, text, font, max_w=tw)
            ascent, descent = font.getmetrics()
            lh_px = int((ascent + descent) * max(0.9, min(2.0, line_height)))
            total_h = lh_px * len(lines)

            # vertical center inside text box
            # cur_y = ty0 + max(0, (th - total_h) // 2)

            # AFTER:
            valign = (style.get("valign") or "center").lower()  # "top" | "center" | "bottom"
            if valign == "top":
                cur_y = ty0
            elif valign == "bottom":
                cur_y = ty0 + max(0, th - total_h)
            else:
                cur_y = ty0 + max(0, (th - total_h) // 2)

            for line in lines:
                line_w = int(draw.textlength(line, font=font))
                if align == "left":
                    cur_x = tx0
                elif align == "right":
                    cur_x = tx0 + (tw - line_w)
                else:
                    cur_x = tx0 + (tw - line_w) // 2

                # shadow (simple offset)
                if shadow_on:
                    sc = _hex_to_rgba(shadow_color, 1.0)
                    draw.text((cur_x + shadow_dx, cur_y + shadow_dy), line, font=font, fill=sc)

                # stroke (PIL supports stroke in draw.text)
                fc = _hex_to_rgba(fill, 1.0)
                if stroke_on and stroke_width > 0:
                    st = _hex_to_rgba(stroke_color, 1.0)
                    draw.text((cur_x, cur_y), line, font=font, fill=fc,
                              stroke_width=stroke_width, stroke_fill=st)
                else:
                    draw.text((cur_x, cur_y), line, font=font, fill=fc)

                cur_y += lh_px

        elif ly.get("type") == "image":
            img_src = (ly.get("imgSrc") or "").strip()
            if not img_src:
                continue

            opacity = _as_float(ly.get("opacity"), 1.0)
            fit = (ly.get("fit") or "contain").lower()

            src_path = Path(img_src)
            if not src_path.is_absolute():
                # relative path is relative to project_json folder caller will ensure
                src_path = Path(img_src)

            if not src_path.exists():
                continue

            try:
                logo = Image.open(src_path).convert("RGBA")
            except Exception:
                continue

            # fit into layer box
            lw, lh = logo.size
            if fit == "cover":
                # cover box
                scale = max(w / lw, h / lh)
            else:
                # contain
                scale = min(w / lw, h / lh)
            nw = max(1, int(lw * scale))
            nh = max(1, int(lh * scale))
            logo = logo.resize((nw, nh), Image.LANCZOS)

            # center in box
            px = x + (w - nw) // 2
            py = y + (h - nh) // 2

            if opacity < 1.0:
                alpha = logo.split()[-1]
                alpha = alpha.point(lambda a: int(a * _clamp(opacity, 0, 1)))
                logo.putalpha(alpha)

            overlay.alpha_composite(logo, (px, py))

    ensure_dir(out_png.parent)
    overlay.save(out_png, "PNG")
    return (W, H)

def _lock_title_subhead_gap(project: dict, gap_px: int = 18,
                           headline_name: str = "headline", subhead_name: str = "subhead") -> dict:
    fmt = project.get("format") or {}
    H = int(fmt.get("h", 720))

    headline = None
    subhead = None
    for ly in project.get("layers", []):
        if ly.get("type") != "text":
            continue
        nm = (ly.get("name") or "").lower()
        if nm == headline_name.lower():
            headline = ly
        elif nm == subhead_name.lower():
            subhead = ly

    if not headline or not subhead:
        return project

    hb = headline.get("box") or {}
    sb = subhead.get("box") or {}

    head_h = float(hb.get("h", 0)) * H
    sub_y  = float(sb.get("y", 0)) * H

    # Make headline box bottom = subhead top - gap
    new_head_y_px = (sub_y - gap_px) - head_h
    new_head_y = max(0.0, min(new_head_y_px / H, 1.0 - float(hb.get("h", 0))))

    hb["y"] = new_head_y
    headline["box"] = hb

    # Use valigns so text sticks to edges (not centered)
    headline.setdefault("style", {})["valign"] = "bottom"
    subhead.setdefault("style", {})["valign"] = "top"

    return project

def render_image_pin(bg_path: Path, project: Dict[str, Any], out_path: Path, fonts_dir: Path):
    fmt = project.get("format") or {}
    W = _as_int(fmt.get("w"), 1280)
    H = _as_int(fmt.get("h"), 720)

    # bgCrop values in source pixels (sx,sy,sw,sh)
    bg_mode = (project.get("bgMode") or "cover").lower()
    bg_dim = _as_float(project.get("bgDim"), 0.0)
    bg_crop = project.get("bgCrop")

    bg = Image.open(bg_path).convert("RGB")
    src_w, src_h = bg.size

    # Create final background frame of size W,H
    bg_frame = Image.new("RGB", (W, H), (0, 0, 0))

    if bg_mode == "crop" and bg_crop:
        sx = int(_as_float(bg_crop.get("sx"), 0))
        sy = int(_as_float(bg_crop.get("sy"), 0))
        sw = int(_as_float(bg_crop.get("sw"), src_w))
        sh = int(_as_float(bg_crop.get("sh"), src_h))
        sx = _clamp(sx, 0, src_w - 1)
        sy = _clamp(sy, 0, src_h - 1)
        sw = _clamp(sw, 1, src_w - sx)
        sh = _clamp(sh, 1, src_h - sy)
        crop = bg.crop((sx, sy, sx + sw, sy + sh))
        crop = crop.resize((W, H), Image.LANCZOS)
        bg_frame.paste(crop, (0, 0))
    elif bg_mode in ("contain", "contain_blur"):
        # contain, with letterbox
        scale = min(W / src_w, H / src_h)
        nw, nh = int(src_w * scale), int(src_h * scale)
        fg = bg.resize((nw, nh), Image.LANCZOS)

        if bg_mode == "contain_blur":
            # crude blur fill: just resize-crop as fill; real blur optional
            fill = bg.copy()
            scale2 = max(W / src_w, H / src_h)
            fw, fh = int(src_w * scale2), int(src_h * scale2)
            fill = fill.resize((fw, fh), Image.LANCZOS)
            # center crop to W,H
            cx = (fw - W) // 2
            cy = (fh - H) // 2
            fill = fill.crop((cx, cy, cx + W, cy + H))
            bg_frame.paste(fill, (0, 0))

        px = (W - nw) // 2
        py = (H - nh) // 2
        bg_frame.paste(fg, (px, py))
    else:
        # cover
        scale = max(W / src_w, H / src_h)
        nw, nh = int(src_w * scale), int(src_h * scale)
        im = bg.resize((nw, nh), Image.LANCZOS)
        cx = (nw - W) // 2
        cy = (nh - H) // 2
        im = im.crop((cx, cy, cx + W, cy + H))
        bg_frame.paste(im, (0, 0))

    # dim background
    if bg_dim > 0:
        dim = Image.new("RGBA", (W, H), (0, 0, 0, int(_clamp(bg_dim, 0, 0.95) * 255)))
        bg_rgba = bg_frame.convert("RGBA")
        bg_rgba.alpha_composite(dim, (0, 0))
        bg_frame = bg_rgba.convert("RGB")

    # overlays
    tmp_overlay = out_path.with_suffix(".overlay.png")
    render_overlay_png(project, tmp_overlay, fonts_dir)
    overlay = Image.open(tmp_overlay).convert("RGBA")

    final = bg_frame.convert("RGBA")
    final.alpha_composite(overlay, (0, 0))

    ensure_dir(out_path.parent)
    # output JPG for Pinterest images (smaller)
    final_rgb = final.convert("RGB")
    final_rgb.save(out_path, "JPEG", quality=92, optimize=True)

    # cleanup overlay
    try:
        tmp_overlay.unlink(missing_ok=True)
    except Exception:
        pass


def run_ffmpeg(cmd: List[str]):
    p = subprocess.run(cmd, capture_output=True, text=True)
    if p.returncode != 0:
        raise RuntimeError(f"FFmpeg failed:\nSTDOUT:\n{p.stdout}\nSTDERR:\n{p.stderr}")


def render_video_pin(bg_video: Path, project: Dict[str, Any], out_path: Path, fonts_dir: Path):
    """
    Approach:
    - render overlay PNG with transparent background
    - ffmpeg:
      1) scale/crop background video to W,H
      2) apply dim overlay via drawbox (or color overlay)
      3) overlay PNG on top
      4) keep audio if present
    """
    fmt = project.get("format") or {}
    W = _as_int(fmt.get("w"), 1080)
    H = _as_int(fmt.get("h"), 1920)

    bg_mode = (project.get("bgMode") or "cover").lower()
    bg_dim = _as_float(project.get("bgDim"), 0.0)

    tmp_overlay = out_path.with_suffix(".overlay.png")
    render_overlay_png(project, tmp_overlay, fonts_dir)

    # video background transform
    # - cover: scale then crop center
    # - contain: scale then pad
    # - crop: if bgCrop exists we can't apply source-pixel crop reliably without probing; treat as cover.
    if bg_mode in ("contain", "contain_blur"):
        vf_bg = (
            f"scale=w={W}:h={H}:force_original_aspect_ratio=decrease,"
            f"pad={W}:{H}:(ow-iw)/2:(oh-ih)/2"
        )
    else:
        vf_bg = f"scale=w={W}:h={H}:force_original_aspect_ratio=increase,crop={W}:{H}"

    # dim via drawbox
    dim_alpha = _clamp(bg_dim, 0.0, 0.95)
    if dim_alpha > 0:
        vf_bg = vf_bg + f",drawbox=x=0:y=0:w=iw:h=ih:color=black@{dim_alpha}:t=fill"

    # overlay
    filter_complex = (
        f"[0:v]{vf_bg}[bg];"
        f"[1:v]format=rgba,scale={W}:{H}[ov];"
        f"[bg][ov]overlay=0:0:format=auto[vout]"
    )

    ensure_dir(out_path.parent)

    cmd = [
            "ffmpeg", "-y",
            "-i", str(bg_video),
            # Removed "-loop", "1", 
            "-i", str(tmp_overlay),
            "-filter_complex", filter_complex,
            "-map", "[vout]",
            "-map", "0:a?",            # audio optional
            "-c:v", "libx264", "-pix_fmt", "yuv420p",
            "-c:a", "aac", "-b:a", "160k",
            "-shortest",
            str(out_path)
        ]
    run_ffmpeg(cmd)

    try:
        tmp_overlay.unlink(missing_ok=True)
    except Exception:
        pass


# -----------------------------
# Excel batch runner
# -----------------------------
REQUIRED_COLS = [
    "headline", "subhead",
    "title", "description",
]

def _read_pin_data_with_rows(excel_path: Path):
    wb = load_workbook(excel_path)
    ws = wb.active

    # header map
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip().lower()] = c

    missing = [c for c in REQUIRED_COLS if c not in headers]
    if missing:
        raise ValueError(f"PIN_DATA.xlsx missing columns: {missing}")

    # ensure status column exists
    if "status" not in headers:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col).value = "status"
        headers["status"] = new_col
        # NOTE: Removed wb.save(excel_path) here. The workbook is saved later
        # in batch_render_from_folder after all header/data is processed.

    rows = []
    for r in range(2, ws.max_row + 1):
        row_obj = {}
        empty = True
        for col_name, cidx in headers.items():
            val = ws.cell(row=r, column=cidx).value
            if col_name != "status" and val not in (None, ""):
                empty = False
            row_obj[col_name] = val

        if empty:
            continue

        # keep excel row number so we can write status back
        rows.append({"_excel_row": r, **row_obj})

    return wb, ws, headers, rows

def _find_media_file_for_title(folder: Path, title: str, exts: set[str]) -> Optional[Path]:
    """
    If a file exists whose stem matches the title (or sanitized title), return it.
    Searches recursively under folder.
    """
    t = (title or "").strip()
    if not t:
        return None

    candidates = {t, sanitize_filename(t)}

    for p in folder.rglob("*"):
        if not p.is_file():
            continue
        if p.suffix.lower() not in exts:
            continue
        if p.stem in candidates:
            return p

    return None


def _read_pin_data(excel_path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(excel_path)
    ws = wb.active

    # header map
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip().lower()] = c

    missing = [c for c in REQUIRED_COLS if c not in headers]
    if missing:
        raise ValueError(f"PIN_DATA.xlsx missing columns: {missing}")

    rows = []
    for r in range(2, ws.max_row + 1):
        row_obj = {}
        empty = True
        for col_name, cidx in headers.items():
            val = ws.cell(row=r, column=cidx).value
            if val not in (None, ""):
                empty = False
            row_obj[col_name] = val
        if not empty:
            rows.append(row_obj)
    return rows


def _append_master_log(master_excel: Path, records: List[Dict[str, Any]]):
    """
    Ensures columns exist, then appends rows:
    media_file, media_type, board_name, title, url
    """
    from openpyxl import Workbook

    cols = ["media_file", "media_type", "board_name", "title", "url"]

    if not master_excel.exists():
        wb = Workbook()
        ws = wb.active
        for i, c in enumerate(cols, start=1):
            ws.cell(row=1, column=i).value = c
        wb.save(master_excel)

    wb = load_workbook(master_excel)
    ws = wb.active

    # map existing headers
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            header_map[str(v).strip().lower()] = c

    # ensure headers
    for c in cols:
        if c not in header_map:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col).value = c
            header_map[c] = new_col

    # append
    for rec in records:
        rr = ws.max_row + 1
        # ws.cell(rr, header_map["media_file"]).value = rec.get("media_file")
        media_path = rec.get("media_file")
        platform_tags = format_tags_for_platforms(rec.get("comma_separated_tags", ""))
        pin_keywords = platform_tags["pinterest"]    # comma keywords
        ig_hashtags  = platform_tags["instagram"]    # #tag #tag
        fb_hashtags  = platform_tags["facebook"]
        tt_hashtags  = platform_tags["tiktok"]
        yt_keywords  = platform_tags["youtube"]
        thumbnail_path = rec.get("thumbnail_img")


        if media_path:
            p = Path(media_path)

            # Try to make path relative to pinterest_uploads
            try:
                rel = p.as_posix().split("pinterest_uploads/", 1)[1]
                media_path = rel
            except IndexError:
                media_path = p.name  # fallback: filename only

        if thumbnail_path:
            tp = Path(thumbnail_path)
            try:
                rel_thumb = tp.as_posix().split("pinterest_uploads/", 1)[1]
                thumbnail_path = rel_thumb
            except IndexError:
                thumbnail_path = tp.name  # fallback: filename only

        ws.cell(rr, header_map["media_file"]).value = media_path
        ws.cell(rr, header_map["media_type"]).value = rec.get("media_type")
        ws.cell(rr, header_map["board_name"]).value = rec.get("board_name")
        ws.cell(rr, header_map["book_title"]).value = rec.get("title")
        ws.cell(rr, header_map["book_url"]).value = rec.get("url")
        ws.cell(rr, header_map["campaign_name"]).value = rec.get("board_name")
        ws.cell(rr, header_map["destination_url"]).value = rec.get("url")
        ws.cell(rr, header_map["destination_type"]).value = rec.get("media_type")
        ws.cell(rr, header_map["future"]).value = rec.get("future")
        ws.cell(rr, header_map["thumbnail_img"]).value = thumbnail_path

        # Pinterest specific columns
        ws.cell(rr, header_map["pin_title"]).value = rec.get("title")

        desc = rec.get("description", "")
        if pin_keywords:
            if desc:
                desc = desc + "\n\n" + pin_keywords
            else:
                desc = pin_keywords
        ws.cell(rr, header_map["pin_description"]).value = desc
        ws.cell(rr, header_map["pin_url_to_link"]).value = rec.get("url")
        ws.cell(rr, header_map["pinterestprofile"]).value = rec.get("board_name")

        # Facebook specific columns        
        ws.cell(rr, header_map["facebookprofile"]).value = rec.get("facebook_profile")
        ws.cell(rr, header_map["fb_caption"]).value = f"{rec.get('title')}\n\n{rec.get('url')}\n\n{rec.get('description', '')}\n\n{fb_hashtags}"
        
        # Instagram specific columns  
        ws.cell(rr, header_map["instagramprofile"]).value = rec.get("instagram_profile")
        ws.cell(rr, header_map["ig_caption"]).value = f"{rec.get('title')}\n\n{rec.get('url')}\n\n{rec.get('description', '')}\n\n{ig_hashtags}"      
        
        # TikTok specific columns
        ws.cell(rr, header_map["tiktokprofile"]).value = rec.get("tiktok_profile")
        ws.cell(rr, header_map["tiktok_caption"]).value = f"{rec.get('title')}\n\n{rec.get('url')}\n\n{rec.get('description', '')}\n\n{tt_hashtags}"

        # YouTube specific columns
        ws.cell(rr, header_map["youtubechannel"]).value = rec.get("youtube_profile")
        ws.cell(rr, header_map["yt_title"]).value = rec.get("title")
        yt_desc = rec.get("description", "")
        if fb_hashtags:
            if yt_desc:
                yt_desc = yt_desc + "\n\n" + fb_hashtags
            else:
                yt_desc = fb_hashtags
        ws.cell(rr, header_map["yt_description"]).value = yt_desc
        ws.cell(rr, header_map["yt_playlist"]).value = rec.get("youtube_playlist")  
        ws.cell(rr, header_map["yt_schedule_date"]).value = rec.get("youtube_schedule_date")
        ws.cell(rr, header_map["yt_tags"]).value = yt_keywords


        # ws.cell(rr, header_map["title"]).value = rec.get("title")
        # ws.cell(rr, header_map["url"]).value = rec.get("url")

    wb.save(master_excel)

def _resolve_template_relative_assets(project: Dict[str, Any], template_dir: Path) -> Dict[str, Any]:
    """Resolve relative image-layer paths (imgSrc) relative to the JSON template folder."""
    for ly in project.get("layers", []):
        if ly.get("type") != "image":
            continue

        img_src = (ly.get("imgSrc") or "").strip()
        if not img_src:
            continue

        p = Path(img_src)
        if p.is_absolute():
            continue

        abs_p = (template_dir / p).resolve()
        if abs_p.exists():
            ly["imgSrc"] = str(abs_p)

    return project


def create_youtube_thumbnail(
    *,
    base_image_path: str,
    json_template_path: str,
    title_text: str,
    output_name: str,
    subhead_text: str = "",
    logo_path: Optional[str] = None,
    fonts_dir: str = "fonts",
) -> Path:
    """
    Create a YouTube thumbnail image using an overlay JSON template.

    - title_text overrides a layer named 'headline' if present,
      otherwise it overrides the first text layer (fallback).
    """

    template_path = Path(json_template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template JSON not found: {template_path}")
    template_dir = template_path.parent.resolve()

    bg_path = Path(base_image_path)
    if not bg_path.exists():
        alt = (template_dir / bg_path).resolve()
        if alt.exists():
            bg_path = alt
        else:
            raise FileNotFoundError(f"Base image not found: {bg_path}")

    out_path = Path(output_name)
    if not out_path.is_absolute():
        out_path = out_path.resolve()
    ensure_dir(out_path.parent)

    fonts_dir_path = Path(fonts_dir)
    if not fonts_dir_path.is_absolute() and not fonts_dir_path.exists():
        alt = (template_dir / fonts_dir_path).resolve()
        if alt.exists():
            fonts_dir_path = alt
        else:
            fonts_dir_path = fonts_dir_path.resolve()

    logo_p: Optional[Path] = None
    if logo_path:
        lp = Path(logo_path)
        if not lp.exists():
            alt = (template_dir / lp).resolve()
            if alt.exists():
                lp = alt
        if lp.exists():
            logo_p = lp.resolve()

    project = load_project_json(template_path)
    project = apply_text_overrides(project, headline=title_text, subhead=subhead_text, logo_path=logo_p)
    project = _lock_title_subhead_gap(project, gap_px=18)
    project = _resolve_template_relative_assets(project, template_dir)

    render_image_pin(bg_path, project, out_path, fonts_dir_path)
    return out_path

def create_youtube_thumbnail_old(base_image_path, json_template_path, title_text, output_name):
    """
    Creates a thumbnail by overlaying text on an image based on a JSON template.
    """
    # 1. Load Template and Background
    with open(json_template_path, 'r', encoding='utf-8') as f:
        project = json.load(f)
    
    bg = Image.open(base_image_path).convert("RGBA")
    W, H = _as_int(project.get("format", {}).get("w", 1280)), _as_int(project.get("format", {}).get("h", 720))
    
    # Resize background to template size if necessary
    bg = bg.resize((W, H), Image.Resampling.LANCZOS)
    
    # 2. Create Overlay Layer
    overlay = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    
    # 3. Process Layers from JSON
    for ly in project.get("layers", []):
        if ly.get("type") == "text":
            # Override text with your provided title_text if the layer is named 'headline'
            # or just use the first text layer found.
            text = title_text if ly.get("name") == "headline" or ly.get("name") == "subhead" else ly.get("text", "")
            
            box = ly.get("box", {})
            style = ly.get("style", {})
            
            # Calculate pixel positions
            tx = int(_as_float(box.get("x")) * W)
            ty = int(_as_float(box.get("y")) * H)
            tw = int(_as_float(box.get("w")) * W)
            th = int(_as_float(box.get("h")) * H)

            # Load Font (Defaults to Arial if font file not found)
            try:
                font_size = _as_int(style.get("fontSize", 60))
                # Update path to where your .ttf files are located
                font = ImageFont.truetype("arial.ttf", font_size) 
            except:
                font = ImageFont.load_default()

            # Apply Styles
            fill_color = _hex_to_rgba(style.get("fill", "#FFFFFF"))
            
            # Draw Text (Simple implementation - fits text in the box)
            # For complex wrapping/alignment, you can import the _wrap_text 
            # and _fit_font_size functions from your original script.
            draw.text((tx, ty), text.upper() if style.get("uppercase") else text, font=font, fill=fill_color)

    # 4. Merge and Save
    final = Image.alpha_composite(bg, overlay).convert("RGB")
    final.save(output_name, "JPEG", quality=95)
    print(f"Thumbnail saved as: {output_name}")


def batch_render_from_folder(
    folder: Path,
    pin_type: str,            # "image" | "video"
    max_pins: int,
    out_dir: Path,
    master_excel: Path,
    skip_overlays: bool = False,
) -> Dict[str, Any]:
    """
    folder must contain PIN_DATA.xlsx + bg media + json templates + logo files.
    """
    folder = folder.resolve()
    excel_path = folder / "PIN_DATA.xlsx"
    if not excel_path.exists():
        raise FileNotFoundError(f"PIN_DATA.xlsx not found in folder: {folder}")

    ensure_dir(out_dir)
    # fonts_dir = folder / "fonts"
    fonts_dir = Path("./fonts")
    ensure_dir(fonts_dir)  # optional

    # Load workbook, worksheet, headers, and data rows
    wb_pin, ws_pin, headers_pin, data_rows = _read_pin_data_with_rows(excel_path)
    
    status_col = headers_pin["status"] # Get the column index for status

    # Pick background media files from the folder
    if pin_type == "video":
        exts = {".mp4", ".mov", ".mkv", ".webm"}
    else:
        exts = {".jpg", ".jpeg", ".png", ".webp"}

    bg_files = sorted([
        p for p in folder.rglob("*")
        if p.is_file() and p.suffix.lower() in exts
    ])

    if not bg_files and not skip_overlays:
        raise FileNotFoundError(
            f"No background {pin_type} files found in {folder}. "
            f"Expected one of: {sorted(exts)}"
        )

    logo_val = "logo.png"
    logo_path = None
    if logo_val not in (None, ""):
        logo_path = Path(str(logo_val)).expanduser()
        if not logo_path.is_absolute():
            logo_path = folder / logo_path
        if not logo_path.exists():
            # don't crash; just ignore
            logo_path = None

    pj = folder / "overlay.json"

    created = []
    processed_count = 0

    for idx, row in enumerate(data_rows, start=1):
        excel_r = row["_excel_row"]

        # 1) Skip already successful rows
        status_val = str(row.get("status") or "").strip().lower()
        if status_val == "success":
            continue

        if max_pins and max_pins > 0 and processed_count >= max_pins:
            break

        filename = str(row.get("filename") or "")
        headline = str(row.get("headline") or "")
        subhead = str(row.get("subhead") or "")
        title = str(row.get("title") or f"pin_{idx}")
        post_title = str(row.get("post_title") or "")
        future = str(row.get("future") or "")
        description = str(row.get("description") or "")
        comma_separated_tags = str(row.get("comma_separated_tags") or "")
        facebook_profile = str(row.get("facebook_profile") or "")
        youtube_profile = str(row.get("youtube_profile") or "")
        youtube_playlist = str(row.get("youtube_playlist") or "")
        youtube_schedule_date = str(row.get("youtube_schedule_date") or "")
        tiktok_profile = str(row.get("tiktok_profile") or "")
        instagram_profile = str(row.get("instagram_profile") or "")
        pin_url = str(row.get("pin_url") or "")
        parent_url = str(row.get("parent_url") or "")
        board_name = str(row.get("pinterest_board_name") or "")
        thumbnail_img = str(row.get("thumbnail_image") or "")

        project = load_project_json(pj)
        project = apply_text_overrides(project, headline, subhead, logo_path)

        matched = None

        if filename.strip():
            matched = _find_media_file_for_filename(folder, filename, exts)

        if not matched:
            matched = _find_media_file_for_title(folder, title, exts)

        if matched:
            bg_path = matched
        else:
            bg_path = bg_files[(processed_count) % len(bg_files)]

        # out_base = sanitize_filename(title)
        out_base = _derive_out_base(filename, title, fallback=f"pin_{idx}")

        candidate = out_base
        suffix_i = 2
        ext = ".mp4" if pin_type == "video" else ".jpg"
        while (out_dir / f"{candidate}{ext}").exists():
            candidate = f"{out_base}_{suffix_i}"
            suffix_i += 1
        out_base = candidate

        try:
            # Always decide extension + default media_type first
            ext = ".mp4" if pin_type == "video" else ".jpg"
            media_type = "video" if pin_type == "video" else "image"

            if skip_overlays:
                # No overlay rendering: just copy the matched/background media into out_dir
                out_path = out_dir / f"{out_base}{ext}"
                ensure_dir(out_path.parent)
                shutil.copy2(bg_path, out_path)

            else:
                # Normal overlay rendering
                if pin_type == "video":
                    out_path = out_dir / f"{out_base}.mp4"
                    render_video_pin(bg_path, project, out_path, fonts_dir)
                    media_type = "video"
                else:
                    out_path = out_dir / f"{out_base}.jpg"
                    render_image_pin(bg_path, project, out_path, fonts_dir)
                    media_type = "image"


            created.append({
                "media_file": str(out_path).replace("\\", "/"),
                "media_type": media_type,
                "board_name": board_name,
                "title": post_title or title,
                "url": pin_url,
                "description": description,
                "facebook_profile": facebook_profile,
                "instagram_profile": instagram_profile,
                "tiktok_profile": tiktok_profile,
                "youtube_profile": youtube_profile,
                "youtube_playlist": youtube_playlist,
                "youtube_schedule_date": youtube_schedule_date,
                "parent_url": parent_url,   
                "comma_separated_tags": comma_separated_tags,     
                "future": future,  
                "thumbnail_img": thumbnail_img,   
            })

            # 3) Mark success immediately and save (so rerun skips)
            print("Updating status in:", excel_path)
            print("Sheet:", ws_pin.title, "Row:", excel_r, "Col:", status_col)
            ws_pin.cell(row=excel_r, column=status_col).value = "success"
            wb_pin.save(excel_path) # <-- Save after individual success

            processed_count += 1

        except Exception as e:
            # Mark failed, save, then re-raise
            ws_pin.cell(row=excel_r, column=status_col).value = f"failed: {type(e).__name__}"
            
            # Save the failure status before exiting or continuing
            try:
                wb_pin.save(excel_path) 
            except Exception as save_e:
                print(f"CRITICAL: Failed to save status update to Excel: {save_e}")
            
            raise

    # 4) Final Save: Ensure any status column creation from _read_pin_data_with_rows 
    # or the last status update is persisted if the loop finished gracefully.
    try:
        wb_pin.save(excel_path)
    except Exception as e:
        # Log this but don't halt the master log update
        print(f"WARNING: Final save of {excel_path} failed: {e}") 

    # log to master
    _append_master_log(master_excel, created)

    return {
        "ok": True,
        "created_count": len(created),
        "output_dir": str(out_dir),
        "master_excel": str(master_excel),
        "sample": created[:3]
    }

# --- EXAMPLE USAGE ---
if __name__ == "__main__":
    create_youtube_thumbnail(
        base_image_path="avatar_thumbnails/female_1.png",
        json_template_path="avatar_thumbnails/female_1_thumbnail.json",
        title_text="Automate Your Money",
        subhead_text="Bills, Savings, and Goals on Autopilot",
        output_name="avatar_thumbnails/final_thumbnail.jpg"
    )