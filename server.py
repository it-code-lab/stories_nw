from __future__ import annotations
import shutil
from time import time
from flask import Flask, redirect, request, jsonify, render_template, send_from_directory, abort, url_for
import subprocess, json, math
from flask_cors import CORS
import os
import traceback
import contentplanner_worker as cpw

import openpyxl  # to print detailed error info
from build_coloring_app_manifest import build_coloring_manifest
from caption_generator import prepare_captions_file_for_notebooklm_audio
from facebook_uploader import upload_facebook_videos
from get_audio import get_audio_file
from instagram_uploader import upload_instagram_posts
from pinterest_uploader import upload_pins
from scene_builder import render_background_and_merge
from scraper import scrape_and_process  # Ensure this exists
from settings import background_music_options, font_settings, tts_engine, voices, sizes
from tiktok_uploader import upload_tiktok_videos
from video_editor import batch_process
from youtube_uploader import upload_shorts_from_master_file, upload_videos
import re
from pathlib import Path
import wave
from urllib.parse import unquote
from polish_audio_auto import polish_audio  # NEW
from auto_mix import mix_files
from quiz import quiz_bp
from flask import send_file
import tempfile
from coloring_upscale import process_coloring_folder
# from google import genai
# from google.genai import types
import os
from flipthrough_video import generate_flipthrough_video, FlipThroughError
from bg_music_video import merge_all_videos_with_bg_music, merge_video_with_bg_music, BgMusicError
import sys
from pdf2image import convert_from_path
import glob
import uuid
from coloring_animation import _create_coloring_animation, _create_coloring_animation_by_color
from sketch_core import build_sketch_from_pil
from PIL import Image
from scene_builder import probe_duration, make_scene, merge_with_heygen
from assemble_from_videos import assemble_videos, assemble_videos_by_titles_if_present

############################
# Imports added for HTML to video maker

# from __future__ import annotations

# from dataclasses import asdict, dataclass, field
# from typing import Any, Literal

# import requests
# from bs4 import BeautifulSoup, NavigableString, Tag
############################

from media_audio import (
    extract_audio_from_video,
    resolve_input_video,
)
# ---- ADD near the top with imports (server.py) ----
import shutil

VIDEO_EXTS = {".mp4", ".mov", ".mkv", ".avi", ".webm", ".m4v"}

def _ensure_ffmpeg():
    if not shutil.which("ffmpeg"):
        raise RuntimeError("FFmpeg not found in PATH. Please install FFmpeg and add it to PATH.")
    if not shutil.which("ffprobe"):
        raise RuntimeError("ffprobe not found in PATH. Please install FFmpeg tools and add to PATH.")


app = Flask(__name__, template_folder='templates')
app.register_blueprint(quiz_bp)  # all quiz endpoints live under /api/quiz
CORS(app)

# Always resolve relative to this file (server.py)
BASE_DIR = Path(__file__).resolve().parent
AUDIO_PATH = BASE_DIR / "audio.wav"   # your file is beside server.py

COLORING_BASE = BASE_DIR / "downloads"
COLORING_BASE.mkdir(exist_ok=True)

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
PAGES_FOLDER = os.path.join(BASE_DIR, "static", "pages")
STATIC_FOLDER = os.path.join(BASE_DIR, "static")

DATA_DIR = BASE_DIR / "data"
UPLOADS_DIR = BASE_DIR / "uploads"
OUT_DIR = BASE_DIR / "out"

UPLOADS_DIR.mkdir(exist_ok=True)
OUT_DIR.mkdir(exist_ok=True)

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PAGES_FOLDER, exist_ok=True)

POPPLER_PATH = r"C:\poppler-25.11.0\Library\bin"  # Adjust this path as needed

# =========================
# Defaults (env-overridable)
# =========================
DEFAULT_TEXT_MODEL  = os.getenv("GEMINI_DEFAULT_TEXT_MODEL",  "gemini-2.0-flash")
DEFAULT_IMAGE_MODEL = os.getenv("GEMINI_DEFAULT_IMAGE_MODEL", "gemini-2.5-flash-image")


# Optional: central place to tweak RPM/paths
GEM_STATE = str(BASE_DIR / ".gemini_pool_state.json")

gemini_pool = None

@app.post("/build_coloring_manifest")
def build_coloring_manifest_route():
    """
    Build coloring manifest + thumbnails under downloads/.

    Expects form-data or JSON:
      - source_folder: relative folder under downloads/
          e.g. "" (for BASE_DIR/downloads),
               "coloring/v2",
               "1.Cute Farm Animals"
      - force: "0" or "1" (optional, default 0)
    """
    try:
        data = request.get_json(silent=True) or request.form
        source_folder = (data.get("source_folder") or "").strip()
        force_raw = (data.get("force") or "0").strip().lower()
        force = force_raw in ("1", "true", "yes", "y", "on")

        result = build_coloring_manifest(
            source_subfolder=source_folder,
            thumb_edge=640,
            force=force,
        )
        result["ok"] = True
        result["source_folder"] = source_folder
        return jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.post("/upload_media")
def upload_media():
    f = request.files.get("file")
    if not f:
        return "missing file", 400

    name = f.filename or "upload.bin"
    ext = os.path.splitext(name)[1].lower()
    safe = f"{os.urandom(8).hex()}{ext}"
    dst = UPLOADS_DIR / safe
    f.save(dst)

    mime = (f.mimetype or "").lower()
    if mime.startswith("image/"):
        t = "image"
    elif mime.startswith("video/"):
        t = "video"
    else:
        return "unsupported type", 400

    return jsonify({"url": f"/uploads/{safe}", "type": t})

@app.post("/planner/populate_images_excel")
def planner_populate_images_excel():
    channel = (request.form.get("youtube_channel_name") or "").strip()
    image_provider = (request.form.get("image_provider") or "").strip()
    image_orientation = (request.form.get("image_orientation") or "").strip()

    if not channel:
        return jsonify({"ok": False, "message": "youtube_channel_name required"}), 400
    res = cpw.populate_image_jobs_excel_for_channel(channel, image_provider=image_provider, image_orientation=image_orientation)
    return jsonify(res)

@app.post("/planner/populate_heygen_excel")
def planner_populate_heygen_excel():
    channel = (request.form.get("youtube_channel_name") or "").strip()
    if not channel:
        return jsonify({"ok": False, "message": "youtube_channel_name required"}), 400
    res = cpw.populate_heygen_submit_excel_for_channel(channel)
    return jsonify(res)


@app.post("/planner/populate_heygen_multipart_excel")
def planner_populate_heygen_multipart_excel():
    channel = (request.form.get("youtube_channel_name") or "").strip()
    if not channel:
        return jsonify({"ok": False, "message": "youtube_channel_name required"}), 400
    res = cpw.populate_heygen_multipart_excel_for_channel(channel)
    return jsonify(res)


@app.post("/planner/populate_upload_excel")
def planner_populate_upload_excel():
    channel = (request.form.get("youtube_channel_name") or "").strip()
    if not channel:
        return jsonify({"ok": False, "message": "youtube_channel_name required"}), 400
    res = cpw.populate_upload_excel_for_channel(channel)
    return jsonify(res)

@app.post("/planner/populate_upload_excel_long")
def planner_populate_upload_excel_long():
    channel = (request.form.get("youtube_channel_name") or "").strip()
    if not channel:
        return jsonify({"ok": False, "message": "youtube_channel_name required"}), 400
    res = cpw.populate_upload_excel_long_for_channel(channel)
    return jsonify(res)

@app.post("/planner/populate_order_excel")
def planner_populate_order_excel():
    channel = (request.form.get("youtube_channel_name") or "").strip()
    if not channel:
        return jsonify({"ok": False, "message": "youtube_channel_name required"}), 400
    return cpw.populate_section_order_excel_from_db(channel)


@app.post("/planner/populate_heygen_bulk_bg_excel")
def populate_heygen_bulk_bg_excel():
    channel = (request.form.get("youtube_channel_name") or "").strip()
    if not channel:
        return jsonify({"ok": False, "message": "youtube_channel_name required"}), 400
    return cpw.populate_heygen_bulk_bg_excel_from_db(channel)


@app.get("/uploads/<path:fn>")
def uploads(fn):
    return send_from_directory(str(UPLOADS_DIR), fn)

@app.post("/save_timeline")
def save_timeline():
    payload = request.get_json(force=True)
    (OUT_DIR / "timeline.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return "Saved to out/timeline.json"

def _safe_name(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^\w\-. ]+", "", s)
    s = s.replace(" ", "_")
    return s or "out"

def _resolve_path(p: str) -> Path:
    p = (p or "").strip().strip('"').strip("'")
    if not p:
        return Path("")
    pp = Path(p)
    return pp if pp.is_absolute() else (BASE_DIR / pp).resolve()

@app.post("/render_bulk_bg")
def render_bulk_bg(orientation="", scale_bg="yes", copy_as_is=True):
    """
    Reads BASE_DIR/heygen_bulk_bg.xlsx with columns:
      - heygen_video : path to HeyGen mp4
      - bg          : path to background image/video

    For each row:
      1) Create background video matching HeyGen duration
      2) Chroma-key merge (keeps HeyGen captions + avatar bubble untouched/unaltered)
      3) Output filename defaults to ORIGINAL HeyGen filename (same name) in OUT_DIR
    """
    try:
        excel_path = (BASE_DIR / "heygen_bulk_bg.xlsx").resolve()
        if not excel_path.exists():
            return jsonify({"ok": False, "error": f"Missing Excel: {excel_path}"}), 400

        if orientation.lower() not in ("landscape", "portrait"):
             out_res = request.form.get("outRes", "1080x1920")
        else:
            out_res = "1920x1080" if orientation.lower() == "landscape" else "1080x1920"


        work_dir = (OUT_DIR / "bulk_work").resolve()
        if work_dir.exists():
            shutil.rmtree(work_dir)

        chroma_key = "0x00FF00"  # HeyGen green export

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # header -> index
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v:
                headers[str(v).strip().lower()] = c

        if "heygen_video" not in headers or "bg" not in headers:
            return jsonify({
                "ok": False,
                "error": "Excel must have header columns: heygen_video, bg"
            }), 400

        # Ensure 'status' column exists
        if "status" not in headers:
            status_col = ws.max_column + 1
            ws.cell(row=1, column=status_col).value = "status"
            headers["status"] = status_col
        else:
            status_col = headers["status"]


        results = []
        for r in range(2, ws.max_row + 1):
            heygen_raw = ws.cell(r, headers["heygen_video"]).value
            bg_raw = ws.cell(r, headers["bg"]).value

            status_val = ws.cell(r, status_col).value
            if status_val and str(status_val).strip().lower() == "success":
                results.append({
                    "row": r,
                    "ok": True,
                    "skipped": True,
                    "reason": "already processed"
                })
                continue

            if status_val and str(status_val).strip().lower() == "success (copied as is)":
                results.append({
                    "row": r,
                    "ok": True,
                    "skipped": True,
                    "reason": "already processed"
                })
                continue

            if not heygen_raw and not bg_raw:
                continue

            heygen_path = _resolve_path(str(heygen_raw))

            if not heygen_path.exists():
                results.append({"row": r, "ok": False, "error": f"HeyGen not found: {heygen_path}"})
                ws.cell(row=r, column=status_col).value = "HeyGen video not found"
                wb.save(excel_path)   # ✅ save immediately (important)
                continue

            if not bg_raw:
                if not copy_as_is:
                    results.append({"row": r, "ok": False, "error": f"BG not found: {bg_raw}"})
                    ws.cell(row=r, column=status_col).value = "BG asset not found"
                    wb.save(excel_path)   # ✅ save immediately (important)
                    continue
                else:
                    bg_asset = None
                    # copy HeyGen as is 
                    final_out = (OUT_DIR / f"{heygen_path.stem}{heygen_path.suffix}").resolve()
                    shutil.copy2(heygen_path, final_out)
                    ws.cell(row=r, column=status_col).value = "success (copied as is)"
                    wb.save(excel_path)   # ✅ save immediately (important)
                    continue

            bg_asset = _resolve_path(str(bg_raw))

            if bg_asset.suffix.lower() == ".mp4" and not bg_asset.exists():
                png_fallback = bg_asset.with_suffix(".png")
                if png_fallback.exists():
                    bg_asset = png_fallback  # use the png instead


            if not bg_asset.exists():
                results.append({"row": r, "ok": False, "error": f"BG not found: {bg_asset}"})
                ws.cell(row=r, column=status_col).value = "BG asset not found"
                wb.save(excel_path)   # ✅ save immediately (important)
                continue

            # 1) duration = HeyGen duration
            dur = probe_duration(heygen_path)

            # 2) build one full background clip (image loops / video holds last frame)
            # If work_dir exists then clean it up first

            work_dir = (OUT_DIR / "bulk_work").resolve()
            # if work_dir.exists():
            #     shutil.rmtree(work_dir)

            work_dir.mkdir(exist_ok=True)



            bg_video = work_dir / f"{_safe_name(heygen_path.stem)}__bg.mp4"

            make_scene(asset=bg_asset, duration=dur, out_path=bg_video, out_res=out_res)

            # 3) output file name = same as original HeyGen file name, but written under OUT_DIR
            final_out = (OUT_DIR / f"{heygen_path.stem}{heygen_path.suffix}").resolve()

            # merge: captions + avatar remain exactly as HeyGen because we don't scale HeyGen layer

            merge_with_heygen(
                background=bg_video,
                heygen=heygen_path,
                out_path=final_out,
                chroma_key_hex=None,  # <-- let the script decide
                scaled_layout=(scale_bg.lower() != "no"),
                auto_detect_chroma=True,
                chroma_detect_hex="0x00FF00",
                chroma_ratio_threshold=0.12,
            )

            ws.cell(row=r, column=status_col).value = "success"
            wb.save(excel_path)   # ✅ save immediately (important)

            results.append({
                "row": r,
                "ok": True,
                "heygen": str(heygen_path),
                "bg": str(bg_asset),
                "output": str(final_out)
            })

        wb.close()
        # return jsonify({"ok": True, "count": len(results), "results": results})
        return jsonify({
        "ok": True,
        "count": len(results),
        "results": results,
        "output": f"Processed/checked {len(results)} rows. See results[] for details."
        })


    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.post("/render")
def render():
    # timeline must exist
    tl_path = OUT_DIR / "timeline.json"
    if not tl_path.exists():
        return jsonify({"error": "Please Save Timeline first."}), 400

    heygen = request.files.get("heygen")
    if not heygen:
        return jsonify({"error": "Missing HeyGen file."}), 400

    out_res = request.form.get("outRes", "1920x1080")

    heygen_path = OUT_DIR / "heygen.mp4"
    meta = json.loads((OUT_DIR / "job_meta.json").read_text())
    orig_name = meta["orig_name"]
    orig_ext  = meta["orig_ext"]

    # heygen_path = OUT_DIR / f"{orig_name}{orig_ext}"
    heygen.save(heygen_path)

    try:
        output_path = render_background_and_merge(
            timeline_json_path=tl_path,
            base_dir=BASE_DIR,
            heygen_path=heygen_path,
            out_dir=OUT_DIR,
            out_res=out_res,
            out_filename=f"{orig_name}{orig_ext}",   # ✅ NEW
        )
        return jsonify({"output": str(output_path)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.post("/api/coloring_sketch")
def api_coloring_sketch():
    """
    Generate sketch SVG + steps for a coloring page.

    Accepts form-data:
      - file: uploaded image
      - mode: auto|cartoon|photo (optional, default: auto)
      - detail: 1..10 (optional, default: 5)
      - vector: outline|centerline (optional, default: outline)
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "error": "Missing file"}), 400

    mode = (request.args.get("mode") or request.form.get("mode") or "auto").strip()
    detail_str = (request.args.get("detail") or request.form.get("detail") or "5").strip()
    vector = (request.args.get("vector") or request.form.get("vector") or "outline").strip()

    try:
        detail = int(detail_str)
    except ValueError:
        detail = 5

    try:
        img = Image.open(file.stream).convert("RGB")
    except Exception as e:
        return jsonify({"ok": False, "error": f"Failed to read image: {e}"}), 400

    try:
        svg, steps = build_sketch_from_pil(
            img,
            mode=mode,
            detail=detail,
            vector=vector,
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": f"Sketch failed: {e}"}), 500

    return jsonify({
        "ok": True,
        "svg": svg,
        "steps": steps,
    })

@app.post("/api/coloring_animation")
def api_coloring_animation():
    """
    Upload a colored coloring-page image and generate a 'before/after' animation:
      - First: clean line-art (desaturated page)
      - Then: sweeping color fill using original colors

    Expects multipart/form-data:
      - image: file (PNG/JPG/WEBP)
      - fps: optional, default 30
      - duration: optional, default 4.0 seconds
      - canvas_preset: "none" | "shorts" | "pinterest" | "custom"
      - canvas_width, canvas_height: used when canvas_preset == "custom"

    Returns JSON:
      {
        ok: true,
        url: "/downloads/coloring_animation/<file>",
        filename: "<file>"
      }
    """
    try:
        _ensure_ffmpeg()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    if "image" not in request.files:
        return jsonify({"ok": False, "error": "Missing 'image' file"}), 400

    file = request.files["image"]
    if not file.filename:
        return jsonify({"ok": False, "error": "Empty filename"}), 400

    # Basic extension check
    fname = file.filename.lower()
    if not any(fname.endswith(ext) for ext in [".png", ".jpg", ".jpeg", ".webp"]):
        return jsonify({"ok": False, "error": "Only PNG/JPG/WEBP are supported"}), 400

    # Parse params
    form = request.form
    try:
        fps = int(form.get("fps") or 30)
        duration = float(form.get("duration") or 4.0)
    except ValueError:
        return jsonify({"ok": False, "error": "Invalid fps/duration"}), 400

    canvas_preset = (form.get("canvas_preset") or "none").lower()
    canvas_width = None
    canvas_height = None

    if canvas_preset == "shorts":
        # Portrait Shorts canvas
        canvas_width, canvas_height = 1080, 1920
    elif canvas_preset == "pinterest":
        canvas_width, canvas_height = 1000, 1500
    elif canvas_preset == "custom":
        try:
            canvas_width = int(form.get("canvas_width") or 0)
            canvas_height = int(form.get("canvas_height") or 0)
        except ValueError:
            return jsonify({"ok": False, "error": "Invalid custom canvas width/height"}), 400

    target_size = None
    if canvas_width and canvas_height and canvas_width > 0 and canvas_height > 0:
        target_size = (canvas_width, canvas_height)

    # Save uploaded original to uploads/
    uploads_dir = BASE_DIR / "uploads" / "coloring_animation"
    uploads_dir.mkdir(parents=True, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}_{os.path.basename(fname)}"
    input_path = uploads_dir / unique_name
    file.save(str(input_path))

    # Prepare output under downloads/coloring_animation so it's served via /downloads/...
    out_dir = COLORING_BASE / "coloring_animation"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = unique_name.rsplit(".", 1)[0] + "_anim.mp4"
    output_path = out_dir / out_name
    mode = (form.get("mode") or "sweep").lower()


    try:
        if mode == "by_color":
            _create_coloring_animation_by_color(
                input_path=input_path,
                output_path=output_path,
                fps=fps,
                num_colors=5,
                brush_steps_per_color=40,
                hold_line_sec=1.2,
                hold_end_sec=1.2,
                target_size=target_size,
                bg_color=(255, 255, 255),
            )
        else:
            _create_coloring_animation(
                input_path=input_path,
                output_path=output_path,
                fps=fps,
                duration_sec=duration,
                target_size=target_size,
                bg_color=(255, 255, 255),
            )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": f"Failed to create animation: {e}"}), 500



    # URL under /downloads
    rel = output_path.resolve().relative_to(COLORING_BASE).as_posix()
    url = f"/downloads/{rel}"

    return jsonify({
        # Old style
        "ok": True,
        "url": url,
        "filename": out_name,
        "fps": fps,
        "duration": duration,
        "canvas": {
            "preset": canvas_preset,
            "width": canvas_width,
            "height": canvas_height,
        },
        # New style (so either JS version works)
        "success": True,
        "video_url": url,
        "video_path": rel,
    })



@app.post("/merge_bg_music")
def merge_bg_music_route():
    """
    Batch mix background music (from edit_vid_audio/) into ALL videos in edit_vid_input/.

    Expects form fields:
      - bg_volume: float, 0.0–2.0 (default 0.3)
      - video_volume: float, 0.0–2.0 (default 1.0)

    Output is written to edit_vid_output/<same filename as input>.
    Returns JSON with list of outputs (web paths) for the UI.
    """
    try:
        bg_volume = float(request.form.get("bg_volume", "0.3") or "0.3")
    except ValueError:
        bg_volume = 0.3

    try:
        video_volume = float(request.form.get("video_volume", "1.0") or "1.0")
    except ValueError:
        video_volume = 1.0

    try:
        result = merge_all_videos_with_bg_music(
            base_dir=BASE_DIR,
            bg_volume=bg_volume,
            video_volume=video_volume,
        )
    except BgMusicError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

    # Build UI-friendly items
    items = []
    for out_path, in_path in result["pairs"]:
        items.append({
            "input_name": in_path.name,
            "output_name": out_path.name,
            "web_path": f"/video/edit_vid_output/{out_path.name}",
        })

    return jsonify({
        "ok": True,
        "total": result["total"],
        "music_file": result["music_file"],
        "bg_volume": bg_volume,
        "video_volume": video_volume,
        "items": items,
    })


@app.post("/api/generate_flipthrough")
def api_generate_flipthrough():
    """
    Generate a flip-through preview video for a given folder under /downloads.
    Expects:
      - folder: subfolder name (e.g. "farm_animals")
      - seconds_per_image (optional, default 0.5)
      - width, height (optional, default 1920x1080)

    Returns:
      { ok: true, url: "/downloads/<folder>/flip_preview.mp4" }
    """
    data = request.get_json(silent=True) or request.form
    folder = (data.get("folder") or "").strip()
    if not folder:
        return jsonify({"ok": False, "error": "Missing 'folder'"}), 400

    try:
        seconds = float(data.get("seconds_per_image") or 1.5)
        width = int(data.get("width") or 1920)
        height = int(data.get("height") or 1080)
        watermark_text = data.get("watermark_text") or "PREVIEW ONLY"
    except ValueError:
        return jsonify({"ok": False, "error": "Invalid numeric parameter"}), 400

    try:
        out_path = generate_flipthrough_video(
            COLORING_BASE,
            folder,
            out_name="flip_preview.mp4",
            seconds_per_image=seconds,
            width=width,
            height=height,
            watermark_text=watermark_text,
        )
    except FlipThroughError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": f"Internal error: {e}"}), 500

    # Build public URL
    rel = out_path.resolve().relative_to(COLORING_BASE).as_posix()
    url = f"/downloads/{rel}"

    return jsonify({
        "ok": True,
        "folder": folder,
        "url": url,
    })


@app.post("/api/upscale_coloring")
def api_upscale_coloring():
    """
    Trigger upscaling for a given folder under /coloring.

    Expects JSON or form-data:
      - folder: subfolder under /coloring (e.g., "farm_animals")
      - page_size: "LETTER" or "EIGHTX10" (optional, default LETTER)
      - threshold: 0-255 (optional, default 200)

    Response:
      {
        ok: true,
        folder: "farm_animals",
        count: 24,
        processed: [
          { file: "page1_upscaled.png", rel_path: "farm_animals/processed_images/page1_upscaled.png", url: "/downloads/farm_animals/processed_images/page1_upscaled.png" },
          ...
        ]
      }
    """
    data = request.get_json(silent=True) or request.form
    folder = (data.get("folder") or "").strip()
    page_size = (data.get("page_size") or "LETTER").strip()
    threshold = int(data.get("threshold") or 200)

    if not folder:
        return jsonify({"ok": False, "error": "Missing 'folder'"}), 400

    try:
        output_files = process_coloring_folder(COLORING_BASE, folder, page_size, threshold)
    except FileNotFoundError as e:
        return jsonify({"ok": False, "error": str(e)}), 404
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": f"Internal error: {e}"}), 500

    processed = []
    for out_path in output_files:
        # path relative to COLORING_BASE for URL mapping
        rel = out_path.relative_to(COLORING_BASE).as_posix()
        processed.append({
            "file": out_path.name,
            "rel_path": rel,
            "url": f"/downloads/{rel}",
        })

    return jsonify({
        "ok": True,
        "folder": folder,
        "count": len(processed),
        "processed": processed,
    })

@app.get("/downloads/<path:subpath>")
def coloring_files(subpath: str):
    """
    Serve files from COLORING_BASE safely.
    Used by the Coloring Book Builder to load processed_images.
    """
    full = (COLORING_BASE / subpath).resolve()
    base = COLORING_BASE.resolve()
    if not str(full).startswith(str(base)) or not full.is_file():
        abort(404)
    return send_from_directory(full.parent, full.name)

@app.post("/generate_meta")
def seo_generate_meta():
    """
    Run get_seo_meta_data.py to generate SEO metadata into pages_with_meta.xlsx.
    Uses pages_input.xlsx as input in the same folder as server.py.
    """
    try:
        base_dir = BASE_DIR  # already defined above
        script_path = base_dir / "get_seo_meta_data.py"

        if not script_path.exists():
            return jsonify({
                "ok": False,
                "error": f"Script not found: {script_path}"
            }), 500

        # Run using the same Python interpreter/venv as the server
        proc = subprocess.run(
            [sys.executable, str(script_path)],
            cwd=str(base_dir),
            capture_output=True,
            text=True,
            timeout=60 * 60,   # up to 1 hour for 400+ pages
        )

        ok = (proc.returncode == 0)

        # Keep logs short-ish for UI
        stdout_tail = (proc.stdout or "")[-4000:]
        stderr_tail = (proc.stderr or "")[-4000:]

        return jsonify({
            "ok": ok,
            "returncode": proc.returncode,
            "stdout": stdout_tail,
            "stderr": stderr_tail,
        }), (200 if ok else 500)

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            "ok": False,
            "error": str(e),
        }), 500


@app.post("/create_images")
def create_images():
    """Run only the image creation job."""
    try:
        from multi_profile_media_agent import createImages
        result = createImages()
        return jsonify({"ok": True, "message": "Images created successfully", "result": str(result)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.post("/bulk_video_maker")
def bulk_video_maker():
    """Run the bulk video maker job (images->video)."""
    try:
        from bulk_video_maker import main as bulkVideoMaker
        result = bulkVideoMaker()
        return jsonify({"ok": True, "message": "Bulk video maker completed successfully", "result": str(result)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/remove_borders")
def remove_borders():
    try:
        from remove_borders import main as remove_borders_main

        data = request.form
        source_folder = (data.get("source_folder") or "").strip()
        border_px_str = (data.get("border_px") or "10").strip()

        try:
            border_px = int(border_px_str)
        except:
            border_px = 10

        result = remove_borders_main(
            source_subfolder=source_folder or None,
            border_px=border_px,
            fill="white"   # or "rgba(255,255,255,0)" for transparent
        )

        status = 200 if result.get("ok") else 500
        return jsonify(result), status

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/create_vector_images")
def create_vector_images():
    """
    Run the vector image creation job.

    Expects (optional) form field:
      - source_folder: relative folder under downloads/
        e.g. "1.Cute Farm Animals/pages"

    The vectorize_images.main() function should then create output under
    vector_images/<source_folder>/...
    """
    try:
        from vectorize_images import main
        data = request.get_json(silent=True) or request.form
        source_folder = (data.get("source_folder") or "").strip()

        # Call your script. Adjust if main() has a different signature.
        if source_folder:
            result = main(source_folder)
            msg = f"Vector images created successfully for '{source_folder}'"
        else:
            result = main()
            msg = "Vector images created successfully (default folder)"

        return jsonify({"ok": True, "message": msg})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

    
@app.post("/create_videos")
def create_videos():
    """Run only the video creation job."""
    try:
        from multi_profile_media_agent import createVideos
        result = createVideos()
        return jsonify({"ok": True, "message": "Videos created successfully", "result": str(result)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

# ---- ADD this route (server.py) ----
@app.post("/upscale")
def upscale_all_videos():
    """
    Batch-only: loop over all videos in 'edit_vid_input/' and upscale each one.

    Optional form fields:
      - width (default 1920)
      - deinterlace ('true'/'false', default 'true')
      - denoise   ('true'/'false', default 'true')
      - crf       (default 18)
      - preset    (default 'slow')
      - keep_audio ('yes'/'no', default 'yes')
    """
    try:
        _ensure_ffmpeg()

        base_dir = BASE_DIR
        in_dir = base_dir / "edit_vid_input"
        out_dir = base_dir / "edit_vid_output"
        out_dir.mkdir(parents=True, exist_ok=True)

        # Options
        width = int(request.form.get("width", "1920") or 1920)
        deinterlace_val = (request.form.get("deinterlace", "true") or "true").lower() in ("1","true","yes","on")
        denoise_val    = (request.form.get("denoise", "true") or "true").lower() in ("1","true","yes","on")
        crf            = int(request.form.get("crf", "18") or 18)
        preset         = (request.form.get("preset", "slow") or "slow").strip()
        keep_audio_req = (request.form.get("keep_audio", "yes") or "yes").lower() in ("1","true","yes","on","y","yes")

        if not in_dir.exists():
            return jsonify({"ok": False, "error": f"Input folder not found: {in_dir}"}), 400

        exts = {".mp4", ".mov", ".mkv", ".avi", ".webm", ".m4v"}
        inputs = [p for p in sorted(in_dir.iterdir()) if p.is_file() and p.suffix.lower() in exts]
        if not inputs:
            return jsonify({"ok": False, "error": "No videos found in edit_vid_input/"}), 400

        # Build filter chain
        vf = []
        if deinterlace_val:
            vf.append("bwdif=mode=1")
        if denoise_val:
            vf.append("atadenoise")
        vf.append(f"scale={width}:-2:flags=lanczos")
        vf.append("unsharp=lx=3:ly=3:la=0.4")
        vf.append("format=yuv420p")
        filter_str = ",".join(vf)

        def has_audio_stream(path: Path) -> bool:
            try:
                # returns 0 if at least one audio stream is present
                subprocess.check_call(
                    ["ffprobe", "-v", "error", "-select_streams", "a",
                     "-show_entries", "stream=index",
                     "-of", "csv=p=0", str(path)],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
                )
                return True
            except subprocess.CalledProcessError:
                return False

        results, errors = [], []

        for src in inputs:
            try:
                dst_name = f"{src.stem}_upscaled_{width}w.mp4"
                dst_path = out_dir / dst_name

                include_audio = keep_audio_req and has_audio_stream(src)

                cmd = [
                    "ffmpeg", "-y",
                    "-i", str(src),
                    "-vf", filter_str,
                    "-c:v", "libx264", "-crf", str(crf), "-preset", preset,
                ]
                if include_audio:
                    cmd += ["-c:a", "aac", "-b:a", "192k"]
                else:
                    cmd += ["-an"]
                cmd += ["-movflags", "+faststart", str(dst_path)]

                print("Upscale cmd:", " ".join(cmd))
                subprocess.check_call(cmd)

                results.append({
                    "input":  str(src.relative_to(base_dir)),
                    "output": f"/video/edit_vid_output/{dst_name}",
                    "kept_audio": include_audio
                })
            except subprocess.CalledProcessError as e:
                errors.append({
                    "input": str(src.relative_to(base_dir)),
                    "error": "FFmpeg failed",
                    "detail": getattr(e, "output", None)
                })
            except Exception as e:
                errors.append({
                    "input": str(src.relative_to(base_dir)),
                    "error": str(e)
                })

        return jsonify({
            "ok": bool(results),
            "mode": "batch",
            "count_processed": len(results),
            "count_errors": len(errors),
            "settings": {
                "width": width,
                "deinterlace": deinterlace_val,
                "denoise": denoise_val,
                "crf": crf,
                "preset": preset,
                "keep_audio_requested": keep_audio_req
            },
            "outputs": results,
            "errors": errors
        }), (200 if results else 500)

    except subprocess.CalledProcessError as e:
        return jsonify({"ok": False, "error": "FFmpeg failed", "detail": getattr(e, "output", None)}), 500
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500



@app.route('/add_auto_captions_to_english_videos', methods=['GET', 'POST'])
def add_captions_to_english_video():
    data = request.form if request.method == "POST" else request.args
    allowed_styles = {
        "premium-gold",
        "clean-cyan",
        "neon-pop",
        "ruby-bold",
        "editorial-white",
    }
    style = (data.get("style") or "premium-gold").strip()
    language = (data.get("language") or "en").strip()
    video_name = (data.get("video") or "composed_video.mp4").strip()
    position = (data.get("position") or "bottom").strip()
    word_transition = (data.get("word_transition") or "none").strip()
    font_size = (data.get("font_size") or "").strip()
    font_scale = (data.get("font_scale") or "100").strip()
    margin_v = (data.get("margin_v") or "").strip()
    words_per_phrase = (data.get("words_per_phrase") or "").strip()
    text_case = (data.get("text_case") or "original").strip()
    caption_box = (data.get("caption_box") or "none").strip()
    outline_scale = (data.get("outline_scale") or "100").strip()
    shadow_scale = (data.get("shadow_scale") or "100").strip()

    if style not in allowed_styles:
        return jsonify({"ok": False, "status": "error", "message": f"Unknown caption style: {style}"}), 400
    if position not in {"top", "middle", "bottom"}:
        return jsonify({"ok": False, "status": "error", "message": f"Unknown caption position: {position}"}), 400
    if word_transition not in {"none", "fade", "pop", "glow"}:
        return jsonify({"ok": False, "status": "error", "message": f"Unknown word transition: {word_transition}"}), 400
    if text_case not in {"original", "uppercase", "title"}:
        return jsonify({"ok": False, "status": "error", "message": f"Unknown text case: {text_case}"}), 400
    if caption_box not in {"none", "soft", "glass", "solid"}:
        return jsonify({"ok": False, "status": "error", "message": f"Unknown caption background: {caption_box}"}), 400

    def optional_int(value, name, low, high):
        if value == "":
            return None
        try:
            parsed = int(value)
        except ValueError:
            raise ValueError(f"{name} must be a number.")
        if parsed < low or parsed > high:
            raise ValueError(f"{name} must be between {low} and {high}.")
        return parsed

    try:
        font_size_value = optional_int(font_size, "Font size", 18, 140)
        margin_v_value = optional_int(margin_v, "Vertical margin", 0, 500)
        words_per_phrase_value = optional_int(words_per_phrase, "Words at a time", 1, 12)
        font_scale_value = float(font_scale)
        if font_scale_value < 50 or font_scale_value > 180:
            raise ValueError("Font scale must be between 50 and 180.")
        outline_scale_value = float(outline_scale)
        if outline_scale_value < 50 or outline_scale_value > 220:
            raise ValueError("Outline strength must be between 50 and 220.")
        shadow_scale_value = float(shadow_scale)
        if shadow_scale_value < 0 or shadow_scale_value > 220:
            raise ValueError("Shadow strength must be between 0 and 220.")
    except ValueError as e:
        return jsonify({"ok": False, "status": "error", "message": str(e)}), 400

    try:
        video_path = (BASE_DIR / video_name).resolve()
        if video_path != BASE_DIR and BASE_DIR not in video_path.parents:
            return jsonify({"ok": False, "status": "error", "message": "Video path must stay inside the application folder."}), 400
        if not video_path.exists() or not video_path.is_file():
            return jsonify({"ok": False, "status": "error", "message": f"Video not found: {video_name}"}), 404
        if video_path.suffix.lower() not in VIDEO_EXTS:
            return jsonify({"ok": False, "status": "error", "message": f"Unsupported video type: {video_path.suffix}"}), 400

        output_path = video_path.with_name(f"{video_path.stem}_captioned{video_path.suffix}")
    except Exception as e:
        return jsonify({"ok": False, "status": "error", "message": str(e)}), 400

    venv_base = BASE_DIR / "venv_whisperx"
    venv_scripts = venv_base / "Scripts"
    venv_python = venv_scripts / "python.exe"
    script_path = BASE_DIR / "whisperx_captions.py"

    env = os.environ.copy()
    env["PATH"] = str(venv_scripts) + os.pathsep + env.get("PATH", "")
    
    command = [
        str(venv_python),
        str(script_path),
        "--video", str(video_path),
        "--language", language,
        "--style", style,
        "--font_scale", str(font_scale_value),
        "--position", position,
        "--word_transition", word_transition,
        "--text_case", text_case,
        "--caption_box", caption_box,
        "--outline_scale", str(outline_scale_value),
        "--shadow_scale", str(shadow_scale_value),
        "--out", str(output_path),
    ]
    if font_size_value is not None:
        command += ["--font_size", str(font_size_value)]
    if margin_v_value is not None:
        command += ["--margin_v", str(margin_v_value)]
    if words_per_phrase_value is not None:
        command += ["--words_per_phrase", str(words_per_phrase_value)]

    try:
        result = subprocess.run(
            command, 
            check=True, 
            capture_output=True, 
            text=True,
            encoding="utf-8", 
            errors="replace",
            env=env,
            cwd=str(BASE_DIR),
        )

        return jsonify({
            "ok": True,
            "status": "success",
            "message": "Premium captions generated successfully!",
            "style": style,
            "settings": {
                "font_size": font_size_value,
                "font_scale": font_scale_value,
                "position": position,
                "margin_v": margin_v_value,
                "words_per_phrase": words_per_phrase_value,
                "word_transition": word_transition,
                "text_case": text_case,
                "caption_box": caption_box,
                "outline_scale": outline_scale_value,
                "shadow_scale": shadow_scale_value,
            },
            "video": str(video_path.name),
            "output": str(output_path.name),
            "logs": result.stdout
        }), 200

    except subprocess.CalledProcessError as e:
        return jsonify({
            "ok": False,
            "status": "error",
            "message": "The caption script failed.",
            "error_details": e.stderr if e.stderr else e.stdout
        }), 500
        
    except FileNotFoundError:
        return jsonify({
            "ok": False,
            "status": "error",
            "message": f"Could not find the Python executable at {venv_python}."
        }), 500

@app.post("/extract_audio")
def extract_audio_route():
    """
    Batch-extract audio for all videos in edit_vid_input/ when no file/path is provided.
    Otherwise (if you ever send "file" or "video"), it behaves like single-file mode.
    Form-data:
      - format: wav|mp3|m4a (default wav)
      - track: audio stream index, default 0
    """
    try:
        base_dir = Path(__file__).resolve().parent
        fmt = (request.form.get("format") or "wav").lower()
        track_raw = request.form.get("track", "0")
        try:
            track = int(track_raw)
        except Exception:
            track = 0

        uploaded = request.files.get("file")
        video_url = (request.form.get("video") or "").strip()

        # If either upload or explicit path is provided, fall back to single-file behavior
        if uploaded or video_url:
            try:
                in_path = resolve_input_video(
                    base_dir=base_dir,
                    uploaded_temp_dir=base_dir / "tmp_upload_video",
                    uploaded_file=uploaded,
                    urlish_path=video_url if video_url else None,
                    fallback_rel="edit_vid_input/bg_video.mp4",
                )
            except Exception as e:
                return jsonify({"ok": False, "error": str(e)}), 400

            result = extract_audio_from_video(
                base_dir=base_dir,
                input_path=in_path,
                fmt=fmt,
                track=track,
                root_output_name="edit_vid_output"
            )

            if not getattr(result, "ok", False):
                return jsonify({
                    "ok": False,
                    "error": result.error or "Failed",
                    "detail": result.detail,
                }), 500

            download_rel = f"/video/{result.output_path.name}" if result.output_path else None
            also_rel = f"/video/edit_vid_audio/{result.output_path.name}" if getattr(result, "copy_path", None) else None

            return jsonify({
                "ok": True,
                "mode": "single",
                "input": result.input_name,
                "format": result.fmt,
                "track": result.track,
                "duration_sec": result.duration_sec,
                "download": download_rel,
                "also_saved_in_edit_vid_audio": also_rel,
            })

        # === Batch mode: scan edit_vid_input/ ===
        in_dir = base_dir / "edit_vid_input"
        out_dir = base_dir / "edit_vid_output"
        out_dir.mkdir(parents=True, exist_ok=True)

        if not in_dir.exists():
            return jsonify({"ok": False, "error": f"Input folder not found: {in_dir}"}), 400

        candidates = sorted([p for p in in_dir.iterdir() if p.is_file() and p.suffix.lower() in VIDEO_EXTS])

        if not candidates:
            return jsonify({"ok": True, "mode": "batch", "items": []})

        items = []
        for vid in candidates:
            try:
                res = extract_audio_from_video(
                    base_dir=base_dir,
                    input_path=vid,
                    fmt=fmt,
                    track=track,
                    root_output_name="edit_vid_output"
                )
                if getattr(res, "ok", False) and getattr(res, "output_path", None):
                    items.append({
                        "ok": True,
                        "input": vid.name,
                        "fmt": res.fmt,
                        "track": res.track,
                        "duration_sec": res.duration_sec,
                        "output_name": res.output_path.name,
                        "download": f"/video/{res.output_path.name}"
                    })
                else:
                    items.append({
                        "ok": False,
                        "input": vid.name,
                        "error": getattr(res, "error", "Failed"),
                        "detail": getattr(res, "detail", None)
                    })
            except Exception as e:
                import traceback
                traceback.print_exc()
                items.append({
                    "ok": False,
                    "input": vid.name,
                    "error": str(e)
                })

        # Always return a compact summary for the UI to render
        return jsonify({
            "ok": True,
            "mode": "batch",
            "format": fmt,
            "track": track,
            "count": len(items),
            "items": items
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500
    


# gemini_pool = GeminiPool(
#     api_keys=None,          # reads GEMINI_API_KEYS from env
#     per_key_rpm=25,         # tune to your safe RPM per key
#     state_path=GEM_STATE,   # persist key usage across restarts
#     autosave_every=3
# )

# =========================
# Minimal, defaulted routes
# =========================

@app.post("/ai/text")
def ai_text():
    """
    UI only needs: {"prompt": "..."}.
    Optional overrides: {"model": "...", "temperature": 0.7, "max_output_tokens": 512}
    """
    data = request.get_json(force=True) if request.is_json else {}
    prompt = (data.get("prompt") or "").strip()
    if not prompt:
        return jsonify({"ok": False, "error": "Missing 'prompt'"}), 400

    # Use server defaults unless client overrides
    model = data.get("model") or DEFAULT_TEXT_MODEL
    temperature = data.get("temperature")
    max_tokens  = data.get("max_output_tokens")

    try:
        text = gemini_pool.generate_text(
            prompt,
            model=model,
            temperature=temperature,
            max_output_tokens=max_tokens
        )
        return jsonify({"ok": True, "model": model, "text": text})
    except Exception as e:
        return jsonify({"ok": False, "model": model, "error": str(e)}), 500


@app.post("/ai/image")
def ai_image():
    """
    UI only needs: {"prompt": "..."}.
    Optional overrides: {"model": "..."} and any extras you pass through later.
    Saves the first image to /edit_vid_input/ and returns a web path for your pipeline.
    """
    data = request.get_json(force=True) if request.is_json else {}
    prompt = (data.get("prompt") or "").strip()
    if not prompt:
        return jsonify({"ok": False, "error": "Missing 'prompt'"}), 400

    model = data.get("model") or DEFAULT_IMAGE_MODEL

    try:
        out_dir = BASE_DIR / "edit_vid_input"
        out_dir.mkdir(exist_ok=True)
        out_name = f"gen_{int(__import__('time').time())}.png"
        out_path = out_dir / out_name

        gemini_pool.generate_image(
            prompt,
            model=model,
            out_path=str(out_path),
            # extra=data.get("extra")  # keep for future size/quality params
        )

        return jsonify({
            "ok": True,
            "model": model,
            "image_path": f"/edit_vid_input/{out_name}"  # relative path for your pipeline
        })
    except Exception as e:
        return jsonify({"ok": False, "model": model, "error": str(e)}), 500


@app.get("/ai/keys")
def ai_keys():
    """Quick peek at rotation state (helpful in logs/dashboards)."""
    return jsonify({"ok": True, "stats": gemini_pool.stats()})

@app.get("/ai/models")
def ai_models():
    try:
        models = gemini_pool.list_models(api_version="v1beta")
        return jsonify({"ok": True, "models": models})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500




@app.post("/mix_vocal_music")
def mix_vocal_music_endpoint():
    """
    Merge vocal + background music into a polished song.
    Accepts optional file uploads (vocal_file, music_file); otherwise
    tries to auto-pick from edit_vid_audio/ (vocal*, music*, etc.) or first two audio files.
    Returns JSON with download link and metadata.
    """
    try:
        base = BASE_DIR
        eva = base / "edit_vid_audio"
        eva.mkdir(exist_ok=True)

        uploaded_v = request.files.get("vocal_file")
        uploaded_m = request.files.get("music_file")

        def _save_upload(upfile, fallback_stem):
            if not upfile or not upfile.filename:
                return None
            ext = os.path.splitext(upfile.filename)[1]
            p = eva / f"{fallback_stem}{ext}"
            upfile.save(str(p))
            return p

        vocal_path = _save_upload(uploaded_v, "vocal_uploaded") if uploaded_v else None
        music_path = _save_upload(uploaded_m, "music_uploaded") if uploaded_m else None

        # Auto-pick from folder if any missing
        def _pick_candidates():
            if not eva.exists():
                return []
            return [
                p for p in sorted(eva.iterdir())
                if p.suffix.lower() in {".wav", ".mp3", ".m4a", ".flac", ".aac", ".ogg"}
            ]

        files = _pick_candidates()
        if not vocal_path:
            # Prefer names
            for p in files:
                if "vocal" in p.name.lower() or "vox" in p.name.lower():
                    vocal_path = p; break
        if not music_path:
            for p in files:
                if p != vocal_path and any(tag in p.name.lower() for tag in ["music", "bg", "instrumental"]):
                    music_path = p; break
        # Fallback to first two
        if not (vocal_path and music_path) and len(files) >= 2:
            vocal_path = vocal_path or files[0]
            music_path = music_path or next(p for p in files if p != vocal_path)

        if not (vocal_path and music_path):
            return jsonify({"ok": False, "error": "Could not find both vocal and music. Upload them or place in edit_vid_audio/."}), 400

        # Read options
        sr = int(request.form.get("sr", 44100))
        music_gain_db = float(request.form.get("music_gain_db", -10.0))
        duck_db = float(request.form.get("duck_db", 10.0))
        duck_floor_db = float(request.form.get("duck_floor_db", -1.0))
        target_lufs = float(request.form.get("target_lufs", -14.0))
        out_fmt = (request.form.get("format", "wav") or "wav").lower()
        out_name = f"final_mix.{ 'mp3' if out_fmt=='mp3' else 'wav'}"
        out_path = base / out_name

        # Run mix
        final_path, meta = mix_files(
            str(vocal_path), str(music_path), str(out_path),
            sr=sr, music_gain_db=music_gain_db, duck_db=duck_db,
            duck_floor_db=duck_floor_db, target_lufs=target_lufs
        )

        # Optional: also copy to edit_vid_audio for downstream steps
        try:
            import shutil
            shutil.copy(str(final_path), str(eva / out_name))
        except Exception:
            pass

        return jsonify({
            "ok": True,
            "download": f"/video/{out_name}",
            "saved_in_edit_vid_audio": f"/video/edit_vid_audio/{out_name}" if (eva / out_name).exists() else None,
            "inputs": {
                "vocal": str(vocal_path.relative_to(base)),
                "music": str(music_path.relative_to(base)),
            },
            "meta": meta
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.post("/polish_audio")
def polish_audio_endpoint():
    """
    Polishes an audio file and writes result to root as 'audio_clean.m4a'.
    Source priority:
      1) uploaded file 'audio_file' (optional)
      2) first file under edit_vid_audio/
      3) ./audio.wav (root)
    Returns JSON with settings + downloadable link: /video/audio_clean.m4a
    """
    try:
        base_dir = Path(__file__).resolve().parent

        # 1) Decide input
        uploaded = request.files.get("audio_file")
        tmp_in = None
        if uploaded and uploaded.filename:
            tmp_in = base_dir / "tmp_upload_in"
            tmp_in.mkdir(exist_ok=True)
            in_path = tmp_in / uploaded.filename
            uploaded.save(str(in_path))
        else:
            # fallbacks: first file in edit_vid_audio, else root audio.wav
            eva = base_dir / "edit_vid_audio"
            candidates = []
            if eva.exists():
                for f in sorted(eva.iterdir()):
                    if f.suffix.lower() in {".wav",".m4a",".mp3",".aac",".flac",".ogg"}:
                        candidates.append(f)
            if not candidates and (base_dir / "audio.wav").exists():
                candidates.append(base_dir / "audio.wav")
            if not candidates:
                return jsonify({"error": "No input audio found. Upload a file or place one in edit_vid_audio/ or audio.wav."}), 400
            in_path = candidates[0]

        # 2) Read options from form (with safe defaults)
        mode = request.form.get("denoise", "auto")
        target = float(request.form.get("target_lufs", -16.0))
        tp = float(request.form.get("tp", -1.5))
        lra = float(request.form.get("lra", 11.0))
        hp = int(request.form.get("hp", 80))
        lp_raw = request.form.get("lp", "12000")
        lp = int(lp_raw) if lp_raw and lp_raw != "0" else None
        deess = int(request.form.get("deess", 5))
        if deess < 0: deess = 0
        if deess > 10: deess = 10
        mono = request.form.get("mono", "false").lower() == "true"
        ar_raw = request.form.get("ar", "").strip()
        ar = int(ar_raw) if ar_raw else None
        speechnorm = request.form.get("speechnorm", "false").lower() == "true"

        out_path = base_dir / "audio_clean.m4a"  # served by /video/<filename>
        meta = polish_audio(
            input_path=str(in_path),
            output_path=str(out_path),
            denoise_mode=mode,
            target_lufs=target,
            tp_limit=tp,
            lra_target=lra,
            highpass_hz=hp,
            lowpass_hz=lp,
            deess_intensity=deess,
            force_mono=mono,
            samplerate=ar,
            use_speechnorm=speechnorm,
        )

        # optional: also copy to edit_vid_audio for downstream steps
        try:
            (base_dir / "edit_vid_audio").mkdir(exist_ok=True)
            shutil.copy(str(out_path), str(base_dir / "edit_vid_audio" / "audio_clean.m4a"))
        except Exception:
            pass

        # Download via existing route /video/<filename> that serves from project root
        # (you already have: send_from_directory(directory='.', path=filename))
        return jsonify({
            "ok": True,
            "download": "/video/audio_clean.m4a",
            "meta": meta
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

def url_to_fs(url_path: str, base_subdir: str) -> Path:
    """
    Map a URL like '/thumbnail_images/a/b.png' to a safe filesystem path
    under BASE_DIR/base_subdir/a/b.png
    """
    if not url_path:
        abort(400, "Missing image path")
    url_path = unquote(url_path).lstrip("/")              # remove leading '/'
    # strip the first segment (should match base_subdir)
    first, _, tail = url_path.partition("/")
    if first != base_subdir:
        abort(400, f"Unexpected base folder: {first}")
    fs_path = (BASE_DIR / base_subdir / tail).resolve()
    allowed_base = (BASE_DIR / base_subdir).resolve()
    if not str(fs_path).startswith(str(allowed_base)):
        abort(400, "Invalid image path")
    return fs_path

def _duration_via_wave(p: Path):
    """Try Python wave for PCM WAV."""
    import wave
    with wave.open(str(p), 'rb') as w:
        frames = w.getnframes()
        rate = w.getframerate()
        if rate == 0:
            raise ValueError("Invalid WAV: sample rate is 0")
        return frames / float(rate)

def _duration_via_ffprobe(p: Path):
    """Fallback: use ffprobe for any codec/container."""
    cmd = [
        "ffprobe", "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        str(p)
    ]
    out = subprocess.check_output(cmd, stderr=subprocess.STDOUT, text=True).strip()
    # ffprobe prints a float in seconds (can be "N/A")
    if not out or out.upper() == "N/A":
        raise ValueError("ffprobe returned no duration")
    dur = float(out)
    if not math.isfinite(dur):
        raise ValueError("Non-finite duration from ffprobe")
    return dur

@app.route("/api/audio-duration")
def audio_duration():
    p = AUDIO_PATH
    if not p.exists():
        return jsonify({"error": f"File not found: {str(p)}", "cwd": str(BASE_DIR)}), 404
    try:
        # 1) Try wave first (fast, if PCM)
        try:
            seconds = _duration_via_wave(p)
        except Exception as e_wave:
            # 2) Fallback to ffprobe
            seconds = _duration_via_ffprobe(p)

        return jsonify({"seconds": round(seconds, 2), "path": str(p)})
    except subprocess.CalledProcessError as e:
        return jsonify({
            "error": "ffprobe failed",
            "detail": e.output
        }), 500
    except Exception as e:
        return jsonify({
            "error": type(e).__name__,
            "detail": str(e),
            "path": str(p)
        }), 500

@app.route('/edit_vid_thumbnail/<path:filename>')
def serve_generated_thumb(filename):
    return send_from_directory(BASE_DIR / 'edit_vid_thumbnail', filename)


@app.route('/generate_thumbnail', methods=['POST'])
def generate_thumbnail():
    try:
        from thumbnail_gen import create_thumbnail

        image_url = request.form.get('image')  # e.g. "/thumbnail_images/foo/bar.png"
        bg_color = request.form.get('bg_color', '#000000')
        text = request.form.get('text', '')
        colors = request.form.get('colors', 'auto')
        print("Generating thumbnail for image:", image_url)
        # Map the URL to a filesystem path safely
        fs_image_path = url_to_fs(image_url, base_subdir='thumbnail_images')
        print("Filesystem image path:", fs_image_path)

        # Ensure output folder exists; return a served URL
        out_dir = BASE_DIR / 'edit_vid_thumbnail'
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / 'thumbnail.png'
        out_url = f"/edit_vid_thumbnail/{out_file.name}"

        create_thumbnail(
            image_path=str(fs_image_path),
            bg_color=bg_color,
            text=text,
            colors=colors,
            output_path=str(out_file)
        )

        return jsonify({"message": "✅ Thumbnail created successfully!", "thumbnail": out_url})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _safe_join(base: Path, sub: str) -> Path:
    """Prevent path traversal; always return a child of base."""
    p = (base / (sub or "")).resolve()
    base = base.resolve()
    if not str(p).startswith(str(base)):
        abort(400, "Invalid folder")
    return p

def _list_current(base_url: str, base_dir: Path, rel_folder: str, exts: set[str]):
    """
    Return files in the current folder and immediate subfolders.
    - base_url: '/thumbnail_images' or '/background_videos'
    - base_dir: BASE_DIR/'thumbnail_images' or 'background_videos'
    - rel_folder: '' or 'krishna/diwali'
    """
    cur = _safe_join(base_dir, rel_folder)
    if not cur.exists():
        return {"cwd": rel_folder, "folders": [], "files": []}

    # immediate subfolders
    folders = []
    for d in sorted([p for p in cur.iterdir() if p.is_dir()]):
        rel = (Path(rel_folder) / d.name).as_posix()
        folders.append(rel)

    # files in current folder only
    files = []
    for f in sorted([p for p in cur.iterdir() if p.is_file()]):
        if f.suffix.lower() in exts:
            rel_file = (Path(rel_folder) / f.name).as_posix()
            files.append(f"{base_url}/{rel_file}")

    # build breadcrumb segments for UI
    crumbs = []
    accum = []
    for part in Path(rel_folder).parts:
        accum.append(part)
        crumbs.append({"name": part, "path": "/".join(accum)})
    return {"cwd": rel_folder, "folders": folders, "files": files, "breadcrumbs": crumbs}

def _list_all_folders(base_dir: Path):
    """Return ALL subfolders (including root '') for the folder dropdown."""
    out = [""]
    for d, subdirs, _ in os.walk(base_dir):
        rel = os.path.relpath(d, base_dir)
        if rel == ".":
            continue
        out.append(rel.replace("\\", "/"))
    return sorted(out)

# --- STATIC file serving (so subpaths are accessible from <img>/<video> tags) ---
@app.route('/thumbnail_images/<path:filename>')
def serve_thumb_image(filename):
    return send_from_directory(BASE_DIR / 'thumbnail_images', filename)

@app.route('/images/<path:filename>')
def serve_reg_image(filename):
    return send_from_directory(BASE_DIR / 'images', filename)

@app.route('/background_videos/<path:filename>')
def serve_bg_video(filename):
    return send_from_directory(BASE_DIR / 'background_videos', filename)

@app.route('/quiz/downloads/<path:filename>')
def serve_quiz_downloads(filename):
    return send_from_directory(BASE_DIR / 'downloads', filename)

@app.get('/trigger_heygen_bulk_shorts')
def trigger_heygen_bulk_shorts():
    """Trigger background creation of HeyGen shorts."""
    from heygen_submit_videos import main as trigger_shorts
    try:
        count = trigger_shorts()
        return jsonify({"ok": True, "message": f"Triggered creation for {count} HeyGen shorts."})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post('/add_heygen_backgrounds')
def add_heygen_backgrounds():
    """Add HeyGen backgrounds to the system."""
    try:
        orientation = request.form.get('orientation', 'landscape')
        scaleBG = request.form.get('scale_bg', 'yes')
        copy_as_is = request.form.get('copy_asis', 'yes') == 'yes'
        render_bulk_bg(orientation, scaleBG, copy_as_is)
        return jsonify({"ok": True, "message": "HeyGen backgrounds added successfully."})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.get('/trigger_heygen_downloads')
def trigger_heygen_downloads():
    """Trigger background download of HeyGen videos."""
    from heygen_download_and_trash import main as trigger_downloads
    try:
        count = trigger_downloads()
        return jsonify({"ok": True, "message": f"Triggered downloads for {count} HeyGen videos."})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

# --- BROWSING APIs ---
@app.get('/list_thumbnail_images')
def list_thumbnail_images():
    folder = request.args.get('folder', '').strip('/')
    base = BASE_DIR / 'thumbnail_images'
    data = _list_current('/thumbnail_images', base, folder, {'.png', '.jpg', '.jpeg', '.webp'})
    data["all_folders"] = _list_all_folders(base)
    return jsonify(data)

@app.get('/list_background_videos')
def list_background_videos():
    folder = request.args.get('folder', '').strip('/')
    base = BASE_DIR / 'background_videos'
    data = _list_current('/background_videos', base, folder, {'.mp4', '.mov', '.mkv', '.webm'})
    data["all_folders"] = _list_all_folders(base)
    return jsonify(data)

@app.post("/populate_media_jobs")
def populate_media_jobs():
    """
    Reads images from a source folder (server path or uploaded 'temp-del'),
    clears 'media_jobs.xlsx' (keeping headers), and populates:
      - prompt: filename
      - image_path: absolute file path
    """
    try:
        data = request.form
        source_type = data.get("source_type") # 'server' or 'upload'
        server_path_input = (data.get("server_path") or "").strip()
        uploaded_files = request.files.getlist("image_files")
        
        base_dir = BASE_DIR
        excel_path = base_dir / "media_jobs.xlsx"
        
        # 1. Determine Target Folder
        target_folder = None
        
        if source_type == "upload":
            # Use 'temp-del' for uploads to generate local paths
            target_folder = base_dir / "temp-del"
            target_folder.mkdir(exist_ok=True)
            
            # Clear existing files in temp-del to avoid mixing batches
            for f in target_folder.iterdir():
                if f.is_file():
                    f.unlink()
            
            # Save uploaded files
            if not uploaded_files or uploaded_files[0].filename == '':
                return jsonify({"ok": False, "error": "No files uploaded."}), 400
                
            for file in uploaded_files:
                # webkitdirectory sends relative paths (folder/file.jpg), we need just filename usually
                # or we keep structure. For this specific job, flat lists are usually better.
                fname = Path(file.filename).name 
                file.save(target_folder / fname)
                
        else:
            # Server path
            if not server_path_input:
                return jsonify({"ok": False, "error": "Server path is empty."}), 400
            
            # Handle absolute vs relative paths
            if os.path.isabs(server_path_input):
                target_folder = Path(server_path_input)
            else:
                target_folder = (base_dir / server_path_input).resolve()
                
            if not target_folder.exists():
                return jsonify({"ok": False, "error": f"Folder not found: {target_folder}"}), 404

        # 2. Get Image Files
        valid_exts = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
        images = [
            f for f in target_folder.iterdir() 
            if f.is_file() and f.suffix.lower() in valid_exts
        ]
        
        if not images:
            return jsonify({"ok": False, "error": "No images found in target folder."}), 400

        # 3. Update Excel
        if not excel_path.exists():
            return jsonify({"ok": False, "error": "media_jobs.xlsx not found."}), 404

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Clear old data (keep header at row 1)
        # delete_rows is efficient; delete from row 2 to end
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
            
        # 4. Populate Data
        # Header assumed: prompt, account_id, image_path, video_cmd, video_path, status, ...
        # We map: prompt -> Col A (1), image_path -> Col C (3)
        
        count = 0
        for img in sorted(images, key=lambda x: x.name):
            ws.append([
                img.name,       # A: prompt
                "",             # B: account_id
                str(img.resolve()), # C: image_path (Absolute path)
                "",             # D: video_cmd
                "",             # E: video_path
                "pending"       # F: status (optional, setting default)
            ])
            count += 1
            
        wb.save(excel_path)
        wb.close()
        
        return jsonify({
            "ok": True, 
            "message": f"Updated media_jobs.xlsx with {count} images.",
            "folder": str(target_folder)
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.post("/render_pinterest_pins_from_pin_data")
def render_pinterest_pins_from_pin_data():
    """
    Reads PIN_DATA.xlsx from a folder, renders image/video pins using project_json overlays,
    writes outputs to pinterest_uploads/pinterest_pins, and appends records to master_shorts_uploader_data.xlsx
    """
    try:
        from pathlib import Path
        from time import time

        from pin_overlay_batch import batch_render_from_folder

        base_dir = BASE_DIR

        data = request.form
        uploaded_files = request.files.getlist("source_dir")  # directory upload (optional)
        source_folder = (data.get("source_folder") or "").strip()
        # print("Source folder:", source_folder)
        pin_type = (data.get("pin_type") or "image").strip().lower()  # image|video
        max_pins = int((data.get("max_pins") or "0").strip() or 0)
        skip_overlays = (data.get("skip_overlays") or "no").strip().lower() == "yes"

        if pin_type not in ("image", "video"):
            pin_type = "image"

        print(f"uploaded_files: {uploaded_files}, source_folder: {source_folder}, pin_type: {pin_type}, max_pins: {max_pins}")
        # Where the folder is
        if source_folder == "" and uploaded_files and len(uploaded_files) > 0:
            upload_root = base_dir / "pinterest_uploads" / "uploads"
            upload_root.mkdir(parents=True, exist_ok=True)

            # Clear previous contents
            for child in upload_root.iterdir():
                if child.is_file():
                    child.unlink()
                elif child.is_dir():
                    import shutil
                    shutil.rmtree(child)

            for fs in uploaded_files:
                filename = Path(fs.filename).name
                dest_path = upload_root / filename
                fs.save(dest_path)

            folder = upload_root
        else:
            # Use server-side folder relative to BASE_DIR or downloads
            # If you prefer downloads/ only, swap BASE_DIR for COLORING_BASE
            folder = (base_dir / source_folder) if source_folder else None
            if not folder or not folder.exists():
                return jsonify({"ok": False, "error": f"Folder not found: {folder}"}), 400

        out_dir = base_dir / "pinterest_uploads" / "pinterest_pins"
        out_dir.mkdir(parents=True, exist_ok=True)

        master_excel = base_dir / "master_shorts_uploader_data.xlsx"

        result = batch_render_from_folder(
            folder=folder,
            pin_type=pin_type,
            max_pins=max_pins,
            out_dir=out_dir,
            master_excel=master_excel,
            skip_overlays=skip_overlays
        )
        return jsonify(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.post("/generate_pinterest_excel_for_coloring_pins")
def generate_pinterest_excel_route():
    """
    Run generate_excel.py to create a Pinterest Pin Excel file
    and Pinterest-ready images from either:
      - an uploaded folder (directory upload from browser), or
      - an existing folder under COLORING_BASE (downloads).
    """
    try:
        base_dir = BASE_DIR
        script_path = base_dir / "for_coloring_pin_load_pinterest_excel.py"

        if not script_path.exists():
            return jsonify({
                "ok": False,
                "error": f"Script not found: {script_path}"
            }), 500

        # --- Read form fields ---
        data = request.form
        uploaded_files = request.files.getlist("source_dir")  # directory upload

        # Optional: existing folder on server (relative to COLORING_BASE)
        source_folder = (data.get("source_folder") or "").strip()

        media_type = (data.get("media_type") or "image").strip()
        # pin_url = (data.get("pin_url") or "").strip()
        # pin_url = ""
        max_pins_str = (data.get("max_pins") or "0").strip()
        output_excel_name = (data.get("output_excel") or "master_shorts_uploader_data.xlsx").strip()

        # These can be blank – generate_excel.py will fall back to pinterest_config.json
        book_title = (data.get("book_title") or "").strip()
        book_url = (data.get("book_url") or "").strip()
        board_name = (data.get("board_name") or "").strip()
        banner_text = (data.get("banner_text") or "").strip()
        watermark_text = (data.get("watermark_text") or "").strip()

        fit_mode = (data.get("fit_mode") or "contain").strip()
        bg_style = (data.get("bg_style") or "white").strip()
        text_shadow = (data.get("text_shadow") or "yes").strip().lower()  # "yes"/"no"

        use_gemini_flag = (data.get("use_gemini") or "no").strip().lower()

        add_bg_music_flag = (data.get("add_bg_music") or "no").strip().lower()

        upload_pinterest_flg = (data.get("upload_pinterest") or "no").strip().lower()

        auto_crop_subject_flag = (data.get("auto_crop_subject") or "yes").strip().lower()
        if auto_crop_subject_flag not in ("yes", "no"):
            auto_crop_subject_flag = "yes"

        try:
            max_pins = int(max_pins_str)
        except ValueError:
            max_pins = 0

        video_duration_str = (data.get("video_duration") or "8").strip()
        video_fps_str = (data.get("video_fps") or "30").strip()

        try:
            video_duration = float(video_duration_str)
        except ValueError:
            video_duration = 8.0

        try:
            video_fps = int(video_fps_str)
        except ValueError:
            video_fps = 30

        video_style = (data.get("video_style") or "single").strip()
        if video_style not in ("single", "flipbook", "slideshow"):
            video_style = "single"

        pages_per_video_str = (data.get("pages_per_video") or "10").strip()
        try:
            pages_per_video = int(pages_per_video_str)
        except ValueError:
            pages_per_video = 10


        # Decide where images_root & source_subfolder will be
        images_root = None
        source_subfolder = None

        # --- CASE 1: user uploaded a directory from their computer ---
        if uploaded_files:
            # Store folder under pinterest_uploads/upload_<timestamp>/
            #upload_root = BASE_DIR / "pinterest_uploads" / f"upload_{int(time())}"

            upload_root = BASE_DIR / "pinterest_uploads/uploads" 
            upload_root.mkdir(parents=True, exist_ok=True)

            # Clear previous contents
            for child in upload_root.iterdir():
                if child.is_file():
                    child.unlink()
                elif child.is_dir():
                    import shutil
                    shutil.rmtree(child)
            
            # With webkitdirectory, file.filename usually contains a relative path like "MyBook/page_01.png"
            for fs in uploaded_files:
                filename = Path(fs.filename).name  # ⬅ strip subfolders
                dest_path = upload_root / filename
                fs.save(dest_path)


            images_root = upload_root
            source_subfolder = None  # we already uploaded the exact folder tree
            print(f"[INFO] Using uploaded folder as images_root: {images_root}")

        # --- CASE 2: existing folder on server under COLORING_BASE (/downloads) ---
        else:
            # base images_root is COLORING_BASE, e.g. BASE_DIR/downloads
            images_root = COLORING_BASE
            source_subfolder = source_folder or None
            print(f"[INFO] Using server folder. images_root={images_root}, source_subfolder={source_subfolder}")

        output_excel = base_dir / output_excel_name

        # Build command for generate_excel.py
        cmd = [
            sys.executable,
            str(script_path),
            "--images-root", str(images_root),
            "--output-excel", str(output_excel),
            "--media-type", media_type,
            # "--pin-url", pin_url,
        ]

        if source_subfolder:
            cmd += ["--source-subfolder", source_subfolder]
        if max_pins > 0:
            cmd += ["--max-pins", str(max_pins)]

        # NEW: video parameters (safe to pass for both image & video)
        cmd += [
            "--video-style", video_style,
            "--pages-per-video", str(pages_per_video),
            "--video-duration", str(video_duration),
            "--video-fps", str(video_fps),
        ]


        # Optional overrides – safe to pass even if empty
        if book_title:
            cmd += ["--book-title", book_title]
        if book_url:
            cmd += ["--book-url", book_url]
        if board_name:
            cmd += ["--board-name", board_name]
        if banner_text:
            cmd += ["--banner-text", banner_text]
        if watermark_text:
            cmd += ["--watermark-text", watermark_text]

        # New options
        cmd += ["--fit-mode", fit_mode]
        cmd += ["--bg-style", bg_style]
        cmd += ["--text-shadow", text_shadow]
        cmd += ["--auto-crop-subject", auto_crop_subject_flag]  # <--- NEW

        cmd += ["--use-gemini", use_gemini_flag]

        cmd += ["--add-bg-music", add_bg_music_flag]

        proc = subprocess.run(
            cmd,
            cwd=str(base_dir),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            # timeout=60 * 60,
        )

        ok = (proc.returncode == 0)
        stdout_tail = (proc.stdout or "")[-4000:]
        stderr_tail = (proc.stderr or "")[-4000:]

        # if media_type == "video":


        if upload_pinterest_flg == "yes" and ok:
            upload_pins() #Pinterest upload
            # if media_type == "video" or media_type == "coloring":
            upload_shorts_from_master_file() #YouTube shorts upload
            upload_facebook_videos() #Facebook videos upload
            upload_tiktok_videos() #TikTok videos upload
            upload_instagram_posts() #Instagram posts upload
        
        return jsonify({
            "ok": ok,
            "returncode": proc.returncode,
            "stdout": stdout_tail,
            "stderr": stderr_tail,
        }), (200 if ok else 500)

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            "ok": False,
            "error": str(e),
        }), 500


@app.route('/convert_landscape_images', methods=['POST'])
def convert_images_route():
    """Convert images from edit_vid_input -> target size and write under edit_vid_output."""
    try:
        print("Processing request.convert_images")

        mode = (request.form.get('mode') or 'portrait').strip().lower()
        fit = (request.form.get('fit') or 'contain').strip().lower()
        bg = (request.form.get('bg') or 'blur').strip().lower()
        quality = request.form.get('quality') or 92

        # Optional custom size
        target_w_in = request.form.get('target_w')
        target_h_in = request.form.get('target_h')

        in_dir = (BASE_DIR / 'edit_vid_input').resolve()
        out_dir = (BASE_DIR / 'edit_vid_output').resolve()

        MODE_TO_SIZE = {
            "portrait": (1080, 1920),
            "1000x1500": (1000, 1500),
            "landscape": (1920, 1080),  # ✅ locked as requested
        }

        if fit not in ("contain", "cover"):
            return jsonify({"ok": False, "error": f"Invalid fit: {fit}"}), 400
        if bg not in ("blur", "white"):
            return jsonify({"ok": False, "error": f"Invalid bg: {bg}"}), 400

        try:
            quality = max(1, min(int(quality), 95))
        except Exception:
            quality = 92

        # Resolve target size
        if mode == "custom":
            try:
                target_w = int(target_w_in)
                target_h = int(target_h_in)
                if target_w <= 0 or target_h <= 0:
                    raise ValueError
            except Exception:
                return jsonify({
                    "ok": False,
                    "error": "Invalid custom size. Provide target_w and target_h."
                }), 400
        else:
            if mode not in MODE_TO_SIZE:
                return jsonify({
                    "ok": False,
                    "error": f"Invalid mode: {mode}. Supported: {list(MODE_TO_SIZE.keys()) + ['custom']}"
                }), 400
            target_w, target_h = MODE_TO_SIZE[mode]

        if not in_dir.exists():
            return jsonify({"ok": False, "error": f"Input folder not found: {in_dir}"}), 400

        from pathlib import Path
        import os
        from convert_landscape_img import convert_one, SUPPORTED_EXTS

        read_count = 0
        wrote_count = 0

        for root, _, files in os.walk(in_dir):
            root_p = Path(root)
            rel = root_p.relative_to(in_dir)

            for fn in files:
                if Path(fn).suffix.lower() not in SUPPORTED_EXTS:
                    continue

                read_count += 1
                in_path = root_p / fn
                out_path = out_dir / rel / Path(fn).with_suffix(".png")

                try:
                    convert_one(
                        in_path=in_path,
                        out_path=out_path,
                        target_w=target_w,
                        target_h=target_h,
                        fit=fit,
                        bg_style=bg,
                        quality=quality,
                    )
                    wrote_count += 1
                except Exception as e:
                    print(f"⚠️ Failed: {in_path} -> {e}")

        return jsonify({
            "ok": True,
            "read_count": read_count,
            "wrote_count": wrote_count,
            "output_dir": str(out_dir),
            "mode": mode,
            "target_w": target_w,
            "target_h": target_h,
            "fit": fit,
            "bg": bg,
            "quality": quality,
        }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500



@app.route('/batchmakevideoimagesfromdir', methods=['POST'])
def batch_make_video_images_from_dir_route():
    try:
        print("Processing request.batch_make_video_images_from_dir")

        # Read directory path from the form
        raw_source = (request.form.get('source_dir') or "").strip()
        target_size = (request.form.get('size_option') or "portrait").strip()
        if not raw_source:
            return "❌ Please provide a source directory.", 400

        # Normalize slashes
        raw_source = raw_source.replace("\\", "/")
        print(f"[DEBUG] raw source_dir from UI: {raw_source}")

        from pathlib import Path

        # If it's not absolute, treat it as relative to BASE_DIR
        p = Path(raw_source)
        if not p.is_absolute():
            full_source_dir = (BASE_DIR / raw_source).resolve()
        else:
            full_source_dir = p.resolve()

        print(f"[DEBUG] Resolved full source_dir: {full_source_dir}")

        # Import your batch helper
        from for_coloring_pin_load_pinterest_excel import batch_make_video_images_from_dir

        # Call the batch function; it returns processed image count
        result = batch_make_video_images_from_dir(str(full_source_dir), target_size)

        if isinstance(result, int):
            msg = f"✅ Batch video generation completed. Processed {result} image(s)."
        else:
            msg = "✅ Batch video generation completed."

        print(msg)
        return msg, 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"❌ Error: {str(e)}", 500



@app.route('/process_master_shorts_file_data', methods=['POST'])
def process_master_shorts_file_data():
    try:
        upload_pins() #Pinterest upload
        upload_shorts_from_master_file() #YouTube shorts upload
        upload_facebook_videos() #Facebook videos upload
        upload_tiktok_videos() #TikTok videos upload
        upload_instagram_posts() #Instagram posts upload
        return "✅ Processing completed successfully!", 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    
@app.route('/select_background_video', methods=['POST'])
def select_background_video():
    """Copy chosen video to edit_vid_input folder for use."""
    try:
        # Be tolerant to different content types
        payload = request.get_json(silent=True) or {}
        src = (payload or {}).get('video') or request.form.get('video') or request.values.get('video')

        if not src:
            return jsonify({"error": "Missing video path"}), 400

        clear_folder("edit_vid_input")  # your helper

        filename = "bg_video.mp4"  # fixed name on destination
        src_path = os.path.join(BASE_DIR, src.strip("/"))
        dest_path = os.path.join(BASE_DIR, "edit_vid_input", filename)

        if not os.path.exists(src_path):
            return jsonify({"error": f"File not found: {src_path}"}), 404

        import shutil
        shutil.copy(src_path, dest_path)
        # return jsonify({"message": "✅ Video selected and ready for editing!", "dest": dest_path})
        return "✅ Processing completed successfully!", 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/sounds/<path:filename>')
def serve_sound(filename):
    sounds_dir = os.path.join(app.root_path, 'sounds')
    return send_from_directory(sounds_dir, filename)


# ------------------------ API ROUTES ------------------------ #

@app.route('/get_full_text', methods=['GET'])
def get_full_text():
    try:
        with open("temp/full_text.txt", "r", encoding="utf-8") as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        traceback.print_exc() 
        return jsonify({"error": str(e)}), 500

@app.route('/get_word_timestamps', methods=['GET'])
def get_word_timestamps():
    try:
        with open("temp/word_timestamps.json", "r", encoding="utf-8") as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        traceback.print_exc() 
        return jsonify({"error": str(e)}), 500

@app.route('/save_word_timestamps', methods=['POST'])
def save_word_timestamps():
    try:
        data = request.json
        with open("temp/word_timestamps.json", "w", encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        return jsonify({"message": "✅ Word timestamps updated successfully!"})
    except Exception as e:
        traceback.print_exc() 
        return jsonify({"error": str(e)}), 500

@app.route('/get_structured_output', methods=['GET'])
def get_structured_output():
    try:
        with open("temp/structured_output.json", "r") as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        traceback.print_exc() 
        return jsonify({"error": str(e)}), 500

# ------------------------ HTML UI ROUTES ------------------------ #

@app.route('/')
def index():
    return render_template('index.html',
        music_options=background_music_options,
        style_options=font_settings,
        tts_options=tts_engine,
        voice_genders=voices,
        voice_map=voices,
        sizes=sizes
    )

@app.route('/thumbnail')
def thumbnail():
    return render_template('thu_index.html')

@app.route('/scene_builder')
def scene_builder():
    return render_template('scene_builder.html')

@app.route('/coloring_book_builder')
def coloringbook():
    return render_template('coloring-book-builder.html')

@app.route('/hindi_caption_builder')
def hindicaptions():
    return render_template('hindi_caption_builder.html')

@app.route('/aivideoprompt')
def aivideoprompt():
    return render_template('aivideoprompt.html')

@app.route('/prepare_captions')
def prep_caption():
    return render_template('index_captions.html')

@app.route('/flipping_book')
def flipping_book():
    return render_template('flipping_book.html')

@app.route('/portrait_website_loader')
def portrait_website_loader():
    # return render_template('portrait_website_loader.html')
    return redirect("https://lab.readernook.com/tools/portrait-website-loader/")

@app.route('/video/<filename>')
def serve_video(filename):
    return send_from_directory(directory='.', path=filename)

@app.route('/create_portrait_n_add_caption')
def create_portrait_n_add_caption():
    try:
        print("Processing request...create_portrait_n_add_caption")
        create_portrait()
        add_caption()
        return "✅ create_portrait_n_add_caption completed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500

@app.route('/apply_caption')
def add_caption():
    try:
        print("Processing request...apply_caption")

        from apply_captions import add_captions_to_video

        # If edit_vid_output/out_landscape.mp4 and edit_vid_output/captions_landscape.ass exists
        

        if os.path.isfile("edit_vid_output/out_landscape.mp4") and os.path.isfile("edit_vid_output/captions_landscape.ass"):
            add_captions_to_video("edit_vid_output/out_landscape.mp4", "edit_vid_output/captions_landscape.ass", "edit_vid_output/landscape_with_captions.mp4")
        if os.path.isfile("edit_vid_output/out_portrait.mp4") and os.path.isfile("edit_vid_output/captions_portrait.ass"):
            add_captions_to_video("edit_vid_output/out_portrait.mp4", "edit_vid_output/captions_portrait.ass", "edit_vid_output/portrait_with_captions.mp4")

        return "✅ apply_caption completed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500  


@app.route('/convert_landscape_to_portrait')
def create_portrait():
    try:
        print("Processing request...convert_landscape_to_portrait")

        from convert_to_portrait import convert_landscape_to_portrait
        if os.path.isfile("edit_vid_output/out_landscape.mp4"):
            convert_landscape_to_portrait("edit_vid_output/out_landscape.mp4", "edit_vid_output/out_portrait.mp4")

        return "✅ convert_landscape_to_portrait completed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500  

@app.route('/process', methods=['POST'])
def process():
    try:
        print("Processing request...")
        urls = request.form.get('urls', '')
        html = request.form.get('html', '')
        excel = request.form.get('excel', 'no')
        notebooklm = request.form.get('notebooklm', 'no')
        language = request.form.get('language', 'english')
        tts = request.form.get('tts', 'google')
        gender = request.form.get('gender', 'Female')
        voice = request.form.get('voice', 'Joanna')
        size = request.form.get('size', 'YouTube Shorts')
        music = request.form.get('music')
        max_words = int(request.form.get('max_words', 4))
        fontsize = int(request.form.get('fontsize', 90))
        y_pos = request.form.get('y_pos', 'center')
        style = request.form.get('style', 'style2')
        skip_puppeteer = request.form.get('skip_puppeteer', 'no')
        skip_captions = request.form.get('skip_captions', 'no')
        pitch = request.form.get('pitch', 'adult')
        disable_subscribe = request.form.get('disable_subscribe', 'no')

        scrape_and_process(urls, excel, size, music, max_words, fontsize, y_pos,
                           style, voice, language, gender, tts,
                           skip_puppeteer, skip_captions, pitch, disable_subscribe, notebooklm, html)

        # return "✅ Processing started!"
        return "✅ Processing completed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500

@app.post("/prepare_captions_from_heygen")
def heygen_init_captions():
    """
    Upload HeyGen mp4, extract audio.wav, generate temp/word_timestamps.json,
    then UI can call /word_timestamps to load it.
    """
    try:
        CAPTION_LANG_OPTIONS = ["english", "hindi"]
        language = (request.form.get("language") or "english").strip().lower()
        if language not in CAPTION_LANG_OPTIONS:
            language = "english"

        music = (request.form.get("music") or "no").strip().lower()  # optional
        is_song = (music == "yes")

        f = request.files.get("heygen")
        if not f or not f.filename:
            return jsonify({"ok": False, "error": "Missing HeyGen video file (form field name: file)."}), 400

        orig_name = Path(f.filename).stem   # "my_heygen_video"
        orig_ext  = Path(f.filename).suffix # ".mp4"

        meta_path = OUT_DIR / "job_meta.json"
        meta_path.write_text(json.dumps({
            "orig_name": orig_name,
            "orig_ext": orig_ext
        }, indent=2))

        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in VIDEO_EXTS:
            return jsonify({"ok": False, "error": f"Unsupported video type: {ext}"}), 400

        # Save uploaded HeyGen video
        uploads_dir = Path(__file__).resolve().parent / "uploads"
        uploads_dir.mkdir(exist_ok=True)
        heygen_path = uploads_dir / f"heygen{ext}"
        f.save(str(heygen_path))

        # Extract audio.wav beside server.py (AUDIO_PATH already points to BASE_DIR/audio.wav)
        base_dir = Path(__file__).resolve().parent
        res = extract_audio_from_video(
            base_dir=base_dir,
            input_path=heygen_path.resolve(),
            fmt="wav",
            track=0,
            root_output_name="audio"  # -> audio.wav
        )

        if not getattr(res, "ok", False):
            return jsonify({"ok": False, "error": res.error or "Audio extract failed", "detail": res.detail}), 500

        # Generate captions/word timestamps (writes temp/word_timestamps.json)
        prepare_captions_file_for_notebooklm_audio(
            audio_path=str(AUDIO_PATH),   # "audio.wav"
            language=language,
            is_song=is_song
        )

        return jsonify({
            "ok": True,
            "language": language,
            "audio": "audio.wav",
            "word_timestamps": "/word_timestamps"
        }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/caption', methods=['POST'])
def caption():
    try:
        print("Processing request...caption")
        music = request.form.get('music', 'no')
        language = request.form.get('language', 'english')
        prepare_captions_file_for_notebooklm_audio(
            audio_path="audio.wav",
            language=language,
            is_song=music == 'yes'
        )

        # return "✅ Processing started!"
        return "✅ Processing completed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500

@app.route('/generatettsaudio', methods=['POST'])
def generatettsaudio():
    try:
        print("Processing request...generatettsaudio")

        ttstext = request.form.get('ttstext', '')

        language = request.form.get('language', 'english')
        tts_engine = request.form.get('tts', 'google')
        gender = request.form.get('gender', 'Female')
        output_audio_file = "audio.wav"
        clear_folder("edit_vid_audio")

        generated_file = get_audio_file(ttstext, output_audio_file, tts_engine, language, gender)
        shutil.copy("audio.wav", "edit_vid_audio/audio.wav")
        
        # return "✅ Processing started!"
        return "✅ Processing completed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500

@app.route("/get_audio", methods=["POST"])
def get_audio():
    """Generate and return TTS audio for question text."""
    data = request.get_json(force=True)
    text = data.get("text", "").strip()
    language = data.get("language", "en").lower()
    if not text:
        return jsonify({"error": "Missing text"}), 400

    try:
        # Temporary file for the output audio
        tmp_file = tempfile.NamedTemporaryFile(suffix=".mp3", delete=False)
        tmp_path = tmp_file.name
        tmp_file.close()

        # Generate audio using the existing TTS pipeline
        get_audio_file(
            text=text,
            audio_file_name=tmp_path,
            tts_engine="google",   # or "amazon" if you prefer
            language="hindi" if language.startswith("hi") else "english",
            gender="Male",
            type="journey",        # "neural" / "journey" / "generative"
            age_group="adult"
        )

        return send_file(tmp_path, mimetype="audio/mpeg")

    except Exception as e:
        print(f"[TTS ERROR] {e}")
        traceback.print_exc() 
        return jsonify({"error": str(e)}), 500

@app.route('/bulkupload', methods=['POST'])
def bulkupload():
    try:
        print("Processing request...upload_videos")
        upload_videos()  # Assuming this function is defined in youtube_uploader.py

        # return "✅ Processing started!"
        return "✅ Processing completed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500
    
@app.route('/runobsrecorder', methods=['POST'])
def run_obs_recorder():
    try:
        print("Processing request...run OBS recorder")
        orientation = request.form.get('orientation', 'landscape')
        duration = request.form.get('duration', '10')  # Default to 10 seconds if not provided
        selectedStyle = request.form.get('captStyle', 'style1')
        captionLength = request.form.get('captLength', '5')
        bgMusicSelected = request.form.get('bgMusicSelect', 'none')
        minLineGapSec = request.form.get('minLineGapSec', '0.40')
        disableSubscribe = request.form.get('disableSubscribe', 'yes')
        outputfile = request.form.get('outputfile', 'test.mp4')
        language = request.form.get('language', 'english')

        if language == "english":
            if selectedStyle == "story-block":
                selectedStyle = "story-block-english"
            elif selectedStyle == "song-block":
                selectedStyle = "song-block-english"


        print("Form payload →", request.form.to_dict(flat=False))
        
        cmd = [
            "node", "puppeteer-launcher.js",
            outputfile, duration, orientation, captionLength, selectedStyle,
            bgMusicSelected, "0.05", "1", disableSubscribe, minLineGapSec
        ]
        print("▶️ Running Puppeteer with:", cmd)
        import subprocess
        subprocess.run(cmd)

        shutil.copy(outputfile, "processed_videos/output.mp4")
        # Here you would call the function to run OBS recorder
        # For example: run_obs_recorder_function()

        # return "✅ Processing started!"
        return "✅ OBS Recorder started successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500
    
@app.route('/editvideos', methods=['POST'])
def run_video_editor():
    try:
        print("Processing request...run edit videos")
        orientation = request.form.get('orientation', 'auto')
        add_music = True
        bg_music_folder = request.form.get('bgmusic')
        if bg_music_folder == 'none':
            add_music = False
        topcut = request.form.get('topcut',0)
        if topcut == '':
            topcut = 0

        bottomcut = request.form.get('bottomcut',0)
        if bottomcut == '':
            bottomcut = 0

        slowfactor = request.form.get('slowfactor',0)
        if slowfactor == '':
            slowfactor = 0

        slow_down = True
        if slowfactor == 0:
            slow_down = False

        add_watermark = True

        watermarkposition = request.form.get('watermarkposition','bottom-left')
        if watermarkposition == "none":
            add_watermark = False

        batch_process(
            input_folder="edit_vid_input",
            output_folder="edit_vid_output",
            bg_music_folder="god_bg",
            remove_top=float(topcut),
            remove_bottom=float(bottomcut),
            add_music=add_music,
            slow_down=slow_down,
            slow_down_factor=float(slowfactor),
            target_orientation=orientation, 
            add_watermark=add_watermark,
            watermark_path="logo.png",
            watermark_position=watermarkposition,
            watermark_scale=0.15
        )
        return "✅ Videos Processed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500    

@app.route('/sunotovideogenerator', methods=['POST'])
def run_sunotovideogenerator():
    try:
        print("*** Processing request sunotovideogenerator: Enlarging clip")
        orientation = 'auto'
        add_music = False
        topcut = 0
        bottomcut = 0
        slow_down = True
        slowfactor = request.form.get('slowfactor',1)
        if slowfactor == '' or slowfactor == '1':
            slowfactor = 1
            slow_down = False

        size = request.form.get('size', 'landscape')

        add_watermark = False
        watermarkposition = 'bottom-left'

        batch_process(
            input_folder="edit_vid_input",
            output_folder="edit_vid_output",
            bg_music_folder="god_bg",
            remove_top=float(topcut),
            remove_bottom=float(bottomcut),
            add_music=add_music,
            slow_down=slow_down,
            slow_down_factor=float(slowfactor),
            target_orientation=orientation, 
            add_watermark=add_watermark,
            watermark_path="logo.png",
            watermark_position=watermarkposition,
            watermark_scale=0.15
        )
        print("✅ Processing request sunotovideogenerator: Enlarging completed successfully")
        # copy files from edit_vid_output to edit_vid_input for next step
        import shutil
        for filename in os.listdir("edit_vid_output"):
            shutil.copy(os.path.join("edit_vid_output", filename), os.path.join("edit_vid_input", filename))

        print("*** Processing request sunotovideogenerator: Assembling clips to make video song")
        assemble_videos(
            video_folder="edit_vid_input",                  # or "edit_vid_output" if you pre-made KB clips
            audio_folder="edit_vid_audio",
            output_path="edit_vid_output/composed_video.mp4",
            fps=30,
            shuffle=True,                                   # different order each run
            prefer_ffmpeg_concat=True                       # auto-uses concat if safe; else MoviePy
        )

        print("✅ Processing request sunotovideogenerator: Assembling completed successfully")
        # place copy of audio file (mp3/wav) from edit_vid_audio to root folder as audio.wav

        audio_folder = "edit_vid_audio"
        import shutil
        from pydub import AudioSegment
        # get the first file found in that folder
        audio_file = os.listdir(audio_folder)[0]
        audio_path = os.path.join(audio_folder, audio_file)

        # output path in root folder
        output_path = "audio.wav"

        # convert mp3 → wav (if needed), else just copy
        if audio_file.lower().endswith(".mp3"):
            # convert using pydub (requires ffmpeg installed)
            sound = AudioSegment.from_mp3(audio_path)
            sound.export(output_path, format="wav")
        else:
            # already wav → overwrite if exists
            shutil.copy(audio_path, output_path)

        print(f"✅ Saved audio file as {output_path}")

        #place copy of composed_video.mp4 from edit_vid_output to root folder as composed_video.mp4. Replace if exists

        shutil.copy("edit_vid_output/composed_video.mp4", "composed_video.mp4")

        # Also place copy of composed_video.mp4 from edit_vid_output to edit_vid_output/out_{size}.mp4. Replace if exists
        shutil.copy("edit_vid_output/composed_video.mp4", f"edit_vid_output/out_{size}.mp4")

        print("✅ Copied composed video to root folder as composed_video.mp4")


        print("*** Processing request sunotovideogenerator...creating captions file ")
        music = request.form.get('music', 'no')
        language = request.form.get('language', 'english')
        if language != 'hindi':
            prepare_captions_file_for_notebooklm_audio(
                audio_path="audio.wav",
                language=language,
                is_song=music == 'yes'
            )
        # else:
        #     if size == 'landscape':
        #         create_portrait()
            # add_caption()
        return "✅ Videos Processed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500    


@app.route('/sunonimagetovideogenerator', methods=['POST'])
def run_sunonimagetovideogenerator():
    try:
        print("*** Processing request sunonimagetovideogenerator: creating clips")
        duration = request.form.get('duration',10)
        if duration == '':
            duration = 10

        video_size = (1920,1080)
        size = request.form.get('size', 'landscape')
        if size == 'portrait':
            video_size = (1080,1920)

        from make_kb_videos import export_kb_videos
        export_kb_videos(
            input_folder="edit_vid_input",   # folder with images
            out_folder="edit_vid_output",    # where to save KB clips
            per_image=int(duration),            # seconds per image
            output_size=video_size,
            zoom_start=1.0, zoom_end=1.05
        )
        print("✅ Processing request sunonimagetovideogenerator: Creating clips completed successfully")
        # copy files from edit_vid_output to edit_vid_input for next step
        import shutil
        for filename in os.listdir("edit_vid_output"):
            shutil.copy(os.path.join("edit_vid_output", filename), os.path.join("edit_vid_input", filename))

        print("*** Processing request sunonimagetovideogenerator: Assembling clips to make video song")

        assemble_videos(
            video_folder="edit_vid_input",                  # or "edit_vid_output" if you pre-made KB clips
            audio_folder="edit_vid_audio",
            output_path="edit_vid_output/composed_video.mp4",
            fps=30,
            shuffle=True,                                   # different order each run
            prefer_ffmpeg_concat=True                       # auto-uses concat if safe; else MoviePy
        )

        print("✅ Processing request sunonimagetovideogenerator: Assembling completed successfully")
        # place copy of audio file (mp3/wav) from edit_vid_audio to root folder as audio.wav

        audio_folder = "edit_vid_audio"
        import shutil
        from pydub import AudioSegment
        # get the first file found in that folder
        audio_file = os.listdir(audio_folder)[0]
        audio_path = os.path.join(audio_folder, audio_file)

        # output path in root folder
        output_path = "audio.wav"

        # convert mp3 → wav (if needed), else just copy
        if audio_file.lower().endswith(".mp3"):
            # convert using pydub (requires ffmpeg installed)
            sound = AudioSegment.from_mp3(audio_path)
            sound.export(output_path, format="wav")
        else:
            # already wav → overwrite if exists
            shutil.copy(audio_path, output_path)

        print(f"✅ Saved audio file as {output_path}")

        #place copy of composed_video.mp4 from edit_vid_output to root folder as composed_video.mp4. Replace if exists

        shutil.copy("edit_vid_output/composed_video.mp4", "composed_video.mp4")
        print("✅ Copied composed video to root folder as composed_video.mp4")

        # Also place copy of composed_video.mp4 from edit_vid_output to edit_vid_output/out_{size}.mp4. Replace if exists
        shutil.copy("edit_vid_output/composed_video.mp4", f"edit_vid_output/out_{size}.mp4")

        print("*** Processing request sunotovideogenerator...creating captions file ")
        language = request.form.get('language', 'english')
        music = request.form.get('music', 'no')

        if language != 'hindi':
            prepare_captions_file_for_notebooklm_audio(
                audio_path="audio.wav",
                language=language,
                is_song=music == 'yes'
            )
        # else:
        #     if size == 'landscape':
        #         create_portrait()
        #     # add_caption()
        return "✅ Videos Processed successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500    

# ------------------------ MAIN ------------------------ #
@app.route('/addoverlays', methods=['POST'])
def add_overlays():
    try:
        print("Processing request...add overlays")
        add_petal_overlay = request.form.get('add_petals', 'no') == 'yes'
        add_sparkle_overlay = request.form.get('add_sparkles', 'no') == 'yes'
        overlay_position = (0, 0)  # Default position, can be modified as needed

        from add_overlays import add_gif_overlays_to_videos
        add_gif_overlays_to_videos(
            input_folder="edit_vid_input",
            output_folder="edit_vid_output",
            add_petal_overlay=add_petal_overlay,
            add_sparkle_overlay=add_sparkle_overlay,
            overlay_position=overlay_position
        )
        return "✅ Overlays added successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500  

@app.route('/multiplyvideo', methods=['POST'])
def multiply_video():
    try:
        print("Processing request...multiplyvideo")
        repeat_factor = request.form.get('repeat_factor', 1)
        if repeat_factor == '':
            repeat_factor = 1

        from multiply_video import multiply_videos
        multiply_videos(
            input_folder="edit_vid_input",
            output_folder="edit_vid_output",
            repeat_factor=int(repeat_factor)
        )
        return "✅ Video multiplied successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500  

@app.route('/makekbvideofromimages', methods=['POST'])
def make_kb_video():
    try:
        print("Processing request...makekbvideo")

        duration = request.form.get('duration',10)
        output_folder = request.form.get('output_folder', 'edit_vid_output')

        input_folder = request.form.get('input_folder', 'edit_vid_input')
        only_select_images_without_video = request.form.get('only_select_images_without_video', 'false') == 'true'  
        if duration == '':
            duration = 10

        video_size = (1920,1080)
        size = request.form.get('size', 'landscape')
        if size == 'portrait':
            video_size = (1080,1920)

        from make_kb_videos import export_kb_videos
        export_kb_videos(
            input_folder=input_folder,   # folder with images
            out_folder=output_folder,    # where to save KB clips
            per_image=int(duration),            # seconds per image
            output_size=video_size,
            zoom_start=1.0, zoom_end=1.05
        )
        return "✅ Ken Burns videos created successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500

VIDEO_EXTS = {".mp4", ".mov", ".mkv", ".avi", ".webm", ".m4v"}

def _files_with_ext(folder: str, exts: set[str]) -> list[str]:
    p = Path(folder)
    if not p.exists() or not p.is_dir():
        return []
    return [
        f.name for f in p.iterdir()
        if f.is_file() and f.suffix.lower() in exts
    ]

@app.route('/assembleclipstomakevideosong', methods=['POST'])
def assemble_clips_to_make_video_song():    
    try:
        print("Processing request...asseleclipstomakevideosong")
        #return with error message if edit_vid_input is empty

        # video_input_folder = "edit_vid_input"
        video_input_folder = request.form.get('input_folder', 'edit_vid_input')
        copyforcaption = request.form.get('copyforcaption','no')
        # --- Validate inputs up front ---
        if not os.path.isdir(video_input_folder):
            print( "❌ Folder 'edit_vid_input' does not exist.")
            return jsonify({"error": "❌ Folder 'edit_vid_input' does not exist."}), 400

        video_files = _files_with_ext(video_input_folder, VIDEO_EXTS)
        if not video_files:
            print("❌ No video files found in {}.", video_input_folder)
            return jsonify({                
                "error": "❌ No video files found in {}.".format(video_input_folder),
                "hint":  "Add at least one of: .mp4, .mov, .mkv, .avi, .webm, .m4v"
            }), 400


        # --- NEW: Validate min duration (>= 4s) for every input video ---
        MIN_SEC = 2.0
        EPS = 0.02  # tiny tolerance for rounding/metadata quirks

        too_short = []
        for vp in video_files:
            dur = probe_duration(os.path.join(video_input_folder, vp))
            if dur + EPS < MIN_SEC:
                too_short.append({
                    "file": os.path.basename(vp),
                    "duration_sec": round(dur, 3),
                })

        if too_short:
            print(f"❌ One or more input videos are shorter than {MIN_SEC:.0f} seconds:", too_short)
            return jsonify({
                "error": f"❌ One or more input videos are shorter than {MIN_SEC:.0f} seconds.",
                "min_seconds": MIN_SEC,
                "files": too_short,
                "hint": "Replace/extend these clips (or remove them) so every clip is at least 4 seconds."
            }), 400
        

        keep_video_audio = request.form.get('keep_video_audio','no')
        video_volume = float(request.form.get('video_volume',0.3))
        bg_volume = float(request.form.get('bg_volume',1.0))
        add_titles = request.form.get('add_titles', 'no') == 'yes'
        enforce_reencoding = request.form.get('enforce_reencoding', 'no') == 'yes'
        input_folder = request.form.get('input_folder', 'edit_vid_input')

        # If input_folder is not edit_vid_input, copy order.xlsx from edit_vid_input to input_folder
        if input_folder != 'edit_vid_input':
            shutil.copy('edit_vid_input/order.xlsx', os.path.join(input_folder, 'order.xlsx'))

        add_transitions = request.form.get('add_transitions', 'no') == 'yes'
        title_sec = float(request.form.get('title_sec', 2.0))
        transition_sec = float(request.form.get('transition_sec', 0.5))

        # --- NEW: If order.xlsx has a Title column (B), build one output per story group (no bg audio) ---
        group_outputs = assemble_videos_by_titles_if_present(
            video_folder=input_folder,
            output_dir="edit_vid_output",
            fps=30,
            prefer_ffmpeg_concat=True,
            keep_video_audio=(keep_video_audio == 'yes'),
            video_volume=video_volume,
            bg_volume=bg_volume,
            add_titles=add_titles,            
            title_sec=title_sec,
            add_transitions=add_transitions,
            transition_sec=transition_sec,
            enforce_reencoding=enforce_reencoding
        )

        if group_outputs:
            return jsonify({
                "ok": True,
                "mode": "story_groups",
                "outputs": group_outputs
            }), 200


        # ---- Derive output filename from first input video ----
        first_video = sorted(video_files)[0]  # deterministic
        base_name = os.path.basename(first_video)  # e.g. my_clip_01.mp4
        output_video_path = os.path.join("edit_vid_output", base_name)


        assemble_videos(
            video_folder=input_folder,                  # or "edit_vid_output" if you pre-made KB clips
            audio_folder="edit_vid_audio",
            output_path=output_video_path,
            fps=30,
            shuffle=True,                                   # different order each run
            prefer_ffmpeg_concat=True,                       # auto-uses concat if safe; else MoviePy
            keep_video_audio = keep_video_audio == 'yes',
            video_volume = video_volume,
            bg_volume = bg_volume,
            add_titles=add_titles,
            title_sec=title_sec,
            add_transitions=add_transitions,
            transition_sec=transition_sec
        )
        if copyforcaption == 'no':
            return "✅ Video song assembled successfully!", 200
        
        shutil.copy(output_video_path, "composed_video.mp4")

        audio_folder = "edit_vid_audio"
        from pydub import AudioSegment
        # get the first file found in that folder
        audio_file = os.listdir(audio_folder)[0]
        audio_path = os.path.join(audio_folder, audio_file)

        # output path in root folder
        output_path = "audio.wav"

        # convert mp3 → wav (if needed), else just copy
        if audio_file.lower().endswith(".mp3"):
            # convert using pydub (requires ffmpeg installed)
            sound = AudioSegment.from_mp3(audio_path)
            sound.export(output_path, format="wav")
        else:
            # already wav → overwrite if exists
            shutil.copy(audio_path, output_path)

        print(f"✅ Saved audio file as {output_path}")

        return "✅ Video song assembled successfully!", 200
    except Exception as e:
        traceback.print_exc() 
        return f"❌ Error: {str(e)}", 500


# server.py
@app.route('/splitvideotoparts', methods=['POST'])
def splitvideotoparts():
    try:
        print("Processing request...splitvideotoparts")
        from assemble_from_videos import split_video, convert_landscape_to_portrait

        # existing field
        max_duration = str(request.form.get('duration', '178')).strip()

        # new fields
        convert_portrait = (request.form.get('convert_portrait', 'no') == 'yes')
        portrait_size    = (request.form.get('portrait_size', '1080x1920') or '1080x1920')
        portrait_layout  = (request.form.get('portrait_layout', 'crop') or 'crop')
        focus            = (request.form.get('focus', 'center') or 'center')
        keep_audio       = (request.form.get('keep_audio', 'yes') == 'yes')

        if portrait_layout not in {"crop", "pad_equal", "pad_caption_bottom"}:
            portrait_layout = "crop"

        out_dir = BASE_DIR / "edit_vid_output"
        out_dir.mkdir(exist_ok=True)

        uploaded_source = None
        uploaded_video = request.files.get("split_video_file")
        if uploaded_video and uploaded_video.filename:
            ext = os.path.splitext(uploaded_video.filename)[1].lower()
            if ext != ".mp4":
                return "❌ Please upload an MP4 file.", 400
            uploaded_source = out_dir / f"split_source_uploaded_{os.urandom(4).hex()}.mp4"
            uploaded_video.save(str(uploaded_source))
            in_path = str(uploaded_source)
        else:
            # default source your UI mentions
            in_path = "edit_vid_output/output.mp4"

        work_src = in_path

        # Optional pre-pass: crop to 9:16 and scale
        if convert_portrait:
            out_portrait = str(out_dir / "output_portrait.mp4")
            convert_landscape_to_portrait(
                input_path=in_path,
                output_path=out_portrait,
                portrait_size=portrait_size,
                fit_mode=portrait_layout,
                focus=focus,
                keep_audio=keep_audio
            )
            work_src = out_portrait

        # Now split whichever source we decided on
        split_video(
            input_path=work_src,
            max_duration=max_duration
        )

        if uploaded_source and uploaded_source.exists():
            uploaded_source.unlink()

        return "✅ splitvideotoparts completed successfully!", 200
    except Exception as e:
        traceback.print_exc()
        return f"❌ Error: {str(e)}", 500


@app.route('/save_word_timestamps_file', methods=['POST'])
def save_word_timestamps_file():
  data = request.get_json(force=True)  # list of {word,start,end,position,matched}
  os.makedirs('temp', exist_ok=True)
  path = os.path.join('temp', 'word_timestamps.json')
  with open(path, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
  return jsonify({"ok": True, "path": "/temp/word_timestamps.json"})    

@app.route('/save_ass', methods=['POST'])
def save_ass():
    """
    Save a .ass subtitle file to the parent folder of this server.py.
    Body: {"filename": "captions_landscape.ass", "content": "...ass text..."}
    """
    try:
        data = request.get_json(force=True, silent=False)
        if not data:
            return jsonify({"ok": False, "error": "Empty payload"}), 400

        filename = str(data.get("filename", "")).strip()
        content  = str(data.get("content", ""))

        # sanitize filename (letters, numbers, ., _, -) and require .ass
        filename = re.sub(r'[^A-Za-z0-9._-]', '_', filename)
        if not filename or not filename.endswith(".ass"):
            return jsonify({"ok": False, "error": "Bad filename"}), 400

        # parent directory of this file's directory
        parent_dir    = os.path.dirname(os.path.abspath(__file__))

        # Go to sub-folder with name edit_vid_output

        # Path to the subfolder 'edit_vid_output'
        output_dir = os.path.join(parent_dir, "edit_vid_output")

        #parent_dir = os.path.abspath(os.path.join(app_dir, os.pardir))

        if not os.path.isdir(output_dir) or not os.access(output_dir, os.W_OK):
            return jsonify({"ok": False, "error": "Parent folder not writable"}), 500

        target_path = os.path.join(output_dir, filename)
        with open(target_path, "w", encoding="utf-8") as f:
            f.write(content)

        return jsonify({"ok": True, "path": target_path})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/render_captions', methods=['POST'])
def render_captions():

    from wordtimestamps_to_ass_captions import build_ass_from_word_json, _ffmpeg_burn_subs

    """
    Body JSON:
    {
      "input_video": "/absolute/or/relative/path.mp4",
      "orientation": "landscape"|"portrait",
      "style": "cinematic"|"pro_pop"|"drift_up"|"typewriter"|"fade"|"softzoom"|"wordpop"|"glowpulse",
      "min_gap_sec": 0.40,
      "words_per_caption": 5,
      "output": "/path/output_with_captions.mp4" (optional)
    }
    """
    try:
        data = request.get_json(force=True)
        #input_video = data.get('input_video')
        orientation = data.get('orientation', 'landscape')
        style = data.get('style', 'cinematic')
        min_gap = float(data.get('min_gap_sec', 0.40))
        wpc = int(data.get('words_per_caption', 5))
        #output = data.get('output')

        if orientation == 'landscape':
            create_portrait()

        # if not input_video or not os.path.isfile(input_video):
        #     return jsonify({"ok": False, "error": "input_video not found"}), 400

        # Where is your word_timestamps.json?  Adjust if needed.
        # (Uses the same JSON you already serve to the front end.)
        app_dir = os.path.dirname(os.path.abspath(__file__))
        temp_dir = os.path.join(app_dir, 'temp')
        word_json_path = os.path.join(temp_dir, 'word_timestamps.json')
        if not os.path.isfile(word_json_path):
            # fallback: try parent folder
            word_json_path = os.path.abspath(os.path.join(app_dir, os.pardir, 'word_timestamps.json'))
        if not os.path.isfile(word_json_path):
            return jsonify({"ok": False, "error": "word_timestamps.json not found"}), 400

        # Build ASS (both orientations saved; use requested orientation for render)
        parent_dir = os.path.abspath(os.path.join(app_dir, os.pardir))
        out_dir = os.path.join(app_dir, 'edit_vid_output')
        os.makedirs(out_dir, exist_ok=True)

        ass_land = build_ass_from_word_json(word_json_path, 'landscape', style, min_gap, wpc)
        ass_port = build_ass_from_word_json(word_json_path, 'portrait',  style, min_gap, wpc)

        path_land = os.path.join(out_dir, 'captions_landscape.ass')
        path_port = os.path.join(out_dir, 'captions_portrait.ass')
        with open(path_land, 'w', encoding='utf-8') as f: f.write(ass_land)
        with open(path_port, 'w', encoding='utf-8') as f: f.write(ass_port)

        # Choose the ASS we render with
        #ass_path = path_port if orientation == 'portrait' else path_land

        # Output path
        # if not output:
        #     base = os.path.splitext(os.path.basename(input_video))[0]
        #     output = os.path.join(parent_dir, f"{base}_with_captions.mp4")

        # Burn it in
        proc = _ffmpeg_burn_subs("edit_vid_output/out_landscape.mp4", path_land, "edit_vid_output/landscape_with_captions.mp4")
        if proc.returncode != 0:
            return jsonify({"ok": False, "error": "ffmpeg failed", "stderr": proc.stderr}), 500

        proc = _ffmpeg_burn_subs("edit_vid_output/out_portrait.mp4", path_port, "edit_vid_output/portrait_with_captions.mp4")
        if proc.returncode != 0:
            return jsonify({"ok": False, "error": "ffmpeg failed", "stderr": proc.stderr}), 500


        return jsonify({
            "ok": True,
            "ass_landscape": path_land,
            "ass_portrait": path_port
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
def bool_from_form(val: str) -> bool:
    """
    Accepts 'true'/'false', 'on'/'off', '1'/'0', or truthy string.
    """
    if val is None:
        return False
    v = str(val).strip().lower()
    return v in ("true", "on", "1", "yes", "y")

@app.post("/upload")
def uploadVid():
    from pathlib import Path
    from datetime import datetime
    from flask import Flask, render_template, request, jsonify
    from playwright.sync_api import sync_playwright
    import youtube_uploader as yu
    try:
        # Collect inputs from form
        video_path            = (request.form.get("video_path") or "").strip()
        youtube_channel_name  = (request.form.get("youtube_channel_name") or "").strip()
        youtube_playlist_name = (request.form.get("youtube_playlist_name") or "").strip()
        youtube_title         = (request.form.get("youtube_title") or "").strip()
        youtube_description   = (request.form.get("youtube_description") or "").strip()
        youtube_tags          = (request.form.get("youtube_tags") or "").strip()
        made_for_kids         = bool_from_form(request.form.get("made_for_kids"))
        schedule_date_raw     = (request.form.get("schedule_date") or "").strip()
        size                  = (request.form.get("size") or "").strip()

        # Build the dict in the exact shape your upload_video expects
        video_info = {
            "video_path": video_path,                         # uploader will append .mp4 and prepend processed_videos
            "youtube_channel_name": youtube_channel_name,
            "youtube_playlist_name": youtube_playlist_name,
            "youtube_title": youtube_title,
            "youtube_description": youtube_description,
            "youtube_tags": youtube_tags,
            "made_for_kids": made_for_kids,
            "schedule_date": schedule_date_raw or None,       # keep raw; your code formats it for YouTube
            "thumbnail_path": "edit_vid_thumbnail/thumbnail.png",
            "size": size
        }

        # Basic validation
        missing = [k for k in ("video_path","youtube_channel_name","youtube_title") if not video_info[k]]
        if missing:
            return jsonify({"ok": False, "error": f"Missing required fields: {', '.join(missing)}"}), 400

        # Launch persistent Chrome with your existing settings
        with sync_playwright() as p:
            browser = p.chromium.launch_persistent_context(
                yu.PROFILE_DIR,
                headless=False,
                executable_path=yu.CHROME_EXECUTABLE,
                args=["--disable-blink-features=AutomationControlled", "--start-maximized"],
            )
            page = browser.new_page()
            page.add_init_script("""Object.defineProperty(navigator, 'webdriver', {get: () => undefined})""")

            # Do the upload (reuses your function)
            video_url = yu.upload_video(page, video_info)

            browser.close()

        return jsonify({"ok": True, "video_url": video_url})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    
def clear_folder(folder_path, extensions=None):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    for file in os.listdir(folder_path):
        full_path = os.path.join(folder_path, file)
        if os.path.isfile(full_path):
            if not extensions or file.lower().endswith(extensions):
                os.remove(full_path)

# def make_segment_id(scene_index: int, segment_index: int, kind: str) -> str:
#     return f"sc{scene_index}-seg{segment_index}-{kind}"

# Replace the current make_segment_id() with this
def make_segment_id(
    scene_index: int,
    kind: str,
    animation_target: str = "",
    logical_index: int | None = None,
) -> str:
    safe_kind = re.sub(r"[^a-zA-Z0-9_]+", "_", (kind or "").strip().lower()).strip("_") or "segment"
    safe_target = re.sub(r"[^a-zA-Z0-9_]+", "_", (animation_target or "").strip().lower()).strip("_")

    if safe_kind == "title":
        return f"sc{scene_index}-title"
    if safe_kind == "subtitle":
        return f"sc{scene_index}-subtitle"
    if safe_kind == "bullet" and safe_target:
        return f"sc{scene_index}-{safe_target}"
    if logical_index is not None:
        return f"sc{scene_index}-{safe_kind}{logical_index}"
    return f"sc{scene_index}-{safe_kind}"

def clear_old_pages():
    """Delete any existing generated page images."""
    for pattern in ("*.png", "*.jpg", "*.jpeg"):
        for f in glob.glob(os.path.join(PAGES_FOLDER, pattern)):
            try:
                os.remove(f)
            except OSError:
                pass


def pdf_to_images(pdf_path):
    """
    Convert a PDF into PNG pages under static/pages/.
    Returns list of static URLs for those pages.
    """
    clear_old_pages()

    pages = convert_from_path(pdf_path, dpi=120, poppler_path=POPPLER_PATH)

    urls = []
    for i, page in enumerate(pages, start=1):
        filename = f"page_{i:03}.png"
        full_path = os.path.join(PAGES_FOLDER, filename)
        page.save(full_path, "PNG")

        # URL for browser
        urls.append(url_for("static", filename=f"pages/{filename}", _external=False))

    return urls

@app.route("/pdf_to_flipping_book")
def pdf_to_flipping_book():
    return render_template("pdf_to_flipping_book.html")

@app.route("/upload_pdf", methods=["POST"])
def upload_pdf():
    if "pdf" not in request.files:
        return jsonify({"success": False, "error": "No file part 'pdf' found."}), 400

    pdf_file = request.files["pdf"]
    if pdf_file.filename == "":
        return jsonify({"success": False, "error": "No file selected."}), 400

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    pdf_path = os.path.join(UPLOAD_FOLDER, "current.pdf")
    pdf_file.save(pdf_path)

    try:
        page_urls = pdf_to_images(pdf_path)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

    if not page_urls:
        return jsonify({"success": False, "error": "No pages generated."}), 500

    return jsonify({"success": True, "pages": page_urls})

####################################################
# START: HTML to video maker enhancements/changes
####################################################

# from __future__ import annotations

from dataclasses import asdict, dataclass, field
from typing import Any, Literal
from http_client import http_get
from bs4 import BeautifulSoup, NavigableString, Tag
from PIL import Image, ImageDraw, ImageFont
import hashlib


BASE_DIR = Path(__file__).resolve().parent
VID_DATA_DIR = OUT_DIR / "video_composer_data"
PROJECTS_DIR = VID_DATA_DIR / "projects"
PROJECTS_DIR.mkdir(parents=True, exist_ok=True)


COMPOSER_DATA_DIR = OUT_DIR / "video_composer_data"
COMPOSER_PROJECTS_DIR = COMPOSER_DATA_DIR / "projects"
COMPOSER_RENDER_DIR = OUT_DIR / "composer_renders"
COMPOSER_SPEECH_DIR = OUT_DIR / "composer_speech"

COMPOSER_PROJECTS_DIR.mkdir(parents=True, exist_ok=True)
COMPOSER_RENDER_DIR.mkdir(parents=True, exist_ok=True)
COMPOSER_SPEECH_DIR.mkdir(parents=True, exist_ok=True)

from datetime import datetime

COMPOSER_PROJECT_ROOT = OUT_DIR / "composer_projects-files"
COMPOSER_PROJECT_ROOT.mkdir(parents=True, exist_ok=True)


def make_audio_cache_key(
    speech_text: str,
    *,
    voice: str,
    language: str,
    engine: str,
    gender: str,
) -> str:
    payload = {
        "speech_text": compact_ws(speech_text),
        "voice": voice,
        "language": language,
        "engine": engine,
        "gender": gender,
    }
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=False)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()

def composer_project_dir(project_id: str) -> Path:
    return COMPOSER_PROJECT_ROOT / project_id

def composer_project_assets_dir(project_id: str) -> Path:
    p = composer_project_dir(project_id) / "assets"
    p.mkdir(parents=True, exist_ok=True)
    return p

SceneType = Literal[
    "title",
    "section_header",
    "bullet_points",
    "paragraph",
    "quote",
    "image_focus",
    "code",
    "table",
    "recap",
]


@dataclass
class ThemePreset:
    id: str
    name: str
    description: str
    font_heading: str
    font_body: str
    palette: dict[str, str]
    motion_style: str
    layout_density: str
    caption_style: str


@dataclass
class MediaAsset:
    id: str
    kind: Literal["image", "video", "audio"]
    url: str | None = None
    local_path: str | None = None
    title: str | None = None
    source: str = "user"
    fit_mode: Literal["cover", "contain", "smart_crop"] = "cover"
    focal_point: dict[str, float] | None = None


@dataclass
class NarrationSettings:
    language: str = "en"
    voice: str = "alloy"
    rate: float = 1.0
    pitch: float = 1.0
    engine: str = "tts"
    pause_after_heading_ms: int = 450
    pause_between_bullets_ms: int = 250


@dataclass
class CaptionSettings:
    enabled: bool = True
    mode: Literal["none", "sentence", "phrase", "word_highlight"] = "phrase"
    max_words_per_line: int = 6
    position: Literal["bottom", "center", "top"] = "bottom"


@dataclass
class SceneTiming:
    estimated_duration_sec: float = 4.0
    manual_duration_sec: float | None = None
    reveal_mode: Literal["all_at_once", "sequential_bullets", "sentence_step"] = "all_at_once"


@dataclass
class SceneLayout:
    template: str = "title_top_content_center"
    text_align: Literal["left", "center", "right"] = "left"
    overlay_position: Literal[
        "left",
        "right",
        "center",
        "bottom_card",
        "top_card",
    ] = "center"
    background_dim: float = 0.35
    background_blur: int = 0


@dataclass
class Scene:
    id: str
    scene_type: SceneType
    title: str = ""
    subtitle: str = ""
    bullets: list[str] = field(default_factory=list)
    body_text: str = ""
    narration_text: str = ""
    on_screen_text: list[str] = field(default_factory=list)
    media_asset_ids: list[str] = field(default_factory=list)
    layout: SceneLayout = field(default_factory=SceneLayout)
    timing: SceneTiming = field(default_factory=SceneTiming)
    notes: list[str] = field(default_factory=list)
    source_refs: list[str] = field(default_factory=list)


@dataclass
class ProjectSettings:
    aspect_ratio: Literal["16:9", "9:16", "1:1"] = "16:9"
    quality_preset: Literal["draft", "standard", "high"] = "standard"
    auto_split_long_sections: bool = True
    max_bullets_per_scene: int = 5
    max_chars_per_scene: int = 320
    enable_smart_suggestions: bool = True


@dataclass
class VideoProject:
    id: str
    title: str
    source_type: Literal["html", "url", "upload"]
    source_value: str
    theme_id: str
    project_settings: ProjectSettings = field(default_factory=ProjectSettings)
    narration: NarrationSettings = field(default_factory=NarrationSettings)
    captions: CaptionSettings = field(default_factory=CaptionSettings)
    scenes: list[Scene] = field(default_factory=list)
    assets: list[MediaAsset] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)

@dataclass
class SpeechSegment:
    id: str
    kind: Literal["title", "subtitle", "bullet", "narration", "quote"]
    display_text: str
    speech_text: str
    show_on_screen: bool = True
    animation: str = "fade"
    animation_target: str = "body"
    order: int = 0


@dataclass
class Scene:
    id: str
    scene_type: SceneType
    title: str = ""
    subtitle: str = ""
    bullets: list[str] = field(default_factory=list)
    body_text: str = ""
    narration_text: str = ""
    on_screen_text: list[str] = field(default_factory=list)
    media_asset_ids: list[str] = field(default_factory=list)
    layout: SceneLayout = field(default_factory=SceneLayout)
    timing: SceneTiming = field(default_factory=SceneTiming)
    notes: list[str] = field(default_factory=list)
    source_refs: list[str] = field(default_factory=list)
    speech_segments: list[SpeechSegment] = field(default_factory=list)

THEMES: dict[str, ThemePreset] = {
    "corporate-clean": ThemePreset(
        id="corporate-clean",
        name="Corporate Clean",
        description="Minimal business presentation style with calm motion.",
        font_heading="Inter",
        font_body="Inter",
        palette={"bg": "#0F172A", "card": "#111827", "text": "#F8FAFC", "accent": "#38BDF8"},
        motion_style="subtle",
        layout_density="comfortable",
        caption_style="clean-bottom",
    ),
    "youtube-explainer": ThemePreset(
        id="youtube-explainer",
        name="YouTube Explainer",
        description="High-contrast explainer look for engagement-focused videos.",
        font_heading="Poppins",
        font_body="Inter",
        palette={"bg": "#111111", "card": "#1F2937", "text": "#FFFFFF", "accent": "#F59E0B"},
        motion_style="dynamic",
        layout_density="medium",
        caption_style="bold-highlight",
    ),
    "premium-dark": ThemePreset(
        id="premium-dark",
        name="Premium Dark",
        description="Cinematic dark layout with elegant cards and slower transitions.",
        font_heading="Playfair Display",
        font_body="Inter",
        palette={"bg": "#030712", "card": "#111827", "text": "#E5E7EB", "accent": "#A78BFA"},
        motion_style="cinematic",
        layout_density="airy",
        caption_style="soft-card",
    ),
}

def _get_existing_segment_display_map(scene: dict) -> dict:
    speech_segments = scene.get("speech_segments", []) or []

    title_display = ""
    bullet_displays = {}

    for seg in speech_segments:
        kind = (seg.get("kind") or "").strip().lower()
        display_text = compact_ws(seg.get("display_text", ""))
        animation_target = (seg.get("animation_target") or "").strip()

        if kind == "title" and display_text and not title_display:
            title_display = display_text

        if kind == "bullet" and animation_target:
            bullet_displays[animation_target] = display_text

    return {
        "title": title_display,
        "bullets": bullet_displays,
    }

def _rebuild_scene_speech_segments(scene: dict, scene_index: int):
    """Rebuild speech segments from edited UI text.

    Important behavior:
    - On initial HTML import, short narration can still be merged into display text.
    - On later preview-editor narration changes, preserve existing title/bullet display_text
      so narration updates do not overwrite what is shown on screen.
    """
    from dataclasses import asdict

    title = compact_ws(scene.get("title", ""))
    bullets = [compact_ws(x) for x in (scene.get("bullets", []) or []) if compact_ws(x)]
    narration_text = scene.get("narration_text", "") or ""

    existing_map = _get_existing_segment_display_map(scene)
    existing_title_display = existing_map.get("title", "")
    existing_bullet_displays = existing_map.get("bullets", {}) or {}

    parsed = _parse_tagged_narration(narration_text)

    def preserve_display_text(new_segments: list[dict]) -> list[dict]:
        for seg in new_segments:
            kind = (seg.get("kind") or "").strip().lower()
            if kind == "title":
                if existing_title_display:
                    seg["display_text"] = existing_title_display
            elif kind == "bullet":
                target = (seg.get("animation_target") or "").strip()
                preserved = existing_bullet_displays.get(target, "")
                if preserved:
                    seg["display_text"] = preserved
        return new_segments

    if not parsed["has_tags"]:
        parts = _split_sentences(narration_text)
        new_segments = build_speech_segments(
            title=title,
            bullets=bullets,
            narration_parts=parts,
            scene_index=scene_index
        )
        scene["speech_segments"] = preserve_display_text([asdict(s) for s in new_segments])
        return


    segments = []
    segment_index = 0
    narration_index = 0

    def add_segment(kind, display_text, speech_text, show_on_screen, animation, animation_target, stable_id=None):
        nonlocal segment_index, narration_index
        speech_text = compact_ws(speech_text)
        display_text = compact_ws(display_text)

        if not speech_text and not display_text:
            return

        segment_index += 1

        if stable_id:
            seg_id = stable_id
        elif kind == "title":
            seg_id = make_segment_id(scene_index, "title", "title")
        elif kind == "bullet":
            seg_id = make_segment_id(scene_index, "bullet", animation_target)
        else:
            narration_index += 1
            seg_id = make_segment_id(scene_index, kind, logical_index=narration_index)

        segments.append({
            "id": seg_id,
            "kind": kind,
            "display_text": display_text,
            "speech_text": speech_text,
            "show_on_screen": show_on_screen,
            "animation": animation,
            "animation_target": animation_target,
            "order": segment_index,
        })

    # segments = []
    # segment_index = 0

    # def add_segment(kind, display_text, speech_text, show_on_screen, animation, animation_target):
    #     nonlocal segment_index
    #     speech_text = compact_ws(speech_text)
    #     display_text = compact_ws(display_text)

    #     if not speech_text and not display_text:
    #         return

    #     segment_index += 1
    #     seg_id_suffix = kind if kind != "bullet" else animation_target
    #     segments.append({
    #         "id": make_segment_id(scene_index, segment_index, seg_id_suffix),
    #         "kind": kind,
    #         "display_text": display_text,
    #         "speech_text": speech_text,
    #         "show_on_screen": show_on_screen,
    #         "animation": animation,
    #         "animation_target": animation_target,
    #         "order": segment_index,
    #     })

    # sequence = parsed.get("sequence") or []
    # for item in sequence:
    #     bucket = item.get("bucket")
    #     idx = item.get("index")
    #     parts = item.get("parts") or []

    #     if bucket == "title":
    #         title_speech = " ".join(parts).strip() or title
    #         title_display = existing_title_display or title
    #         add_segment("title", title_display, title_speech, True, "titleReveal", "title")

    #     elif bucket == "bullet":
    #         bullet_text = bullets[idx - 1] if idx and 0 < idx <= len(bullets) else ""
    #         bullet_speech = " ".join(parts).strip() or bullet_text
    #         target = f"bullet_{idx}"
    #         bullet_display = existing_bullet_displays.get(target, "")
    #         if not bullet_display:
    #             bullet_display = merged_short_display_text(bullet_text, bullet_speech, max_words=10)
    #         add_segment("bullet", bullet_display, bullet_speech, True, "bulletReveal", target)

    #     else:
    #         for part in parts:
    #             add_segment("narration", "", part, False, "none", "none")

    # if title and not any(seg.get("kind") == "title" for seg in segments):
    #     add_segment("title", existing_title_display or title, title, True, "titleReveal", "title")

    sequence = parsed.get("sequence") or []
    has_title_bucket = any((item.get("bucket") == "title") for item in sequence)

    for item in sequence:
        bucket = item.get("bucket")
        idx = item.get("index")
        parts = item.get("parts") or []

        if bucket == "title":
            # Blank [TITLE] should mean: show title, do NOT speak it
            title_speech = " ".join(parts).strip()
            title_display = existing_title_display or title
            add_segment("title", title_display, title_speech, True, "titleReveal", "title")

        elif bucket == "bullet":
            bullet_text = bullets[idx - 1] if idx and 0 < idx <= len(bullets) else ""
            bullet_speech = " ".join(parts).strip() or bullet_text
            target = f"bullet_{idx}"
            bullet_display = existing_bullet_displays.get(target, "")
            if not bullet_display:
                bullet_display = merged_short_display_text(bullet_text, bullet_speech, max_words=10)
            add_segment("bullet", bullet_display, bullet_speech, True, "bulletReveal", target)

        else:
            for part in parts:
                add_segment("narration", "", part, False, "none", "none")

    # Only add spoken title fallback when there was NO [TITLE] tag at all
    if title and not has_title_bucket and not any(seg.get("kind") == "title" for seg in segments):
        add_segment("title", existing_title_display or title, title, True, "titleReveal", "title")
    scene["speech_segments"] = segments

# def _rebuild_scene_speech_segments(scene: dict, scene_index: int):
#     """Rebuild speech segments from edited UI text, supporting tagged narration in the same order as written."""
#     from dataclasses import asdict

#     title = compact_ws(scene.get("title", ""))
#     bullets = [compact_ws(x) for x in (scene.get("bullets", []) or []) if compact_ws(x)]
#     narration_text = scene.get("narration_text", "") or ""

#     parsed = _parse_tagged_narration(narration_text)

#     if not parsed["has_tags"]:
#         parts = _split_sentences(narration_text)
#         new_segments = build_speech_segments(
#             title=title,
#             bullets=bullets,
#             narration_parts=parts,
#             scene_index=scene_index
#         )
#         scene["speech_segments"] = [asdict(s) for s in new_segments]
#         return

#     segments = []
#     segment_index = 0

#     def add_segment(kind, display_text, speech_text, show_on_screen, animation, animation_target):
#         nonlocal segment_index
#         speech_text = compact_ws(speech_text)
#         display_text = compact_ws(display_text)

#         if not speech_text and not display_text:
#             return

#         segment_index += 1
#         seg_id_suffix = kind if kind != "bullet" else animation_target
#         segments.append({
#             "id": make_segment_id(scene_index, segment_index, seg_id_suffix),
#             "kind": kind,
#             "display_text": display_text,
#             "speech_text": speech_text,
#             "show_on_screen": show_on_screen,
#             "animation": animation,
#             "animation_target": animation_target,
#             "order": segment_index,
#         })

#     sequence = parsed.get("sequence") or []
#     for item in sequence:
#         bucket = item.get("bucket")
#         idx = item.get("index")
#         parts = item.get("parts") or []

#         if bucket == "title":
#             title_speech = " ".join(parts).strip() or title
#             add_segment("title", title, title_speech, True, "titleReveal", "title")
#         elif bucket == "bullet":
#             bullet_text = bullets[idx - 1] if idx and 0 < idx <= len(bullets) else ""
#             bullet_speech = " ".join(parts).strip() or bullet_text
#             bullet_display = merged_short_display_text(bullet_text, bullet_speech, max_words=10)
#             add_segment("bullet", bullet_display, bullet_speech, True, "bulletReveal", f"bullet_{idx}")
#         else:
#             for part in parts:
#                 add_segment("narration", "", part, False, "none", "none")

#     if title and not any(seg.get("kind") == "title" for seg in segments):
#         add_segment("title", title, title, True, "titleReveal", "title")

#     scene["speech_segments"] = segments

# def _rebuild_scene_speech_segments(scene: dict, scene_index: int):
#     """Helper to rebuild speech segments if the user edited the text in the UI."""
#     import re
#     from dataclasses import asdict

#     title = scene.get("title", "")
#     bullets = scene.get("bullets", [])
#     narration_text = scene.get("narration_text", "")

#     # Split narration safely into sentences
#     parts = [p.strip() for p in re.split(r'(?<=[.!?])\s+', narration_text) if p.strip()]
#     if not parts and narration_text.strip():
#         parts = [narration_text.strip()]

#     new_segments = build_speech_segments(
#         title=title,
#         bullets=bullets,
#         narration_parts=parts,
#         scene_index=scene_index
#     )
#     scene["speech_segments"] = [asdict(s) for s in new_segments]

@app.post("/api/projects/<project_id>/paste-background")
def paste_background_for_project(project_id: str):
    try:
        project = load_project(project_id)
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404

    file = request.files.get("file")
    scene_id = (request.form.get("scene_id") or "").strip()

    if not file:
        return jsonify({"ok": False, "error": "Missing file"}), 400
    if not scene_id:
        return jsonify({"ok": False, "error": "Missing scene_id"}), 400

    ext = Path(file.filename or "clipboard.png").suffix.lower() or ".png"
    if ext not in {".png", ".jpg", ".jpeg", ".webp"}:
        ext = ".png"

    scene_index = None
    scene_ref = None
    for idx, scene in enumerate(project.get("scenes", []), start=1):
        if scene.get("id") == scene_id:
            scene_index = idx
            scene_ref = scene
            break

    if scene_index is None or scene_ref is None:
        return jsonify({"ok": False, "error": "Scene not found"}), 404

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"sc{scene_index}-bg-{timestamp}{ext}"
    save_path = composer_project_assets_dir(project_id) / filename
    file.save(save_path)

    asset_id = f"sc{scene_index}-bg"
    asset_url = f"/out/composer-project-files/{project_id}/assets/{filename}"

    project_assets = project.get("assets", []) or []

    # remove older scene background asset with same id if present
    project_assets = [a for a in project_assets if a.get("id") != asset_id]

    new_asset = {
        "id": asset_id,
        "kind": "image",
        "url": asset_url,
        "local_path": str(save_path),
        "title": f"Scene {scene_index} background",
        "source": "clipboard",
        "fit_mode": "cover",
        "focal_point": None,
    }
    project_assets.append(new_asset)
    project["assets"] = project_assets

    scene_ref["media_asset_ids"] = [asset_id]

    project_path(project_id).write_text(
        json.dumps(project, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )

    return jsonify({
        "ok": True,
        "asset": new_asset,
        "project": project,
    })

# @app.get("/html-video-composer/player/<project_id>")
# def html_video_composer_player(project_id: str):
#     return render_template("html_video_composer_player.html", project_id=project_id)



def _split_sentences(text: str) -> list[str]:
    text = compact_ws(text)
    if not text:
        return []
    parts = [compact_ws(p) for p in re.split(r'(?<=[.!?])\s+', text) if compact_ws(p)]
    return parts or [text]

def build_tagged_narration_text(
    *,
    title: str = "",
    bullets: list[str] | None = None,
    narration_parts: list[str] | None = None,
    body_text: str = "",
) -> str:
    title = compact_ws(title)
    bullets = [compact_ws(x) for x in (bullets or []) if compact_ws(x)]
    narration_parts = [compact_ws(x) for x in (narration_parts or []) if compact_ws(x)]
    body_text = compact_ws(body_text)

    lines: list[str] = []

    # Add title block
    if title:
        lines.append("[TITLE]")
        lines.append(title)
        lines.append("")

    # Bullet scenes
    if bullets:
        for i, bullet in enumerate(bullets, start=1):
            spoken = narration_parts[i - 1] if i - 1 < len(narration_parts) else bullet
            lines.append(f"[B{i}]")
            lines.append(spoken or bullet)
            lines.append("")

        extra_parts = narration_parts[len(bullets):]
        if extra_parts:
            lines.append("[OUTRO]")
            lines.append(" ".join(extra_parts))
            lines.append("")
        return "\n".join(lines).strip()

    # Non-bullet scenes
    main_text = " ".join(narration_parts) if narration_parts else body_text
    if main_text:
        lines.append("[INTRO]" if title else "[TITLE]")
        lines.append(main_text)

    return "\n".join(lines).strip()
def _parse_tagged_narration(narration_text: str) -> dict:
    text = (narration_text or "").replace("\r\n", "\n").strip()

    parsed = {
        "intro": [],
        "title": [],
        "outro": [],
        "bullets": {},
        "has_tags": False,
        "sequence": [],
    }

    if not text:
        return parsed

    current_bucket = ("intro", None)
    current_lines = []
    current_bucket_explicit = False
    tag_re = re.compile(r'^\[(TITLE|INTRO|OUTRO|B(\d+))\]\s*(.*)$', re.IGNORECASE)

    def flush_bucket():
        nonlocal current_lines, current_bucket_explicit
        bucket, idx = current_bucket

        content = compact_ws(" ".join(current_lines))
        current_lines = []

        # keep explicitly tagged buckets in sequence even when blank
        if not current_bucket_explicit and not content:
            return

        parts = _split_sentences(content) if content else []

        if parts:
            if bucket == "bullet":
                parsed["bullets"].setdefault(idx, []).extend(parts)
            else:
                parsed[bucket].extend(parts)

        parsed["sequence"].append({
            "bucket": bucket,
            "index": idx,
            "parts": parts,
        })

        current_bucket_explicit = False

    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue

        m = tag_re.match(line)
        if m:
            parsed["has_tags"] = True
            flush_bucket()
            tag_name = m.group(1).upper()
            rest = (m.group(3) or "").strip()

            if tag_name == "TITLE":
                current_bucket = ("title", None)
            elif tag_name == "INTRO":
                current_bucket = ("intro", None)
            elif tag_name == "OUTRO":
                current_bucket = ("outro", None)
            else:
                current_bucket = ("bullet", int(m.group(2)))

            current_bucket_explicit = True

            if rest:
                current_lines.append(rest)
            continue

        current_lines.append(line)

    flush_bucket()
    return parsed

# def _parse_tagged_narration(narration_text: str) -> dict:
#     text = (narration_text or "").replace("\r\n", "\n").strip()

#     parsed = {
#         "intro": [],
#         "title": [],
#         "outro": [],
#         "bullets": {},
#         "has_tags": False,
#         "sequence": [],
#     }

#     if not text:
#         return parsed

#     current_bucket = ("intro", None)
#     current_lines = []
#     tag_re = re.compile(r'^\[(TITLE|INTRO|OUTRO|B(\d+))\]\s*(.*)$', re.IGNORECASE)

#     def flush_bucket():
#         nonlocal current_lines
#         bucket, idx = current_bucket
#         if not current_lines:
#             return
#         content = compact_ws(" ".join(current_lines))
#         current_lines = []
#         if not content:
#             return
#         parts = _split_sentences(content)
#         if bucket == "bullet":
#             parsed["bullets"].setdefault(idx, []).extend(parts)
#         else:
#             parsed[bucket].extend(parts)
#         parsed["sequence"].append({
#             "bucket": bucket,
#             "index": idx,
#             "parts": parts,
#         })

#     for raw_line in text.split("\n"):
#         line = raw_line.strip()
#         if not line:
#             continue

#         m = tag_re.match(line)
#         if m:
#             parsed["has_tags"] = True
#             flush_bucket()
#             tag_name = m.group(1).upper()
#             rest = (m.group(3) or "").strip()

#             if tag_name == "TITLE":
#                 current_bucket = ("title", None)
#             elif tag_name == "INTRO":
#                 current_bucket = ("intro", None)
#             elif tag_name == "OUTRO":
#                 current_bucket = ("outro", None)
#             else:
#                 current_bucket = ("bullet", int(m.group(2)))

#             if rest:
#                 current_lines.append(rest)
#             continue

#         current_lines.append(line)

#     flush_bucket()
#     return parsed
def slugify(text: str) -> str:
    text = (text or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "untitled"

def _load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "arial.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            pass
    return ImageFont.load_default()

def render_composer_frame(
    *,
    out_path: Path,
    scene: dict,
    width: int = 1920,
    height: int = 1080,
    visible_bullets: int = 0,
) -> None:
    img = Image.new("RGB", (width, height), "#071225")
    draw = ImageDraw.Draw(img)

    # Background asset if present and image
    bg_applied = False
    media_ids = scene.get("media_asset_ids", []) or []
    project_assets = scene.get("_project_assets", []) or []
    if media_ids and project_assets:
        asset_map = {a.get("id"): a for a in project_assets}
        first_asset = asset_map.get(media_ids[0])
        if first_asset and first_asset.get("kind") == "image":
            try:
                bg_path = resolve_asset_to_local_file(first_asset)
                print("DEBUG bg asset:", first_asset)
                print("DEBUG resolved bg path:", bg_path)

                if bg_path and bg_path.exists():
                    bg = Image.open(bg_path).convert("RGB")
                    bg = bg.resize((width, height))
                    img.paste(bg, (0, 0))
                    overlay = Image.new("RGBA", (width, height), (3, 10, 24, 130))
                    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
                    draw = ImageDraw.Draw(img)
                    bg_applied = True
                else:
                    print("DEBUG background path missing or unresolved")
            except Exception as e:
                print("DEBUG background load failed:", e)

    if not bg_applied:
        # subtle gradient fallback
        for y in range(height):
            ratio = y / max(1, height - 1)
            r = int(7 * (1 - ratio) + 14 * ratio)
            g = int(18 * (1 - ratio) + 28 * ratio)
            b = int(37 * (1 - ratio) + 58 * ratio)
            draw.line([(0, y), (width, y)], fill=(r, g, b))

    # Card
    # card_left = int(width * 0.16)
    # card_top = int(height * 0.14)
    # card_right = int(width * 0.84)
    # card_bottom = int(height * 0.72)

    title = scene.get("title", "")
    bullets = scene.get("bullets", [])[:visible_bullets]

    # Transparent overlay layers need RGBA
    base = img.convert("RGBA")
    overlay_layer = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    odraw = ImageDraw.Draw(overlay_layer)

    # Main heading glass card - closer to preview
    title_left = int(width * 0.16)
    title_top = int(height * 0.14)
    title_right = int(width * 0.84)
    title_bottom = int(height * 0.34)

    # Bullet glass card - separate box like preview feel
    bul_left = int(width * 0.16)
    bul_top = int(height * 0.33)
    bul_right = int(width * 0.84)
    bul_bottom = int(height * 0.70)

    # soft shadow
    odraw.rounded_rectangle(
        [title_left + 8, title_top + 8, title_right + 8, title_bottom + 8],
        radius=28,
        fill=(0, 0, 0, 70),
    )
    odraw.rounded_rectangle(
        [bul_left + 8, bul_top + 8, bul_right + 8, bul_bottom + 8],
        radius=28,
        fill=(0, 0, 0, 60),
    )

    # glass panels
    odraw.rounded_rectangle(
        [title_left, title_top, title_right, title_bottom],
        radius=28,
        fill=(15, 23, 42, 118),   # semi-transparent
        outline=(255, 255, 255, 26),
        width=2,
    )
    odraw.rounded_rectangle(
        [bul_left, bul_top, bul_right, bul_bottom],
        radius=28,
        fill=(15, 23, 42, 82),    # more transparent than title card
        outline=(255, 255, 255, 18),
        width=1,
    )

    # subtle top highlight line
    odraw.rounded_rectangle(
        [title_left + 28, title_top + 22, title_left + 150, title_top + 28],
        radius=3,
        fill=(56, 189, 248, 180),
    )

    # merge overlay with background
    img = Image.alpha_composite(base, overlay_layer).convert("RGBA")
    draw = ImageDraw.Draw(img)

    title_font = _load_font(54)
    bullet_font = _load_font(28)

    # Title text
    tx = title_left + 34
    ty = title_top + 28
    draw.multiline_text(
        (tx, ty),
        title,
        font=title_font,
        fill=(245, 248, 255, 255),
        spacing=10,
    )

    # Bullets
    btx = bul_left + 34
    bty = bul_top + 26

    for bullet in bullets:
        # bullet row background
        row_h = 44
        draw.rounded_rectangle(
            [btx - 8, bty - 6, bul_right - 34, bty + row_h],
            radius=10,
            fill=(255, 255, 255, 10),
        )

        draw.text(
            (btx, bty),
            "•",
            font=bullet_font,
            fill=(255, 255, 255, 230),
        )
        draw.multiline_text(
            (btx + 24, bty),
            bullet,
            font=bullet_font,
            fill=(230, 236, 245, 245),
            spacing=6,
        )
        bty += 58

    img = img.convert("RGB")

    img.save(out_path)


def render_scene_segments_to_video(
    *,
    project_id: str,
    scene: dict,
    project_assets: list[dict],
    render_dir: Path,
    out_res: str = "1920x1080",
) -> Path:
    width, height = [int(x) for x in out_res.split("x")]
    
    # Use fallback to "id" if "scene_id" isn't present
    scene_dir = render_dir / _safe_filename(scene.get("scene_id", scene.get("id", "scene")))
    scene_dir.mkdir(parents=True, exist_ok=True)

    speech_plan = scene.get("speech_plan", {}) or {}
    segments = speech_plan.get("segments", []) or []

    # FALLBACK: If a scene has no text/segments, hold the frame for the estimated duration
    if not segments:
        fallback_dur = scene.get("timing", {}).get("manual_duration_sec")
        if not fallback_dur:
            fallback_dur = scene.get("timing", {}).get("estimated_duration_sec", 4.0)
        segments = [{
            "id": f"dummy_{uuid.uuid4().hex[:8]}",
            "kind": "narration",
            "duration_ms": int(float(fallback_dur) * 1000)
        }]

    segment_mp4s = []
    visible_bullets = 0

    for idx, seg in enumerate(segments, start=1):
        seg_kind = seg.get("kind", "narration")
        seg_id = seg.get("id") or f"seg_{idx}"
        audio_path = COMPOSER_SPEECH_DIR / project_id / f"{seg_id}.mp3"

        if seg_kind == "bullet":
            visible_bullets += 1

        frame_path = scene_dir / f"{idx:03d}_{seg_id}.png"
        temp_scene = dict(scene)
        temp_scene["_project_assets"] = project_assets
        render_composer_frame(
            out_path=frame_path,
            scene=temp_scene,
            width=width,
            height=height,
            visible_bullets=visible_bullets,
        )

        clip_path = scene_dir / f"{idx:03d}_{seg_id}.mp4"
        duration_sec = max(0.3, float(seg.get("duration_ms", 1000)) / 1000.0)

        # FIX: Explicitly enforce 44.1kHz, Stereo, and square pixels (setsar=1) across ALL clips
        if audio_path.exists() and audio_path.stat().st_size > 0:
            _run([
                "ffmpeg", "-y",
                "-loop", "1",
                "-i", str(frame_path),
                "-i", str(audio_path),
                "-c:v", "libx264",
                "-tune", "stillimage",
                "-pix_fmt", "yuv420p",
                "-c:a", "aac",
                "-ac", "2",           # Force stereo
                "-ar", "44100",       # Force 44.1kHz
                "-af", "aresample=async=1:min_hard_comp=0.100:first_pts=0",
                "-shortest",
                "-r", "30",
                "-vf", f"scale={width}:{height},setsar=1", # Force square pixels
                str(clip_path),
            ])
        else:
            _run([
                "ffmpeg", "-y",
                "-loop", "1",
                "-i", str(frame_path),
                "-f", "lavfi",
                "-i", "anullsrc=r=44100:cl=stereo",
                "-t", str(duration_sec),
                "-c:v", "libx264",
                "-tune", "stillimage",
                "-pix_fmt", "yuv420p",
                "-c:a", "aac",
                "-ac", "2",
                "-ar", "44100",
                "-r", "30",
                "-vf", f"scale={width}:{height},setsar=1",
                str(clip_path),
            ])

        segment_mp4s.append(clip_path)

    concat_list = scene_dir / "segments.txt"
    concat_list.write_text(
        "\n".join([f"file '{p.as_posix()}'" for p in segment_mp4s]),
        encoding="utf-8"
    )

    scene_video = scene_dir / "scene.mp4"
    _run([
        "ffmpeg", "-y",
        "-f", "concat",
        "-safe", "0",
        "-i", str(concat_list),
        "-c:v", "libx264",
        "-pix_fmt", "yuv420p",
        "-c:a", "aac",
        "-ar", "44100",
        "-ac", "2",
        str(scene_video),
    ])

    return scene_video

# def render_scene_segments_to_video(
#     *,
#     project_id: str,
#     scene: dict,
#     project_assets: list[dict],
#     render_dir: Path,
#     out_res: str = "1920x1080",
# ) -> Path:
#     width, height = [int(x) for x in out_res.split("x")]
#     scene_dir = render_dir / _safe_filename(scene.get("scene_id", "scene"))
#     scene_dir.mkdir(parents=True, exist_ok=True)

#     speech_plan = scene.get("speech_plan", {}) or {}
#     segments = speech_plan.get("segments", []) or []
#     title = scene.get("title", "")
#     bullets = scene.get("bullets", []) or []

#     segment_mp4s = []

#     visible_bullets = 0
#     for idx, seg in enumerate(segments, start=1):
#         seg_kind = seg.get("kind", "narration")
#         seg_id = seg.get("id") or f"seg_{idx}"
#         audio_path = COMPOSER_SPEECH_DIR / project_id / f"{seg_id}.mp3"

#         # visible state
#         if seg_kind == "bullet":
#             visible_bullets += 1

#         frame_path = scene_dir / f"{idx:03d}_{seg_id}.png"
#         temp_scene = dict(scene)
#         temp_scene["_project_assets"] = project_assets
#         render_composer_frame(
#             out_path=frame_path,
#             scene=temp_scene,
#             width=width,
#             height=height,
#             visible_bullets=visible_bullets,
#         )

#         clip_path = scene_dir / f"{idx:03d}_{seg_id}.mp4"

#         duration_sec = max(0.3, float(seg.get("duration_ms", 1000)) / 1000.0)

#         if audio_path.exists():
#             _run([
#                 "ffmpeg", "-y",
#                 "-loop", "1",
#                 "-i", str(frame_path),
#                 "-i", str(audio_path),
#                 "-c:v", "libx264",
#                 "-tune", "stillimage",
#                 "-pix_fmt", "yuv420p",
#                 "-c:a", "aac",
#                 "-af", "aresample=async=1:min_hard_comp=0.100:first_pts=0",
#                 "-shortest",
#                 "-r", "30",
#                 "-vf", f"scale={width}:{height}",
#                 str(clip_path),
#             ])
#         else:
#             _run([
#                 "ffmpeg", "-y",
#                 "-loop", "1",
#                 "-i", str(frame_path),
#                 "-f", "lavfi",
#                 "-i", f"anullsrc=r=44100:cl=stereo",
#                 "-t", str(duration_sec),
#                 "-c:v", "libx264",
#                 "-tune", "stillimage",
#                 "-pix_fmt", "yuv420p",
#                 "-c:a", "aac",
#                 "-r", "30",
#                 "-vf", f"scale={width}:{height}",
#                 str(clip_path),
#             ])

#         segment_mp4s.append(clip_path)

#     concat_list = scene_dir / "segments.txt"
#     concat_list.write_text(
#         "\n".join([f"file '{p.as_posix()}'" for p in segment_mp4s]),
#         encoding="utf-8"
#     )

#     scene_video = scene_dir / "scene.mp4"
#     # _run([
#     #     "ffmpeg", "-y",
#     #     "-f", "concat",
#     #     "-safe", "0",
#     #     "-i", str(concat_list),
#     #     "-c", "copy",
#     #     str(scene_video),
#     # ])

#     _run([
#         "ffmpeg", "-y",
#         "-f", "concat",
#         "-safe", "0",
#         "-i", str(concat_list),
#         "-c:v", "libx264",
#         "-pix_fmt", "yuv420p",
#         "-c:a", "aac",
#         "-ar", "44100",
#         "-ac", "2",
#         str(scene_video),
#     ])


#     return scene_video

# @app.post("/api/projects/cleanup-all")
# def cleanup_all_composer_projects():
#     removed = {
#         "projects": 0,
#         "speech_dirs": 0,
#         "render_dirs": 0,
#     }

#     if COMPOSER_PROJECTS_DIR.exists():
#         for p in COMPOSER_PROJECTS_DIR.glob("*.json"):
#             p.unlink(missing_ok=True)
#             removed["projects"] += 1

#     if COMPOSER_SPEECH_DIR.exists():
#         for p in COMPOSER_SPEECH_DIR.iterdir():
#             if p.is_dir():
#                 shutil.rmtree(p, ignore_errors=True)
#                 removed["speech_dirs"] += 1

#     if COMPOSER_RENDER_DIR.exists():
#         for p in COMPOSER_RENDER_DIR.iterdir():
#             if p.is_dir():
#                 shutil.rmtree(p, ignore_errors=True)
#                 removed["render_dirs"] += 1

#     return jsonify({
#         "ok": True,
#         "message": "All composer projects and generated files were removed.",
#         "removed": removed,
#     })


from io import BytesIO

def resolve_asset_to_local_file(asset: dict) -> Path | None:
    url = (asset.get("url") or "").strip()
    if not url:
        return None

    # Local uploaded file served by Flask
    if url.startswith("/uploads/"):
        p = BASE_DIR / url.lstrip("/")
        return p if p.exists() else None

    # Absolute or relative filesystem path
    p = _resolve_path(url)
    if p.exists():
        return p

    if url.startswith("/out/composer-project-files/"):
        parts = url.strip("/").split("/")
        # expected: out / composer-project-files / <project_id> / assets / <filename>
        if len(parts) >= 5 and parts[0] == "out" and parts[1] == "composer-project-files" and parts[3] == "assets":
            pid = parts[2]
            filename = parts[4]
            p = composer_project_assets_dir(pid) / filename
            return p if p.exists() else None
    
    # Remote image URL
    if url.startswith("http://") or url.startswith("https://"):
        tmp_dir = OUT_DIR / "composer_temp_assets"
        tmp_dir.mkdir(parents=True, exist_ok=True)
        ext = os.path.splitext(url.split("?")[0])[1].lower() or ".jpg"
        tmp_path = tmp_dir / f"{uuid.uuid4().hex}{ext}"
        try:
            r = http_get(url, timeout=20)
            r.raise_for_status()
            tmp_path.write_bytes(r.content)
            return tmp_path
        except Exception:
            return None

    return None

@app.post("/api/projects/cleanup-all")
def cleanup_all_composer_projects():
    removed = {
        "project_dirs": 0,
        "project_jsons": 0,
        "speech_dirs": 0,
        "render_dirs": 0,
    }

    if COMPOSER_PROJECT_ROOT.exists():
        for p in COMPOSER_PROJECT_ROOT.iterdir():
            if p.is_dir():
                shutil.rmtree(p, ignore_errors=True)
                removed["project_dirs"] += 1

    if COMPOSER_PROJECTS_DIR.exists():
        for p in COMPOSER_PROJECTS_DIR.glob("*.json"):
            p.unlink(missing_ok=True)
            removed["project_jsons"] += 1

    if COMPOSER_SPEECH_DIR.exists():
        for p in COMPOSER_SPEECH_DIR.iterdir():
            if p.is_dir():
                shutil.rmtree(p, ignore_errors=True)
                removed["speech_dirs"] += 1

    if COMPOSER_RENDER_DIR.exists():
        for p in COMPOSER_RENDER_DIR.iterdir():
            if p.is_dir():
                shutil.rmtree(p, ignore_errors=True)
                removed["render_dirs"] += 1

    return jsonify({
        "ok": True,
        "message": "All composer projects and generated files were removed.",
        "removed": removed,
    })


@app.post("/api/projects/<project_id>/render-composer-video")
def render_composer_video(project_id: str):
    try:
        project = load_project(project_id)
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404

    out_res = request.get_json(silent=True) or {}
    out_res = out_res.get("outRes", "1920x1080")

    render_root = COMPOSER_RENDER_DIR / project_id
    if render_root.exists():
        shutil.rmtree(render_root)
    render_root.mkdir(parents=True, exist_ok=True)

    project_assets = project.get("assets", []) or []
    scene_videos = []

    for scene in project.get("scenes", []):
        scene_video = render_scene_segments_to_video(
            project_id=project_id,
            scene=scene,
            project_assets=project_assets,
            render_dir=render_root,
            out_res=out_res,
        )
        scene_videos.append(scene_video)

    if not scene_videos:
        return jsonify({"ok": False, "error": "No scene videos were created."}), 400

    concat_list = render_root / "all_scenes.txt"
    concat_list.write_text(
        "\n".join([f"file '{p.as_posix()}'" for p in scene_videos]),
        encoding="utf-8"
    )

    final_video = render_root / "final_composer_video.mp4"
    # _run([
    #     "ffmpeg", "-y",
    #     "-f", "concat",
    #     "-safe", "0",
    #     "-i", str(concat_list),
    #     "-c", "copy",
    #     str(final_video),
    # ])

    _run([
        "ffmpeg", "-y",
        "-f", "concat",
        "-safe", "0",
        "-i", str(concat_list),
        "-c:v", "libx264",
        "-pix_fmt", "yuv420p",
        "-c:a", "aac",
        "-ar", "44100",
        "-ac", "2",
        str(final_video),
    ])

    return jsonify({
        "ok": True,
        "project_id": project_id,
        "video_path": str(final_video),
        "video_url": f"/composer-renders/{project_id}/final_composer_video.mp4",
    })


@app.get("/composer-renders/<path:fn>")
def composer_renders(fn):
    return send_from_directory(str(COMPOSER_RENDER_DIR), fn)

@app.get("/out/composer-project-files/<path:fn>")
def composer_project_files(fn):
    return send_from_directory(str(COMPOSER_PROJECT_ROOT), fn)

def _duration_via_ffprobe(path: Path) -> float:
    _ensure_ffmpeg()
    cmd = [
        "ffprobe", "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        str(path),
    ]
    out = subprocess.check_output(cmd, text=True).strip()
    return float(out)

def word_count(text: str) -> int:
    text = compact_ws(text)
    return len(text.split()) if text else 0

def fact_display_text(label: str, value: str, max_words: int = 10) -> str:
    label = compact_ws(label)
    value = compact_ws(value)
    full = f"{label}: {value}" if value else label
    return full if word_count(full) <= max_words else label

def merged_short_display_text(display_text: str, speech_text: str, max_words: int = 10) -> str:
    display_text = compact_ws(display_text)
    speech_text = compact_ws(speech_text)

    if not display_text:
        return speech_text
    if not speech_text:
        return display_text
    if display_text == speech_text:
        return display_text

    # If spoken text already contains the visible bullet, show the spoken line
    if display_text.lower() in speech_text.lower() and word_count(speech_text) <= max_words:
        return speech_text

    merged = f"{display_text} — {speech_text}"
    if word_count(merged) <= max_words:
        return merged

    return display_text

def _run(cmd: list[str]) -> None:
    subprocess.run(cmd, check=True)

def _safe_filename(text: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9._-]+", "_", (text or "").strip())
    return text[:120] or "item"

def compact_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def text_from_node(node: Tag | NavigableString | None) -> str:
    if node is None:
        return ""
    if isinstance(node, NavigableString):
        return compact_ws(str(node))
    return compact_ws(node.get_text(" ", strip=True))


def estimate_duration(narration_text: str, bullets: list[str]) -> float:
    text = compact_ws(narration_text)
    words = len(text.split()) if text else 0
    bullet_bonus = max(0, len(bullets) - 1) * 0.8
    return round(max(3.0, words / 2.7 + bullet_bonus), 2)

def build_speech_segments(title: str, bullets: list[str], narration_parts: list[str], scene_index: int) -> list[SpeechSegment]:
    segments: list[SpeechSegment] = []
    segment_index = 0

    if compact_ws(title):
        segment_index += 1
        segments.append(SpeechSegment(
            id=make_segment_id(scene_index, "title", "title"),
            kind="title",
            display_text=compact_ws(title),
            speech_text=compact_ws(title),
            show_on_screen=True,
            animation="titleReveal",
            animation_target="title",
            order=segment_index,
        ))

    clean_bullets = [compact_ws(x) for x in bullets if compact_ws(x)]
    for i, bullet in enumerate(clean_bullets, start=1):
        segment_index += 1
        target = f"bullet_{i}"
        segments.append(SpeechSegment(
            id=make_segment_id(scene_index, "bullet", target),
            kind="bullet",
            display_text=bullet,
            speech_text=bullet,
            show_on_screen=True,
            animation="bulletReveal",
            animation_target=target,
            order=segment_index,
        ))

    narration_index = 0
    used = {compact_ws(title), *clean_bullets}
    for part in [compact_ws(x) for x in narration_parts if compact_ws(x)]:
        if part in used:
            continue
        segment_index += 1
        narration_index += 1
        segments.append(SpeechSegment(
            id=make_segment_id(scene_index, "narration", logical_index=narration_index),
            kind="narration",
            display_text="",
            speech_text=part,
            show_on_screen=False,
            animation="none",
            animation_target="none",
            order=segment_index,
        ))

    return segments

# def build_speech_segments(title: str, bullets: list[str], narration_parts: list[str], scene_index: int) -> list[SpeechSegment]:
#     segments: list[SpeechSegment] = []
#     segment_index = 0

#     if compact_ws(title):
#         segment_index += 1
#         segments.append(SpeechSegment(
#             id=make_segment_id(scene_index, segment_index, "title"),
#             kind="title",
#             display_text=compact_ws(title),
#             speech_text=compact_ws(title),
#             show_on_screen=True,
#             animation="titleReveal",
#             animation_target="title",
#             order=segment_index,
#         ))

#     clean_bullets = [compact_ws(x) for x in bullets if compact_ws(x)]
#     for i, bullet in enumerate(clean_bullets, start=1):
#         segment_index += 1
#         segments.append(SpeechSegment(
#             id=make_segment_id(scene_index, segment_index, f"bullet{i}"),
#             kind="bullet",
#             display_text=bullet,
#             speech_text=bullet,
#             show_on_screen=True,
#             animation="bulletReveal",
#             animation_target=f"bullet_{i}",
#             order=segment_index,
#         ))

#     used = {compact_ws(title), *clean_bullets}
#     for part in [compact_ws(x) for x in narration_parts if compact_ws(x)]:
#         if part in used:
#             continue
#         segment_index += 1
#         segments.append(SpeechSegment(
#             id=make_segment_id(scene_index, segment_index, "narration"),
#             kind="narration",
#             display_text="",
#             speech_text=part,
#             show_on_screen=False,
#             animation="none",
#             animation_target="none",
#             order=segment_index,
#         ))

#     return segments


# def build_speech_segments(title: str, bullets: list[str], narration_parts: list[str]) -> list[SpeechSegment]:
#     segments: list[SpeechSegment] = []
#     order = 0

#     if compact_ws(title):
#         order += 1
#         segments.append(SpeechSegment(
#             id=str(uuid.uuid4()),
#             kind="title",
#             display_text=compact_ws(title),
#             speech_text=compact_ws(title),
#             show_on_screen=True,
#             animation="titleReveal",
#             animation_target="title",
#             order=order,
#         ))

#     clean_bullets = [compact_ws(x) for x in bullets if compact_ws(x)]
#     for i, bullet in enumerate(clean_bullets, start=1):
#         order += 1
#         segments.append(SpeechSegment(
#             id=str(uuid.uuid4()),
#             kind="bullet",
#             display_text=bullet,
#             speech_text=bullet,
#             show_on_screen=True,
#             animation="bulletReveal",
#             animation_target=f"bullet_{i}",
#             order=order,
#         ))

#     # narration-only support text that should not appear on screen
#     extra_parts = [compact_ws(x) for x in narration_parts if compact_ws(x)]
#     used = {compact_ws(title), *clean_bullets}
#     for part in extra_parts:
#         if part in used:
#             continue
#         order += 1
#         segments.append(SpeechSegment(
#             id=str(uuid.uuid4()),
#             kind="narration",
#             display_text="",
#             speech_text=part,
#             show_on_screen=False,
#             animation="none",
#             animation_target="none",
#             order=order,
#         ))

#     return segments

def build_narration(scene_type: SceneType, title: str, body: str, bullets: list[str]) -> str:
    title = compact_ws(title)
    body = compact_ws(body)
    bullets = [compact_ws(x) for x in bullets if compact_ws(x)]

    if scene_type == "title":
        return title
    if scene_type == "section_header":
        return title
    if scene_type == "bullet_points":
        parts: list[str] = []
        if title:
            parts.append(title)
        parts.extend(bullets)
        return ". ".join(parts)
    if body:
        if title:
            return f"{title}. {body}"
        return body
    return title


HEADER_TAGS = {"h1", "h2", "h3", "h4", "h5", "h6"}

def pick_best_root(soup: BeautifulSoup) -> Tag:
    candidates = []

    # Prefer semantic containers first
    for node in soup.find_all(["article", "main", "section", "body"]):
        if not isinstance(node, Tag):
            continue

        text_len = len(compact_ws(node.get_text(" ", strip=True)))
        block_count = len(node.find_all(["h1", "h2", "h3", "p", "ul", "ol", "blockquote", "img", "pre", "code", "table"]))

        # Skip containers that are effectively empty
        if text_len < 40 and block_count == 0:
            continue

        score = 0
        if node.name == "article":
            score += 50
        elif node.name == "main":
            score += 40
        elif node.name == "section":
            score += 25
        elif node.name == "body":
            score += 10

        # Prefer content-rich containers
        score += min(text_len // 100, 50)
        score += min(block_count * 3, 40)

        # Prefer likely content classes
        class_text = " ".join(node.get("class", []))
        if any(x in class_text for x in ["detail", "content", "article", "post", "main"]):
            score += 20

        candidates.append((score, node))

    if not candidates:
        return soup.body or soup

    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]

def remove_non_content_blocks(root: Tag) -> None:
    selectors = [
        "header",
        "nav",
        "footer",
        "script",
        "style",
        "noscript",
        ".site-header",
        ".main-nav",
    ]

    for selector in selectors:
        for node in root.select(selector):
            node.decompose()

def extract_strong_lead_text(node: Tag) -> tuple[str, str] | None:
    strong = node.find("strong")
    if not strong:
        return None

    strong_text = compact_ws(strong.get_text(" ", strip=True)).rstrip(":")
    full_text = compact_ws(node.get_text(" ", strip=True))
    if not strong_text or not full_text:
        return None

    remainder = full_text
    if full_text.startswith(strong_text):
        remainder = compact_ws(full_text[len(strong_text):].lstrip(": ").strip())

    return strong_text, remainder


def is_short_fact_line(text: str) -> bool:
    if not text:
        return False
    if ":" not in text:
        return False
    left, right = text.split(":", 1)
    return len(left.strip()) <= 40 and len(right.strip()) > 0


def split_fact_line(text: str) -> tuple[str, str] | None:
    if not is_short_fact_line(text):
        return None
    left, right = text.split(":", 1)
    return compact_ws(left), compact_ws(right)


def build_display_and_narration(title: str, body: str, bullets: list[str]) -> tuple[list[str], str]:
    display_items = []
    if title:
        display_items.append(title)
    display_items.extend([b for b in bullets if b])

    narration_parts = []
    if title:
        narration_parts.append(title)
    if body:
        narration_parts.append(body)
    if bullets:
        narration_parts.extend(bullets)

    narration = ". ".join([compact_ws(x) for x in narration_parts if compact_ws(x)])
    return display_items, narration

def finalize_grouped_scene(
    scenes: list[Scene],
    warnings: list[str],
    *,
    title: str,
    subtitle: str,
    bullets: list[str],
    narration_parts: list[str],
    scene_type: str = "bullet_points",
    layout_template: str = "bullet_focus",
    text_align: str = "left",
    overlay_position: str = "left",
) -> None:
    clean_title = compact_ws(title)
    clean_subtitle = compact_ws(subtitle)
    clean_bullets = [compact_ws(x) for x in bullets if compact_ws(x)]
    clean_narration_parts = [compact_ws(x) for x in narration_parts if compact_ws(x)]

    if not clean_title and not clean_bullets and not clean_narration_parts:
        return

    # narration_text = ". ".join(clean_narration_parts)

    narration_text = build_tagged_narration_text(
        title=clean_title,
        bullets=clean_bullets,
        narration_parts=clean_narration_parts,
    )

    scene = Scene(
        id=str(uuid.uuid4()),
        scene_type=scene_type,
        title=clean_title,
        subtitle=clean_subtitle,
        bullets=clean_bullets,
        body_text="",
        narration_text=narration_text,
        on_screen_text=[clean_title, *clean_bullets] if clean_title else clean_bullets[:],
        source_refs=["grouped"],
        layout=SceneLayout(
            template=layout_template,
            text_align=text_align,
            overlay_position=overlay_position,
        ),
        timing=SceneTiming(reveal_mode="sequential_bullets" if clean_bullets else "all_at_once"),
    )

    scene_index = len(scenes) + 1
    scene.speech_segments = build_speech_segments(
        title=clean_title,
        bullets=clean_bullets,
        narration_parts=clean_narration_parts,
        scene_index=scene_index,
    )

    # scene.speech_segments = build_speech_segments(
    #     title=clean_title,
    #     bullets=clean_bullets,
    #     narration_parts=clean_narration_parts,
    # )

    if not compact_ws(scene.narration_text):
        scene.narration_text = build_narration(
            scene.scene_type,
            scene.title,
            scene.body_text,
            scene.bullets,
        )

    scene.timing.estimated_duration_sec = estimate_duration(scene.narration_text, scene.bullets)

    if len(scene.narration_text) > 500:
        scene.notes.append("Narration may be too long for a single scene.")
        warnings.append(f"Long narration in scene: {scene.title[:60]}")

    if sum(len(x) for x in scene.on_screen_text) > 260:
        scene.notes.append("Too much text on screen. Consider splitting the scene.")

    scenes.append(scene)


def split_bullets_for_scenes(title: str, bullets: list[str], narration_parts: list[str], warnings: list[str], scenes: list[Scene]) -> None:
    clean_bullets = [compact_ws(x) for x in bullets if compact_ws(x)]
    clean_narration_parts = [compact_ws(x) for x in narration_parts if compact_ws(x)]

    if not clean_bullets and not clean_narration_parts:
        return

    if len(clean_bullets) <= 5:
        finalize_grouped_scene(
            scenes,
            warnings,
            title=title,
            subtitle="",
            bullets=clean_bullets,
            narration_parts=clean_narration_parts,
        )
        return

    bullet_chunks = [clean_bullets[i:i + 5] for i in range(0, len(clean_bullets), 5)]
    narration_chunks = [clean_narration_parts[i:i + 5] for i in range(0, len(clean_narration_parts), 5)]

    for idx, chunk in enumerate(bullet_chunks, start=1):
        chunk_title = title if idx == 1 else f"{title} ({idx})"
        finalize_grouped_scene(
            scenes,
            warnings,
            title=chunk_title,
            subtitle="",
            bullets=chunk,
            narration_parts=narration_chunks[idx - 1] if idx - 1 < len(narration_chunks) else chunk,
        )

def parse_html_to_project(html: str, *, source_type: str, source_value: str, theme_id: str) -> VideoProject:
    soup = BeautifulSoup(html, "html.parser")

    doc_title = compact_ws(text_from_node(soup.title))
    root = pick_best_root(soup)
    remove_non_content_blocks(root)

    block_nodes = root.find_all(
        ["h1", "h2", "h3", "h4", "h5", "h6", "p", "ul", "ol", "blockquote", "img", "pre", "code", "table"],
        recursive=True
    )

    scenes: list[Scene] = []
    assets: list[MediaAsset] = []
    warnings: list[str] = []

    project_title = doc_title or "HTML Video Project"

    current_heading = ""
    current_subtitle = ""
    current_bullets: list[str] = []
    current_narration_parts: list[str] = []
    first_heading_seen = False

    def flush_current_scene(is_final: bool = False) -> None:
        nonlocal current_heading, current_subtitle, current_bullets, current_narration_parts, first_heading_seen

        if not current_heading and not current_bullets and not current_narration_parts:
            return

        # Only create a heading-only scene if:
        # 1) it is the first title scene, or
        # 2) there are visible bullets, or
        # 3) this is the final flush and we still want to preserve content
        has_meaningful_narration = any(
            compact_ws(part) and compact_ws(part) != compact_ws(current_heading)
            for part in current_narration_parts
        )

        should_create_scene = bool(current_bullets) or has_meaningful_narration or is_final or (current_heading == project_title)

        if should_create_scene:
            split_bullets_for_scenes(
                title=current_heading or "Scene",
                bullets=current_bullets,
                narration_parts=current_narration_parts,
                warnings=warnings,
                scenes=scenes,
            )

        current_heading = ""
        current_subtitle = ""
        current_bullets = []
        current_narration_parts = []

    # def flush_current_scene() -> None:
    #     nonlocal current_heading, current_subtitle, current_bullets, current_narration_parts

    #     if not current_heading and not current_bullets and not current_narration_parts:
    #         return

    #     split_bullets_for_scenes(
    #         title=current_heading or "Scene",
    #         bullets=current_bullets,
    #         narration_parts=current_narration_parts,
    #         warnings=warnings,
    #         scenes=scenes,
    #     )

    #     if not current_bullets and current_narration_parts:
    #         finalize_grouped_scene(
    #             scenes,
    #             warnings,
    #             title=current_heading or "Scene",
    #             subtitle=current_subtitle,
    #             bullets=[],
    #             narration_parts=current_narration_parts,
    #             scene_type="section_header" if current_heading else "paragraph",
    #             layout_template="section_divider" if current_heading else "text_card",
    #             text_align="left",
    #             overlay_position="left" if current_heading else "bottom_card",
    #         )

    #     current_heading = ""
    #     current_subtitle = ""
    #     current_bullets = []
    #     current_narration_parts = []

    for node in block_nodes:
        tag = node.name.lower()

        if tag in HEADER_TAGS:
            flush_current_scene(is_final=False)

            heading = text_from_node(node)
            if not heading:
                continue

            current_heading = heading
            current_subtitle = ""
            current_bullets = []
            current_narration_parts = []

            if not first_heading_seen:
                first_heading_seen = True
                project_title = heading

            continue

        if tag == "p":
            text = text_from_node(node)
            if not text:
                continue

            strong_pair = extract_strong_lead_text(node)
            fact_pair = split_fact_line(text)

            if strong_pair:
                # bullet_title, bullet_desc = strong_pair
                # current_bullets.append(bullet_title)
                # current_narration_parts.append(f"{bullet_title}. {bullet_desc}" if bullet_desc else bullet_title)
                bullet_title, bullet_desc = strong_pair
                bullet_text = fact_display_text(bullet_title, bullet_desc, max_words=10)
                narration_text = f"{bullet_title}: {bullet_desc}" if bullet_desc else bullet_title

                current_bullets.append(bullet_text)
                current_narration_parts.append(narration_text)

                continue

            if fact_pair:
                fact_label, fact_value = fact_pair
                bullet_text = f"{fact_label}: {fact_value}"
                current_bullets.append(bullet_text)
                current_narration_parts.append(f"{fact_label}. {fact_value}.")
                continue

            # plain paragraphs go to narration only
            current_narration_parts.append(text)
            if not current_heading and not first_heading_seen:
                current_heading = project_title
                first_heading_seen = True

            continue

        if tag in {"ul", "ol"}:
            li_nodes = node.find_all("li", recursive=False)

            for li in li_nodes:
                pair = extract_strong_lead_text(li)
                if pair:
                    bullet_title, bullet_desc = pair
                    bullet_text = fact_display_text(bullet_title, bullet_desc, max_words=10)
                    narration_text = f"{bullet_title}: {bullet_desc}" if bullet_desc else bullet_title

                    current_bullets.append(bullet_text)
                    current_narration_parts.append(narration_text)
                else:
                    li_text = text_from_node(li)
                    if li_text:
                        current_bullets.append(li_text)
                        current_narration_parts.append(li_text)
            continue

        if tag == "blockquote":
            quote = text_from_node(node)
            if quote:
                current_narration_parts.append(quote)
            continue

        if tag == "img":
            src = (node.get("src") or "").strip()
            if not src:
                continue

            asset_id = str(uuid.uuid4())
            assets.append(MediaAsset(
                id=asset_id,
                kind="image",
                url=src,
                title=node.get("alt") or "Image",
                source="html",
            ))

            # attach to the most recent grouped scene if possible, else create one
            if scenes:
                scenes[-1].media_asset_ids.append(asset_id)
            else:
                finalize_grouped_scene(
                    scenes,
                    warnings,
                    title=node.get("alt") or project_title,
                    subtitle="",
                    bullets=[],
                    narration_parts=[node.get("alt") or "Visual reference"],
                    scene_type="image_focus",
                    layout_template="image_with_caption",
                    text_align="left",
                    overlay_position="bottom_card",
                )
                scenes[-1].media_asset_ids.append(asset_id)
            continue

        if tag in {"pre", "code"}:
            code_text = text_from_node(node)
            if code_text:
                current_narration_parts.append(code_text[:360])
            continue

        if tag == "table":
            rows = []
            for tr in node.find_all("tr"):
                cols = [text_from_node(td) for td in tr.find_all(["th", "td"])]
                cols = [c for c in cols if c]
                if cols:
                    rows.append(" | ".join(cols))

            if rows:
                current_narration_parts.extend(rows[:6])
            continue

    flush_current_scene(is_final=True)

    return VideoProject(
        id=str(uuid.uuid4()),
        title=project_title,
        source_type=source_type,  # type: ignore[arg-type]
        source_value=source_value,
        theme_id=theme_id,
        scenes=scenes,
        assets=assets,
        warnings=warnings,
        metadata={
            "scene_count": len(scenes),
            "asset_count": len(assets),
            "slug": slugify(project_title),
        },
    )

def parse_html_to_project_delit(html: str, *, source_type: str, source_value: str, theme_id: str) -> VideoProject:
    soup = BeautifulSoup(html, "html.parser")

    doc_title = compact_ws(text_from_node(soup.title))
    # article = soup.find("article") or soup.find("main") or soup.body or soup
    # block_nodes = article.find_all([*HEADER_TAGS, "p", "ul", "ol", "blockquote", "img", "pre", "code", "table"], recursive=True)

    root = pick_best_root(soup)
    remove_non_content_blocks(root)
    
    block_nodes = root.find_all(
        ["h1", "h2", "h3", "h4", "h5", "h6", "p", "ul", "ol", "blockquote", "img", "pre", "code", "table"],
        recursive=True
    )

    scenes: list[Scene] = []
    assets: list[MediaAsset] = []
    warnings: list[str] = []

    project_title = doc_title or "HTML Video Project"

    def add_scene(scene: Scene) -> None:
        # scene.narration_text = build_narration(scene.scene_type, scene.title, scene.body_text, scene.bullets)
        if not compact_ws(scene.narration_text):
            scene.narration_text = build_narration(
                scene.scene_type,
                scene.title,
                scene.body_text,
                scene.bullets
            )
        scene.timing.estimated_duration_sec = estimate_duration(scene.narration_text, scene.bullets)
        if len(scene.narration_text) > 500:
            scene.notes.append("Narration may be too long for a single scene.")
            warnings.append(f"Long narration in scene: {scene.title[:60]}")
        if sum(len(x) for x in scene.on_screen_text) > 260:
            scene.notes.append("Too much text on screen. Consider splitting the scene.")
        scenes.append(scene)

    first_title_used = False

    for node in block_nodes:
        tag = node.name.lower()
        if tag in HEADER_TAGS:
            heading = text_from_node(node)
            if not heading:
                continue
            if not first_title_used:
                first_title_used = True
                project_title = heading
                add_scene(Scene(
                    id=str(uuid.uuid4()),
                    scene_type="title",
                    title=heading,
                    on_screen_text=[heading],
                    source_refs=[tag],
                    layout=SceneLayout(template="hero_centered", text_align="center", overlay_position="center"),
                ))
            else:
                add_scene(Scene(
                    id=str(uuid.uuid4()),
                    scene_type="section_header",
                    title=heading,
                    on_screen_text=[heading],
                    source_refs=[tag],
                    layout=SceneLayout(template="section_divider", text_align="left", overlay_position="left"),
                ))
            continue

        if tag == "p":
            text = text_from_node(node)
            if not text:
                continue

            strong_pair = extract_strong_lead_text(node)
            fact_pair = split_fact_line(text)

            if strong_pair:
                bullet_title, bullet_desc = strong_pair
                display_items, narration = build_display_and_narration(
                    title="Key point",
                    body="",
                    bullets=[bullet_title]
                )
                add_scene(Scene(
                    id=str(uuid.uuid4()),
                    scene_type="bullet_points",
                    title="Key point",
                    subtitle="",
                    bullets=[bullet_title],
                    body_text="",
                    narration_text = build_tagged_narration_text(
                        title="Key point",
                        bullets=[bullet_title],
                        narration_parts=[f"{bullet_title}. {bullet_desc}" if bullet_desc else bullet_title],
                    ),
                    on_screen_text=display_items,
                    source_refs=[tag, "strong"],
                    layout=SceneLayout(template="bullet_focus", text_align="left", overlay_position="left"),
                    timing=SceneTiming(reveal_mode="sequential_bullets"),
                ))
                continue

            if fact_pair:
                fact_label, fact_value = fact_pair
                add_scene(Scene(
                    id=str(uuid.uuid4()),
                    scene_type="bullet_points",
                    title="Key fact",
                    subtitle="",
                    bullets=[f"{fact_label}: {fact_value}"],
                    body_text="",
                    narration_text = build_tagged_narration_text(
                        title="Key fact",
                        bullets=[f"{fact_label}: {fact_value}"],
                        narration_parts=[f"{fact_label}. {fact_value}."],
                    ),
                    on_screen_text=["Key fact", f"{fact_label}: {fact_value}"],
                    source_refs=[tag, "fact_line"],
                    layout=SceneLayout(template="bullet_focus", text_align="left", overlay_position="left"),
                    timing=SceneTiming(reveal_mode="sequential_bullets"),
                ))
                continue

            add_scene(Scene(
                id=str(uuid.uuid4()),
                scene_type="paragraph",
                title="",
                subtitle=text[:140] if len(text) <= 140 else "",
                body_text="",
                narration_text = build_tagged_narration_text(
                    title="",
                    bullets=[],
                    narration_parts=[text],
                ),
                on_screen_text=[],
                source_refs=[tag],
                layout=SceneLayout(template="text_card", text_align="left", overlay_position="bottom_card"),
            ))
            continue

        # if tag == "p":
        #     text = text_from_node(node)
        #     if not text:
        #         continue
        #     add_scene(Scene(
        #         id=str(uuid.uuid4()),
        #         scene_type="paragraph",
        #         body_text=text,
        #         on_screen_text=[text],
        #         source_refs=[tag],
        #         layout=SceneLayout(template="text_card", text_align="left", overlay_position="bottom_card"),
        #     ))
        #     continue


        if tag in {"ul", "ol"}:
            li_nodes = node.find_all("li", recursive=False)
            bullet_titles = []
            bullet_narrations = []

            for li in li_nodes:
                pair = extract_strong_lead_text(li)
                if pair:
                    bullet_title, bullet_desc = pair
                    bullet_titles.append(bullet_title)
                    bullet_narrations.append(f"{bullet_title}. {bullet_desc}" if bullet_desc else bullet_title)
                else:
                    li_text = text_from_node(li)
                    if li_text:
                        bullet_titles.append(li_text)
                        bullet_narrations.append(li_text)

            bullet_titles = [x for x in bullet_titles if x]
            bullet_narrations = [x for x in bullet_narrations if x]

            if not bullet_titles:
                continue

            chunks = [bullet_titles[i:i + 5] for i in range(0, len(bullet_titles), 5)]
            narration_chunks = [bullet_narrations[i:i + 5] for i in range(0, len(bullet_narrations), 5)]

            for idx, chunk in enumerate(chunks, start=1):
                title = "Key points" if len(chunks) == 1 else f"Key points ({idx})"
                # narration_text = ". ".join(narration_chunks[idx - 1])

                chunk_narrations = narration_chunks[idx - 1] if idx - 1 < len(narration_chunks) else chunk

                narration_text = build_tagged_narration_text(
                    title=title,
                    bullets=chunk,
                    narration_parts=chunk_narrations,
                )

                add_scene(Scene(
                    id=str(uuid.uuid4()),
                    scene_type="bullet_points",
                    title=title,
                    subtitle="",
                    bullets=chunk,
                    body_text="",
                    narration_text=narration_text,
                    on_screen_text=[title, *chunk],
                    source_refs=[tag],
                    timing=SceneTiming(reveal_mode="sequential_bullets"),
                    layout=SceneLayout(template="bullet_focus", text_align="left", overlay_position="left"),
                ))
            continue

        # if tag in {"ul", "ol"}:
        #     bullets = [text_from_node(li) for li in node.find_all("li", recursive=False)]
        #     bullets = [x for x in bullets if x]
        #     if not bullets:
        #         continue
        #     chunks = [bullets[i:i + 5] for i in range(0, len(bullets), 5)]
        #     for idx, chunk in enumerate(chunks, start=1):
        #         title = "Key points" if len(chunks) == 1 else f"Key points ({idx})"
        #         add_scene(Scene(
        #             id=str(uuid.uuid4()),
        #             scene_type="bullet_points",
        #             title=title,
        #             bullets=chunk,
        #             on_screen_text=[title, *chunk],
        #             source_refs=[tag],
        #             timing=SceneTiming(reveal_mode="sequential_bullets"),
        #             layout=SceneLayout(template="bullet_focus", text_align="left", overlay_position="left"),
        #         ))
        #     continue

        if tag == "blockquote":
            quote = text_from_node(node)
            if not quote:
                continue
            add_scene(Scene(
                id=str(uuid.uuid4()),
                scene_type="quote",
                body_text=quote,
                on_screen_text=[quote],
                source_refs=[tag],
                layout=SceneLayout(template="quote_spotlight", text_align="center", overlay_position="center"),
            ))
            continue

        if tag == "img":
            src = (node.get("src") or "").strip()
            if not src:
                continue
            asset_id = str(uuid.uuid4())
            assets.append(MediaAsset(
                id=asset_id,
                kind="image",
                url=src,
                title=node.get("alt") or "Image",
                source="html",
            ))
            add_scene(Scene(
                id=str(uuid.uuid4()),
                scene_type="image_focus",
                title=node.get("alt") or "Visual",
                on_screen_text=[node.get("alt") or "Visual"],
                media_asset_ids=[asset_id],
                source_refs=[tag],
                layout=SceneLayout(template="image_with_caption", text_align="left", overlay_position="bottom_card"),
            ))
            continue

        if tag in {"pre", "code"}:
            code_text = text_from_node(node)
            if not code_text:
                continue
            snippet = code_text[:360]
            add_scene(Scene(
                id=str(uuid.uuid4()),
                scene_type="code",
                title="Code example",
                body_text=snippet,
                on_screen_text=["Code example", snippet],
                source_refs=[tag],
                layout=SceneLayout(template="code_panel", text_align="left", overlay_position="center"),
            ))
            continue

        if tag == "table":
            rows = []
            for tr in node.find_all("tr"):
                cols = [text_from_node(td) for td in tr.find_all(["th", "td"])]
                cols = [c for c in cols if c]
                if cols:
                    rows.append(" | ".join(cols))
            if not rows:
                continue
            preview = rows[:6]
            add_scene(Scene(
                id=str(uuid.uuid4()),
                scene_type="table",
                title="Table",
                body_text="\n".join(preview),
                on_screen_text=["Table", *preview[:4]],
                source_refs=[tag],
                layout=SceneLayout(template="table_card", text_align="left", overlay_position="center"),
            ))

    if not scenes:
        warnings.append("No supported content blocks were found in the HTML.")

    recap_points = []
    for scene in scenes:
        if scene.scene_type == "section_header" and scene.title:
            recap_points.append(scene.title)
        if len(recap_points) == 4:
            break
    if recap_points:
        add_scene(Scene(
            id=str(uuid.uuid4()),
            scene_type="recap",
            title="Recap",
            bullets=recap_points,
            on_screen_text=["Recap", *recap_points],
            layout=SceneLayout(template="bullet_focus", text_align="left", overlay_position="right"),
            timing=SceneTiming(reveal_mode="sequential_bullets"),
        ))

    return VideoProject(
        id=str(uuid.uuid4()),
        title=project_title,
        source_type=source_type,  # type: ignore[arg-type]
        source_value=source_value,
        theme_id=theme_id,
        scenes=scenes,
        assets=assets,
        warnings=warnings,
        metadata={
            "scene_count": len(scenes),
            "asset_count": len(assets),
            "slug": slugify(project_title),
        },
    )


def project_path(project_id: str) -> Path:
    return PROJECTS_DIR / f"{project_id}.json"


def save_project(project: VideoProject | dict[str, Any] | str, project_data: dict[str, Any] | None = None) -> None:
    """Persist a composer project from either a dataclass instance or a raw dict."""
    if project_data is not None:
        project_id = str(project)
        payload = project_data
    elif isinstance(project, dict):
        project_id = str(project.get("id") or "")
        payload = project
    else:
        project_id = str(getattr(project, "id", "") or "")
        payload = asdict(project)

    if not project_id:
        raise ValueError("Project id is required to save the project.")

    project_path(project_id).write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def load_project(project_id: str) -> dict[str, Any]:
    path = project_path(project_id)
    if not path.exists():
        raise FileNotFoundError(f"Project not found: {project_id}")
    return json.loads(path.read_text(encoding="utf-8"))


@app.get("/api/themes")
def get_themes():
    return jsonify({"themes": [asdict(theme) for theme in THEMES.values()]})


@app.post("/api/projects/from-html")
def create_project_from_html():
    data = request.get_json(force=True)
    html = (data.get("html") or "").strip()
    if not html:
        return jsonify({"ok": False, "error": "Missing 'html'"}), 400

    theme_id = data.get("theme_id") or "corporate-clean"
    if theme_id not in THEMES:
        return jsonify({"ok": False, "error": f"Unknown theme_id: {theme_id}"}), 400

    source_value = data.get("source_value") or "inline-html"
    project = parse_html_to_project(
        html,
        source_type="html",
        source_value=source_value,
        theme_id=theme_id,
    )
    save_project(project)
    return jsonify({"ok": True, "project": asdict(project)})


@app.post("/api/projects/from-url")
def create_project_from_url():
    data = request.get_json(force=True)
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"ok": False, "error": "Missing 'url'"}), 400

    theme_id = data.get("theme_id") or "corporate-clean"
    if theme_id not in THEMES:
        return jsonify({"ok": False, "error": f"Unknown theme_id: {theme_id}"}), 400

    try:
        response = http_get(url, timeout=20)
        response.raise_for_status()
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Failed to fetch URL: {exc}"}), 400

    project = parse_html_to_project(
        response.text,
        source_type="url",
        source_value=url,
        theme_id=theme_id,
    )
    save_project(project)
    return jsonify({"ok": True, "project": asdict(project)})


@app.get("/api/projects/<project_id>")
def get_project(project_id: str):
    try:
        return jsonify({"ok": True, "project": load_project(project_id)})
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404


@app.put("/api/projects/<project_id>")
def update_project(project_id: str):
    data = request.get_json(force=True)
    try:
        current = load_project(project_id)
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404

    allowed_top_level = {
        "title",
        "theme_id",
        "project_settings",
        "narration",
        "captions",
        "scenes",
        "assets",
        "warnings",
        "metadata",
    }
    for key in allowed_top_level:
        if key in data:
            current[key] = data[key]

    scenes = current.get("scenes", []) or []
    for idx, scene in enumerate(scenes, start=1):
        _rebuild_scene_speech_segments(scene, idx)
        scene["speech_plan"] = build_scene_speech_plan(
            speech_segments=scene.get("speech_segments", []) or [],
            pause_after_title_ms=450,
            pause_between_segments_ms=200,
            outro_hold_ms=500,
        )

    project_path(project_id).write_text(
        json.dumps(current, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )
    return jsonify({"ok": True, "project": current, "message": "Project saved."})

@app.post("/api/projects/<project_id>/render-plan")
def build_render_plan(project_id: str):
    try:
        project = load_project(project_id)
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404

    scenes = project.get("scenes", [])
    total_estimated = 0.0
    render_scenes = []

    for idx, scene in enumerate(scenes, start=1):
        _rebuild_scene_speech_segments(scene, idx)
        speech_segments = scene.get("speech_segments", []) or []
        timing = scene.get("timing", {}) or {}

        speech_plan = build_scene_speech_plan(
            speech_segments=speech_segments,
            pause_after_title_ms=450,
            pause_between_segments_ms=200,
            outro_hold_ms=500,
        )

        scene_duration = round((speech_plan["total_duration_ms"] or 0) / 1000.0, 2)
        if timing.get("manual_duration_sec"):
            scene_duration = float(timing["manual_duration_sec"])

        total_estimated += scene_duration

        render_scenes.append({
            "index": idx,
            "scene_id": scene.get("id"),
            "scene_type": scene.get("scene_type"),
            "title": scene.get("title", ""),
            "subtitle": scene.get("subtitle", ""),
            "bullets": scene.get("bullets", []),
            "layout": scene.get("layout", {}).get("template"),
            "media_asset_ids": scene.get("media_asset_ids", []),
            "speech_segments": speech_segments,
            "speech_plan": speech_plan,
            "duration_sec": scene_duration,
        })

    return jsonify({
        "ok": True,
        "project_id": project_id,
        "theme": THEMES.get(project.get("theme_id", "corporate-clean")).__dict__ if project.get("theme_id", "corporate-clean") in THEMES else None,
        "total_estimated_duration_sec": round(total_estimated, 2),
        "render_scenes": render_scenes,
    })

# @app.post("/api/projects/<project_id>/render-plan")
# def build_render_plan(project_id: str):
#     try:
#         project = load_project(project_id)
#     except FileNotFoundError as exc:
#         return jsonify({"ok": False, "error": str(exc)}), 404

#     scenes = project.get("scenes", [])
#     total_estimated = 0.0
#     render_scenes = []

#     for idx, scene in enumerate(scenes, start=1):
#         timing = scene.get("timing", {})
#         scene_duration = timing.get("manual_duration_sec") or timing.get("estimated_duration_sec") or 4.0
#         total_estimated += float(scene_duration)
#         render_scenes.append({
#             "index": idx,
#             "scene_id": scene.get("id"),
#             "scene_type": scene.get("scene_type"),
#             "duration_sec": scene_duration,
#             "layout": scene.get("layout", {}).get("template"),
#             "narration_text": scene.get("narration_text", ""),
#             "on_screen_text": scene.get("on_screen_text", []),
#             "media_asset_ids": scene.get("media_asset_ids", []),
#             "reveal_mode": timing.get("reveal_mode", "all_at_once"),
#         })

#     return jsonify({
#         "ok": True,
#         "project_id": project_id,
#         "theme": THEMES.get(project.get("theme_id", "corporate-clean")).__dict__ if project.get("theme_id", "corporate-clean") in THEMES else None,
#         "total_estimated_duration_sec": round(total_estimated, 2),
#         "render_scenes": render_scenes,
#     })

def estimate_text_duration_ms(text: str, words_per_minute: int = 145, min_ms: int = 700) -> int:
    text = compact_ws(text)
    if not text:
        return 0
    words = max(1, len(text.split()))
    ms = int((words / float(words_per_minute)) * 60_000)
    return max(min_ms, ms)


def build_scene_speech_plan(
    *,
    speech_segments: list[dict],
    pause_after_title_ms: int = 450,
    pause_between_segments_ms: int = 200,
    outro_hold_ms: int = 500,
) -> dict:
    current_ms = 0
    segments_out = []

    for idx, seg in enumerate(speech_segments):
        seg_kind = seg.get("kind", "narration")
        speech_text = compact_ws(seg.get("speech_text", ""))
        display_text = compact_ws(seg.get("display_text", ""))
        timing_text = speech_text or display_text

        if not timing_text:
            continue

        duration_ms = seg.get("audio_duration_ms")
        if not duration_ms:
            duration_ms = estimate_text_duration_ms(timing_text)

        start_ms = current_ms
        end_ms = start_ms + int(duration_ms)

        segments_out.append({
            "id": seg.get("id") or str(uuid.uuid4()),
            "kind": seg_kind,
            "display_text": display_text,
            "speech_text": speech_text,
            "show_on_screen": bool(seg.get("show_on_screen", True)),
            "animation": seg.get("animation", "fade"),
            "animation_target": seg.get("animation_target", "body"),
            "start_ms": start_ms,
            "end_ms": end_ms,
            "duration_ms": int(duration_ms),
            "order": seg.get("order", idx + 1),
        })

        current_ms = end_ms
        if seg_kind == "title":
            current_ms += pause_after_title_ms
        else:
            current_ms += pause_between_segments_ms

    total_duration_ms = current_ms + outro_hold_ms if segments_out else 0

    return {
        "mode": "sequential",
        "segments": segments_out,
        "total_duration_ms": total_duration_ms,
    }

# def build_scene_speech_plan(
#     *,
#     speech_segments: list[dict],
#     pause_after_title_ms: int = 450,
#     pause_between_segments_ms: int = 200,
#     outro_hold_ms: int = 500,
# ) -> dict:
#     current_ms = 0
#     segments_out = []

#     for idx, seg in enumerate(speech_segments):
#         seg_kind = seg.get("kind", "narration")
#         speech_text = compact_ws(seg.get("speech_text", ""))
#         if not speech_text:
#             continue

#         duration_ms = seg.get("audio_duration_ms")
#         if not duration_ms:
#             duration_ms = estimate_text_duration_ms(speech_text)

#         start_ms = current_ms
#         end_ms = start_ms + int(duration_ms)

#         segments_out.append({
#             "id": seg.get("id") or str(uuid.uuid4()),
#             "kind": seg_kind,
#             "display_text": seg.get("display_text", ""),
#             "speech_text": speech_text,
#             "show_on_screen": bool(seg.get("show_on_screen", True)),
#             "animation": seg.get("animation", "fade"),
#             "animation_target": seg.get("animation_target", "body"),
#             "start_ms": start_ms,
#             "end_ms": end_ms,
#             "duration_ms": int(duration_ms),
#             "order": seg.get("order", idx + 1),
#         })

#         current_ms = end_ms
#         if seg_kind == "title":
#             current_ms += pause_after_title_ms
#         else:
#             current_ms += pause_between_segments_ms

#     total_duration_ms = current_ms + outro_hold_ms if segments_out else 0

#     return {
#         "mode": "sequential",
#         "segments": segments_out,
#         "total_duration_ms": total_duration_ms,
#     }


def get_audio_duration_ms(audio_path: Path) -> int:
    try:
        return int(_duration_via_ffprobe(audio_path) * 1000)
    except Exception:
        try:
            return int(_duration_via_wave(audio_path) * 1000)
        except Exception:
            return 0


# @app.post("/api/projects/<project_id>/synthesize-speech")
# def synthesize_project_speech(project_id: str):
#     try:
#         project = load_project(project_id)
#     except FileNotFoundError as exc:
#         return jsonify({"ok": False, "error": str(exc)}), 404

#     speech_root = OUT_DIR / "composer_speech" / project_id
#     speech_root.mkdir(parents=True, exist_ok=True)

#     narration_cfg = project.get("narration", {}) or {}
#     voice = narration_cfg.get("voice", "alloy")
#     language = narration_cfg.get("language", "english")
#     engine = narration_cfg.get("engine", "google")
#     gender = narration_cfg.get("gender", "Male")

#     for idx, scene in enumerate(project.get("scenes", []), start=1):
#         _rebuild_scene_speech_segments(scene, idx)
#         speech_segments = scene.get("speech_segments", []) or []

#         for seg in speech_segments:
#             speech_text = compact_ws(seg.get("speech_text", ""))
#             if not speech_text:
#                 continue

#             seg_id = seg.get("id") or str(uuid.uuid4())
#             audio_file_name = f"{seg_id}.mp3"
#             audio_path = speech_root / audio_file_name
            
#             # ADAPT THIS LINE to your existing get_audio_file() signature
#             # Example idea only:
#             get_audio_file(text=speech_text, audio_file_name=audio_file_name, tts_engine="google", language="english", gender="Male")

#             # Copy audio_file_name file to speech_root
#             shutil.move(audio_file_name, audio_path)

#             if not audio_path.exists():
#                 return jsonify({
#                     "ok": False,
#                     "error": "TTS adapter line needs to be connected to your existing get_audio_file() signature."
#                 }), 500

#             seg["audio_path"] = str(audio_path)
#             seg["audio_duration_ms"] = get_audio_duration_ms(audio_path)

#         scene["speech_plan"] = build_scene_speech_plan(
#             speech_segments=speech_segments,
#             pause_after_title_ms=450,
#             pause_between_segments_ms=200,
#             outro_hold_ms=500,
#         )
#         scene["timing"]["estimated_duration_sec"] = round(scene["speech_plan"]["total_duration_ms"] / 1000.0, 2)

#     project_path(project_id).write_text(json.dumps(project, indent=2, ensure_ascii=False), encoding="utf-8")

#     return jsonify({"ok": True, "project": project})


@app.post("/api/projects/<project_id>/synthesize-speech")
def synthesize_project_speech(project_id: str):
    try:
        project = load_project(project_id)
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404

    speech_root = OUT_DIR / "composer_speech" / project_id
    speech_root.mkdir(parents=True, exist_ok=True)

    narration_cfg = project.get("narration", {}) or {}
    voice = (narration_cfg.get("voice") or "alloy").strip() or "alloy"

    language = (narration_cfg.get("language") or "english").strip().lower() or "english"
    if language == "en":
        language = "english"

    engine = (narration_cfg.get("engine") or "google").strip().lower() or "google"
    if engine == "tts":
        engine = "google"

    gender = (narration_cfg.get("gender") or "Male").strip() or "Male"

    meta_path = speech_root / "_audio_meta.json"
    if meta_path.exists():
        try:
            audio_meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            audio_meta = {}
    else:
        audio_meta = {}

    generated_count = 0
    reused_count = 0
    display_only_count = 0

    for idx, scene in enumerate(project.get("scenes", []), start=1):
        _rebuild_scene_speech_segments(scene, idx)
        speech_segments = scene.get("speech_segments", []) or []

        for seg in speech_segments:
            speech_text = compact_ws(seg.get("speech_text", ""))
            seg_id = seg.get("id") or str(uuid.uuid4())
            seg["id"] = seg_id
            audio_file_name = f"{seg_id}.mp3"
            audio_path = speech_root / audio_file_name

            # display-only segment: keep timing but do not generate TTS

            if not speech_text:
                seg.pop("audio_file", None)
                seg.pop("audio_cache_key", None)
                seg["audio_duration_ms"] = 0
                audio_meta.pop(seg_id, None)   # add this
                display_only_count += 1
                continue

            # if not speech_text:
            #     seg.pop("audio_file", None)
            #     seg.pop("audio_cache_key", None)
            #     seg["audio_duration_ms"] = 0
            #     display_only_count += 1
            #     continue


            cache_key = make_audio_cache_key(
                speech_text,
                voice=voice,
                language=language,
                engine=engine,
                gender=gender,
            )

            meta_entry = audio_meta.get(seg_id)
            if isinstance(meta_entry, dict):
                existing_cache_key = meta_entry.get("cache_key")
                existing_duration_ms = int(meta_entry.get("audio_duration_ms") or 0)
            else:
                existing_cache_key = meta_entry
                existing_duration_ms = 0

            if audio_path.exists() and existing_cache_key == cache_key:
                seg["audio_file"] = str(audio_path)
                seg["audio_cache_key"] = cache_key
                seg["audio_duration_ms"] = existing_duration_ms or get_audio_duration_ms(audio_path)
                audio_meta[seg_id] = {
                    "cache_key": cache_key,
                    "audio_duration_ms": int(seg.get("audio_duration_ms") or 0),
                }
                reused_count += 1
                continue

            temp_audio_path = Path(audio_file_name)
            try:
                if temp_audio_path.exists():
                    temp_audio_path.unlink()
            except Exception:
                pass

            get_audio_file(
                text=speech_text,
                audio_file_name=audio_file_name,
                tts_engine=engine,
                language=language,
                gender=gender,
            )

            if temp_audio_path.resolve() != audio_path.resolve():
                if audio_path.exists():
                    try:
                        audio_path.unlink()
                    except Exception:
                        pass
                shutil.move(str(temp_audio_path), str(audio_path))

            if not audio_path.exists():
                return jsonify({
                    "ok": False,
                    "error": "TTS audio file was not created.",
                }), 500

            duration_ms = get_audio_duration_ms(audio_path)
            seg["audio_file"] = str(audio_path)
            seg["audio_cache_key"] = cache_key
            seg["audio_duration_ms"] = duration_ms
            audio_meta[seg_id] = {
                "cache_key": cache_key,
                "audio_duration_ms": duration_ms,
            }
            generated_count += 1

        speech_plan = build_scene_speech_plan(
            speech_segments=speech_segments,
            pause_after_title_ms=450,
            pause_between_segments_ms=200,
            outro_hold_ms=500,
        )

        timing = scene.get("timing", {}) or {}
        timing["estimated_duration_sec"] = round(speech_plan["total_duration_ms"] / 1000.0, 2)
        scene["timing"] = timing
        scene["speech_plan"] = speech_plan

    meta_path.write_text(json.dumps(audio_meta, indent=2, ensure_ascii=False), encoding="utf-8")
    save_project(project_id, project)

    summary = {
        "generated": generated_count,
        "reused": reused_count,
        "display_only": display_only_count,
    }
    message = f"TTS completed. Generated {generated_count}, reused {reused_count}."

    return jsonify({"ok": True, "project": project, "summary": summary, "message": message})


@app.post("/api/projects/<project_id>/speech-plan")
def update_project_speech_plan(project_id: str):
    try:
        project = load_project(project_id)
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404

    scenes = project.get("scenes", [])
    updated_scenes = []

    for idx, scene in enumerate(scenes, start=1):
        _rebuild_scene_speech_segments(scene, idx)
        speech_segments = scene.get("speech_segments", []) or []
        speech_plan = build_scene_speech_plan(
            speech_segments=speech_segments,
            pause_after_title_ms=450,
            pause_between_segments_ms=200,
            outro_hold_ms=500,
        )

        timing = scene.get("timing", {}) or {}
        timing["estimated_duration_sec"] = round(speech_plan["total_duration_ms"] / 1000.0, 2)
        scene["timing"] = timing
        scene["speech_plan"] = speech_plan
        updated_scenes.append(scene)

    project["scenes"] = updated_scenes
    project_path(project_id).write_text(json.dumps(project, indent=2, ensure_ascii=False), encoding="utf-8")

    return jsonify({"ok": True, "project": project})

@app.get("/api/projects")
def list_projects():
    items = []
    for path in sorted(PROJECTS_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            items.append({
                "id": data.get("id"),
                "title": data.get("title"),
                "theme_id": data.get("theme_id"),
                "scene_count": len(data.get("scenes", [])),
                "source_type": data.get("source_type"),
                "source_value": data.get("source_value"),
            })
        except Exception:
            continue
    return jsonify({"ok": True, "projects": items})

@app.get("/html-video-composer")
def html_video_composer():
    return render_template("html_video_composer.html")

@app.get("/html-video-composer_deep")
def html_video_composer_deep():
    return render_template("html_video_composer_deep.html")

from flask import render_template, send_from_directory, jsonify


def composer_speech_dir(project_id: str) -> Path:
    return OUT_DIR / "composer_speech" / project_id


def build_playback_payload(project: dict) -> dict:
    assets = project.get("assets", []) or []
    asset_map = {a.get("id"): a for a in assets if a.get("id")}

    scenes_out = []
    for scene_index, scene in enumerate(project.get("scenes", []), start=1):
        media_asset_ids = scene.get("media_asset_ids", []) or []
        bg_asset = asset_map.get(media_asset_ids[0]) if media_asset_ids else None

        speech_plan = scene.get("speech_plan", {}) or {}
        segments = speech_plan.get("segments", []) or []

        if not segments:
            raw_segments = scene.get("speech_segments", []) or []
            if raw_segments:
                rebuilt_plan = build_scene_speech_plan(
                    speech_segments=raw_segments,
                    pause_after_title_ms=450,
                    pause_between_segments_ms=200,
                    outro_hold_ms=500,
                )
                speech_plan = rebuilt_plan
                segments = rebuilt_plan.get("segments", []) or []

        playback_segments = []
        total_duration_ms = int(speech_plan.get("total_duration_ms") or 0)
        for seg in segments:
            seg_id = seg.get("id")
            seg_kind = seg.get("kind", "narration")
            audio_url = None
            if seg_id and seg.get("speech_text", "").strip():
                audio_url = f"/composer-speech/{project.get('id')}/{seg_id}.mp3"

            duration_ms = int(seg.get("duration_ms") or 0)
            end_ms = int(seg.get("end_ms") or 0)
            total_duration_ms = max(total_duration_ms, end_ms)

            playback_segments.append({
                "id": seg_id,
                "kind": seg_kind,
                "display_text": seg.get("display_text", ""),
                "speech_text": seg.get("speech_text", ""),
                "show_on_screen": bool(seg.get("show_on_screen", True)),
                "animation": seg.get("animation", "fade"),
                "animation_target": seg.get("animation_target", "body"),
                "start_ms": int(seg.get("start_ms") or 0),
                "end_ms": end_ms,
                "duration_ms": duration_ms,
                "audio_url": audio_url,
            })

        scenes_out.append({
            "index": scene_index,
            "scene_id": scene.get("id"),
            "scene_type": scene.get("scene_type"),
            "title": scene.get("title", ""),
            "subtitle": scene.get("subtitle", ""),
            "bullets": scene.get("bullets", []),
            "layout": scene.get("layout", {}),
            "timing": scene.get("timing", {}),
            "background_asset": bg_asset,
            "segments": playback_segments,
            "total_duration_ms": total_duration_ms,
        })

    return {
        "project_id": project.get("id"),
        "title": project.get("title", "HTML Video Composer Project"),
        "theme_id": project.get("theme_id", "corporate-clean"),
        "project_settings": project.get("project_settings", {}) or {},
        "scenes": scenes_out,
    }


@app.get("/html-video-composer/player/<project_id>")
def html_video_composer_player(project_id: str):
    return render_template("html_video_composer_player.html", project_id=project_id)


@app.get("/api/projects/<project_id>/playback")
def get_project_playback(project_id: str):
    try:
        project = load_project(project_id)
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404

    return jsonify({
        "ok": True,
        "playback": build_playback_payload(project),
    })


@app.get("/composer-speech/<project_id>/<path:filename>")
def composer_speech_files(project_id: str, filename: str):
    return send_from_directory(str(composer_speech_dir(project_id)), filename)

####################################################
# END: HTML to video maker enhancements/changes
####################################################

if __name__ == '__main__':
    # app.run(debug=True, port=5000)

    # DND - Working but not in use
    # from gemini_pool import GeminiPool
    # GEM_STATE = str((Path(__file__).resolve().parent / ".gemini_pool_state.json"))
    # gemini_pool = GeminiPool(
    #     api_keys=None,
    #     per_key_rpm=25,
    #     state_path=GEM_STATE,
    #     autosave_every=3,
    # )   
    #  
    app.run(debug=True, host='0.0.0.0', port=5000)  # Use host='
