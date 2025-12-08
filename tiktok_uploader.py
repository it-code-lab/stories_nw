# tiktok_uploader.py

import os
import time
from pathlib import Path
from datetime import datetime

from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

# ================== CONFIG ==================

TIKTOK_UPLOAD_URL = "https://www.tiktok.com/tiktokstudio/upload?from=webapp"
#TIKTOK_UPLOAD_URL = "https://www.tiktok.com"

# Re-use the same Chrome profile you use for Pinterest
PROFILE_DIR = r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 21"
CHROME_EXECUTABLE = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

# Re-use the same Excel for now
EXCEL_FILE = "master_shorts_uploader_data.xlsx"

BASE_DIR = Path(__file__).resolve().parent
MEDIA_BASE = BASE_DIR / "pinterest_uploads"   # where media_file paths are relative to

VIDEO_EXTS = {".mp4", ".mov", ".mkv", ".webm"}


# ================== EXCEL HELPERS ==================

def ensure_tiktok_status_columns(ws, headers):
    """
    Ensure these columns exist:
      - tiktok_video_url
      - tiktok_upload_status
      - tiktok_uploaded_at
      - tiktok_caption   (optional override per row)
    """
    extra_cols = [
        "tiktok_video_url",
        "tiktok_upload_status",
        "tiktok_uploaded_at",
        "tiktok_caption",
    ]
    changed = False
    for col_name in extra_cols:
        if col_name not in headers:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers), value=col_name)
            changed = True

    header_map = {name: idx + 1 for idx, name in enumerate(headers)}
    return headers, header_map, changed


def load_videos_from_excel():
    """
    Load rows from EXCEL_FILE that have a video in 'media_file'.

    Returns: wb, ws, rows, header_map
    """
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    headers = [h if h is not None else "" for h in headers]

    headers, header_map, _ = ensure_tiktok_status_columns(ws, headers)

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)}
        record["_row_idx"] = row_idx

        media_file = (record.get("media_file") or "").strip()
        status_val = (record.get("tiktok_upload_status") or "").strip()

        if not media_file:
            continue

        if status_val == "success":
            continue

        ext = Path(media_file).suffix.lower()
        media_type = (record.get("media_type") or "").strip().lower()

        # Only take video rows
        if media_type and media_type != "video":
            continue
        if ext and ext not in VIDEO_EXTS:
            continue

        rows.append(record)

    return wb, ws, rows, header_map


def save_tiktok_status(ws, header_map, row_idx, url, status, caption=""):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    url_col = header_map.get("tiktok_video_url")
    status_col = header_map.get("tiktok_upload_status")
    ts_col = header_map.get("tiktok_uploaded_at")
    cap_col = header_map.get("tiktok_caption")

    if url_col:
        ws.cell(row=row_idx, column=url_col, value=url)
    if status_col:
        ws.cell(row=row_idx, column=status_col, value=status)
    if ts_col:
        ws.cell(row=row_idx, column=ts_col, value=now_str)
    if cap_col and caption:
        # Save the final caption we actually used (nice for reference)
        ws.cell(row=row_idx, column=cap_col, value=caption)


def is_excel_file_locked(file_path):
    try:
        with open(file_path, "a"):
            pass
        return False
    except PermissionError:
        return True


# ================== TIKTOK HELPERS ==================

def resolve_media_path(media_file: str) -> str:
    """
    Resolve media_file path relative to MEDIA_BASE.
    """
    rel = media_file.replace("\\", "/")
    full_path = (MEDIA_BASE / rel).resolve()
    return str(full_path)


def open_tiktok_upload(page):
    page.goto(TIKTOK_UPLOAD_URL)
    page.wait_for_load_state("networkidle")
    time.sleep(3)


def upload_video(page, media_path: str):
    """
    Upload video file on TikTok Studio upload page.

    NOTE: Selectors may need small adjustment using Playwright inspector.
    """
    print(f"  -> Uploading video: {media_path}")

    # Common pattern: input[type=file][accept*="video"]
    file_input = page.locator('input[type="file"][accept*="video"]')
    if file_input.count() == 0:
        # Fallback: any file input
        file_input = page.locator('input[type="file"]')

    count = file_input.count()
    if count == 0:
        raise RuntimeError("Could not find video upload input on TikTok page.")

    target = file_input.first
    if count > 1:
        # Prefer visible one
        for i in range(count):
            el = file_input.nth(i)
            if el.is_visible():
                target = el
                break

    target.set_input_files(media_path)

    # Give TikTok some time to upload & process
    # (you can later replace with a smarter wait e.g. for a thumbnail or 'Caption' to appear)
    time.sleep(20)


def build_caption(row: dict) -> str:
    """
    Build a TikTok caption from row data if tiktok_caption is not provided.
    TikTok limit is 2,200 chars; we keep it well under that.
    """
    manual = (row.get("tiktok_caption") or "").strip()
    if manual:
        caption = manual
        link = (
            (row.get("pin_url_to_link") or "") or
            (row.get("book_url") or "")
        ).strip()

        if link:
            if caption:
                caption += "\n\n"
            caption += link
    else:
        title = (row.get("pin_title") or "").strip()
        desc = (row.get("pin_description") or "").strip()
        tags = (row.get("pin_tags") or "").strip()
        url = (
            (row.get("pin_url_to_link") or "")
            or (row.get("book_url") or "")
        ).strip()

        parts = []
        if title:
            parts.append(title)
        if desc:
            parts.append(desc)
        if tags:
            parts.append(tags)
        if url:
            parts.append(url)

        caption = "\n\n".join(parts)

    # Safety: limit to 2,200 chars
    if len(caption) > 2200:
        caption = caption[:2190].rstrip() + "…"

    return caption

def get_caption_editor(page, caption_sample: str | None = None):
    """
    Try several locators to find the TikTok caption editor.

    caption_sample: optional text we expect to be inside the combobox
                    (e.g., current auto-filled filename or caption).
    """
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

    candidates = []

    # 1) Your recorded locator: getByRole('combobox').filter({ hasText: '...' }).locator('div').nth(2)
    if caption_sample:
        snippet = caption_sample[:60]  # use a shorter piece to avoid exact-match issues
        candidates.append(
            page.get_by_role("combobox")
                .filter(has_text=snippet)
                .locator("div")
                .nth(2)
        )

    # 2) Element with .caption-editor class (from your screenshot)
    candidates.append(page.locator(".caption-editor").first)

    # 3) data-e2e variant TikTok often uses
    candidates.append(
        page.locator('[data-e2e="caption-editor"] [contenteditable="true"]').first
    )
    candidates.append(
        page.locator('[data-e2e="caption-editor"]').first
    )

    # 4) Fallback: any visible contenteditable div (last resort)
    candidates.append(page.locator('div[contenteditable="true"]').first)

    for loc in candidates:
        try:
            loc.wait_for(state="visible", timeout=10000)
            if loc.count() > 0 and loc.is_visible():
                return loc
        except PlaywrightTimeoutError:
            continue
        except Exception:
            continue

    return None


def fill_caption(page, caption_text: str):
    """
    Fill the caption on TikTok upload page.

    Uses generic selectors because TikTok UI changes frequently.
    Adjust once with Playwright inspector if needed.
    """
    caption_text = (caption_text or "").strip()
    if not caption_text:
        print("  [INFO] No caption text provided; skipping caption fill.")
        return

    print("  -> Filling caption...")

    # We pass a small sample so the combobox filter can match the current text (e.g., filename)
    sample = caption_text[:60]
    editor = get_caption_editor(page, caption_sample=sample)

    if not editor:
        print("❌ Could not find caption editor on TikTok page.")
        return

    editor.click()
    time.sleep(1)

    # Clear existing text
    try:
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
    except Exception:
        pass

    time.sleep(0.2)
    page.keyboard.type(caption_text, delay=5)


def click_post(page):
    """
    Click the 'Post' button to publish the video.
    """
    print("  -> Clicking Post...")

    import re

    try:
        btn = page.get_by_role("button", name=re.compile(r"Post", re.I))
        btn.click()
    except Exception as e:
        print(f"[ERROR] Could not click Post button: {e}")
        raise

    # Give TikTok some time to publish & redirect
    time.sleep(20)


def extract_video_url(page) -> str:
    """
    After posting, try to grab the video URL.

    We look at current URL first, then any link containing '/video/'.
    """
    url = page.url
    if "/video/" in url:
        return url

    try:
        link = page.locator('a[href*="/video/"]').first
        if link:
            href = link.get_attribute("href")
            if href:
                if href.startswith("http"):
                    return href
                else:
                    return "https://www.tiktok.com" + href
    except Exception:
        pass

    return "Unknown"


# ================== MAIN UPLOADER ==================


def upload_tiktok_videos():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file not found: {EXCEL_FILE}")
        return

    if is_excel_file_locked(EXCEL_FILE):
        print(f"Error: Please close '{EXCEL_FILE}' before running the uploader.")
        return

    wb, ws, rows, header_map = load_videos_from_excel()

    if not rows:
        print("No valid video rows found in Excel.")
        return

    print(f"Loaded {len(rows)} TikTok rows from {EXCEL_FILE}")

    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            PROFILE_DIR,
            headless=False,
            executable_path=CHROME_EXECUTABLE,
            args=["--disable-blink-features=AutomationControlled", "--start-maximized"],
        )
        page = browser.new_page()

        # Hide automation fingerprint
        page.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )

        # You should already be logged in to TikTok in this Chrome profile.

        # total = len(rows)
        for row in rows: 
            # row_idx = total - rev_idx + 2  # Excel row index (header is row 1)
            row_idx = row["_row_idx"]

            status_val = str(row.get("tiktok_upload_status") or "").strip().lower()
            if status_val == "success":
                print(f"Skipping already uploaded TikTok: {row.get('pin_title')}")
                continue

            try:
                print(f"\n=== TikTok upload for media: {row.get('media_file')} (row {row_idx}) ===")

                open_tiktok_upload(page)
                # time.sleep(3000)
                media_path = resolve_media_path(str(row["media_file"]))
                if not os.path.exists(media_path):
                    raise FileNotFoundError(f"Media file not found: {media_path}")


                # Wait for the main upload button/area to be visible
                # The 'Select file' button is often the key stable element
                upload_selector = page.get_by_text("Select video", exact=True)
                upload_selector.wait_for(state="visible", timeout=30000)
                print("Upload page loaded successfully.")

                # --- STEP 2: FILE UPLOAD ---
                # TikTok Studio uses an HTML input element hidden behind the UI, 
                # which we target directly using set_input_files on the relevant area.
                
                # Find the hidden file input element (usually <input type="file">)
                # This is one of the most stable selectors for the TikTok upload interface.
                file_input_locator = page.locator('input[type="file"]')
                
                print(f"Uploading file: {media_path}")
                
                # Use set_input_files directly on the file input
                file_input_locator.set_input_files(media_path)
                
                # Wait for the video to finish uploading/processing (Title/Desc fields appear)
                # page.get_by_role("textbox", name="Title").wait_for(state="visible", timeout=60000)
                # print("Video uploaded and processing started.")

                # --- STEP 3: FILL DETAILS ---
                caption = build_caption(row)
                fill_caption(page, caption)
                print("Title and Description filled.")

                # Optional: Select the video cover (this step is complex and skipped for simplicity)

                # --- STEP 4: POST THE VIDEO ---
                
                # Locate the 'Post' button. It's usually enabled after processing is done.
                # print("Waiting for Post button to become active...")
                # post_button = page.get_by_role("button", name="Post", exact=True)
                
                post_button = page.locator('[data-e2e="post_video_button"]').first
                # Wait for the button to be clickable (not disabled)
                # post_button.wait_for(state="enabled", timeout=60000)
                
                print("Clicking Post...")
                post_button.click()
                
                # --- STEP 5: VERIFICATION ---
                
                # After clicking Post, a success modal should appear
                # page.get_by_role("heading", name="Your videos have been uploaded").wait_for(timeout=15000)
                # print("✅ Upload Successful! Videos have been uploaded.")

                # upload_video(page, media_path)



                # click_post(page)
                # video_url = extract_video_url(page)
                time.sleep(5)
                save_tiktok_status(ws, header_map, row_idx, "", "Success", caption)
                print(f"✅ TikTok video posted & recorded for row {row_idx}")

            except Exception as e:
                err_msg = str(e)[:500]
                print(f"❌ Error uploading TikTok for row {row_idx}: {err_msg}")
                save_tiktok_status(ws, header_map, row_idx, "", err_msg)

            # Gentle pause between posts
            time.sleep(5)

        wb.save(EXCEL_FILE)
        browser.close()

    print("All TikTok uploads done.")


if __name__ == "__main__":
    upload_tiktok_videos()
