# assemble_from_videos.py
import shutil
import os, random, math, subprocess, tempfile, json
from glob import glob
from typing import List, Dict, Tuple
from moviepy.editor import AudioFileClip, VideoFileClip, concatenate_videoclips, CompositeAudioClip
import re   
import tempfile
import os
import subprocess
import shlex
import tempfile, os

def _escape_drawtext(s: str) -> str:
    # For Windows paths in filtergraphs: 
    # 1. Replace backslash with 4 backslashes (for FFmpeg's nested parsing)
    # 2. Escape the colon (C\: instead of C:)
    return (s.replace("\\", "\\\\")
             .replace(":", "\\:")
             .replace("'", "'\\''")
             .replace("%", "\\%"))

def _escape_drawtext_old(s: str) -> str:
    # drawtext escaping: \, :, ', and %
    return (s.replace("\\", "\\\\")
             .replace(":", "\\:")
             .replace("'", "\\'")
             .replace("%", "\\%"))

def make_title_card(out_path, title, w, h, fps=30, dur=3.5, font="C:/Windows/Fonts/arial.ttf"):
    import subprocess
    # IMPORTANT: use forward slashes for ffmpeg drawtext fontfile on Windows
    font = font.replace("\\", "/")
    # escape single quotes in title for drawtext
    safe_title = title.replace("'", r"\'")
    cmd = [
        "ffmpeg", "-y",
        "-f", "lavfi", "-i", f"color=c=#101010:s={w}x{h}:d={dur}",
        "-f", "lavfi", "-i", f"anullsrc=r=48000:cl=stereo:d={dur}",
        "-vf",
        (
            "drawtext="
            f"fontfile='{font}':"
            f"text='Next: {safe_title}':"
            "fontsize=64:fontcolor=white:"
            "x=(w-text_w)/2:y=(h-text_h)/2:"
            "box=1:boxcolor=black@0.35:boxborderw=40,"
            "fade=t=in:st=0:d=0.25,fade=t=out:st=3.25:d=0.25"
        ),
        "-r", str(fps),
        "-c:v", "libx264", "-pix_fmt", "yuv420p",
        "-c:a", "aac",
        "-shortest",
        out_path
    ]
    subprocess.run(cmd, check=True)

def _make_title_card_ffmpeg(
    out_path: str,
    title: str,
    w: int,
    h: int,
    fps: int,
    dur: float,
    fontfile: str | None = None,
    bg: str = "#101010",              # ✅ accepts "bg" now
    fontsize: int = 64,
    fontcolor: str = "white",         # ✅ accepts "fontcolor"
    box_alpha: float = 0.35,          # ✅ accepts "box_alpha"
    box_border: int = 40,             # ✅ accepts "box_border"
    fade_sec: float = 0.25,           # ✅ accepts "fade_sec"
):
    """
    Create a short silent title card MP4 using FFmpeg (lavfi color + anullsrc).
    """
    if not _bin_exists("ffmpeg"):
        raise RuntimeError("ffmpeg not found in PATH; cannot generate title cards.")

    # Default font on Windows
    if not fontfile:
        fontfile = r"C:\Windows\Fonts\arial.ttf"

    # FFmpeg drawtext: forward slashes and escape drive colon (C\:)
    ff_font = fontfile.replace("\\", "/").replace(":", "\\:")

    # Escape title for drawtext (keep it conservative)
    # safe_title = title.replace("'", r"\'").replace("%", r"\%")
    safe_title = _escape_drawtext(title)

    # Fade in/out inside the title card
    fade_in = max(0.0, min(fade_sec, dur / 2.0))
    fade_out = fade_in
    fade_out_start = max(dur - fade_out, 0.0)

    # write title to a UTF-8 file (FFmpeg reads it correctly)
    tmp_txt = os.path.join(tempfile.gettempdir(), f"title_{os.getpid()}.txt")
    with open(tmp_txt, "w", encoding="utf-8") as f:
        f.write(title)

    # escape only for *paths* inside filtergraph
    ff_txt = tmp_txt.replace("\\", "/").replace(":", "\\:")

    vf = (
        f"drawtext=fontfile='{ff_font}':"
        f"textfile='{ff_txt}':"
        f"fontsize={int(fontsize)}:"
        f"fontcolor={fontcolor}:"
        f"x=(w-text_w)/2:y=(h-text_h)/2:"
        f"box=1:boxcolor=black@{box_alpha}:boxborderw={int(box_border)},"
        f"fade=t=in:st=0:d={fade_in:.3f},"
        f"fade=t=out:st={fade_out_start:.3f}:d={fade_out:.3f}"
    )

    cmd = [
        "ffmpeg", "-y",
        "-f", "lavfi", "-i", f"color=c={bg}:s={w}x{h}:d={dur}",
        "-f", "lavfi", "-i", f"anullsrc=r=48000:cl=stereo:d={dur}",
        "-vf", vf,
        "-r", str(fps),
        "-c:v", "libx264", "-pix_fmt", "yuv420p",
        "-c:a", "aac",
        "-shortest",
        out_path
    ]
    print("▶ Title card cmd:", " ".join(cmd))
    subprocess.run(cmd, check=True)

    # (optional cleanup at end)
    try:
        os.remove(tmp_txt)
    except:
        pass
    
def _ffmpeg_concat_reencode(
    plan,
    out_path: str,
    fps: int,
    w: int,
    h: int,
    keep_audio: bool,
    crf: int = 18,
    preset: str = "veryfast",
):
    """
    Concatenate clips (and optional silent title cards) via FFmpeg filter concat.
    This avoids timing tricks (adelay/xfade) and keeps each clip's internal sync intact.
    plan items: (path, full_d, use_d)
    """
    if not _bin_exists("ffmpeg"):
        raise RuntimeError("ffmpeg not found in PATH; cannot concatenate with re-encode.")

    n = len(plan)
    cmd = ["ffmpeg", "-y"]

    for (p, _full_d, use_d) in plan:
        cmd += ["-t", f"{use_d:.3f}", "-i", p]

    fc = []
    v_in = []
    a_in = []

    for i in range(n):
        fc.append(
            f"[{i}:v]"
            f"scale={w}:{h}:force_original_aspect_ratio=decrease,"
            f"pad={w}:{h}:(ow-iw)/2:(oh-ih)/2,"
            f"setsar=1,format=yuv420p"
            f"[v{i}]"
        )
        v_in.append(f"[v{i}]")
        if keep_audio:
            # If an input has no audio stream, FFmpeg will error. Guard earlier in assemble_videos.
            fc.append(f"[{i}:a]aformat=sample_rates=48000:channel_layouts=stereo,aresample=48000[a{i}]")
            a_in.append(f"[a{i}]")

    if keep_audio:
        # Interleave video and audio labels: [v0][a0][v1][a1]...
        interleaved_labels = ""
        for i in range(n):
            interleaved_labels += f"[v{i}][a{i}]"
        fc.append(interleaved_labels + f"concat=n={n}:v=1:a=1[vout][aout]")
    else:
        fc.append("".join(v_in) + f"concat=n={n}:v=1:a=0[vout]")

    cmd += [
        "-filter_complex", ";".join(fc),
        "-map", "[vout]",
    ]
    if keep_audio:
        cmd += ["-map", "[aout]"]

    cmd += [
        "-r", str(fps),
        "-c:v", "libx264", "-crf", str(crf), "-preset", preset,
    ]
    if keep_audio:
        cmd += ["-c:a", "aac", "-b:a", "192k"]
    else:
        cmd += ["-an"]

    cmd += [out_path]
    print("▶ Concat re-encode cmd:", " ".join(cmd))
    subprocess.run(cmd, check=True)

def _build_titles_transitions_cmd(
    plan,
    output_path: str,
    audio_path: str | None,
    fps: int,
    add_titles: bool,
    title_sec: float,
    use_transitions: bool,
    trans_sec: float,
    keep_video_audio: bool,
    video_volume: float,
    bg_volume: float,
    crf: int = 18,
    preset: str = "veryfast",
    fontfile: str | None = None,
):
    n = len(plan)
    cmd = ["ffmpeg", "-y"]
    for (p, _full_d, use_d) in plan:
        cmd += ["-t", f"{use_d:.3f}", "-i", p]

    bg_audio_idx = None
    if audio_path:
        bg_audio_idx = n
        cmd += ["-i", audio_path]

    fc = []
    vlabels = []
    alabels = []
    
    # 1. Setup raw streams
    for i in range(n):
        v, a = f"v{i}", f"a{i}"
        vlabels.append(v)
        alabels.append(a)
        fc.append(f"[{i}:v]setpts=PTS-STARTPTS,fps={fps},format=yuv420p[{v}]")
        # Apply the "breathing room" delay to the audio source directly
        delay_ms = 1000 if i > 0 else 0
        fc.append(f"[{i}:a]asetpts=PTS-STARTPTS,adelay={delay_ms}|{delay_ms}[{a}]")

    # 2. Synchronized Transitions (Video AND Audio)
    if use_transitions and n > 1:
        cur_v, cur_a = vlabels[0], alabels[0]
        total_time = plan[0][2]
        
        for i in range(1, n):
            nxt_v, nxt_a = vlabels[i], alabels[i]
            offset = max(total_time - trans_sec, 0.0)
            
            out_v, out_a = f"vx{i}", f"ax{i}"
            
            # Video Fade
            fc.append(f"[{cur_v}][{nxt_v}]xfade=transition=fadeblack:duration={trans_sec}:offset={offset:.3f}[{out_v}]")
            
            # Audio Alignment: Use amix with start_time to match the video offset
            # This ensures Clip 2's audio starts exactly when the video transition starts
            fc.append(f"[{cur_a}][{nxt_a}]amix=inputs=2:duration=first:dropout_transition=0,atrim=start=0[temp_a{i}];")
            # We use a more robust way to chain audio for xfade: 
            # For simplicity in complex scripts, we can use the 'acrossfade' filter
            fc.append(f"[{cur_a}][{nxt_a}]acrossfade=d={trans_sec}:c1=fade:c2=fade[{out_a}]")
            
            cur_v, cur_a = out_v, out_a
            total_time += plan[i][2] - trans_sec
            
        vout, aout_final = cur_v, cur_a
    else:
        # Simple concat if no transitions
        v_ins = "".join([f"[{v}]" for v in vlabels])
        a_ins = "".join([f"[{a}]" for a in alabels])
        fc.append(f"{v_ins}concat=n={n}:v=1:a=0[vout]")
        fc.append(f"{a_ins}concat=n={n}:v=0:a=1[aout_final]")
        vout, aout_final = "vout", "aout_final"

    # 3. Titles (Aligned to the combined timeline)
    final_v = vout
    if add_titles:
        t_cursor = 0.0
        for i, (p, _full_d, use_d) in enumerate(plan):
            title = _escape_drawtext(os.path.splitext(os.path.basename(p))[0])
            
            # Start title after transition + small buffer
            start = t_cursor + (trans_sec if i > 0 else 0.5)
            end = start + title_sec
            
            font = f":fontfile={_escape_drawtext(fontfile)}" if fontfile else ""
            outv = f"vt{i}"
            fc.append(f"[{final_v}]drawtext=text='{title}'{font}:x=60:y=80:fontsize=56:fontcolor=white:box=1:boxcolor=black@0.45:boxborderw=20:enable='between(t,{start:.3f},{end:.3f})'[{outv}]")
            final_v = outv
            
            t_cursor += (use_d - trans_sec) if i < n-1 else use_d

    # 4. Final Mixing
    map_args = ["-map", f"[{final_v}]"]
    if audio_path:
        fc.append(f"[{aout_final}]volume={video_volume}[va];[{bg_audio_idx}:a]volume={bg_volume}[ba];[va][ba]amix=inputs=2:normalize=1[afinal]")
        map_args += ["-map", "[afinal]"]
    else:
        map_args += ["-map", f"[{aout_final}]"]

    cmd += ["-filter_complex", ";".join(fc)] + map_args + ["-c:v", "libx264", "-crf", str(crf), "-preset", preset, "-c:a", "aac", "-shortest", output_path]
    return cmd

def _build_titles_transitions_cmd_old(
    plan,                 # [(path, full_d, use_d), ...]
    output_path: str,
    audio_path: str | None,
    fps: int,
    add_titles: bool,
    title_sec: float,
    use_transitions: bool,
    trans_sec: float,
    keep_video_audio: bool,
    video_volume: float,
    bg_volume: float,
    crf: int = 18,
    preset: str = "veryfast",
    fontfile: str | None = None,
):
    """
    FFmpeg filter graph that:
    - trims each input to use_d (we pass -t per input)
    - applies xfade between clips (optional)
    - overlays filename titles for first title_sec of each clip (optional)
    - audio:
        - if bg audio exists => uses bg audio, optionally mixes original audio
        - if bg audio missing => optionally keeps original audio, otherwise no audio
    NOTE: This path re-encodes (required for titles/transitions).
    """

    n = len(plan)
    if n == 0:
        raise RuntimeError("Empty plan")

    cmd = ["ffmpeg", "-y"]

    # inputs: each clip trimmed to use_d via -t
    for (p, _full_d, use_d) in plan:
        cmd += ["-t", f"{use_d:.3f}", "-i", p]

        # cmd += ["-i", p, "-t", f"{use_d:.3f}"]

    # optional bg audio as last input
    bg_audio_idx = None
    if audio_path:
        bg_audio_idx = n
        cmd += ["-i", audio_path]

    # --- filter_complex for video ---
    fc = []

    # normalize each clip video stream
    vlabels = []
    for i in range(n):
        v = f"v{i}"
        vlabels.append(v)
        fc.append(f"[{i}:v]setpts=PTS-STARTPTS,fps={fps},format=yuv420p[{v}]")

    # chain video with or without transitions
    if use_transitions and n > 1:
        cur = vlabels[0]
        total = plan[0][2]  # accumulated timeline length
        for i in range(1, n):
            nxt = vlabels[i]
            offset = max(total - trans_sec, 0.0)
            out = f"vx{i}"
            fc.append(
                f"[{cur}][{nxt}]xfade=transition=fade:duration={trans_sec}:offset={offset:.3f}[{out}]"
            )
            cur = out
            total += plan[i][2] - trans_sec  # overlap removes trans duration
        vout = cur
        timeline_total = total
    else:
        ins = "".join([f"[{vlabels[i]}]" for i in range(n)])
        fc.append(f"{ins}concat=n={n}:v=1:a=0[vcat]")
        vout = "vcat"
        timeline_total = sum(x[2] for x in plan)

    # --- optional titles ---
    final_v = vout
    if add_titles:
        t_cursor = 0.0
        for i, (p, _full_d, use_d) in enumerate(plan):
            title = os.path.splitext(os.path.basename(p))[0]
            title = _escape_drawtext(title)

            start = t_cursor
            end = t_cursor + min(title_sec, use_d)

            font = f":fontfile={_escape_drawtext(fontfile)}" if fontfile else ""
            outv = f"vt{i}"

            # lower-third style
            fc.append(
                f"[{final_v}]drawtext=text='{title}'{font}:"
                f"x=60:y=80:fontsize=56:fontcolor=white:"
                f"box=1:boxcolor=black@0.45:boxborderw=20:"
                f"enable='between(t,{start:.3f},{end:.3f})'[{outv}]"
            )
            final_v = outv

            t_cursor += use_d
            if use_transitions and n > 1:
                t_cursor -= trans_sec  # account for overlap

    # --- audio handling ---
    # cases:
    # 1) bg audio exists:
    #    - if keep_video_audio and clips have audio => mix clip audio + bg audio
    #    - else => bg audio only
    # 2) bg audio missing:
    #    - if keep_video_audio => concat original audio (best effort)
    #    - else => no audio

    map_args = []
    out_args = []

    map_args += ["-map", f"[{final_v}]"]

    if audio_path:
        # bg audio only OR mix
        if keep_video_audio:
            # concat clip audios (if any). If a clip has no audio, ffmpeg may fail;
            # your workflow usually has audio in clips when keep_video_audio=yes.
            # If needed, we can add an "anullsrc" fallback later.
            alabels = []
            for i in range(n):
                a = f"a{i}"
                alabels.append(a)
                fc.append(f"[{i}:a]asetpts=PTS-STARTPTS[{a}]")

            # concat audio from clips
            insa = "".join([f"[{a}]" for a in alabels])
            fc.append(f"{insa}concat=n={n}:v=0:a=1[acat]")

            # mix with bg audio
            fc.append(
                f"[acat]volume={video_volume}[va];"
                f"[{bg_audio_idx}:a]volume={bg_volume}[ba];"
                f"[va][ba]amix=inputs=2:normalize=1[aout]"
            )
            map_args += ["-map", "[aout]"]
        else:
            map_args += ["-map", f"{bg_audio_idx}:a:0"]

        out_args += ["-c:a", "aac", "-b:a", "192k", "-shortest"]

    else:
        if keep_video_audio:
            # best-effort concat original audio tracks
            alabels = []
            for i in range(n):
                a = f"a{i}"
                alabels.append(a)
                fc.append(f"[{i}:a]asetpts=PTS-STARTPTS[{a}]")
            insa = "".join([f"[{a}]" for a in alabels])
            fc.append(f"{insa}concat=n={n}:v=0:a=1[acat]")
            map_args += ["-map", "[acat]"]
            out_args += ["-c:a", "aac", "-b:a", "192k"]
        else:
            out_args += ["-an"]

    cmd += ["-filter_complex", ";".join(fc)]
    cmd += map_args
    cmd += ["-c:v", "libx264", "-crf", str(crf), "-preset", preset]
    cmd += out_args
    cmd += [output_path]

    print("▶ Titles/Transitions cmd:\n", " ".join(shlex.quote(x) for x in cmd))
    return cmd

# --------------------------
# FFmpeg / FFprobe utilities
# --------------------------
def _bin_exists(name: str) -> bool:
    try:
        subprocess.run([name, "-version"], capture_output=True, check=True)
        return True
    except Exception:
        return False

def _ffprobe_stream_info(path: str) -> Dict:
    """
    Return primary video stream info as a dict:
    {codec_name, width, height, avg_frame_rate (string), pix_fmt}
    """
    try:
        out = subprocess.check_output([
            "ffprobe", "-v", "error",
            "-select_streams", "v:0",
            "-show_entries", "stream=codec_name,width,height,avg_frame_rate,pix_fmt",
            "-of", "json",
            path
        ], universal_newlines=True)
        data = json.loads(out)
        streams = data.get("streams", [])
        if not streams:
            return {}
        s = streams[0]
        return {
            "codec_name": s.get("codec_name"),
            "width": s.get("width"),
            "height": s.get("height"),
            "avg_frame_rate": s.get("avg_frame_rate"),
            "pix_fmt": s.get("pix_fmt"),
        }
    except Exception:
        return {}

def split_video(input_path,  max_duration=180):
    print("✅ Received split_video Arguments:", locals())

    output_dir="edit_vid_output"
    os.makedirs(output_dir, exist_ok=True)
    output_pattern = os.path.join(output_dir, "part_%02d.mp4")

    cmd = [
        "ffmpeg", "-y", "-i", input_path,
        "-c", "copy", "-map", "0",
        "-f", "segment",
        "-segment_time", str(max_duration),
        output_pattern
    ]
    subprocess.run(cmd, check=True)
    print(f"✅ Video split into parts saved under: {output_dir}")


# assemble_from_videos.py
def convert_landscape_to_portrait(
    input_path: str,
    output_path: str,
    portrait_size: str = "1080x1920",
    focus: str = "center",            # "left" | "center" | "right"
    keep_audio: bool = True
):
    """
    Crop a landscape video to 9:16 portrait by trimming left/right edges,
    then scale to the requested resolution. Optionally keep audio.
    """
    # Parse WxH like "1080x1920"
    try:
        tw, th = portrait_size.lower().split("x")
        tw, th = int(tw), int(th)
    except Exception:
        tw, th = 1080, 1920

    # Center/left/right anchor for crop X position
    # Crop width is ih*9/16 (portrait AR) when source is landscape
    # - center: x = (iw - ih*9/16)/2
    # - left:   x = 0
    # - right:  x = (iw - ih*9/16)
    if focus == "left":
        x_expr = "0"
    elif focus == "right":
        x_expr = "iw - ih*9/16"
    else:
        x_expr = "(iw - ih*9/16)/2"

    # If source is taller-than-wide (unlikely for landscape), this still behaves,
    # but we primarily target landscape→portrait cropping.
    crop_filter = f"crop=ih*9/16:ih:{x_expr}:0"

    # Build ffmpeg cmd
    cmd = [
        "ffmpeg", "-y",
        "-i", input_path,
        "-vf", f"{crop_filter},scale={tw}:{th}",
    ]

    if keep_audio:
        cmd += ["-c:v", "libx264", "-crf", "18", "-preset", "veryfast", "-c:a", "aac", "-b:a", "192k"]
    else:
        cmd += ["-an", "-c:v", "libx264", "-crf", "18", "-preset", "veryfast"]

    cmd += [output_path]

    subprocess.run(cmd, check=True)
    print(f"✅ Portrait video created at: {output_path}")


def _ffprobe_duration(path: str) -> float:
    """Return duration in seconds using ffprobe (format duration)."""
    try:
        out = subprocess.check_output([
            "ffprobe", "-v", "error",
            "-show_entries", "format=duration",
            "-of", "default=noprint_wrappers=1:nokey=1",
            path
        ], universal_newlines=True).strip()
        return float(out)
    except Exception:
        # MoviePy fallback (slower but robust)
        try:
            return float(VideoFileClip(path).duration)
        except Exception:
            return 0.0

def _has_audio_stream(path: str) -> bool:
    try:
        out = subprocess.check_output([
            "ffprobe", "-v", "error",
            "-select_streams", "a:0",
            "-show_entries", "stream=codec_type",
            "-of", "json",
            path
        ], universal_newlines=True)
        data = json.loads(out)
        return bool(data.get("streams"))
    except Exception:
        return False
    
def _can_safe_concat(video_paths: List[str]) -> Tuple[bool, str]:
    """
    Check if all videos share the same codec/resolution/fps/pix_fmt,
    which is required for FFmpeg concat with -c:v copy.
    """
    if not _bin_exists("ffmpeg") or not _bin_exists("ffprobe"):
        return False, "FFmpeg/FFprobe not available"

    ref = None
    for p in video_paths:
        info = _ffprobe_stream_info(p)
        if not info or not all(info.get(k) for k in ["codec_name","width","height","avg_frame_rate","pix_fmt"]):
            return False, f"Missing stream info for: {os.path.basename(p)}"

        # Normalize avg_frame_rate textual forms (e.g., "30000/1001" vs "29.97")
        afr = info["avg_frame_rate"]
        if afr and "/" in afr:
            n, d = afr.split("/")
            try:
                afr_float = float(n) / float(d) if float(d) != 0 else 0.0
            except Exception:
                afr_float = 0.0
        else:
            try:
                afr_float = float(afr)
            except Exception:
                afr_float = 0.0
        info["_afr_float"] = afr_float

        if ref is None:
            ref = info
            continue

        same = (
            info["codec_name"] == ref["codec_name"] and
            info["width"] == ref["width"] and
            info["height"] == ref["height"] and
            info["pix_fmt"] == ref["pix_fmt"] and
            abs(info["_afr_float"] - ref["_afr_float"]) < 1e-3
        )
        if not same:
            reason = (
                f"Mismatch: {os.path.basename(p)} "
                f"(codec={info['codec_name']}, size={info['width']}x{info['height']}, "
                f"fps≈{info['_afr_float']:.3f}, pix_fmt={info['pix_fmt']}) "
                f"vs ref (codec={ref['codec_name']}, size={ref['width']}x{ref['height']}, "
                f"fps≈{ref['_afr_float']:.3f}, pix_fmt={ref['pix_fmt']})"
            )
            return False, reason

    return True, "All inputs match (codec/size/fps/pix_fmt)"

# --------------------------
# Discovery helpers
# --------------------------
def _find_audio(audio_folder: str) -> str:
    audio_exts = ("*.mp3","*.wav","*.m4a","*.aac","*.flac","*.ogg")
    files = []
    for e in audio_exts:
        files.extend(glob(os.path.join(audio_folder, e)))
    if not files:
        raise RuntimeError(f"No audio file found in {audio_folder}")
    return os.path.abspath(sorted(files)[0])

def _find_videos(video_folder: str) -> List[str]:
    video_exts = ("*.mp4","*.mov","*.mkv","*.webm")
    files = []
    for e in video_exts:
        files.extend(glob(os.path.join(video_folder, e)))
    if not files:
        raise RuntimeError(f"No video files found in {video_folder}")
    return [os.path.abspath(p) for p in sorted(files)]

# ---------- Single-video fast path ----------

def _assemble_single_video_fast(
    video_path: str,
    audio_path: str,
    output_path: str,
    keep_video_audio: bool,
    video_volume: float,
    bg_volume: float,
):
    """
    Best-case: 1 video + 1 bg audio.
    - Loops video with -stream_loop to cover bg audio duration.
    - If keep_video_audio=False → use bg audio only.
    - If keep_video_audio=True → mix original audio + bg audio with volumes.
    - Uses -c:v copy → max quality, very fast.
    """
    audio_duration = _ffprobe_duration(audio_path)
    video_duration = _ffprobe_duration(video_path)
    if audio_duration <= 0 or video_duration <= 0:
        raise RuntimeError("Invalid durations for fast path")

    loop_count = max(int(math.ceil(audio_duration / video_duration) - 1), 0)
    has_v_audio = _has_audio_stream(video_path)

    # Only background audio (default behavior)
    if not keep_video_audio or not has_v_audio:
        cmd = [
            "ffmpeg", "-y",
            "-stream_loop", str(loop_count), "-i", video_path,
            "-i", audio_path,
            "-map", "0:v:0", "-map", "1:a:0",
            "-c:v", "copy",
            "-c:a", "aac", "-b:a", "192k",
            "-shortest",
            output_path,
        ]
    else:
        # Mix original audio + bg audio
        fc = (
            f"[0:a]volume={video_volume}[v0];"
            f"[1:a]volume={bg_volume}[a1];"
            f"[v0][a1]amix=inputs=2:normalize=1[aout]"
        )
        cmd = [
            "ffmpeg", "-y",
            "-stream_loop", str(loop_count), "-i", video_path,
            "-i", audio_path,
            "-filter_complex", fc,
            "-map", "0:v:0", "-map", "[aout]",
            "-c:v", "copy",
            "-c:a", "aac", "-b:a", "192k",
            "-shortest",
            output_path,
        ]

    print("▶ Single-video fast path:", " ".join(cmd))
    subprocess.run(cmd, check=True)

def _try_find_audio(audio_folder: str) -> str | None:
    """Return first audio file path, or None if folder missing/empty."""
    if not audio_folder or not os.path.isdir(audio_folder):
        return None
    try:
        return _find_audio(audio_folder)
    except Exception:
        return None

def _infer_title_from_path(p: str) -> str:
    """Infer a human-friendly title from a file path."""
    base = os.path.basename(p)
    name, _ = os.path.splitext(base)
    # Replace separators with spaces, collapse whitespace
    name = re.sub(r"[_\-]+", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name or base



def _run_cmd(cmd):
    """Run ffmpeg/ffprobe command with logging."""
    print("▶ FFmpeg cmd:")
    print(" ".join(cmd))
    proc = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True
    )
    if proc.returncode != 0:
        raise RuntimeError(f"Command failed:\n{proc.stdout}")
    return proc.stdout

def _make_looped_audio(
    input_audio: str,
    out_audio: str,
    loops: int,
    target_sec: float,
    bg_volume: float = 0.5,
):
    """
    Create a looped background audio track long enough to cover target_sec,
    then trim it exactly to target_sec.
    """
    # concat the same audio N times
    inputs = []
    for _ in range(loops):
        inputs.extend(["-i", input_audio])

    filter_inputs = "".join(f"[{i}:a]" for i in range(loops))
    filter_complex = (
        f"{filter_inputs}concat=n={loops}:v=0:a=1,"
        f"volume={bg_volume},atrim=0:{target_sec}"
    )

    cmd = [
        "ffmpeg", "-y",
        *inputs,
        "-filter_complex", filter_complex,
        "-c:a", "aac",
        out_audio
    ]
    _run_cmd(cmd)


def _ffmpeg_concat_demuxer(
    video_paths,
    output_path,
    keep_video_audio=True,
    bg_audio_path=None,
    video_volume=1.0,
    bg_volume=0.5,
    fps=30,
):
    """
    Fast concat using FFmpeg concat demuxer.
    Requires all clips to be stream-compatible.
    """
    with tempfile.NamedTemporaryFile("w", delete=False, suffix=".txt") as f:
        list_file = f.name
        for p in video_paths:
            f.write(f"file '{os.path.abspath(p)}'\n")

    cmd = [
        "ffmpeg", "-y",
        "-f", "concat",
        "-safe", "0",
        "-i", list_file,
        "-c", "copy",
        output_path
    ]
    _run_cmd(cmd)

    try:
        os.remove(list_file)
    except:
        pass

def assemble_videos_new(
    video_folder: str,
    audio_folder: str,
    output_path: str,
    fps: int = 30,
    shuffle: bool = True,
    prefer_ffmpeg_concat: bool = True,
    keep_video_audio: bool = False,
    video_volume: float = 1.0,
    bg_volume: float = 0.5,

    # NEW: breathing mode
    breathing_mode: str = "none",     # "none" | "title_card"
    breathing_sec: float = 3.5,

    # Legacy (kept for compatibility; NOT used when breathing_mode == "title_card")
    add_titles: bool = False,
    title_sec: float = 5.0,
    add_transitions: bool = False,
    transition_sec: float = 0.5,

    # Title card styling options (optional)
    title_prefix: str = "Next:",
    title_bg: str = "#101010",
    title_font_size: int = 64,
    title_box_alpha: float = 0.35,
    title_box_border: int = 40,
    title_fade_sec: float = 0.25,
    title_fontfile: str = None,       # Windows: "C:/Windows/Fonts/arial.ttf"
    title_text_color: str = "white",
):
    """
    Assemble multiple short clips into a single long video.

    NEW BEHAVIOR:
      If breathing_mode == "title_card":
        - Generates a short silent title-card MP4 between clips (3–5s recommended)
        - Concatenates using FFmpeg filter concat (re-encode) for reliability
        - Avoids audio delays / transitions that often cause A/V drift or captions desync

    NOTE:
      - If keep_video_audio=True, every input clip MUST have an audio stream (or you should set keep_video_audio=False).
      - Background audio (if present) is mixed after concat, keeping the timeline stable.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    if add_titles or add_transitions:
        add_titles = False
        add_transitions = False
        breathing_mode="title_card"

    # --- Collect video files ---
    exts = (".mp4", ".mov", ".m4v", ".mkv", ".webm", ".avi")
    video_paths = [
        os.path.join(video_folder, f)
        for f in os.listdir(video_folder)
        if f.lower().endswith(exts)
    ]
    video_paths.sort()
    if shuffle:
        random.shuffle(video_paths)

    if not video_paths:
        raise RuntimeError(f"No video files found in '{video_folder}'.")

    # --- Collect background audio files (optional) ---
    audio_exts = (".mp3", ".wav", ".m4a", ".aac", ".flac", ".ogg")
    bg_audio_paths = []
    if audio_folder and os.path.isdir(audio_folder):
        bg_audio_paths = [
            os.path.join(audio_folder, f)
            for f in os.listdir(audio_folder)
            if f.lower().endswith(audio_exts)
        ]
        bg_audio_paths.sort()

    # Pick a background audio (first) if present; keep existing behavior
    bg_audio_path = bg_audio_paths[0] if bg_audio_paths else None
    if not bg_audio_path:
        print(f"[Info] No background audio found in '{audio_folder}'. Proceeding with video-only merge.")

    # -------------------------------------------------------------------------
    # NEW PATH: breathing_mode == "title_card"
    # -------------------------------------------------------------------------
    if breathing_mode == "title_card":
        # Probe first clip to set output canvas (w/h). We keep a stable canvas for all title cards.
        # v0 = _ffprobe_stream_info(video_paths[0], stream_type="v:0")
        v0 = _ffprobe_stream_info(video_paths[0])

        if not v0:
            raise RuntimeError(f"Could not ffprobe first video: {video_paths[0]}")
        out_w = int(v0.get("width") or 1920)
        out_h = int(v0.get("height") or 1080)

        # Optionally auto-pick a default font file on Windows if not supplied
        if title_fontfile is None:
            # Will work on Windows; on Linux/mac you should set this explicitly if drawtext fails.
            title_fontfile = "C:/Windows/Fonts/arial.ttf"

        # Build an expanded list of segments:
        # clip1, title(for clip2), clip2, title(for clip3), clip3, ...
        temp_dir = os.path.join(os.path.dirname(output_path) or ".", "__tmp_titlecards")
        os.makedirs(temp_dir, exist_ok=True)

        segments = [video_paths[0]]
        for i in range(1, len(video_paths)):
            next_clip = video_paths[i]
            next_title = _infer_title_from_path(next_clip)
            card_path = os.path.join(temp_dir, f"title_{i:03d}.mp4")

            _make_title_card_ffmpeg(
                out_path=card_path,
                title=f"{title_prefix} {next_title}".strip(),
                w=out_w,
                h=out_h,
                fps=fps,
                dur=breathing_sec,
                bg=title_bg,
                fontfile=title_fontfile,
                fontsize=title_font_size,
                fontcolor=title_text_color,
                box_alpha=title_box_alpha,
                box_border=title_box_border,
                fade_sec=title_fade_sec,
            )
            segments.append(card_path)
            segments.append(next_clip)

        # If we are keeping clip audio, ensure each segment has audio.
        # Title cards DO have silent audio, but real clips might not.
        # if keep_video_audio:
        #     for p in segments:
        #         ainfo = _ffprobe_stream_info(p, stream_type="a:0")
        #         if not ainfo:
        #             raise RuntimeError(
        #                 f"Input segment has no audio stream but keep_video_audio=True:\n  {p}\n"
        #                 "Fix: re-export that clip with audio, OR set keep_video_audio=False."
        #             )
        if keep_video_audio:
            for p in segments:
                if not _has_audio_stream(p):
                    raise RuntimeError(
                        f"Input segment has no audio stream but keep_video_audio=True:\n  {p}\n"
                        "Fix: re-export that clip with audio, OR set keep_video_audio=False."
            )

        # Concat (re-encode) to a temp output first
        tmp_concat = os.path.join(os.path.dirname(output_path) or ".", "__tmp_concat.mp4")
        plan = []
        for p in segments:
            d = _ffprobe_duration(p)
            if d <= 0.02:
                continue
            plan.append((p, d, d))

        _ffmpeg_concat_reencode(
            plan=plan,
            out_path=tmp_concat,
            fps=fps,
            w=out_w,
            h=out_h,
            keep_audio=keep_video_audio,
        )


        # If no BG audio, finalize by moving tmp_concat to output_path
        if not bg_audio_path:
            try:
                if os.path.abspath(tmp_concat) != os.path.abspath(output_path):
                    if os.path.exists(output_path):
                        os.remove(output_path)
                    os.replace(tmp_concat, output_path)
            finally:
                # Cleanup temp title cards
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                except:
                    pass
            print(f"[OK] Saved: {output_path}")
            return output_path

        # If BG audio exists, mix it on top (without changing video timing)
        # 1) compute durations (video and bg)
        vid_dur = _ffprobe_duration(tmp_concat) or 0.0
        bg_dur = _ffprobe_duration(bg_audio_path) or 0.0

        # 2) loop bg audio if it's shorter than video
        # We'll create a temp bg track that is >= vid_dur then trim to vid_dur
        tmp_bg = os.path.join(os.path.dirname(output_path) or ".", "__tmp_bg.wav")
        if bg_dur <= 0:
            # fallback: just output video as-is
            if os.path.exists(output_path):
                os.remove(output_path)
            os.replace(tmp_concat, output_path)
            print(f"[Warn] BG audio had zero duration. Saved video-only: {output_path}")
            return output_path

        loops = int(math.ceil(vid_dur / bg_dur)) if bg_dur > 0 else 1
        # Create looped bg (concat filter on audio)
        _make_looped_audio(
            input_audio=bg_audio_path,
            out_audio=tmp_bg,
            loops=max(1, loops),
            target_sec=vid_dur,
            bg_volume=bg_volume,
        )

        # 3) mix: clip audio (volume video_volume) + bg audio (already at bg_volume)
        cmd = [
            "ffmpeg", "-y",
            "-i", tmp_concat,
            "-i", tmp_bg,
            "-filter_complex",
            (
                f"[0:a]volume={video_volume}[va];"
                f"[1:a]volume=1.0[ba];"
                f"[va][ba]amix=inputs=2:duration=first:dropout_transition=0[aout]"
            ),
            "-map", "0:v:0",
            "-map", "[aout]",
            "-c:v", "copy",
            "-c:a", "aac",
            "-shortest",
            output_path
        ]
        _run_cmd(cmd)

        # Cleanup
        for p in [tmp_concat, tmp_bg]:
            try:
                if os.path.exists(p):
                    os.remove(p)
            except:
                pass
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except:
            pass

        print(f"[OK] Saved: {output_path}")
        return output_path

    # -------------------------------------------------------------------------
    # OLD / LEGACY PATH (kept as-is; you can keep using it if breathing_mode != title_card)
    # -------------------------------------------------------------------------


    # Otherwise, do the original plain concat logic
    # (Prefer concat demuxer if possible, else fallback to moviepy)
    if prefer_ffmpeg_concat:
        try:
            _ffmpeg_concat_demuxer(
                video_paths=video_paths,
                output_path=output_path,
                keep_video_audio=keep_video_audio,
                bg_audio_path=bg_audio_path,
                video_volume=video_volume,
                bg_volume=bg_volume,
                fps=fps,
            )
            return output_path
        except Exception as e:
            print(f"[Warn] FFmpeg concat demuxer failed, falling back to MoviePy. Reason: {e}")

    # MoviePy fallback
    from moviepy.editor import VideoFileClip, concatenate_videoclips, AudioFileClip, CompositeAudioClip

    clips = []
    for p in video_paths:
        c = VideoFileClip(p)
        # Normalize FPS if needed (not strictly required, but helps with export)
        if fps:
            c = c.set_fps(fps)
        # Apply video volume if clip has audio
        if keep_video_audio and c.audio is not None:
            c = c.volumex(video_volume)
        else:
            c = c.without_audio()
        clips.append(c)

    final_video = concatenate_videoclips(clips, method="compose")

    if bg_audio_path:
        bg = AudioFileClip(bg_audio_path).volumex(bg_volume)
        # loop bg to match video duration
        if bg.duration < final_video.duration:
            n = int(math.ceil(final_video.duration / bg.duration))
            bg = concatenate_videoclips([bg] * n)  # AudioFileClip doesn't support this directly in some versions
            # safer approach is to subclip after looping; but keep existing behavior where possible

        bg = bg.subclip(0, final_video.duration)

        if keep_video_audio and final_video.audio is not None:
            final_audio = CompositeAudioClip([final_video.audio, bg])
        else:
            final_audio = bg
        final_video = final_video.set_audio(final_audio)

    final_video.write_videofile(
        output_path,
        fps=fps,
        codec="libx264",
        audio_codec="aac",
        threads=4,
        preset="medium",
        bitrate=None,
        audio_bitrate="192k",
        temp_audiofile="__temp_audio.m4a",
        remove_temp=True
    )

    # Close clips
    for c in clips:
        try: c.close()
        except: pass
    try: final_video.close()
    except: pass
    try:
        if bg_audio_path:
            try: bg.close()
            except: pass
    except:
        pass

    print(f"[OK] Saved: {output_path}")
    return output_path

def _find_videos_with_manifest(video_folder: str) -> list[str]:
    """
    Manifest-driven video ordering.
    If order.txt or order.xlsx exists:
      - ONLY files listed in manifest are merged
      - Extra files in folder are ignored
      - Missing listed files cause a hard error
    Otherwise:
      - Fall back to sorted auto-discovery
    """
    import os
    from glob import glob

    txt_path = os.path.join(video_folder, "order.txt")
    xlsx_path = os.path.join(video_folder, "order.xlsx")

    def resolve(name: str) -> str:
        name = name.strip().strip('"').strip("'")
        if not name:
            return ""
        if name.startswith("#") or name.startswith("//"):
            return ""
        p = name if os.path.isabs(name) else os.path.join(video_folder, name)
        return os.path.abspath(p)

    # ------------------------
    # order.txt (highest priority)
    # ------------------------
    if os.path.isfile(txt_path):
        ordered = []
        with open(txt_path, "r", encoding="utf-8") as f:
            for line in f:
                p = resolve(line)
                if not p:
                    continue
                if not os.path.isfile(p):
                    raise RuntimeError(f"[ORDER ERROR] Missing file listed in order.txt: {p}")
                ordered.append(p)

        if not ordered:
            raise RuntimeError("order.txt exists but contains no valid video entries.")

        print(f"[INFO] Using order.txt with {len(ordered)} videos. Unlisted files will be skipped.")
        return ordered

    # ------------------------
    # order.xlsx
    # ------------------------
    if os.path.isfile(xlsx_path):
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        ordered = []
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            val = row[0]
            if val is None:
                continue
            p = resolve(str(val))
            if not p:
                continue
            if not os.path.isfile(p):
                raise RuntimeError(f"[ORDER ERROR] Missing file listed in order.xlsx: {p}")
            ordered.append(p)

        if not ordered:
            raise RuntimeError("order.xlsx exists but column A contains no valid filenames.")

        print(f"[INFO] Using order.xlsx with {len(ordered)} videos. Unlisted files will be skipped.")
        return ordered

    # ------------------------
    # Fallback: auto-discovery
    # ------------------------
    video_exts = ("*.mp4", "*.mov", "*.mkv", "*.webm")
    files = []
    for e in video_exts:
        files.extend(glob(os.path.join(video_folder, e)))

    if not files:
        raise RuntimeError(f"No video files found in {video_folder}")

    print("[INFO] No order file found. Using alphabetical order.")
    return [os.path.abspath(p) for p in sorted(files)]

# --------------------------
# Main assembly
# --------------------------
def assemble_videos(
    video_folder: str,
    audio_folder: str,
    output_path: str,
    fps: int = 30,
    shuffle: bool = True,
    prefer_ffmpeg_concat: bool = True,
    keep_video_audio: bool = False,
    video_volume: float = 0.4,
    bg_volume: float = 1.0,
    add_titles: bool = False,
    title_sec: float = 2.0,
    add_transitions: bool = False,
    transition_sec: float = 0.5,
):
    """
    If bg audio exists -> match bg-audio duration (existing behavior).
    If bg audio is missing/empty -> just merge the clips (still supports ffmpeg concat fallback).
    """

    print("Received assemble_videos Arguments:", locals())
    clear_folder("edit_vid_output")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if os.path.isfile(os.path.join(video_folder, "order.txt")) or \
    os.path.isfile(os.path.join(video_folder, "order.xlsx")):
        shuffle = False

    breathing_mode = "none"

    if add_titles or add_transitions:
        add_titles = False
        add_transitions = False
        breathing_mode="title_card"

    # Collect videos
    # video_paths = _find_videos(video_folder)
    video_paths = _find_videos_with_manifest(video_folder)

    if shuffle and len(video_paths) > 1:
        random.shuffle(video_paths)

    # Try to load bg audio (optional now)
    audio_path = _try_find_audio(audio_folder)
    audio = None
    audio_duration = None

    if audio_path:
        audio = AudioFileClip(audio_path)
        audio_duration = float(audio.duration)
    else:
        print(f"[Info] No background audio found in '{audio_folder}'. Proceeding with video-only merge.")



    # -------------------------------------------------------------------------
    # NEW PATH: breathing_mode == "title_card"
    # -------------------------------------------------------------------------
    if breathing_mode == "title_card":
        # Probe first clip to set output canvas (w/h). We keep a stable canvas for all title cards.
        # v0 = _ffprobe_stream_info(video_paths[0], stream_type="v:0")
        v0 = _ffprobe_stream_info(video_paths[0])

        if not v0:
            raise RuntimeError(f"Could not ffprobe first video: {video_paths[0]}")
        out_w = int(v0.get("width") or 1920)
        out_h = int(v0.get("height") or 1080)

        # Optionally auto-pick a default font file on Windows if not supplied
        title_prefix = "Next:"
        title_bg = "#101010"
        title_font_size = 64
        title_box_alpha = 0.35
        title_box_border = 40
        title_fade_sec = 0.25
        # title_fontfile = "C:/Windows/Fonts/arial.ttf"    
        title_fontfile = "C:/Windows/Fonts/mangal.ttf"   
        title_text_color = "white"
        breathing_sec = 5

        # Build an expanded list of segments:
        # clip1, title(for clip2), clip2, title(for clip3), clip3, ...
        temp_dir = os.path.join(os.path.dirname(output_path) or ".", "__tmp_titlecards")
        os.makedirs(temp_dir, exist_ok=True)

        segments = [video_paths[0]]
        for i in range(1, len(video_paths)):
            next_clip = video_paths[i]
            next_title = _infer_title_from_path(next_clip)
            card_path = os.path.join(temp_dir, f"title_{i:03d}.mp4")          


            _make_title_card_ffmpeg(
                out_path=card_path,
                title=f"{title_prefix} {next_title}".strip(),
                w=out_w,
                h=out_h,
                fps=fps,
                dur=breathing_sec,
                bg=title_bg,
                fontfile=title_fontfile,
                fontsize=title_font_size,
                fontcolor=title_text_color,
                box_alpha=title_box_alpha,
                box_border=title_box_border,
                fade_sec=title_fade_sec,
            )
            segments.append(card_path)
            segments.append(next_clip)

        # If we are keeping clip audio, ensure each segment has audio.
        # Title cards DO have silent audio, but real clips might not.
        # if keep_video_audio:
        #     for p in segments:
        #         ainfo = _ffprobe_stream_info(p, stream_type="a:0")
        #         if not ainfo:
        #             raise RuntimeError(
        #                 f"Input segment has no audio stream but keep_video_audio=True:\n  {p}\n"
        #                 "Fix: re-export that clip with audio, OR set keep_video_audio=False."
        #             )
        if keep_video_audio:
            for p in segments:
                if not _has_audio_stream(p):
                    raise RuntimeError(
                        f"Input segment has no audio stream but keep_video_audio=True:\n  {p}\n"
                        "Fix: re-export that clip with audio, OR set keep_video_audio=False."
            )

        # Concat (re-encode) to a temp output first
        tmp_concat = os.path.join(os.path.dirname(output_path) or ".", "__tmp_concat.mp4")
        plan = []
        for p in segments:
            d = _ffprobe_duration(p)
            if d <= 0.02:
                continue
            plan.append((p, d, d))

        _ffmpeg_concat_reencode(
            plan=plan,
            out_path=tmp_concat,
            fps=fps,
            w=out_w,
            h=out_h,
            keep_audio=keep_video_audio,
        )


        # If no BG audio, finalize by moving tmp_concat to output_path
        if not audio_path:
            try:
                if os.path.abspath(tmp_concat) != os.path.abspath(output_path):
                    if os.path.exists(output_path):
                        os.remove(output_path)
                    os.replace(tmp_concat, output_path)
            finally:
                # Cleanup temp title cards
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                except:
                    pass
            print(f"[OK] Saved: {output_path}")
            return output_path

        # If BG audio exists, mix it on top (without changing video timing)
        # 1) compute durations (video and bg)
        vid_dur = _ffprobe_duration(tmp_concat) or 0.0
        bg_dur = _ffprobe_duration(audio_path) or 0.0

        # 2) loop bg audio if it's shorter than video
        # We'll create a temp bg track that is >= vid_dur then trim to vid_dur
        tmp_bg = os.path.join(os.path.dirname(output_path) or ".", "__tmp_bg.wav")
        if bg_dur <= 0:
            # fallback: just output video as-is
            if os.path.exists(output_path):
                os.remove(output_path)
            os.replace(tmp_concat, output_path)
            print(f"[Warn] BG audio had zero duration. Saved video-only: {output_path}")
            return output_path

        loops = int(math.ceil(vid_dur / bg_dur)) if bg_dur > 0 else 1
        # Create looped bg (concat filter on audio)
        _make_looped_audio(
            input_audio=audio_path,
            out_audio=tmp_bg,
            loops=max(1, loops),
            target_sec=vid_dur,
            bg_volume=bg_volume,
        )

        # 3) mix: clip audio (volume video_volume) + bg audio (already at bg_volume)
        cmd = [
            "ffmpeg", "-y",
            "-i", tmp_concat,
            "-i", tmp_bg,
            "-filter_complex",
            (
                f"[0:a]volume={video_volume}[va];"
                f"[1:a]volume=1.0[ba];"
                f"[va][ba]amix=inputs=2:duration=first:dropout_transition=0[aout]"
            ),
            "-map", "0:v:0",
            "-map", "[aout]",
            "-c:v", "copy",
            "-c:a", "aac",
            "-shortest",
            output_path
        ]
        _run_cmd(cmd)

        # Cleanup
        for p in [tmp_concat, tmp_bg]:
            try:
                if os.path.exists(p):
                    os.remove(p)
            except:
                pass
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except:
            pass

        print(f"[OK] Saved: {output_path}")
        return output_path
    # -------------------------
    # VIDEO-ONLY path (no bg audio)
    # -------------------------
    if not audio_path:
        # Single video fast path (copy, optionally strip audio)
        if prefer_ffmpeg_concat and len(video_paths) == 1 and _bin_exists("ffmpeg"):
            cmd = ["ffmpeg", "-y", "-i", video_paths[0], "-c", "copy"]
            if not keep_video_audio:
                cmd += ["-an"]
            cmd += [output_path]
            print("▶ Single-video (no bg audio):", " ".join(cmd))
            subprocess.run(cmd, check=True)
            return

        # ---- NEW: titles/transitions even when bg audio is missing ----
        if (add_titles or add_transitions) and _bin_exists("ffmpeg"):
            # Build a plan that uses full clip durations (no trimming needed)
            tiny = 0.02
            plan = []
            for p in video_paths:
                d = _ffprobe_duration(p)
                if d > tiny:
                    plan.append((p, d, d))

            if not plan:
                raise RuntimeError("All candidate videos are zero-length or unreadable.")

            # Locate this section in assemble_videos (around line 535)
            raw_font = r"C:\Windows\Fonts\arial.ttf"
            # We escape it specifically for the drawtext filter's requirements
            escaped_font = raw_font.replace("\\", "/").replace(":", "\\:")

            cmd = _build_titles_transitions_cmd(
                plan=plan,
                output_path=output_path,
                audio_path=None,                 # no bg audio
                fps=fps,
                add_titles=add_titles,
                title_sec=title_sec,
                use_transitions=add_transitions,
                trans_sec=transition_sec,
                keep_video_audio=keep_video_audio,
                video_volume=video_volume,       # used only if mixing (bg audio present); harmless here
                bg_volume=bg_volume,
                fontfile=escaped_font ,  # Pass the escaped version
            )
            subprocess.run(cmd, check=True)
            return

        # Multi-clip: attempt ffmpeg concat if safe (stream copy)
        if prefer_ffmpeg_concat and _bin_exists("ffmpeg"):
            can_concat, reason = _can_safe_concat(video_paths)
            if can_concat:
                with tempfile.TemporaryDirectory() as td:
                    list_txt = os.path.join(td, "list.txt")
                    with open(list_txt, "w", encoding="utf-8") as f:
                        for p in video_paths:
                            safe_p = p.replace("'", r"'\''")
                            f.write(f"file '{safe_p}'\n")

                    cmd = [
                        "ffmpeg", "-y",
                        "-f", "concat", "-safe", "0",
                        "-i", list_txt,
                        "-c", "copy",
                    ]
                    if not keep_video_audio:
                        cmd += ["-an"]
                    cmd += [output_path]

                    print("▶ Concat (no bg audio):", " ".join(cmd))
                    subprocess.run(cmd, check=True)
                return
            else:
                print(f"[Info] Falling back to MoviePy (concat not safe): {reason}")

        # MoviePy fallback (robust)
        clips = []
        try:
            tiny = 0.02
            for p in video_paths:
                c = VideoFileClip(p)
                if c.duration <= tiny:
                    c.close()
                    continue
                if not keep_video_audio:
                    c = c.without_audio()
                clips.append(c)

            if not clips:
                raise RuntimeError("All candidate videos are zero-length or unreadable.")

            video = concatenate_videoclips(clips, method="compose")
            video.write_videofile(
                output_path,
                fps=fps,
                codec="libx264",
                audio_codec="aac",
                threads=4,
                temp_audiofile="__temp_audio.m4a",
                remove_temp=True
            )
        finally:
            for c in clips:
                try: c.close()
                except: pass
        return

    # -------------------------
    # EXISTING behavior (bg audio present)
    # -------------------------
    # Fast path: exactly one video file
    if prefer_ffmpeg_concat and len(video_paths) == 1 and _bin_exists("ffmpeg"):
        audio.close()
        _assemble_single_video_fast(
            video_paths[0],
            audio_path,
            output_path,
            keep_video_audio=keep_video_audio,
            video_volume=video_volume,
            bg_volume=bg_volume,
        )
        return

    # Multi-clip plan (existing logic)
    tiny = 0.02
    durations, valid_paths = [], []
    for p in video_paths:
        d = _ffprobe_duration(p)
        if d > tiny:
            valid_paths.append(p)
            durations.append(d)

    if not valid_paths:
        audio.close()
        raise RuntimeError("All candidate videos are zero-length or unreadable.")

    remaining = audio_duration
    plan = []
    idx = 0
    while remaining > tiny:
        p = valid_paths[idx % len(valid_paths)]
        d = durations[idx % len(durations)]
        use_d = min(d, remaining)
        plan.append((p, d, use_d))
        remaining -= use_d
        idx += 1

    # ---- NEW: titles/transitions path (forces re-encode) ----
    if (add_titles or add_transitions) and _bin_exists("ffmpeg"):
        try:
            audio.close()
        except: 
            pass

        cmd = _build_titles_transitions_cmd(
            plan=plan,
            output_path=output_path,
            audio_path=audio_path,              # may be None, but here it's present
            fps=fps,
            add_titles=add_titles,
            title_sec=title_sec,
            use_transitions=add_transitions,
            trans_sec=transition_sec,
            keep_video_audio=keep_video_audio,
            video_volume=video_volume,
            bg_volume=bg_volume,
            # fontfile=r"C:\Windows\Fonts\arial.ttf",  # optional if drawtext font issues
        )
        subprocess.run(cmd, check=True)
        return

    # Try high-performance ffmpeg concat path
    if prefer_ffmpeg_concat:
        can_concat, reason = _can_safe_concat(valid_paths)
        if can_concat:
            audio.close()
            with tempfile.TemporaryDirectory() as td:
                list_txt = os.path.join(td, "list.txt")
                with open(list_txt, "w", encoding="utf-8") as f:
                    for (p, _, _) in plan:
                        safe_p = p.replace("'", r"'\''")
                        f.write(f"file '{safe_p}'\n")

                temp_concat = os.path.join(td, "concat.mp4")

                # Concat with both video+audio (if present), stream copy
                cmd_concat = [
                    "ffmpeg", "-y",
                    "-f", "concat", "-safe", "0",
                    "-i", list_txt,
                    "-c", "copy",
                    temp_concat,
                ]
                subprocess.run(cmd_concat, check=True)

                has_v_audio = _has_audio_stream(temp_concat)

                if not keep_video_audio or not has_v_audio:
                    # Only bg audio
                    cmd_mux = [
                        "ffmpeg", "-y",
                        "-i", temp_concat,
                        "-i", audio_path,
                        "-map", "0:v:0", "-map", "1:a:0",
                        "-c:v", "copy",
                        "-c:a", "aac", "-b:a", "192k",
                        "-shortest",
                        output_path,
                    ]
                else:
                    # Mix concat audio + bg
                    fc = (
                        f"[0:a]volume={video_volume}[v0];"
                        f"[1:a]volume={bg_volume}[a1];"
                        f"[v0][a1]amix=inputs=2:normalize=1[aout]"
                    )
                    cmd_mux = [
                        "ffmpeg", "-y",
                        "-i", temp_concat,
                        "-i", audio_path,
                        "-filter_complex", fc,
                        "-map", "0:v:0", "-map", "[aout]",
                        "-c:v", "copy",
                        "-c:a", "aac", "-b:a", "192k",
                        "-shortest",
                        output_path,
                    ]

                print("▶ Concat+Mux:", " ".join(cmd_mux))
                subprocess.run(cmd_mux, check=True)
            return
        else:
            print(f"[Info] Falling back to MoviePy (concat not safe): {reason}")

    # MoviePy fallback (non-uniform inputs)
    clips = []
    try:
        for (p, full_d, use_d) in plan:
            c = VideoFileClip(p)
            if use_d < (c.duration - tiny):
                c = c.subclip(0, use_d)
            if not keep_video_audio:
                c = c.without_audio()
            clips.append(c)

        video = concatenate_videoclips(clips, method="compose")

        bg_clip = audio.volumex(bg_volume)
        if keep_video_audio and video.audio is not None:
            v_clip = video.audio.volumex(video_volume)
            final_audio = CompositeAudioClip([v_clip, bg_clip])
        else:
            final_audio = bg_clip

        video = video.set_audio(final_audio).set_duration(audio_duration)

        video.write_videofile(
            output_path,
            fps=fps,
            codec="libx264",
            audio_codec="aac",
            threads=4,
            temp_audiofile="__temp_audio.m4a",
            remove_temp=True
        )
    finally:
        for c in clips:
            try:
                c.close()
            except:
                pass
        try:
            audio.close()
        except:
            pass


def assemble_videos_old1(
    video_folder: str,
    audio_folder: str,
    output_path: str,
    fps: int = 30,
    shuffle: bool = True,
    prefer_ffmpeg_concat: bool = True,
    keep_video_audio: bool = False,
    video_volume: float = 0.4,
    bg_volume: float = 1.0,
):
    """
    Build a final video matching the audio duration.

    Optimizations:
    - If only 1 video -> fast ffmpeg path with -stream_loop & optional audio mix.
    - If multiple videos & uniform -> ffmpeg concat + optional audio mix, -c:v copy.
    - Else -> MoviePy fallback with proper trimming & optional audio mix.
    """

    print("Received assemble_videos Arguments:", locals())
    clear_folder("edit_vid_output")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Load bg audio (for duration reference)
    audio_path = _find_audio(audio_folder)
    audio = AudioFileClip(audio_path)
    audio_duration = float(audio.duration)

    # Collect videos
    video_paths = _find_videos(video_folder)
    if shuffle and len(video_paths) > 1:
        random.shuffle(video_paths)

    # Fast path: exactly one video file
    if prefer_ffmpeg_concat and len(video_paths) == 1 and _bin_exists("ffmpeg"):
        audio.close()
        _assemble_single_video_fast(
            video_paths[0],
            audio_path,
            output_path,
            keep_video_audio=keep_video_audio,
            video_volume=video_volume,
            bg_volume=bg_volume,
        )
        return

    # Multi-clip plan
    tiny = 0.02
    durations, valid_paths = [], []
    for p in video_paths:
        d = _ffprobe_duration(p)
        if d > tiny:
            valid_paths.append(p)
            durations.append(d)

    if not valid_paths:
        audio.close()
        raise RuntimeError("All candidate videos are zero-length or unreadable.")

    remaining = audio_duration
    plan = []
    idx = 0
    while remaining > tiny:
        p = valid_paths[idx % len(valid_paths)]
        d = durations[idx % len(durations)]
        use_d = min(d, remaining)
        plan.append((p, d, use_d))
        remaining -= use_d
        idx += 1

    # Try high-performance ffmpeg concat path
    if prefer_ffmpeg_concat:
        can_concat, reason = _can_safe_concat(valid_paths)
        if can_concat:
            audio.close()
            with tempfile.TemporaryDirectory() as td:
                list_txt = os.path.join(td, "list.txt")
                with open(list_txt, "w", encoding="utf-8") as f:
                    for (p, _, _) in plan:
                        safe_p = p.replace("'", r"'\''")
                        f.write(f"file '{safe_p}'\n")

                temp_concat = os.path.join(td, "concat.mp4")

                # Concat with both video+audio (if present), stream copy
                cmd_concat = [
                    "ffmpeg", "-y",
                    "-f", "concat", "-safe", "0",
                    "-i", list_txt,
                    "-c", "copy",
                    temp_concat,
                ]
                subprocess.run(cmd_concat, check=True)

                has_v_audio = _has_audio_stream(temp_concat)

                if not keep_video_audio or not has_v_audio:
                    # Only bg audio
                    cmd_mux = [
                        "ffmpeg", "-y",
                        "-i", temp_concat,
                        "-i", audio_path,
                        "-map", "0:v:0", "-map", "1:a:0",
                        "-c:v", "copy",
                        "-c:a", "aac", "-b:a", "192k",
                        "-shortest",
                        output_path,
                    ]
                else:
                    # Mix concat audio + bg
                    fc = (
                        f"[0:a]volume={video_volume}[v0];"
                        f"[1:a]volume={bg_volume}[a1];"
                        f"[v0][a1]amix=inputs=2:normalize=1[aout]"
                    )
                    cmd_mux = [
                        "ffmpeg", "-y",
                        "-i", temp_concat,
                        "-i", audio_path,
                        "-filter_complex", fc,
                        "-map", "0:v:0", "-map", "[aout]",
                        "-c:v", "copy",
                        "-c:a", "aac", "-b:a", "192k",
                        "-shortest",
                        output_path,
                    ]

                print("▶ Concat+Mux:", " ".join(cmd_mux))
                subprocess.run(cmd_mux, check=True)
            return
        else:
            print(f"[Info] Falling back to MoviePy (concat not safe): {reason}")

    # MoviePy fallback (non-uniform inputs)
    clips = []
    try:
        for (p, full_d, use_d) in plan:
            c = VideoFileClip(p)
            if use_d < (c.duration - tiny):
                c = c.subclip(0, use_d)
            if not keep_video_audio:
                c = c.without_audio()
            clips.append(c)

        video = concatenate_videoclips(clips, method="compose")

        bg_clip = audio.volumex(bg_volume)
        if keep_video_audio and video.audio is not None:
            v_clip = video.audio.volumex(video_volume)
            final_audio = CompositeAudioClip([v_clip, bg_clip])
        else:
            final_audio = bg_clip

        video = video.set_audio(final_audio).set_duration(audio_duration)

        video.write_videofile(
            output_path,
            fps=fps,
            codec="libx264",
            audio_codec="aac",
            threads=4,
            temp_audiofile="__temp_audio.m4a",
            remove_temp=True
        )
    finally:
        for c in clips:
            try:
                c.close()
            except:
                pass
        try:
            audio.close()
        except:
            pass

# DND - Working
def assemble_videos_OLD(
    video_folder: str,
    audio_folder: str,
    output_path: str,
    fps: int = 30,
    shuffle: bool = True,
    prefer_ffmpeg_concat: bool = True,  # will auto-fallback if not safe
):
    """
    Auto-selects FFmpeg concat (stream-copy) if safe; otherwise falls back to MoviePy.

    - Reads the *real* duration of each clip.
    - Repeats clips (loop through list) until sum >= audio duration.
    - If using MoviePy: trims the last clip exactly to fit.
    - If using FFmpeg concat: concatenates whole clips, then muxes audio with -shortest.
    """
    
    print("Received assemble_videos Arguments:", locals())
    clear_folder("edit_vid_output")
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 1) Load audio + duration
    audio_path = _find_audio(audio_folder)
    audio = AudioFileClip(audio_path)
    audio_duration = float(audio.duration)

    # 2) Collect videos
    video_paths = _find_videos(video_folder)
    if shuffle:
        random.shuffle(video_paths)

    # 3) Precompute durations (skip empties)
    durations = []
    valid_paths = []
    tiny = 0.02
    for p in video_paths:
        d = _ffprobe_duration(p)
        if d > tiny:
            valid_paths.append(p)
            durations.append(d)
    if not valid_paths:
        audio.close()
        raise RuntimeError("All candidate videos are zero-length or unreadable.")

    # 4) Build a plan (path, full_duration, use_duration) to cover >= audio length
    remaining = audio_duration
    plan: List[Tuple[str, float, float]] = []
    idx = 0
    while remaining > tiny:
        p = valid_paths[idx % len(valid_paths)]
        d = durations[idx % len(durations)]
        use_d = min(d, remaining)
        plan.append((p, d, use_d))
        remaining -= use_d
        idx += 1

    # 5) Decide path: FFmpeg concat if safe and preferred, else MoviePy
    if prefer_ffmpeg_concat:
        can_concat, reason = _can_safe_concat(valid_paths)
        if can_concat:
            # ---- FFmpeg concat (no re-encode) ----
            audio.close()  # we'll remux with ffmpeg
            with tempfile.TemporaryDirectory() as td:
                list_txt = os.path.join(td, "list.txt")
                # Write whole files (concat demuxer can't trim mid-file)
                with open(list_txt, "w", encoding="utf-8") as f:
                    for (p, full_d, use_d) in plan:
                        # Always list the *entire* file
                        safe_p = p.replace("'", r"'\''")
                        f.write(f"file '{safe_p}'\n")

                # 1) Concat (video only), stream copy
                temp_concat = os.path.join(td, "concat.mp4")
                cmd_concat = [
                    "ffmpeg", "-y",
                    "-f", "concat", "-safe", "0",
                    "-i", list_txt,
                    "-c:v", "copy",
                    "-an",
                    temp_concat
                ]
                subprocess.run(cmd_concat, check=True)

                # 2) Mux audio, end at audio length
                cmd_mux = [
                    "ffmpeg", "-y",
                    "-i", temp_concat,
                    "-i", audio_path,
                    "-map", "0:v:0", "-map", "1:a:0",
                    "-c:v", "copy",
                    "-c:a", "aac", "-b:a", "192k",
                    "-shortest",
                    output_path
                ]
                subprocess.run(cmd_mux, check=True)

            return
        else:
            print(f"[Info] Falling back to MoviePy (concat not safe): {reason}")

    # ---- MoviePy re-encode path (robust, trims last clip) ----
    clips = []
    try:
        for (p, full_d, use_d) in plan:
            c = VideoFileClip(p).without_audio()
            if use_d < (c.duration - tiny):
                c = c.subclip(0, use_d)
            clips.append(c)

        video = concatenate_videoclips(clips, method="compose")
        video = video.set_audio(audio).set_duration(audio_duration)

        video.write_videofile(
            output_path,
            fps=fps,
            codec="libx264",
            audio_codec="aac",
            threads=4,
            temp_audiofile="__temp_audio.m4a",
            remove_temp=True
        )
    finally:
        for c in clips:
            try: c.close()
            except: pass
        try: audio.close()
        except: pass



def clear_folder(folder_path, extensions=None):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    for file in os.listdir(folder_path):
        full_path = os.path.join(folder_path, file)
        if os.path.isfile(full_path):
            if not extensions or file.lower().endswith(extensions):
                os.remove(full_path)

# --------------------------
# Example usage (parameters)
# --------------------------
if __name__ == "__main__":
    assemble_videos(
        video_folder="edit_vid_input",                  # or "edit_vid_output" if you pre-made KB clips
        audio_folder="edit_vid_audio",
        output_path="edit_vid_output/final_video.mp4",
        fps=30,
        shuffle=True,                                   # different order each run
        prefer_ffmpeg_concat=True                       # auto-uses concat if safe; else MoviePy
    )



# clear_folder("edit_vid_output")

# def clear_folder(folder_path, extensions=None):
#     if not os.path.exists(folder_path):
#         os.makedirs(folder_path)
#     for file in os.listdir(folder_path):
#         full_path = os.path.join(folder_path, file)
#         if os.path.isfile(full_path):
#             if not extensions or file.lower().endswith(extensions):
#                 os.remove(full_path)
