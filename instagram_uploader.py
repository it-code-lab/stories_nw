import os
import time
from pathlib import Path
from datetime import datetime

from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

# ============== CONFIG ==============

INSTAGRAM_BASE_URL = "https://www.instagram.com/"

# IMPORTANT: set this to your public profile URL so we can read the latest post
INSTAGRAM_PROFILE_URL = "https://www.instagram.com/creativecubs173/"

PROFILE_DIR = r"C:\Users\mail2\AppData\Local\Google\Chrome\User Data\Profile 21"
CHROME_EXECUTABLE = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

EXCEL_FILE = "master_shorts_uploader_data.xlsx"

BASE_DIR = Path(__file__).resolve().parent
MEDIA_BASE = BASE_DIR / "pinterest_uploads"   # same folder you use for Pinterest pins

# ============== EXCEL HELPERS ==============


def ensure_instagram_columns(ws, headers):
    """
    Ensure these Instagram columns exist:
    - instagram_post_url
    - instagram_upload_status
    - instagram_uploaded_at

    Returns updated headers list + header_map.
    """
    extra_cols = [
        "instagram_post_url",
        "instagram_upload_status",
        "instagram_uploaded_at",
    ]
    changed = False

    for col in extra_cols:
        if col not in headers:
            headers.append(col)
            ws.cell(row=1, column=len(headers), value=col)
            changed = True

    header_map = {name: idx + 1 for idx, name in enumerate(headers)}
    return headers, header_map, changed


def load_rows_from_excel():
    """
    Load rows from EXCEL_FILE, ensuring IG status columns are present.

    Returns: wb, ws, rows, header_map
    """
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [cell.value or "" for cell in ws[1]]
    headers, header_map, changed = ensure_instagram_columns(ws, headers)

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rec = {h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)}
        rec["_row_idx"] = row_idx

        media_file = (rec.get("media_file") or "").strip()
        status_val = (rec.get("instagram_upload_status") or "").strip()
        media_type = (rec.get("media_type") or "").strip()

        if media_type and media_type != "video":
            continue

        if status_val == "success":
            continue

        title = (rec.get("pin_title") or "").strip()
        if not media_file or not title:
            continue

        rows.append(rec)

    return wb, ws, rows, header_map


def save_instagram_status(ws, header_map, row_idx, url, status):
    """
    Save Instagram status back into Excel.
    """
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    url_col = header_map.get("instagram_post_url")
    status_col = header_map.get("instagram_upload_status")
    ts_col = header_map.get("instagram_uploaded_at")

    if url_col:
        ws.cell(row=row_idx, column=url_col, value=url)
    if status_col:
        ws.cell(row=row_idx, column=status_col, value=status)
    if ts_col:
        ws.cell(row=row_idx, column=ts_col, value=now_str)


def is_excel_file_locked(file_path):
    try:
        with open(file_path, "a"):
            pass
        return False
    except PermissionError:
        return True


# ============== INSTAGRAM HELPERS ==============


def resolve_media_path(media_file: str) -> str:
    """
    Resolve media_file relative to MEDIA_BASE.
    """
    rel = media_file.replace("\\", "/")
    full_path = (MEDIA_BASE / rel).resolve()
    return str(full_path)


def open_instagram_home(page):
    # page.goto(INSTAGRAM_BASE_URL, wait_until="networkidle")
    page.goto(INSTAGRAM_BASE_URL)
    time.sleep(2)


def start_new_post(page, media_path: str):
    """
    Click the "Create" / "New post" button and upload the video.
    We assume you're already logged in in this Chrome profile.
    """
    print(f"  -> Starting IG post with: {media_path}")

    # Click the "Create" / "New post" button.
    # Different UIs use different labels; try a few.
    clicked = False
    # try:
    #     page.get_by_role("button", name="Create").click()
    #     print("  -> Clicked 'Create' button.")
    #     clicked = True
    # except Exception:
    #     print("  [INFO] 'Create' button not found, trying 'New post'...")
    #     pass

    # if not clicked:
    #     try:
    #         page.get_by_role("button", name="New post").click()
    #         print("  -> Clicked 'New post' button.")
    #         clicked = True
    #     except Exception:
    #         print("  [INFO] 'New post' button not found, trying SVG button...")
    #         pass

    if not clicked:
        try:
            page.locator('svg[aria-label="New post"]').locator("xpath=..").click()
            print("  -> Clicked SVG 'New post' button.")
            clicked = True
        except Exception as e:
            print(f"  [WARN] Could not click Create/New post button automatically: {e}")
            print("  Please adjust the selector in start_new_post().")
            raise
            
    page.locator('svg[aria-label="Post"]').locator("xpath=..").click()
    print("  -> Clicked SVG 'Post' button.")
        
    # Click the button with name "Select from computer"
    # try:
    #     page.get_by_role("button", name="Select from computer").click()
    #     print("  -> Clicked 'Select from computer' button.")
    # except Exception as e:
    #     print(f"  [WARN] Could not click 'Select from computer' button: {e}")    
    # click_next_buttons(page)

    # Wait for file input, then upload
    # file_input = page.locator('input[type="file"]')
    # file_input.wait_for(state="visible", timeout=15000)
    # file_input.set_input_files(media_path)

    # Setup the file chooser listener BEFORE clicking the upload button
    with page.expect_file_chooser() as fc_info:
        page.get_by_role("button", name="Select from computer").click()
    
    file_chooser = fc_info.value
    file_chooser.set_files(media_path)


    # Give IG time to process the video
    time.sleep(5)

    # # 1. Crop Screen -> Click "Next"
    # print("Step 1: Crop -> Next")
    # page.get_by_role("button", name="Next").first.click()

def click_next_buttons(page):
    """
    On desktop, IG usually shows:
      - first screen: crop/ratio -> 'Next'
      - second screen: filters -> 'Next'
      - third screen: caption, details -> 'Share'
    We click 'Next' up to 2 times if present.
    """
    # There might be one or two "Next" steps
    for i in range(2):
        try:
            btn = page.get_by_role("button", name="Next")
            if btn.is_enabled():
                btn.click()
                time.sleep(3)
            else:
                break
        except Exception:
            break


def fill_caption(page, caption: str):
    """
    Fill the caption textarea ("Write a caption...").
    """
    caption = (caption or "").strip()
    if not caption:
        return

    try:
        # Common pattern on web: textarea with aria-label "Write a caption..."
        textarea = page.locator('textarea[aria-label="Write a caption..."]')
        if textarea.count() == 0:
            # Fallback: any textarea visible
            textarea = page.locator("textarea").filter(has_text="").first

        textarea.wait_for(state="visible", timeout=15000)
        textarea.click()
        textarea.fill(caption)
        print("  -> Caption filled.")
    except Exception as e:
        print(f"  [WARN] Could not fill caption: {e}")


def click_share(page):
    """
    Click the 'Share' button to publish the post.
    """
    print("  -> Clicking Share...")
    try:
        share_btn = page.get_by_role("button", name="Share")
        share_btn.click()
    except Exception as e:
        print(f"  [ERROR] Could not click Share button: {e}")
        raise

    # Wait for upload to complete; this may need tuning
    time.sleep(15)


def extract_latest_post_url(page) -> str:
    """
    After sharing, go to your profile and read the first post URL
    (reel or photo). This assumes the newly created post is now first.
    """
    try:
        page.goto(INSTAGRAM_PROFILE_URL, wait_until="networkidle")
        time.sleep(3)

        # Prefer reels; fall back to /p/ posts
        thumb = page.locator('article a[href*="/reel/"]').first
        if thumb.count() == 0:
            thumb = page.locator('article a[href*="/p/"]').first

        if thumb.count() == 0:
            print("  [WARN] Could not find any post links on profile.")
            return "Unknown"

        href = thumb.get_attribute("href") or ""
        href = href.strip()
        if not href:
            return "Unknown"

        if href.startswith("http"):
            return href
        else:
            return INSTAGRAM_BASE_URL.rstrip("/") + href
    except Exception as e:
        print(f"  [WARN] Failed to extract latest post URL: {e}")
        return "Unknown"


def build_instagram_caption(row: dict) -> str:
    """
    Build caption for this row.

    Uses:
      - instagram_caption column if present and non-empty
      - else pin_description + URL
    """
    insta_cap = (row.get("instagram_caption") or "").strip()
    if insta_cap:
        return insta_cap

    desc = (row.get("pin_description") or "").strip()
    link = (
        (row.get("pin_url_to_link") or "") or
        (row.get("book_url") or "")
    ).strip()

    caption = desc
    if link:
        if caption:
            caption += "\n\n"
        caption += link

    return caption


def create_instagram_post(page, row: dict) -> str:
    """
    Full flow for a single row:
    - open IG home
    - start new post
    - upload video
    - click Next screens
    - fill caption
    - share
    - get URL of latest post
    """
    media_file = (row.get("media_file") or "").strip()
    media_path = resolve_media_path(media_file)

    print(f"\n=== Creating IG post for media: {media_file} ===")

    if not os.path.exists(media_path):
        raise FileNotFoundError(f"Media file not found: {media_path}")

    open_instagram_home(page)
    start_new_post(page, media_path)

    # Instagram sidebar "Create" button. Usually named "Create" or "New post"
    # We try "New post" first (common for desktop), then fallback to "Create"
    time.sleep(1)
    # 1. Crop Screen -> Click "Next"
    print("Step 1: Crop -> Next")
    page.get_by_role("button", name="Next").first.click()
    time.sleep(1)
    # 2. Edit/Filter Screen -> Click "Next"
    print("Step 2: Edit -> Next")
    time.sleep(2) # Wait briefly for video buffer/filter preview
    page.get_by_role("button", name="Next").first.click()
    time.sleep(1)
    caption = build_instagram_caption(row)
    # fill_caption(page, caption)
    # click_share(page)


    # 3. Caption Screen
    print("Step 3: Adding Caption...")
    # Locator for caption area
    caption_area = page.get_by_role("textbox", name="Write a caption...")
    if caption_area.is_visible():
        caption_area.fill(caption)
    time.sleep(1)
    # 4. Share
    print("Sharing post...")
    page.get_by_role("button", name="Share").first.click()
    # --- STEP 5: VERIFY UPLOAD ---
            
    print("Waiting for success message...")
    
    # We use a text locator because the class names (x1lliihq...) are dynamic 
    # and will break the script if Instagram updates them.
    
    # This looks for an <h3> specifically containing the text from your div
    success_locator = page.locator("h3", has_text="Your reel has been shared.")
    
    # Wait up to 60 seconds for the upload/processing to finish
    try:
        success_locator.wait_for(state="visible", timeout=60000)
        print("✅ Success: 'Your reel has been shared.' message detected.")
    except:
        # Fallback: Sometimes it says "Your post has been shared." instead
        print("Checking for alternative success message...")
        page.get_by_role("heading", name="Your post has been shared.").wait_for(timeout=5000)
        print("✅ Success: 'Your post has been shared.' message detected.")

    post_url = ""        
    # post_url = extract_latest_post_url(page)
    # print(f"  -> IG post URL: {post_url}")
    return post_url


# ============== MAIN ==============


def upload_instagram_posts():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file not found: {EXCEL_FILE}")
        return

    if is_excel_file_locked(EXCEL_FILE):
        print(f"Error: Please close '{EXCEL_FILE}' before running the uploader.")
        return

    wb, ws, rows, header_map = load_rows_from_excel()
    if not rows:
        print("No valid rows found in Excel.")
        return

    print(f"Loaded {len(rows)} rows from {EXCEL_FILE}")

    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            PROFILE_DIR,
            headless=False,
            executable_path=CHROME_EXECUTABLE,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--start-maximized",
            ],
        )
        page = browser.new_page()
        page.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )

        # total = len(rows)
        # Go bottom-to-top, like Pinterest uploader
        for row in rows: 
            # excel_row_idx = total - rev_idx + 2
            excel_row_idx = row["_row_idx"]

            status_val = str(row.get("instagram_upload_status") or "").strip().lower()
            if status_val == "success":
                print(f"Skipping already uploaded IG row {excel_row_idx}")
                continue

            try:
                post_url = create_instagram_post(page, row)
                save_instagram_status(ws, header_map, excel_row_idx, post_url, "Success")
                print(f"✅ IG post created & recorded for row {excel_row_idx}")
            except Exception as e:
                err_msg = str(e)[:500]
                print(f"❌ Error creating IG post for row {excel_row_idx}: {err_msg}")
                save_instagram_status(ws, header_map, excel_row_idx, "", err_msg)

            # Be gentle with IG
            time.sleep(5)

        wb.save(EXCEL_FILE)
        browser.close()

    print("All done with Instagram uploads.")


if __name__ == "__main__":
    upload_instagram_posts()
