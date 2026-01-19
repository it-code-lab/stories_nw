import time
from datetime import datetime
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError
import os

# ================== CONFIG ==================
EXCEL_FILE = "heygen_submit_videos.xlsx"
SHEET_NAME = "Sheet1"

STORAGE_STATE = "heygen_state.json"   # <-- from successful login
CHROME_EXECUTABLE = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

HEADLESS = False
NAV_TIMEOUT = 120_000
UI_TIMEOUT = 60_000

COL_URL = "HeyGen_Template_url"
COL_TEXT = "story_text"
COL_NAME = "video_name"
COL_STATUS = "status"
COL_MESSAGE = "message"
COL_SUBMITTED_AT = "submitted_at"


# ================== HELPERS ==================

def norm(v):
    return "" if v is None else str(v).strip()


def ensure_columns(ws, headers):
    existing = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    for h in headers:
        if h not in existing:
            ws.cell(1, len(existing) + 1, h)
            existing.append(h)
    return {h: i + 1 for i, h in enumerate(existing)}


def set_status(ws, colmap, row, status, msg=""):
    ws.cell(row, colmap[COL_STATUS]).value = status
    ws.cell(row, colmap[COL_MESSAGE]).value = msg
    ws.cell(row, colmap[COL_SUBMITTED_AT]).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ================== HEYGEN ACTIONS ==================

def click_story_editor(page):
    for locator in [
        lambda: page.get_by_text("text", exact=True),
        lambda: page.locator("[contenteditable='true']").first,
        lambda: page.locator("textarea").first,
    ]:
        try:
            el = locator()
            el.wait_for(state="visible", timeout=10_000)
            el.click()
            return
        except Exception:
            pass
    raise RuntimeError("Story editor not found")

def click_story_editor_n_enter_text(page, text):
    for locator in [
        lambda: page.locator("[contenteditable='true']").first,
        lambda: page.get_by_text("text", exact=True),        
        lambda: page.locator("textarea").first,
    ]:
        try:
            print("Trying locator...")
            el = locator()
            el.wait_for(state="visible", timeout=10_000)
            el.click()


            page.wait_for_timeout(1000)
            page.keyboard.press("ControlOrMeta+A")
            page.keyboard.press("Backspace")
            el.type(text, delay=20)
            page.keyboard.press("Tab") 
            page.wait_for_timeout(2000) 
            return
        except Exception:
            pass
    raise RuntimeError("Story editor not found")

def wait_for_script_processing(page):
    # This searches for the duration indicator (e.g., "00:15") 
    # and waits for it to not be "00:00"
    try:
        # Adjust the selector based on the actual duration element in the HeyGen UI
        duration_locator = page.locator("text=/\\d{2}:\\d{2}/").first
        duration_locator.wait_for(state="visible", timeout=10000)
        print("Script processed: Duration detected.")
    except:
        print("Warning: Could not confirm duration update.")

def fill_story_new(page, text):

    for locator in [
        lambda: page.get_by_text("text", exact=True),
        lambda: page.locator("[contenteditable='true']").first,
        lambda: page.locator("textarea").first,
    ]:
        try:
            el = locator()
            el.wait_for(state="visible", timeout=10_000)
            el.click()
            
            page.wait_for_timeout(1000)
            page.keyboard.press("ControlOrMeta+A")
            page.keyboard.press("Backspace")
            # page.keyboard.insert_text(text)
            # Type with a 20ms delay between keys
            el.type(text, delay=20)
            
            # NEW: Press Tab to "blur" the field and trigger an update event
            page.keyboard.press("Tab") 
            page.wait_for_timeout(2000) 
            
            # NEW: Click a neutral area of the UI to ensure focus is lost
            page.mouse.click(0, 0) 
            
            # Wait longer for the "Script processing" or "Preview" to update
            page.wait_for_timeout(8000)

            return
        except Exception:
            pass
    raise RuntimeError("Story editor not found")



def fill_story(page, text):
    click_story_editor_n_enter_text(page, text)

    # page.wait_for_timeout(10000)


def fill_story_Jan16_working(page, text):
    click_story_editor(page)
    page.wait_for_timeout(1000)
    page.keyboard.press("ControlOrMeta+A")
    page.keyboard.press("Backspace")
    page.keyboard.insert_text(text)
    page.wait_for_timeout(10000)


def click_generate(page):
    page.get_by_role("button", name="Generate").click(timeout=UI_TIMEOUT)
    page.wait_for_timeout(1000)


def rename_video(page, name):
    box = page.get_by_role("textbox", name="Untitled Video")
    page.wait_for_timeout(1000)
    box.click(timeout=10_000)
    box.press("ControlOrMeta+A")
    box.fill(name)
    page.wait_for_timeout(1000)


def click_submit(page):
    page.get_by_role("button", name="Submit").click(timeout=UI_TIMEOUT)
    page.wait_for_timeout(1000)


def ensure_logged_in(page, context, target_url: str, storage_state_path: str):
    """
    If redirected to login, pause for manual login, then re-save storage_state.
    """
    # If we are on login, or page contains obvious login cues
    if "login" in page.url.lower():
        print("\n⚠️ HeyGen login required (session expired).")
        print("👉 Please login in the opened browser window.")
        print("👉 After you reach the HeyGen home/dashboard, press ENTER here...")
        input()

        # Re-save updated session so next rows don't require login again
        context.storage_state(path=storage_state_path)
        print(f"✅ Updated storage_state saved to: {storage_state_path}")

        # Go back to the intended page
        page.goto(target_url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)


def submit_one_new(page, context, url, text, name):

    MAX_RETRIES = 3
    
    for attempt in range(1, MAX_RETRIES + 1):

        page.goto(url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)

        ensure_logged_in(page, context, url, STORAGE_STATE)

        duration_loc = page.locator("span.tw-text-textDisable").last
        old_duration = duration_loc.inner_text() if duration_loc.is_visible() else "00:00"

        fill_story(page, text)
        page.wait_for_timeout(1000)

        print("Waiting for HeyGen to process script...")
        try:
            page.wait_for_function(
                """(oldVal) => {
                    const spans = document.querySelectorAll('span.tw-text-textDisable');
                    if (spans.length < 2) return false;
                    const totalDuration = spans[1].innerText; // Grab the second span
                    return totalDuration !== oldVal && totalDuration !== '00:00';
                }""",
                old_duration,
                timeout=30000 # Give it up to 30s for long scripts
            )

            print(f"✅ Script processed! ")
            # If we reach here, validation passed. Proceed to finish.
            click_generate(page)
            page.wait_for_timeout(1000)  # wait a bit for UI to update
            rename_video(page, name)
            page.wait_for_timeout(1000)

            click_submit(page)
            page.wait_for_timeout(1000)
            return # Exit the function successfully

        except Exception:
            print(f"   ⚠️ Attempt {attempt} failed: Duration didn't update.")
            if attempt == MAX_RETRIES:
                raise RuntimeError("Failed to process story after maximum retries.")
            print("   Re-trying...")
            page.wait_for_timeout(2000)



def submit_one(page, context, url, text, name):
    page.goto(url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)

    ensure_logged_in(page, context, url, STORAGE_STATE)

    # duration_loc = page.locator("span.tw-text-textDisable").last
    # old_duration = duration_loc.inner_text() if duration_loc.is_visible() else "00:00"

    fill_story(page, text)
    page.wait_for_timeout(1000)

    # print("Waiting for HeyGen to process script...")
    # try:
    #     page.wait_for_function(
    #         """(oldVal) => {
    #             const spans = document.querySelectorAll('span.tw-text-textDisable');
    #             if (spans.length < 2) return false;
    #             const totalDuration = spans[1].innerText; // Grab the second span
    #             return totalDuration !== oldVal && totalDuration !== '00:00';
    #         }""",
    #         old_duration,
    #         timeout=30000 # Give it up to 30s for long scripts
    #     )

    #     print(f"✅ Script processed! ")
    # except Exception:
    #     print("⚠️ Timeout: Duration didn't update. HeyGen might be slow or script is identical.")

    click_generate(page)
    page.wait_for_timeout(1000)  # wait a bit for UI to update
    rename_video(page, name)
    page.wait_for_timeout(1000)

    click_submit(page)
    page.wait_for_timeout(1000)


def ensure_storage_state(p):
    """
    Ensures heygen_state.json exists.
    If missing, launches browser for manual login and saves it.
    """
    if os.path.exists(STORAGE_STATE):
        print(f"✅ Found storage state: {STORAGE_STATE}")
        return

    print("⚠️ heygen_state.json not found.")
    print("👉 Opening HeyGen for one-time manual login...")

    browser = p.chromium.launch(
        headless=False,
        executable_path=CHROME_EXECUTABLE,
        args=["--start-maximized"],
    )

    context = browser.new_context()
    page = context.new_page()

    page.goto("https://app.heygen.com/home", wait_until="domcontentloaded")

    print("\n=================================================")
    print("PLEASE LOGIN MANUALLY IN THE OPENED BROWSER")
    print("After you see the HeyGen dashboard, press ENTER here")
    print("=================================================\n")
    input()

    # Save login session
    context.storage_state(path=STORAGE_STATE)
    print(f"✅ Storage state saved to {STORAGE_STATE}")

    browser.close()


# ================== MAIN ==================

def main():
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    colmap = ensure_columns(
        ws,
        [COL_URL, COL_TEXT, COL_NAME, COL_STATUS, COL_MESSAGE, COL_SUBMITTED_AT],
    )
    wb.save(EXCEL_FILE)

    rows = []
    for r in range(2, ws.max_row + 1):
        url = norm(ws.cell(r, colmap[COL_URL]).value)
        text = ws.cell(r, colmap[COL_TEXT]).value
        name = norm(ws.cell(r, colmap[COL_NAME]).value)
        status = norm(ws.cell(r, colmap[COL_STATUS]).value).lower()

        if url and name and norm(text) and status not in ("success", "done", "future"):
            rows.append((r, url, str(text), name))

    if not rows:
        print("No pending rows.")
        return

    with sync_playwright() as p:

        # 🔐 Ensure login session exists (auto-login bootstrap)
        ensure_storage_state(p)    

        browser = p.chromium.launch(
            headless=HEADLESS,
            executable_path=CHROME_EXECUTABLE,
            args=["--start-maximized"],
        )

        context = browser.new_context(storage_state=STORAGE_STATE)
        page = context.new_page()
        page.set_default_timeout(UI_TIMEOUT)

        for r, url, text, name in rows:
            print(f"\nSubmitting row {r}: {name}")
            try:
                set_status(ws, colmap, r, "processing")
                wb.save(EXCEL_FILE)

                submit_one(page, context, url, text, name)

                set_status(ws, colmap, r, "Success", "Submitted")
                wb.save(EXCEL_FILE)
                print("✅ Success")

                time.sleep(2)

            except PWTimeoutError as e:
                set_status(ws, colmap, r, "Error", "Timeout")
                wb.save(EXCEL_FILE)
                print("❌ Timeout")

            except Exception as e:
                set_status(ws, colmap, r, "Error", str(e)[:300])
                wb.save(EXCEL_FILE)
                print("❌ Error:", e)

        browser.close()

    print("\nAll done.")


if __name__ == "__main__":
    main()
